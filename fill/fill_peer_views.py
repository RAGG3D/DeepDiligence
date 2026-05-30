#!/usr/bin/env python3
"""
fill_peer_views.py -- Parse Gemini research .md reports and fill Peer Views sheet.

Parses PEER_VIEW_START/END blocks from Gemini output, extracts drug readout data,
and writes to the Peer Views sheet via surgical zip patching (NEVER openpyxl .save()).

Usage:
    python fill_peer_views.py --ticker CMPX [--report-dir path] [--dry-run]
"""

import argparse
import logging
import re
import shutil
import zipfile
from dataclasses import dataclass, field, fields
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


_EMPTY_CALC_CHAIN = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
    '<calcChain xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"></calcChain>'
)
logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

# ══════════════════════════════════════════════════════════════════════════════
#  CONSTANTS
# ══════════════════════════════════════════════════════════════════════════════

_NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

SHEET_NAME = "Peer Views"

# Date epoch for Excel serial dates (1900-based)
_EXCEL_EPOCH = datetime(1899, 12, 30)


# ══════════════════════════════════════════════════════════════════════════════
#  DRUG READOUT DATACLASS — 39 fields
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class DrugReadout:
    """One clinical data readout for one drug in one indication."""
    drug_name: str = ""
    company: str = ""
    ticker: str = ""
    innovation: str = ""
    target: str = ""
    result: str = ""
    nct: str = ""
    treatment_line: str = ""
    phase: str = ""
    stage: str = ""
    data_date: str = ""
    conference: str = ""
    n_patients: str = ""
    orr: str = ""
    bicr_orr: str = ""
    cr: str = ""
    pr: str = ""
    dcr: str = ""
    median_pfs: str = ""
    median_rpfs: str = ""
    median_os: str = ""
    pfs_6mo: str = ""
    pfs_12mo: str = ""
    pfs_24mo: str = ""
    os_18mo: str = ""
    os_24mo: str = ""
    median_dfs: str = ""
    median_followup: str = ""
    geq_g3_sae_pct: str = ""
    geq_g3_clinical_ae: str = ""
    route: str = ""
    dosing_schedule: str = ""
    latest_annual_sale: str = ""
    first_yr_sale: str = ""
    stock_price_before: str = ""
    stock_price_after: str = ""
    stock_change_1d: str = ""
    stock_change_3d: str = ""
    source: str = ""

    indication: str = ""  # Set from the PEER_VIEW_START block

    def dedup_key(self) -> str:
        """Unique key for deduplication: drug+nct+date."""
        return f"{self.drug_name}|{self.nct}|{self.data_date}"


# Mapping from Gemini key-value labels → DrugReadout field names
_GEMINI_KEY_MAP: Dict[str, str] = {
    "drug name": "drug_name",
    "company": "company",
    "ticker": "ticker",
    "innovation": "innovation",
    "target": "target",
    "result": "result",
    "nct#": "nct",
    "nct": "nct",
    "treatment line": "treatment_line",
    "phase": "phase",
    "stage": "stage",
    "data date": "data_date",
    "conference": "conference",
    "n": "n_patients",
    "orr": "orr",
    "bicr orr": "bicr_orr",
    "cr": "cr",
    "pr": "pr",
    "dcr": "dcr",
    "median pfs": "median_pfs",
    "median rpfs": "median_rpfs",
    "median os": "median_os",
    "6 mo pfs rate": "pfs_6mo",
    "12 mo pfs rate": "pfs_12mo",
    "24 mo pfs rate": "pfs_24mo",
    "18 mo os rate": "os_18mo",
    "24 mo os rate": "os_24mo",
    "median dfs": "median_dfs",
    "median follow-up": "median_followup",
    "geq g3 sae pct": "geq_g3_sae_pct",
    "geq g3 clinical ae": "geq_g3_clinical_ae",
    "route": "route",
    "dosing schedule": "dosing_schedule",
    "latest annual sale": "latest_annual_sale",
    "1st yr sale": "first_yr_sale",
    "stock price day before": "stock_price_before",
    "stock price day after": "stock_price_after",
    "stock change 1d": "stock_change_1d",
    "stock change 3d": "stock_change_3d",
    "source": "source",
}

# Mapping from Excel D-column labels → DrugReadout field names
# Handles the variations across different sections
_EXCEL_LABEL_MAP: Dict[str, str] = {
    # Drug identity
    "innovation": "innovation",
    "target": "target",
    "result": "result",
    # Dates & context
    "date": "data_date",
    "readout phase": "phase",
    "stage": "stage",
    "treatment line": "treatment_line",
    "total treatment line": "treatment_line",
    "median treatment line": "treatment_line",
    "indication": "indication",
    "conference": "conference",
    # Safety
    "\u2265g3 sae/patients": "geq_g3_sae_pct",
    "≥g3 sae/patients": "geq_g3_sae_pct",
    "\u2265g3 clinical ae": "geq_g3_clinical_ae",
    "≥g3 clinical ae": "geq_g3_clinical_ae",
    # Efficacy
    "evaluable patients": "n_patients",
    "patient number": "n_patients",
    "n": "n_patients",
    "orr": "orr",
    "orr (confirmed)": "orr",
    "bicr orr": "bicr_orr",
    "cr": "cr",
    "cr (confirmed)": "cr",
    "pr": "pr",
    "dcr": "dcr",
    "median pfs": "median_pfs",
    "pfs >= 6 mo": "pfs_6mo",
    "pfs ≥ 6 mo": "pfs_6mo",
    "6 mo pfs": "pfs_6mo",
    "12 mo pfs": "pfs_12mo",
    "24 mo pfs": "pfs_24mo",
    "median rpfs": "median_rpfs",
    "median os": "median_os",
    "18 mo os": "os_18mo",
    "24 mo os": "os_24mo",
    "median dfs": "median_dfs",
    "median follow-up": "median_followup",
    # Sales
    "latest sale (mm usd)": "latest_annual_sale",
    "1st yr sale (mm usd)": "first_yr_sale",
    # Stock
    "market reaction (-1d)": "stock_price_before",
    "market reaction (+1d)": "stock_price_after",
    # mCRPC-specific
    "\u226550% psa decline rate": "orr",  # Map PSA decline to ORR field
    "≥50% psa decline rate": "orr",
    "\u226590% psa decline rate": "cr",  # Map to CR field
    "≥90% psa decline rate": "cr",
    "response evaluable patient": "n_patients",
    "stable disease rate": "dcr",
    # Route/dosing
    "route": "route",
    "dosing schedule": "dosing_schedule",
}

# Standard new section row template (field labels for D column)
# Used when creating a brand new section
_NEW_SECTION_FIELDS: List[Tuple[str, str]] = [
    # (d_label, readout_field)
    ("Innovation", "innovation"),
    ("Target", "target"),
    ("Result", "result"),
    ("Date", "data_date"),
    ("Readout Phase", "phase"),
    ("Treatment Line", "treatment_line"),
    ("Evaluable Patients", "n_patients"),
    ("ORR", "orr"),
    ("BICR ORR", "bicr_orr"),
    ("CR", "cr"),
    ("PR", "pr"),
    ("DCR", "dcr"),
    ("Median PFS", "median_pfs"),
    ("Median rPFS", "median_rpfs"),
    ("Median OS", "median_os"),
    ("6 Mo PFS", "pfs_6mo"),
    ("12 Mo PFS", "pfs_12mo"),
    ("24 Mo PFS", "pfs_24mo"),
    ("18 Mo OS", "os_18mo"),
    ("24 Mo OS", "os_24mo"),
    ("Median DFS", "median_dfs"),
    ("Median Follow-Up", "median_followup"),
    ("\u2265G3 SAE/Patients", "geq_g3_sae_pct"),
    ("\u2265G3 Clinical AE", "geq_g3_clinical_ae"),
    ("Route", "route"),
    ("Dosing Schedule", "dosing_schedule"),
    ("Latest Sale (MM USD)", "latest_annual_sale"),
    ("1st Yr Sale (MM USD)", "first_yr_sale"),
]


# ══════════════════════════════════════════════════════════════════════════════
#  MARKDOWN PARSING
# ══════════════════════════════════════════════════════════════════════════════

def parse_peer_view_blocks(text: str) -> Dict[str, List[DrugReadout]]:
    """Parse PEER_VIEW_START/END blocks from markdown text.

    Returns: {indication: [DrugReadout, ...]}
    """
    result: Dict[str, List[DrugReadout]] = {}

    # Find all PEER_VIEW_START/END blocks
    block_re = re.compile(
        r'####\s*PEER_VIEW_START:\s*(.+?)\s*\n(.*?)\n####\s*PEER_VIEW_END:\s*\1',
        re.DOTALL | re.IGNORECASE,
    )

    for block_m in block_re.finditer(text):
        indication = block_m.group(1).strip()
        block_body = block_m.group(2)

        readouts = _parse_readouts(block_body, indication)
        if readouts:
            if indication not in result:
                result[indication] = []
            result[indication].extend(readouts)

    return result


def _parse_readouts(block_body: str, indication: str) -> List[DrugReadout]:
    """Parse individual drug readouts from a PEER_VIEW block body."""
    readouts: List[DrugReadout] = []

    # Split by readout headers: "##### Drug: CTX-009 — Readout 1"
    readout_re = re.compile(
        r'#####\s+Drug:\s+.+?(?:—|--|-)\s*Readout\s+\d+',
        re.IGNORECASE,
    )
    parts = readout_re.split(block_body)
    # First part is before any readout header (skip)
    # Also get the headers for drug name extraction
    headers = readout_re.findall(block_body)

    for header, body in zip(headers, parts[1:]):
        readout = DrugReadout(indication=indication)

        # Parse key-value pairs: "- Key: Value"
        kv_re = re.compile(r'^-\s+(.+?):\s+(.+)$', re.MULTILINE)
        for kv_m in kv_re.finditer(body):
            key = kv_m.group(1).strip().lower()
            value = kv_m.group(2).strip()

            if value in ("/", "N/A", "n/a", "-", "—", ""):
                value = "/"

            field_name = _GEMINI_KEY_MAP.get(key)
            if field_name and hasattr(readout, field_name):
                setattr(readout, field_name, value)

        # Fall back: extract drug name from header if not in body
        if not readout.drug_name:
            hdr_m = re.search(r'Drug:\s+(.+?)(?:\s*(?:—|--|-)\s*Readout)', header)
            if hdr_m:
                readout.drug_name = hdr_m.group(1).strip()

        if readout.drug_name:
            readouts.append(readout)

    return readouts


def scan_report_files(report_dir: Path) -> str:
    """Concatenate all .md report files in a directory."""
    md_files = sorted(report_dir.glob("*.md"))
    if not md_files:
        logger.warning(f"No .md files found in {report_dir}")
        return ""

    logger.info(f"Found {len(md_files)} .md file(s) in {report_dir}")
    parts = []
    for f in md_files:
        logger.info(f"  Reading: {f.name}")
        parts.append(f.read_text(encoding="utf-8"))
    return "\n\n".join(parts)


# ══════════════════════════════════════════════════════════════════════════════
#  XML HELPERS (adapted from fill_tam.py)
# ══════════════════════════════════════════════════════════════════════════════

def _get_sheet_zip_paths(xlsx_path: Path) -> Dict[str, str]:
    """Return {sheet_name: zip_entry_path} for all worksheets."""
    import xml.etree.ElementTree as ET
    with zipfile.ZipFile(xlsx_path) as zf:
        wb_xml = ET.fromstring(zf.read("xl/workbook.xml"))
        rels_xml = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))

    rid_to_path: Dict[str, str] = {}
    for rel in rels_xml:
        if "worksheet" in rel.get("Type", ""):
            rid = rel.get("Id", "")
            target = rel.get("Target", "")
            rid_to_path[rid] = (
                f"xl/{target}" if not target.startswith("/") else target.lstrip("/")
            )

    sheet_map: Dict[str, str] = {}
    for sheet_elem in wb_xml.findall(f".//{{{_NS_MAIN}}}sheet"):
        name = sheet_elem.get("name", "")
        rid = sheet_elem.get(f"{{{_NS_R}}}id", "")
        if rid in rid_to_path:
            sheet_map[name] = rid_to_path[rid]
    return sheet_map


def _find_cell(xml: str, addr: str) -> Optional[Tuple[int, int, str, str]]:
    """Find a cell by address. Returns (lt, end_pos, open_tag, inner) or None."""
    search = f'r="{addr}"'
    start = 0
    while True:
        pos = xml.find(search, start)
        if pos == -1:
            return None
        lt = xml.rfind("<", 0, pos)
        if lt == -1 or xml[lt + 1] != "c" or xml[lt + 2] not in (" ", "\t", "\n", "/", ">"):
            start = pos + 1
            continue
        tag_end = xml.index(">", lt) + 1
        open_tag = xml[lt:tag_end]
        if xml[tag_end - 2: tag_end] == "/>":
            return (lt, tag_end, open_tag, "")
        c_end = xml.index("</c>", tag_end) + 4
        inner = xml[tag_end:c_end - 4]
        return (lt, c_end, open_tag, inner)


def _get_cell_style(xml: str, addr: str) -> Optional[str]:
    """Extract s="N" style attribute from a cell."""
    found = _find_cell(xml, addr)
    if not found:
        return None
    _, _, open_tag, _ = found
    m = re.search(r's="(\d+)"', open_tag)
    return m.group(1) if m else None


def _col_of(addr: str) -> int:
    col_str = "".join(c for c in addr if c.isalpha())
    return column_index_from_string(col_str)


def _xml_escape(text: str) -> str:
    """XML-escape text for inlineStr content."""
    text = text.replace("&", "&amp;")
    text = text.replace("<", "&lt;")
    text = text.replace(">", "&gt;")
    text = text.replace('"', "&quot;")
    return text


def _patch_text_cell(xml: str, addr: str, text: str,
                     style: Optional[str] = None) -> str:
    """Write text cell as inlineStr. Preserves xml[:lt] prefix."""
    escaped = _xml_escape(text)
    found = _find_cell(xml, addr)
    if found:
        lt, end_pos, open_tag, inner = found
        # Build new open tag with t="inlineStr"
        open_tag_clean = re.sub(r'\s+t="[^"]*"', '', open_tag)
        open_tag_clean = open_tag_clean.replace("/>", ">")
        if not open_tag_clean.endswith(">"):
            open_tag_clean += ">"
        # Insert t="inlineStr" after r="..."
        open_tag_clean = re.sub(
            r'(r="[^"]*")',
            r'\1 t="inlineStr"',
            open_tag_clean,
        )
        if style:
            if re.search(r's="[^"]*"', open_tag_clean):
                open_tag_clean = re.sub(r's="[^"]*"', f's="{style}"', open_tag_clean)
            else:
                open_tag_clean = re.sub(
                    r'(r="[^"]*")', rf'\1 s="{style}"', open_tag_clean
                )
        new_cell = f'{open_tag_clean}<is><t>{escaped}</t></is></c>'
        return xml[:lt] + new_cell + xml[end_pos:]

    # Cell not found: insert new
    row_num = int("".join(c for c in addr if c.isdigit()))
    s_attr = f' s="{style}"' if style else ""
    new_cell = f'<c r="{addr}"{s_attr} t="inlineStr"><is><t>{escaped}</t></is></c>'
    return _insert_cell_xml(xml, row_num, addr, new_cell)


def _patch_numeric_cell(xml: str, addr: str, val_str: str,
                        style: Optional[str] = None) -> str:
    """Find cell r="ADDR" and update its <v> content, or insert new cell."""
    found = _find_cell(xml, addr)
    if found:
        lt, end_pos, open_tag, inner = found
        open_tag_clean = re.sub(r'\s+t="[^"]*"', '', open_tag)
        if style:
            if re.search(r's="[^"]*"', open_tag_clean):
                open_tag_clean = re.sub(r's="[^"]*"', f's="{style}"', open_tag_clean)
            else:
                open_tag_clean = re.sub(
                    r'(r="[^"]*")', rf'\1 s="{style}"', open_tag_clean
                )
        if open_tag_clean.endswith("/>"):
            new_cell = open_tag_clean[:-2] + f"><v>{val_str}</v></c>"
            return xml[:lt] + new_cell + xml[end_pos:]
        return xml[:lt] + open_tag_clean + f"<v>{val_str}</v></c>" + xml[end_pos:]

    # Insert new cell
    row_num = int("".join(c for c in addr if c.isdigit()))
    s_attr = f' s="{style}"' if style else ""
    new_cell = f'<c r="{addr}"{s_attr}><v>{val_str}</v></c>'
    return _insert_cell_xml(xml, row_num, addr, new_cell)


def _patch_formula_cell(xml: str, addr: str, formula: str,
                        style: Optional[str] = None) -> str:
    """Write a formula cell."""
    found = _find_cell(xml, addr)
    if found:
        lt, end_pos, open_tag, inner = found
        open_tag_clean = re.sub(r'\s+t="[^"]*"', '', open_tag)
        if style:
            if re.search(r's="[^"]*"', open_tag_clean):
                open_tag_clean = re.sub(r's="[^"]*"', f's="{style}"', open_tag_clean)
            else:
                open_tag_clean = re.sub(
                    r'(r="[^"]*")', rf'\1 s="{style}"', open_tag_clean
                )
        if open_tag_clean.endswith("/>"):
            new_cell = open_tag_clean[:-2] + f"><f>{formula}</f></c>"
        else:
            new_cell = open_tag_clean + f"<f>{formula}</f></c>"
        return xml[:lt] + new_cell + xml[end_pos:]

    # Insert new formula cell
    row_num = int("".join(c for c in addr if c.isdigit()))
    s_attr = f' s="{style}"' if style else ""
    new_cell = f'<c r="{addr}"{s_attr}><f>{formula}</f></c>'
    return _insert_cell_xml(xml, row_num, addr, new_cell)


def _insert_cell_xml(xml: str, row_num: int, addr: str, cell_xml: str) -> str:
    """Insert a cell XML string into the correct position in a row."""
    row_search = f'r="{row_num}"'
    row_pos = 0
    while True:
        rp = xml.find(row_search, row_pos)
        if rp == -1:
            logger.warning(f"  Row {row_num} not found; cannot insert {addr}")
            return xml
        lt = xml.rfind("<", 0, rp)
        if lt != -1 and xml[lt + 1: lt + 4] == "row":
            break
        row_pos = rp + 1

    row_tag_end = xml.index(">", lt) + 1

    # Self-closing row?
    if xml[row_tag_end - 2: row_tag_end] == "/>":
        return (
            xml[:row_tag_end - 2]
            + f">{cell_xml}</row>"
            + xml[row_tag_end:]
        )

    row_end = xml.index("</row>", row_tag_end)
    row_body = xml[row_tag_end:row_end]
    col_idx = _col_of(addr)

    # Insert in column order
    cells = list(re.finditer(r'<c\b[^>]*\br="([A-Z]+\d+)"', row_body))
    insert_at = len(row_body)
    for m in cells:
        if _col_of(m.group(1)) > col_idx:
            insert_at = m.start()
            break

    new_body = row_body[:insert_at] + cell_xml + row_body[insert_at:]
    return xml[:row_tag_end] + new_body + xml[row_end:]


# ══════════════════════════════════════════════════════════════════════════════
#  PEER VIEWS SHEET DISCOVERY
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class SectionInfo:
    """Describes one indication section in the Peer Views sheet."""
    indication: str
    header_row: int      # Row with "X" in col A, section name in col D
    first_data_row: int  # First field row (e.g., Innovation)
    last_data_row: int   # Last field row (e.g., 1st Yr Sale)
    # Ticker row: typically header_row + 2 (for Peer View style) or +3
    ticker_row: int
    drug_name_row: int   # Row below ticker row
    # Map: row_num → DrugReadout field name
    field_rows: Dict[int, str]
    # Existing drug columns: set of col letters with data
    existing_cols: Set[str]
    # Bloomberg formula rows (Market Reaction)
    mr_minus1_row: Optional[int] = None
    mr_plus1_row: Optional[int] = None
    # Styles for data cells (row → style from reference column)
    ref_styles: Dict[int, str] = field(default_factory=dict)


def discover_sections(xml: str, ws) -> List[SectionInfo]:
    """Discover all indication sections in the Peer Views sheet."""
    sections: List[SectionInfo] = []
    max_row = ws.max_row or 200

    # Step 1: Find all section header rows (col A = "X")
    header_rows: List[Tuple[int, str]] = []
    for row in range(1, max_row + 1):
        a_val = ws.cell(row=row, column=1).value
        if a_val and str(a_val).strip().upper() == "X":
            d_val = ws.cell(row=row, column=4).value
            if d_val:
                header_rows.append((row, str(d_val).strip()))

    if not header_rows:
        logger.warning("No sections found in Peer Views sheet")
        return sections

    # Step 2: For each section, discover structure
    for idx, (hdr_row, indication) in enumerate(header_rows):
        # Section ends at next header or max_row
        next_hdr = header_rows[idx + 1][0] if idx + 1 < len(header_rows) else max_row + 1

        # Find field rows by scanning D column
        field_rows: Dict[int, str] = {}
        ticker_row = 0
        drug_name_row = 0
        mr_minus1 = None
        mr_plus1 = None
        first_field_row = 0
        last_field_row = 0

        for row in range(hdr_row + 1, min(next_hdr, hdr_row + 40)):
            d_val = ws.cell(row=row, column=4).value
            if d_val is None:
                continue
            d_str = str(d_val).strip()
            if not d_str:
                continue

            d_lower = d_str.lower()

            # Check for Market Reaction rows
            if "market reaction" in d_lower and "-1d" in d_lower:
                mr_minus1 = row
                continue
            if "market reaction" in d_lower and "+1d" in d_lower:
                mr_plus1 = row
                continue

            # Check for ticker/drug name header rows
            if d_lower in ("ticker name:", "ticker name"):
                continue

            # Map to readout field
            mapped = _EXCEL_LABEL_MAP.get(d_lower)
            if mapped:
                field_rows[row] = mapped
                if not first_field_row:
                    first_field_row = row
                last_field_row = row

        # Find ticker and drug name rows
        # Pattern: ticker row has Bloomberg tickers in cols E-K
        # Drug name row is typically the row right after ticker row
        for row in range(hdr_row + 1, min(next_hdr, hdr_row + 8)):
            for col_idx in range(5, 12):  # E through K
                val = ws.cell(row=row, column=col_idx).value
                if val and "Equity" in str(val):
                    ticker_row = row
                    break
            if ticker_row:
                break

        # Drug name row: row after ticker with non-field content in E+ cols
        if ticker_row:
            for row in range(ticker_row + 1, ticker_row + 3):
                for col_idx in range(5, 12):
                    val = ws.cell(row=row, column=col_idx).value
                    if val and row not in field_rows:
                        drug_name_row = row
                        break
                if drug_name_row:
                    break
            if not drug_name_row:
                drug_name_row = ticker_row + 1

        if not first_field_row:
            logger.warning(f"  Section '{indication}' R{hdr_row}: no field rows found")
            continue

        # Find existing drug columns (any col with data in the field rows)
        existing_cols: Set[str] = set()
        max_col = ws.max_column or 40
        for col_idx in range(5, max_col + 1):
            col_letter = get_column_letter(col_idx)
            for row in list(field_rows.keys())[:5]:
                val = ws.cell(row=row, column=col_idx).value
                if val is not None and str(val).strip() not in ("", "/"):
                    existing_cols.add(col_letter)
                    break

        # Get reference styles from existing data column
        ref_styles: Dict[int, str] = {}
        if existing_cols:
            ref_col = sorted(existing_cols, key=lambda c: column_index_from_string(c))[0]
            for row in field_rows:
                s = _get_cell_style(xml, f"{ref_col}{row}")
                if s:
                    ref_styles[row] = s
            # Also get styles for ticker and drug name rows
            if ticker_row:
                s = _get_cell_style(xml, f"{ref_col}{ticker_row}")
                if s:
                    ref_styles[ticker_row] = s
            if drug_name_row:
                s = _get_cell_style(xml, f"{ref_col}{drug_name_row}")
                if s:
                    ref_styles[drug_name_row] = s

        section = SectionInfo(
            indication=indication,
            header_row=hdr_row,
            first_data_row=first_field_row,
            last_data_row=last_field_row,
            ticker_row=ticker_row,
            drug_name_row=drug_name_row,
            field_rows=field_rows,
            existing_cols=existing_cols,
            mr_minus1_row=mr_minus1,
            mr_plus1_row=mr_plus1,
            ref_styles=ref_styles,
        )
        sections.append(section)
        logger.info(
            f"  Section: '{indication}' R{hdr_row} "
            f"fields={len(field_rows)} cols={len(existing_cols)} "
            f"ticker_row={ticker_row} drug_row={drug_name_row}"
        )

    return sections


def _match_section(sections: List[SectionInfo], indication: str) -> Optional[SectionInfo]:
    """Find the best matching section for a given indication.

    Tries exact match first, then substring match, then keyword match.
    """
    ind_lower = indication.strip().lower()

    # Exact match
    for s in sections:
        if s.indication.strip().lower() == ind_lower:
            return s

    # Substring match
    for s in sections:
        s_lower = s.indication.strip().lower()
        if ind_lower in s_lower or s_lower in ind_lower:
            return s

    # Keyword match (e.g., "mUC" matches "mUC Drug")
    ind_words = set(re.findall(r'[a-zA-Z0-9]+', ind_lower))
    for s in sections:
        s_words = set(re.findall(r'[a-zA-Z0-9]+', s.indication.lower()))
        overlap = ind_words & s_words
        if overlap and len(overlap) >= min(len(ind_words), len(s_words)) * 0.5:
            return s

    return None


# ══════════════════════════════════════════════════════════════════════════════
#  VALUE CONVERSION
# ══════════════════════════════════════════════════════════════════════════════

def _is_numeric(val: str) -> bool:
    """Check if a value string can be written as a number."""
    if not val or val in ("/", "-", "—", "N/A"):
        return False
    # Strip % sign
    cleaned = val.rstrip("%").strip()
    try:
        float(cleaned)
        return True
    except ValueError:
        return False


def _to_numeric(val: str) -> str:
    """Convert value to numeric string for Excel."""
    cleaned = val.rstrip("%").strip()
    return cleaned


def _is_date_field(field_name: str) -> bool:
    return field_name == "data_date"


def _date_to_serial(date_str: str) -> Optional[int]:
    """Convert date string to Excel serial number."""
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%d-%b-%Y", "%B %d, %Y"):
        try:
            dt = datetime.strptime(date_str, fmt)
            delta = dt - _EXCEL_EPOCH
            return delta.days
        except ValueError:
            continue
    return None


# ══════════════════════════════════════════════════════════════════════════════
#  DEDUPLICATION
# ══════════════════════════════════════════════════════════════════════════════

def _get_existing_readout_keys(xml: str, section: SectionInfo) -> Set[str]:
    """Build dedup keys from existing columns in a section."""
    keys: Set[str] = set()

    # Find which field rows map to drug_name, nct, data_date
    drug_name_rows = [r for r, f in section.field_rows.items() if f == "drug_name"]
    nct_rows = [r for r, f in section.field_rows.items() if f == "nct"]
    date_rows = [r for r, f in section.field_rows.items() if f == "data_date"]

    # Drug name is in drug_name_row (not a field row)
    drug_row = section.drug_name_row

    for col in section.existing_cols:
        col_idx = column_index_from_string(col)
        parts = []

        # Drug name
        if drug_row:
            found = _find_cell(xml, f"{col}{drug_row}")
            if found:
                _, _, _, inner = found
                t_m = re.search(r'<t[^>]*>(.*?)</t>', inner)
                v_m = re.search(r'<v>(.*?)</v>', inner)
                parts.append(t_m.group(1) if t_m else (v_m.group(1) if v_m else ""))
            else:
                parts.append("")
        else:
            parts.append("")

        # NCT (often not a field in existing sections)
        parts.append("")

        # Date
        for row in date_rows:
            found = _find_cell(xml, f"{col}{row}")
            if found:
                _, _, _, inner = found
                v_m = re.search(r'<v>(.*?)</v>', inner)
                parts.append(v_m.group(1) if v_m else "")
                break
        else:
            parts.append("")

        key = "|".join(parts)
        if any(p for p in parts):
            keys.add(key)

    return keys


# ══════════════════════════════════════════════════════════════════════════════
#  WRITE READOUTS TO SECTION
# ══════════════════════════════════════════════════════════════════════════════

def _write_readouts_to_section(
    xml: str,
    section: SectionInfo,
    readouts: List[DrugReadout],
    dry_run: bool,
) -> Tuple[str, int]:
    """Write readouts to an existing section. Returns (xml, count_written)."""
    if not readouts:
        return xml, 0

    # Find next empty column
    if section.existing_cols:
        max_existing = max(
            column_index_from_string(c) for c in section.existing_cols
        )
    else:
        max_existing = 4  # D column

    count = 0
    col_idx = max_existing + 1

    for readout in readouts:
        col = get_column_letter(col_idx)
        logger.info(f"    Writing {readout.drug_name} readout → col {col}")

        # 1. Ticker row
        if section.ticker_row:
            ticker_val = readout.ticker if readout.ticker and readout.ticker != "/" else ""
            if ticker_val:
                style = section.ref_styles.get(section.ticker_row)
                if dry_run:
                    logger.info(f"      [DRY-RUN] {col}{section.ticker_row} = '{ticker_val}'")
                else:
                    xml = _patch_text_cell(xml, f"{col}{section.ticker_row}", ticker_val, style)

        # 2. Drug name row
        if section.drug_name_row:
            style = section.ref_styles.get(section.drug_name_row)
            if dry_run:
                logger.info(f"      [DRY-RUN] {col}{section.drug_name_row} = '{readout.drug_name}'")
            else:
                xml = _patch_text_cell(xml, f"{col}{section.drug_name_row}",
                                       readout.drug_name, style)

        # 3. Field rows
        for row, field_name in section.field_rows.items():
            val = getattr(readout, field_name, "")
            if not val or val == "/":
                # Write "/" as text for missing data (matching existing convention)
                style = section.ref_styles.get(row)
                if dry_run:
                    logger.info(f"      [DRY-RUN] {col}{row} = '/'")
                else:
                    xml = _patch_text_cell(xml, f"{col}{row}", "/", style)
                continue

            style = section.ref_styles.get(row)
            addr = f"{col}{row}"

            if _is_date_field(field_name):
                serial = _date_to_serial(val)
                if serial:
                    # Use date style (s=850 from reference)
                    date_style = section.ref_styles.get(row, style)
                    if dry_run:
                        logger.info(f"      [DRY-RUN] {addr} = date({val})")
                    else:
                        xml = _patch_numeric_cell(xml, addr, str(serial), date_style)
                else:
                    if dry_run:
                        logger.info(f"      [DRY-RUN] {addr} = '{val}' (text)")
                    else:
                        xml = _patch_text_cell(xml, addr, val, style)
            elif _is_numeric(val):
                if dry_run:
                    logger.info(f"      [DRY-RUN] {addr} = {_to_numeric(val)}")
                else:
                    xml = _patch_numeric_cell(xml, addr, _to_numeric(val), style)
            else:
                if dry_run:
                    logger.info(f"      [DRY-RUN] {addr} = '{val}'")
                else:
                    xml = _patch_text_cell(xml, addr, val, style)

        # 4. Bloomberg market reaction formulas
        # These reference header rows: stock prices day before/after
        # Pattern from existing: =IFERROR(E13/E12-1, "")
        # where row 12 = price before readout date, row 13 = price on readout date
        # Only if we have the formula rows
        if section.mr_minus1_row:
            # Copy formula pattern from existing column
            ref_col = sorted(section.existing_cols)[0] if section.existing_cols else None
            if ref_col:
                found = _find_cell(xml, f"{ref_col}{section.mr_minus1_row}")
                if found:
                    _, _, _, inner = found
                    f_m = re.search(r'<f>(.*?)</f>', inner)
                    if f_m:
                        old_formula = f_m.group(1)
                        # Shift column reference
                        new_formula = re.sub(
                            rf'(?<![A-Z]){ref_col}(\d+)',
                            rf'{col}\1',
                            old_formula,
                        )
                        mr_style = _get_cell_style(xml, f"{ref_col}{section.mr_minus1_row}")
                        if dry_run:
                            logger.info(f"      [DRY-RUN] {col}{section.mr_minus1_row} = f({new_formula})")
                        else:
                            xml = _patch_formula_cell(
                                xml, f"{col}{section.mr_minus1_row}",
                                new_formula, mr_style
                            )

        if section.mr_plus1_row:
            ref_col = sorted(section.existing_cols)[0] if section.existing_cols else None
            if ref_col:
                found = _find_cell(xml, f"{ref_col}{section.mr_plus1_row}")
                if found:
                    _, _, _, inner = found
                    f_m = re.search(r'<f>(.*?)</f>', inner)
                    if f_m:
                        old_formula = f_m.group(1)
                        new_formula = re.sub(
                            rf'(?<![A-Z]){ref_col}(\d+)',
                            rf'{col}\1',
                            old_formula,
                        )
                        mr_style = _get_cell_style(xml, f"{ref_col}{section.mr_plus1_row}")
                        if dry_run:
                            logger.info(f"      [DRY-RUN] {col}{section.mr_plus1_row} = f({new_formula})")
                        else:
                            xml = _patch_formula_cell(
                                xml, f"{col}{section.mr_plus1_row}",
                                new_formula, mr_style
                            )

        col_idx += 1
        count += 1

    return xml, count


# ══════════════════════════════════════════════════════════════════════════════
#  NEW SECTION CREATION
# ══════════════════════════════════════════════════════════════════════════════

def _add_rows_to_xml(xml: str, after_row: int, count: int) -> str:
    """Add empty rows after a given row by inserting <row> elements.

    NOTE: This is a simplified approach. We insert row elements with
    renumbered row attributes. Existing rows after `after_row` are NOT
    renumbered (new sections go at the end of the sheet data).
    """
    # Find the </sheetData> tag to insert before
    sd_end = xml.find("</sheetData>")
    if sd_end == -1:
        logger.error("Cannot find </sheetData> in XML")
        return xml

    new_rows = []
    for i in range(count):
        row_num = after_row + 1 + i
        new_rows.append(f'<row r="{row_num}"/>')

    insert_xml = "".join(new_rows)
    return xml[:sd_end] + insert_xml + xml[sd_end:]


def _create_new_section(
    xml: str,
    indication: str,
    readouts: List[DrugReadout],
    last_row: int,
    ref_section: Optional[SectionInfo],
    dry_run: bool,
) -> Tuple[str, int, int]:
    """Create a brand new section at the end of the sheet.

    Returns: (xml, rows_added, readouts_written)
    """
    # Section layout:
    # Row 0: blank spacer
    # Row 1: header (X in A, indication name in D)
    # Row 2: blank spacer
    # Row 3: ticker row
    # Row 4: drug name row
    # Row 5+: field rows (from _NEW_SECTION_FIELDS)

    n_fields = len(_NEW_SECTION_FIELDS)
    section_rows = 5 + n_fields  # spacer + header + spacer + ticker + drugname + fields
    start_row = last_row + 2  # Leave a gap

    logger.info(f"  Creating new section '{indication}' starting at R{start_row}")

    if dry_run:
        logger.info(f"    [DRY-RUN] Would create {section_rows} rows at R{start_row}")
        for i, (label, _) in enumerate(_NEW_SECTION_FIELDS):
            logger.info(f"    [DRY-RUN] D{start_row + 5 + i} = '{label}'")
        return xml, section_rows, 0

    # Add row elements
    xml = _add_rows_to_xml(xml, start_row - 1, section_rows)

    # Get reference styles from an existing section
    hdr_style = "794"  # Default section header style
    label_style = "714"  # Default label style
    ticker_style = "638"
    drug_style = "712"
    data_text_style = "847"
    data_num_style = "852"
    marker_style = "798"

    if ref_section and ref_section.ref_styles:
        # Use styles from reference section
        if ref_section.field_rows:
            first_field = min(ref_section.field_rows.keys())
            label_ref = _get_cell_style(xml, f"D{first_field}")
            if label_ref:
                label_style = label_ref

    # Write section structure
    # Header row
    hdr_row = start_row
    xml = _patch_text_cell(xml, f"A{hdr_row}", "X", marker_style)
    xml = _patch_text_cell(xml, f"D{hdr_row}", indication, hdr_style)

    # Ticker row
    ticker_row = start_row + 3
    # Drug name row
    drug_name_row = start_row + 4

    # Field rows
    field_rows: Dict[int, str] = {}
    for i, (label, field_name) in enumerate(_NEW_SECTION_FIELDS):
        row = start_row + 5 + i
        xml = _patch_text_cell(xml, f"D{row}", label, label_style)
        field_rows[row] = field_name

    # Build a temporary SectionInfo for writing
    temp_section = SectionInfo(
        indication=indication,
        header_row=hdr_row,
        first_data_row=start_row + 5,
        last_data_row=start_row + 5 + n_fields - 1,
        ticker_row=ticker_row,
        drug_name_row=drug_name_row,
        field_rows=field_rows,
        existing_cols=set(),
        ref_styles={
            ticker_row: ticker_style,
            drug_name_row: drug_style,
            **{r: data_text_style for r in field_rows},
        },
    )

    # Now write readouts
    xml, count = _write_readouts_to_section(xml, temp_section, readouts, dry_run)

    return xml, section_rows, count


# ══════════════════════════════════════════════════════════════════════════════
#  SORT READOUTS
# ══════════════════════════════════════════════════════════════════════════════

def _sort_readouts(readouts: List[DrugReadout], study_drug: str) -> List[DrugReadout]:
    """Sort readouts: study drug first, then marketed (have sales), then R&D.

    Within each group, sort by drug name, then date.
    """
    def sort_key(r: DrugReadout) -> Tuple[int, str, str]:
        # Priority: 0=study drug, 1=marketed, 2=R&D
        if r.drug_name.lower() == study_drug.lower():
            priority = 0
        elif r.result and r.result.lower() == "approved":
            priority = 1
        elif r.latest_annual_sale and r.latest_annual_sale != "/":
            priority = 1
        else:
            priority = 2
        return (priority, r.drug_name.lower(), r.data_date or "")

    return sorted(readouts, key=sort_key)


# ══════════════════════════════════════════════════════════════════════════════
#  FIND DCF FILE
# ══════════════════════════════════════════════════════════════════════════════

def find_dcf_file(ticker: str) -> Optional[Path]:
    """Find the DCF Excel file for a ticker."""
    dd_dir = Path(f"/mnt/c/Users/yzsun/Desktop/DD/{ticker}")
    if not dd_dir.exists():
        return None

    candidates = []
    for f in dd_dir.iterdir():
        if not f.suffix == ".xlsx":
            continue
        name = f.name
        # Skip backups, lock files, temp files
        if name.startswith("~") or name.startswith("."):
            continue
        if "_pre_" in name or "_backup" in name:
            continue
        if "DCF" in name and ticker in name:
            candidates.append(f)

    if not candidates:
        return None

    # Prefer newest by mtime
    candidates.sort(key=lambda f: f.stat().st_mtime, reverse=True)
    return candidates[0]


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN PATCHING
# ══════════════════════════════════════════════════════════════════════════════

def patch_peer_views(
    xlsx_path: Path,
    readouts_by_ind: Dict[str, List[DrugReadout]],
    ticker: str,
    dry_run: bool,
) -> Dict[str, int]:
    """Patch the Peer Views sheet with parsed readout data.

    Returns: {indication: count_written}
    """
    # 1. Discover sheet structure using openpyxl (read-only)
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        logger.error(f"Sheet '{SHEET_NAME}' not found in workbook")
        wb.close()
        return {}

    ws = wb[SHEET_NAME]
    max_row = ws.max_row or 200

    # 2. Read sheet XML for style discovery
    sheet_zip_paths = _get_sheet_zip_paths(xlsx_path)
    sheet_zip = sheet_zip_paths.get(SHEET_NAME)
    if not sheet_zip:
        logger.error(f"Cannot find zip path for '{SHEET_NAME}'")
        wb.close()
        return {}

    with zipfile.ZipFile(xlsx_path) as zf:
        xml = zf.read(sheet_zip).decode("utf-8")

    logger.info(f"Peer Views sheet: {sheet_zip} ({len(xml):,} chars)")

    # 3. Discover sections
    sections = discover_sections(xml, ws)
    wb.close()

    if not sections:
        logger.warning("No sections discovered; will create new sections")

    # 4. Process each indication
    results: Dict[str, int] = {}
    last_row = max_row

    # Determine study drug name (first drug in first indication's readouts)
    study_drug = ""
    for ind_readouts in readouts_by_ind.values():
        if ind_readouts:
            study_drug = ind_readouts[0].drug_name
            break

    for indication, readouts in readouts_by_ind.items():
        logger.info(f"\n  Processing indication: {indication} ({len(readouts)} readouts)")

        # Sort readouts
        readouts = _sort_readouts(readouts, study_drug)

        # Deduplicate
        seen_keys: Set[str] = set()
        unique_readouts: List[DrugReadout] = []
        for r in readouts:
            key = r.dedup_key()
            if key not in seen_keys:
                seen_keys.add(key)
                unique_readouts.append(r)
            else:
                logger.info(f"    Skipping duplicate: {r.drug_name} {r.data_date}")

        readouts = unique_readouts

        # Find matching section
        section = _match_section(sections, indication)

        if section:
            logger.info(f"    Matched section: '{section.indication}' R{section.header_row}")

            # Check existing readouts for dedup
            existing_keys = _get_existing_readout_keys(xml, section)
            new_readouts = []
            for r in readouts:
                key = r.dedup_key()
                if key not in existing_keys:
                    new_readouts.append(r)
                else:
                    logger.info(f"    Already exists: {r.drug_name} {r.data_date}")

            if not new_readouts:
                logger.info(f"    No new readouts for '{indication}'")
                results[indication] = 0
                continue

            xml, count = _write_readouts_to_section(xml, section, new_readouts, dry_run)
            results[indication] = count
        else:
            logger.info(f"    No matching section found — creating new section")
            ref = sections[0] if sections else None
            xml, rows_added, count = _create_new_section(
                xml, indication, readouts, last_row, ref, dry_run,
            )
            last_row += rows_added + 2
            results[indication] = count

    # 5. Write back via zip patching
    if dry_run:
        logger.info("\nDry run complete — no changes written.")
        return results

    modified: Dict[str, bytes] = {sheet_zip: xml.encode("utf-8")}

    # Add fullCalcOnLoad
    with zipfile.ZipFile(xlsx_path) as zf:
        wb_xml_bytes = zf.read("xl/workbook.xml")
    wb_xml_str = wb_xml_bytes.decode("utf-8")
    if "fullCalcOnLoad" not in wb_xml_str:
        wb_xml_str = wb_xml_str.replace("<calcPr", '<calcPr fullCalcOnLoad="1"', 1)
        logger.info("Added fullCalcOnLoad='1' to workbook.xml")
    modified["xl/workbook.xml"] = wb_xml_str.encode("utf-8")

    # Write new zip
    tmp_path = xlsx_path.with_suffix(".~peer_patch.xlsx")
    with zipfile.ZipFile(xlsx_path, "r") as zin:
        with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                if item.filename == "xl/calcChain.xml":
                    continue  # removed — references stripped below
                if item.filename in modified:
                    zout.writestr(item, modified[item.filename])
                else:
                    zout.writestr(item, zin.read(item.filename))

    try:
        tmp_path.replace(xlsx_path)
    except PermissionError:
        import os
        os.remove(str(xlsx_path))
        tmp_path.rename(xlsx_path)

    logger.info(f"\nPatch applied -> {xlsx_path}")
    return results


# ══════════════════════════════════════════════════════════════════════════════
#  PRINT SUMMARY
# ══════════════════════════════════════════════════════════════════════════════

def print_summary(
    readouts_by_ind: Dict[str, List[DrugReadout]],
    results: Dict[str, int],
    dry_run: bool,
):
    """Print a summary table."""
    print(f"\n{'=' * 70}")
    print(f"Peer Views Update {'(DRY RUN)' if dry_run else 'COMPLETE'}")
    print(f"{'=' * 70}")
    print(f"{'Indication':<30} {'Parsed':>8} {'Written':>8} {'Drugs':>8}")
    print(f"{'-' * 30} {'-' * 8} {'-' * 8} {'-' * 8}")

    total_parsed = 0
    total_written = 0
    for ind, readouts in readouts_by_ind.items():
        n_parsed = len(readouts)
        n_written = results.get(ind, 0)
        drugs = len(set(r.drug_name for r in readouts))
        total_parsed += n_parsed
        total_written += n_written
        print(f"{ind:<30} {n_parsed:>8} {n_written:>8} {drugs:>8}")

    print(f"{'-' * 30} {'-' * 8} {'-' * 8}")
    print(f"{'TOTAL':<30} {total_parsed:>8} {total_written:>8}")
    print(f"{'=' * 70}\n")


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Parse Gemini research reports and fill Peer Views sheet"
    )
    parser.add_argument("--ticker", required=True, help="Stock ticker (e.g. CMPX)")
    parser.add_argument("--report-dir",
                        help="Directory containing .md reports (default: DD/{TICKER}/pipeline_base4/)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Preview changes without writing")
    parser.add_argument("--file", help="Path to DCF Excel file (auto-detected if not specified)")

    args = parser.parse_args()

    # Locate report directory
    if args.report_dir:
        report_dir = Path(args.report_dir)
    else:
        report_dir = Path(f"/mnt/c/Users/yzsun/Desktop/DD/{args.ticker}/pipeline_base4")

    if not report_dir.exists():
        logger.error(f"Report directory not found: {report_dir}")
        return

    # Locate DCF file
    if args.file:
        xlsx_path = Path(args.file)
    else:
        xlsx_path = find_dcf_file(args.ticker)

    if not xlsx_path or not xlsx_path.exists():
        logger.error(f"DCF file not found for {args.ticker}")
        return

    logger.info(f"Ticker: {args.ticker}")
    logger.info(f"Report dir: {report_dir}")
    logger.info(f"DCF file: {xlsx_path}")
    logger.info(f"Mode: {'DRY RUN' if args.dry_run else 'LIVE'}")

    # Step 1: Parse reports
    logger.info("\n" + "=" * 70)
    logger.info("STEP 1: Parsing Gemini reports")
    logger.info("=" * 70)

    text = scan_report_files(report_dir)
    if not text:
        logger.error("No report text found")
        return

    readouts_by_ind = parse_peer_view_blocks(text)
    if not readouts_by_ind:
        logger.error("No PEER_VIEW_START/END blocks found in reports")
        return

    total = sum(len(v) for v in readouts_by_ind.values())
    logger.info(f"Parsed {total} readouts across {len(readouts_by_ind)} indications")
    for ind, readouts in readouts_by_ind.items():
        drugs = set(r.drug_name for r in readouts)
        logger.info(f"  {ind}: {len(readouts)} readouts, {len(drugs)} drugs")

    # Step 2: Backup
    if not args.dry_run:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = xlsx_path.with_name(
            f"{xlsx_path.stem}_pre_peerviews_{ts}.xlsx"
        )
        shutil.copy2(xlsx_path, backup_path)
        logger.info(f"Backup: {backup_path}")

    # Step 3: Patch
    logger.info("\n" + "=" * 70)
    logger.info("STEP 2: Patching Peer Views sheet")
    logger.info("=" * 70)

    results = patch_peer_views(xlsx_path, readouts_by_ind, args.ticker, args.dry_run)

    # Step 4: Summary
    print_summary(readouts_by_ind, results, args.dry_run)


if __name__ == "__main__":
    main()
