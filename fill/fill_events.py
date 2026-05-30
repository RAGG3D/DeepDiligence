#!/usr/bin/env python3
"""
Biotech Catalyst Analyst — auto-fills Historical Events sheet in DCF Excel files.

Usage:
    python fill_events.py TICKER
    python fill_events.py "Fill for ticker: TICKER"

What it does:
    1. Opens  C:\\Users\\yzsun\\Desktop\\DD\\{TICKER}\\DCF*.xlsx
    2. Locates the "Historical Events" sheet (4 side-by-side yearly blocks)
    3. Fetches closing prices from Yahoo Finance
    4. Fetches press releases from SEC EDGAR (8-K/6-K), GlobeNewswire, Yahoo Finance news, IR RSS
    5. Writes prices → Share Price column; ≤15-word summaries → EVT column
    6. Saves and closes the workbook

Environment:
    ANTHROPIC_API_KEY — used by Claude Haiku for EVT summarisation
"""

import os
import sys
import glob
import re
import json
import time
import platform
from datetime import datetime, date, timedelta
from typing import Optional

# ── API-key bootstrap ─────────────────────────────────────────────────────────
def _load_api_key() -> Optional[str]:
    """Find ANTHROPIC_API_KEY from env → .env file → ~/.anthropic/api_key."""
    key = os.environ.get("ANTHROPIC_API_KEY")
    if key:
        return key
    # .env file next to this script
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    if os.path.exists(env_path):
        with open(env_path) as f:
            for line in f:
                line = line.strip()
                if line.startswith("ANTHROPIC_API_KEY="):
                    return line.split("=", 1)[1].strip().strip('"').strip("'")
    # ~/.anthropic/api_key
    alt = os.path.expanduser("~/.anthropic/api_key")
    if os.path.exists(alt):
        return open(alt).read().strip()
    return None

import requests
from bs4 import BeautifulSoup
import yfinance as yf
import openpyxl
import anthropic
from dotenv import load_dotenv

load_dotenv()

# ── Path resolution ───────────────────────────────────────────────────────────

def get_dd_path() -> str:
    """Return the base DD folder path appropriate for the running OS."""
    if platform.system() == "Windows":
        return r"C:\Users\yzsun\Desktop\DD"
    # WSL2 or Linux — Windows C: drive is mounted at /mnt/c/
    wsl_path = "/mnt/c/Users/yzsun/Desktop/DD"
    if os.path.isdir(wsl_path):
        return wsl_path
    raise EnvironmentError(
        "Cannot locate DD folder. Expected C:\\Users\\yzsun\\Desktop\\DD "
        "(Windows) or /mnt/c/Users/yzsun/Desktop/DD (WSL2)."
    )


# ── Excel sheet layout ────────────────────────────────────────────────────────
# The "Historical Events" sheet has four side-by-side yearly blocks.
# Each block: Date | Share Price | DoD Chg | EVT | Category  (5 columns)
# Blank separator column between blocks.
#   FA 2022 → cols B-F  (2-6)
#   FA 2023 → cols H-L  (8-12)
#   FA 2024 → cols N-R  (14-18)
#   FA 2025 → cols T-X  (20-24)
BLOCKS = [
    # (label,    date_col, price_col, dod_col, evt_col, cat_col)  — 1-indexed
    ("FA 2022",  2,  3,  4,  5,  6),
    ("FA 2023",  8,  9, 10, 11, 12),
    ("FA 2024", 14, 15, 16, 17, 18),
    ("FA 2025", 20, 21, 22, 23, 24),
]

# Values that mean "no valid data here"
BLANK_VALUES = {None, "", "#N/A N/A", "#VALUE!", "#REF!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!"}


def is_blank(v) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and v.strip() in BLANK_VALUES:
        return True
    return False


# ── File discovery ────────────────────────────────────────────────────────────

def find_dcf_file(ticker: str) -> str:
    base   = get_dd_path()
    folder = os.path.join(base, ticker)
    if not os.path.isdir(folder):
        raise FileNotFoundError(f"Ticker folder not found: {folder}")
    files = glob.glob(os.path.join(folder, "DCF*.xlsx"))
    # Filter out backup, lock, and pre-events files
    files = [f for f in files
             if "backup" not in f and "~$" not in f and "pre_events" not in f]
    if not files:
        raise FileNotFoundError(f"No DCF*.xlsx file in: {folder}")
    # Prefer newest file (by modification time)
    files.sort(key=lambda f: os.path.getmtime(f), reverse=True)
    return files[0]


# ── Date parsing ──────────────────────────────────────────────────────────────

def parse_cell_date(v) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y/%m/%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(v.strip(), fmt).date()
            except ValueError:
                pass
    return None


def _parse_date_str(s: str) -> Optional[date]:
    s = s.strip()
    for fmt in ("%Y-%m-%d", "%B %d, %Y", "%b %d, %Y", "%m/%d/%Y", "%d %B %Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(s[:10]).date()
    except (ValueError, IndexError):
        pass
    return None


# ── Load date map from workbook ───────────────────────────────────────────────

def load_date_map(ws) -> dict:
    """
    Returns {date_obj: (row_number, block_index)} for every date row
    across all four yearly blocks (rows 9+).

    Strategy: row 9 of each block holds a direct datetime value (anchor).
    Rows 10+ use formula =prev+1, whose cached result is gone after an
    openpyxl save. We therefore compute all subsequent dates in Python by
    incrementing timedelta(days=1) from the anchor — no formula evaluation.
    """
    result: dict[date, tuple[int, int]] = {}
    for bi, (label, dcol, *_) in enumerate(BLOCKS):
        # Row 9 is always the year's Jan-1 anchor (direct datetime value)
        anchor_val = ws.cell(row=9, column=dcol).value
        anchor = parse_cell_date(anchor_val)
        if anchor is None:
            continue   # block not present / unexpected layout

        # Walk rows 9..500; stop when the EVT/price block clearly ends
        # (we expect exactly 365 rows per year)
        for offset in range(365):
            row_idx = 9 + offset
            d = anchor + timedelta(days=offset)
            result[d] = (row_idx, bi)
    return result


# ── Yahoo Finance prices ──────────────────────────────────────────────────────

def fetch_prices(ticker: str, all_dates: list) -> tuple:
    """
    Returns (calendar_prices, trading_prices) where:
      calendar_prices  — {date: price} for every requested date, forward-filled
                         for weekends / holidays (written to Excel)
      trading_prices   — {date: price} for actual trading days only
                         (used for accurate DoD % change calculation)
    """
    today = date.today()
    past_dates = [d for d in all_dates if d <= today]
    if not past_dates:
        return {}, {}

    start = min(past_dates) - timedelta(days=7)
    end   = min(max(past_dates) + timedelta(days=3), today)

    print(f"    Downloading {ticker} prices {start} → {end} …")
    try:
        stock = yf.Ticker(ticker)
        hist  = stock.history(start=start.isoformat(), end=end.isoformat(), auto_adjust=True)
    except Exception as e:
        print(f"    [yfinance] {e}")
        return {}, {}

    # Raw trading-day prices (no forward-fill)
    trading_prices: dict[date, float] = {}
    for ts, row in hist.iterrows():
        trading_prices[ts.date()] = round(float(row["Close"]), 4)

    # Calendar prices: forward-fill non-trading days for Excel writing
    available = sorted(trading_prices)
    calendar_prices: dict[date, float] = {}
    for d in past_dates:
        if d in trading_prices:
            calendar_prices[d] = trading_prices[d]
        else:
            past = [p for p in available if p <= d]
            if past:
                calendar_prices[d] = trading_prices[past[-1]]

    return calendar_prices, trading_prices


# ── Event sources ─────────────────────────────────────────────────────────────

_REQ_HEADERS = {
    "User-Agent": (
        "AutoEvtPriceSheet/1.0 "
        "(biotech catalyst research tool; github.com/user/auto_evt_price_sheet)"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}


def fetch_yahoo_news(ticker: str) -> list:
    """Return list of (date, title) from Yahoo Finance news feed."""
    events: list[tuple[date, str]] = []
    try:
        stock = yf.Ticker(ticker)
        news  = stock.news or []
        for item in news:
            ts    = item.get("providerPublishTime", 0)
            title = item.get("title", "").strip()
            if ts and title:
                d = datetime.utcfromtimestamp(ts).date()
                events.append((d, title))
        print(f"    Yahoo Finance news: {len(events)} items")
    except Exception as e:
        print(f"    [Yahoo news] {e}")
    return events


def fetch_globenewswire(ticker: str) -> list:
    """Scrape GlobeNewswire search results for the ticker."""
    events: list[tuple[date, str]] = []
    urls_to_try = [
        f"https://www.globenewswire.com/search/keyword/{ticker}?page=1",
        f"https://www.globenewswire.com/search/keyword/{ticker}",
    ]
    for url in urls_to_try:
        try:
            r = requests.get(url, headers=_REQ_HEADERS, timeout=15)
            if r.status_code != 200:
                continue
            soup = BeautifulSoup(r.text, "html.parser")
            articles = (
                soup.select("article")
                or soup.select(".mainItem")
                or soup.select("[data-story-id]")
                or soup.select(".padd-20")
            )
            for art in articles[:60]:
                time_tag = art.find("time")
                link_tag = art.find("a")
                if not time_tag or not link_tag:
                    continue
                dt_str = time_tag.get("datetime") or time_tag.get_text()
                d      = _parse_date_str(dt_str)
                title  = link_tag.get_text(" ", strip=True)
                if d and title and len(title) > 5:
                    events.append((d, title))
            if events:
                break
        except Exception as e:
            print(f"    [GlobeNewswire] {e}")
    print(f"    GlobeNewswire: {len(events)} items")
    return events


def fetch_google_news(ticker: str, company_name: str = None) -> list:
    """Fetch news headlines from Google News RSS feed.
    When company_name is provided, filters results to only include items
    mentioning the company or ticker in a financial context (avoids noise
    for generic tickers like TARA, SAGE, RARE, etc.).
    """
    import xml.etree.ElementTree as ET
    events: list[tuple[date, str]] = []
    url = f"https://news.google.com/rss/search?q={ticker}&hl=en-US&gl=US&ceid=US:en"
    # Build filter terms from company name (e.g. "Protara Therapeutics, Inc." → "PROTARA")
    filter_terms: list[str] = []
    if company_name:
        short = company_name.split(",")[0].strip()   # "Protara Therapeutics"
        filter_terms.append(short.upper())
        # Also add first word if it's distinctive (>4 chars, not "THE")
        first_word = short.split()[0].upper() if short else ""
        if len(first_word) > 4 and first_word not in {"THE", "GROUP", "GLOBAL"}:
            filter_terms.append(first_word)
    try:
        r = requests.get(url, headers=_REQ_HEADERS, timeout=15)
        if r.status_code != 200:
            print(f"    [Google News] HTTP {r.status_code}")
            return events
        root = ET.fromstring(r.content)
        for item in root.iter("item"):
            title_el = item.find("title")
            date_el = item.find("pubDate")
            if title_el is None or date_el is None:
                continue
            raw_title = (title_el.text or "").strip()
            # Strip source suffix like " - CNBC", " - Reuters"
            if " - " in raw_title:
                raw_title = raw_title.rsplit(" - ", 1)[0].strip()
            if len(raw_title) < 10:
                continue
            # Filter: only keep items relevant to the company
            if filter_terms:
                upper = raw_title.upper()
                relevant = (
                    any(ft in upper for ft in filter_terms)
                    or f"({ticker})" in raw_title
                    or f"${ticker}" in raw_title
                    or f"NASDAQ:{ticker}" in upper
                    or f"{ticker}-" in upper    # drug name pattern (TARA-002)
                )
                if not relevant:
                    continue
            # Parse RFC 2822 date: "Wed, 21 Jan 2026 08:00:00 GMT"
            d = _parse_date_str(date_el.text or "")
            if d is None:
                try:
                    from email.utils import parsedate_to_datetime
                    d = parsedate_to_datetime(date_el.text).date()
                except Exception:
                    continue
            events.append((d, raw_title))
        print(f"    Google News: {len(events)} items")
    except Exception as e:
        print(f"    [Google News] {e}")
    return events


def fetch_google_news_conferences(ticker: str, company_name: str) -> list:
    """
    Dedicated conference-specific Google News search using the company's full name.
    This catches conference participation announcements that generic ticker searches miss
    (especially for tickers that are common words like TARA, SAGE, RARE, etc.).
    Results are filtered to only include items that mention the company.
    """
    import xml.etree.ElementTree as ET
    events: list[tuple[date, str]] = []
    seen_titles: set[str] = set()
    short_name = company_name.split(",")[0].strip()   # "Protara Therapeutics"
    first_word = short_name.split()[0].upper() if short_name else ""

    # Queries use QUOTED company name to ensure relevance — no broad ticker query
    queries = [
        f'"{short_name}" conference OR presentation',
        f'"{short_name}" poster OR "presented data" OR "annual meeting"',
    ]
    for query in queries:
        try:
            url = (
                f"https://news.google.com/rss/search?"
                f"q={requests.utils.quote(query)}&hl=en-US&gl=US&ceid=US:en"
            )
            r = requests.get(url, headers=_REQ_HEADERS, timeout=15)
            if r.status_code != 200:
                continue
            root = ET.fromstring(r.content)
            for item in root.iter("item"):
                title_el = item.find("title")
                date_el = item.find("pubDate")
                if title_el is None or date_el is None:
                    continue
                raw = (title_el.text or "").strip()
                if " - " in raw:
                    raw = raw.rsplit(" - ", 1)[0].strip()
                if len(raw) < 10 or raw in seen_titles:
                    continue
                # Filter: must mention company name, ticker, or drug name
                upper = raw.upper()
                if (short_name.upper() not in upper
                        and (len(first_word) < 5 or first_word not in upper)
                        and f"({ticker})" not in raw
                        and f"${ticker}" not in raw):
                    continue
                seen_titles.add(raw)
                d = _parse_date_str(date_el.text or "")
                if d is None:
                    try:
                        from email.utils import parsedate_to_datetime
                        d = parsedate_to_datetime(date_el.text).date()
                    except Exception:
                        continue
                events.append((d, raw))
        except Exception:
            continue
    print(f"    Google News conferences: {len(events)} items")
    return events


def _get_cik(ticker: str) -> Optional[str]:
    """Look up SEC EDGAR CIK for a ticker symbol."""
    try:
        headers = {**_REQ_HEADERS, "User-Agent": "AutoEvtResearch/1.0 noreply@example.com"}
        r = requests.get(
            "https://www.sec.gov/files/company_tickers.json",
            headers=headers, timeout=10,
        )
        tickers_json = r.json()
        for entry in tickers_json.values():
            if entry["ticker"].upper() == ticker.upper():
                return str(entry["cik_str"]).zfill(10)
    except Exception as e:
        print(f"    [EDGAR CIK lookup] {e}")
    return None


# Cache for company name lookups (avoids duplicate SEC API calls)
_company_name_cache: dict[str, Optional[str]] = {}


def _get_company_name(ticker: str) -> Optional[str]:
    """Look up the full company name for a ticker from SEC EDGAR."""
    if ticker.upper() in _company_name_cache:
        return _company_name_cache[ticker.upper()]
    try:
        headers = {**_REQ_HEADERS, "User-Agent": "AutoEvtResearch/1.0 noreply@example.com"}
        r = requests.get(
            "https://www.sec.gov/files/company_tickers.json",
            headers=headers, timeout=10,
        )
        for entry in r.json().values():
            if entry["ticker"].upper() == ticker.upper():
                name = entry.get("title", "")
                _company_name_cache[ticker.upper()] = name
                return name
    except Exception:
        pass
    _company_name_cache[ticker.upper()] = None
    return None


# ── IR RSS feed ──────────────────────────────────────────────────────────────

_LEGAL_SUFFIXES = re.compile(
    r'\s*,?\s*\b(Inc|Corp|Ltd|AG|SA|SE|NV|PLC|LP|LLC|Co)\b\.?\s*$',
    re.IGNORECASE,
)


def _discover_ir_domain(company_name: str) -> Optional[str]:
    """Guess IR page domain from company name. Returns domain or None."""
    if not company_name:
        return None
    # Remove legal suffixes: "Molecular Partners AG" → "Molecular Partners"
    clean = _LEGAL_SUFFIXES.sub('', company_name).strip()
    # Join words, lowercase: "Molecular Partners" → "molecularpartners"
    slug = ''.join(clean.lower().split())
    # Remove non-alphanumeric
    slug = re.sub(r'[^a-z0-9]', '', slug)

    # Try common IR subdomain patterns
    candidates = [
        f"investors.{slug}.com",
        f"ir.{slug}.com",
    ]
    for domain in candidates:
        try:
            r = requests.head(
                f"https://{domain}/",
                headers=_REQ_HEADERS, timeout=5,
                allow_redirects=True,
            )
            if r.status_code < 400:
                return domain
        except Exception:
            continue
    return None


def fetch_ir_rss(ticker: str, company_name: str = None) -> list:
    """Fetch press releases from company IR page RSS feed.
    Many biotech IR pages (Notified/Q4 platform) offer RSS at
    /rss/news-releases.xml. Gracefully returns empty list on failure.
    """
    events: list[tuple[date, str]] = []
    ir_domain = _discover_ir_domain(company_name)
    if ir_domain is None:
        print(f"    IR RSS: domain not found for {company_name or ticker}")
        return events

    url = f"https://{ir_domain}/rss/news-releases.xml"
    try:
        r = requests.get(url, headers=_REQ_HEADERS, timeout=15)
        if r.status_code != 200:
            print(f"    IR RSS: HTTP {r.status_code} for {url}")
            return events

        import xml.etree.ElementTree as ET
        root = ET.fromstring(r.content)
        for item in root.findall('.//item'):
            title_el = item.find('title')
            date_el = item.find('pubDate')
            if title_el is None or date_el is None:
                continue
            title = (title_el.text or '').strip()
            if not title or len(title) < 10:
                continue
            d = _parse_date_str(date_el.text or '')
            if d is None:
                try:
                    from email.utils import parsedate_to_datetime
                    d = parsedate_to_datetime(date_el.text).date()
                except Exception:
                    continue
            events.append((d, title))
        print(f"    IR RSS ({ir_domain}): {len(events)} items")
    except requests.exceptions.Timeout:
        print(f"    IR RSS: timeout for {ir_domain} (may be unreachable from this network)")
    except Exception as e:
        print(f"    IR RSS: {e}")
    return events


# 8-K item codes → human-readable labels (fallback when no press-release title found)
_ITEM_LABELS = {
    "1.01": "Material Agreement",    "1.02": "Agreement Termination",
    "2.02": "Quarterly/Annual Results", "3.02": "Securities Sale",
    "4.01": "Auditor Change",        "5.02": "Director/Officer Change",
    "5.07": "Compensation Disclosure", "7.01": "Reg-FD Disclosure",
    "8.01": "Other Material Events", "9.01": "Financial Exhibits",
}

_SKIP_TEXTS = {
    "EXHIBIT 99.1", "EXHIBIT 99", "EDGAR", "FORM 8-K", "8-K", "SEC",
    "DOCUMENT AND ENTITY INFORMATION", "COVER PAGE",
    "UNITED STATES SECURITIES AND EXCHANGE COMMISSION",
    "CURRENT REPORT", "PURSUANT TO SECTION", "COMMISSION FILE NUMBER",
}


def _edgar_pr_title(cik_int: int, accn: str, headers: dict,
                    primary_doc: str = None) -> Optional[str]:
    """
    Fetch a filing index for one 8-K, find the best document, and
    return:  "<Headline> | <first 2500 chars of body text>"

    Three-tier document fallback:
      1. EX-99.1 press release (ideal for earnings, material events)
      2. Any EX-99.x exhibit (some companies use EX-99.2 etc.)
      3. Primary 8-K document itself (always exists — contains actual disclosure)

    The body snippet lets _has_conference_keyword() detect keywords like
    "ASCO" / "SITC" that are buried in the press release body, and gives
    the LLM clinical metrics (ORR, PFS, OS) for structured extraction.
    """
    accn_clean = accn.replace("-", "")
    idx_url = (
        f"https://www.sec.gov/Archives/edgar/data/"
        f"{cik_int}/{accn_clean}/{accn}-index.htm"
    )
    try:
        r = requests.get(idx_url, headers=headers, timeout=10)
        soup = BeautifulSoup(r.text, "html.parser")

        # ── Tier 1: Find EX-99.1 press release ──────────────────────────
        pr_doc = None
        any_exhibit = None   # Track any EX-99.x as Tier 2 fallback
        for row in soup.find_all("tr"):
            cells = [td.get_text(" ", strip=True) for td in row.find_all("td")]
            # Tier 1: exact EX-99.1 or PRESS RELEASE
            if any("EX-99.1" in c or "PRESS RELEASE" in c.upper() for c in cells):
                for cell in cells:
                    if cell.lower().endswith((".htm", ".html")) and cell not in {"", "index.htm"}:
                        pr_doc = cell
                        break
                if pr_doc is None:
                    for a in row.find_all("a", href=True):
                        href = a["href"]
                        if href.lower().endswith((".htm", ".html")):
                            pr_doc = href.split("/")[-1]
                            break
            # Tier 2: track any EX-99.x exhibit
            if pr_doc is None and any_exhibit is None:
                if any("EX-99" in c for c in cells):
                    for cell in cells:
                        if cell.lower().endswith((".htm", ".html")) and cell not in {"", "index.htm"}:
                            any_exhibit = cell
                            break
                    if any_exhibit is None:
                        for a in row.find_all("a", href=True):
                            href = a["href"]
                            if href.lower().endswith((".htm", ".html")):
                                any_exhibit = href.split("/")[-1]
                                break
            if pr_doc:
                break

        # ── Tier 2: Fall back to any EX-99.x exhibit ────────────────────
        if pr_doc is None and any_exhibit is not None:
            pr_doc = any_exhibit

        # ── Tier 3: Fall back to primary 8-K document ───────────────────
        if pr_doc is None and primary_doc:
            pr_doc = primary_doc

        if pr_doc is None:
            return None

        pr_url = (
            f"https://www.sec.gov/Archives/edgar/data/"
            f"{cik_int}/{accn_clean}/{pr_doc}"
        )
        r2 = requests.get(pr_url, headers=headers, timeout=10)
        soup2 = BeautifulSoup(r2.text, "html.parser")

        # ── Slide-deck detection: presentations filed as EX-99.1 contain
        # images with hidden 1pt OCR text — useless for summarization.
        # The primary 8-K document has the actual Item 7.01 disclosure.
        if primary_doc and pr_doc != primary_doc:
            img_count = len(soup2.find_all("img"))
            if img_count > 3:
                pr_url = (
                    f"https://www.sec.gov/Archives/edgar/data/"
                    f"{cik_int}/{accn_clean}/{primary_doc}"
                )
                r2 = requests.get(pr_url, headers=headers, timeout=10)
                soup2 = BeautifulSoup(r2.text, "html.parser")

        # ── Headline: first short, non-boilerplate tag ─────────────────────
        headline = None
        for tag in soup2.find_all(["title", "h1", "h2", "h3", "p", "b", "strong"])[:30]:
            text = tag.get_text(" ", strip=True)
            upper = text.upper()
            if (15 < len(text) < 250
                    and not text.startswith('"')
                    and not any(skip in upper for skip in _SKIP_TEXTS)
                    and "FORWARD-LOOKING" not in upper
                    and "INVESTOR CONTACT" not in upper
                    and "GLOBE NEWSWIRE" not in upper
                    and not upper.startswith("EXHIBIT")):
                headline = text
                break

        # ── Body: concatenate <p> paragraphs up to 2500 chars ─────────────
        _BODY_SKIP = {
            "FORWARD-LOOKING", "INVESTOR CONTACT", "GLOBE NEWSWIRE",
            "SAFE HARBOR", "SOURCE:", "ABOUT ",
            "SIGNATURES", "PURSUANT TO THE REQUIREMENTS",
            "POWER OF ATTORNEY", "DATE:",
        }
        body_parts: list[str] = []
        body_len = 0
        for p in soup2.find_all("p"):
            text = p.get_text(" ", strip=True)
            if len(text) < 20:
                continue
            upper = text.upper()
            if any(skip in upper for skip in _BODY_SKIP):
                continue
            body_parts.append(text)
            body_len += len(text)
            if body_len >= 2500:
                break
        body_snippet = " ".join(body_parts)[:2500]

        # ── Full-text fallback for iXBRL or non-standard HTML ────────────
        # Some documents (iXBRL 8-K, presentation-style press releases) put
        # content in <div>/<span>/tables instead of <p>.  Fall back to
        # get_text() and extract the meaningful content.
        if not headline and not body_snippet:
            import re as _re
            full_text = soup2.get_text(" ", strip=True)
            # Strategy 1: Find "Item X.XX" section (for 8-K primary docs)
            item_match = _re.search(
                r'(Item\s+\d+\.\d+[^.]*?\.)\s*(.*?)(?=\bItem\s+\d+\.\d+\b|\bSIGNATURES?\b|$)',
                full_text, _re.DOTALL | _re.IGNORECASE
            )
            if item_match:
                headline = item_match.group(1).strip()[:200]
                raw_body = item_match.group(2).strip()
            else:
                # Strategy 2: Skip boilerplate prefix, take first content
                # (for press releases in non-standard format)
                raw_body = full_text
                # Skip known boilerplate prefixes
                for marker in ["Exhibit 99.1", "Exhibit 99", "EX-99.1"]:
                    idx = raw_body.find(marker)
                    if idx >= 0:
                        raw_body = raw_body[idx + len(marker):].strip()
                        break
            # Filter boilerplate from body
            if raw_body:
                body_sents = []
                for sent in raw_body.split(". "):
                    s = sent.strip()
                    if len(s) < 15:
                        continue
                    upper_s = s.upper()
                    if any(skip in upper_s for skip in _BODY_SKIP):
                        break
                    body_sents.append(s)
                    if sum(len(x) for x in body_sents) > 2500:
                        break
                body_snippet = ". ".join(body_sents)[:2500]
                # Extract headline from first sentence if still missing
                if not headline and body_snippet:
                    first_period = body_snippet.find(". ")
                    if first_period > 0 and first_period < 250:
                        headline = body_snippet[:first_period + 1]
                        body_snippet = body_snippet[first_period + 2:]
                    else:
                        headline = body_snippet[:200]

        if headline and body_snippet:
            return f"{headline} | {body_snippet}"
        return headline or body_snippet or None

    except Exception as e:
        # Log but don't crash — some filings may be unavailable
        print(f"    [EDGAR PR] accn={accn}: {e}")
    return None


def fetch_edgar_8k(ticker: str) -> list:
    """
    Pull 8-K press release titles and dates from SEC EDGAR.
    Strategy:
      1. Look up CIK via company_tickers.json
      2. Pull submissions JSON for the company
      3. For each 8-K, fetch the filing index → find EX-99.1 → extract headline
      4. Fall back to item-code description if no press release found
    """
    events: list[tuple[date, str]] = []
    headers = {**_REQ_HEADERS, "User-Agent": "AutoEvtResearch/1.0 noreply@example.com"}

    cik = _get_cik(ticker)
    if cik is None:
        print(f"    [EDGAR] CIK not found for {ticker}, skipping.")
        return events

    try:
        r = requests.get(
            f"https://data.sec.gov/submissions/CIK{cik}.json",
            headers=headers, timeout=15,
        )
        sub     = r.json()
        company = sub.get("name", ticker)
        recent  = sub.get("filings", {}).get("recent", {})
        forms   = recent.get("form", [])
        f_dates = recent.get("filingDate", [])
        accns   = recent.get("accessionNumber", [])
        items   = recent.get("items", [])   # 8-K item codes, e.g. "2.02,9.01"
        primary_docs = recent.get("primaryDocument", [])

        cik_int = int(cik)

        # Detect foreign private issuer: has 6-K filings but no 8-K
        has_8k = any("8-K" in f and f != "8-K/A" for f in forms)
        has_6k = any("6-K" in f for f in forms)
        target_forms = set()
        if has_8k:
            target_forms.add("8-K")
        if has_6k:
            target_forms.add("6-K")

        for i, (form, fdate, accn, item_code) in enumerate(zip(forms, f_dates, accns, items)):
            if form not in target_forms:
                continue
            d = _parse_date_str(fdate)
            if d is None:
                continue

            # Get the primary document filename from submissions JSON
            prim_doc = primary_docs[i] if i < len(primary_docs) else None

            # Try EX-99.1, then any EX-99.x, then primary 8-K document
            pr_title = _edgar_pr_title(cik_int, accn, headers, primary_doc=prim_doc)
            time.sleep(0.1)   # polite rate-limiting

            if pr_title:
                title = pr_title
            else:
                # Fallback: use item-code descriptions
                codes = [c.strip() for c in str(item_code).split(",") if c.strip()]
                labels = [_ITEM_LABELS.get(c) for c in codes if c in _ITEM_LABELS]
                meaningful = [lb for lb in labels if lb and lb != "Financial Exhibits"]
                if meaningful:
                    title = f"{company}: {'; '.join(meaningful[:2])}"
                elif form == "6-K":
                    title = f"{company}: 6-K Report"
                else:
                    title = f"{company}: 8-K Filing"

            events.append((d, title))

        print(f"    SEC EDGAR 8-K/6-K: {len(events)} filings")
    except Exception as e:
        print(f"    [EDGAR submissions] {e}")

    return events


def fetch_edgar_efts_conferences(ticker: str, company_name: str = None) -> list:
    """
    Search SEC EDGAR full-text search (EFTS) for conference/presentation
    mentions in 8-K filings.  Uses COMPANY NAME (not just ticker) because
    small-cap tickers like TARA/SAGE/RARE are common words that produce noise.
    """
    events: list[tuple[date, str]] = []
    headers = {**_REQ_HEADERS, "User-Agent": "AutoEvtResearch/1.0 noreply@example.com"}
    cik = _get_cik(ticker)
    if cik is None:
        return events

    # Use short company name for more accurate search; fall back to ticker
    # Strip ", Inc." / ", Ltd." etc. that break exact-match in filing text
    if company_name:
        short_name = company_name.split(",")[0].strip()
        search_entity = f'"{short_name}"'
    else:
        search_entity = f'"{ticker}"'
    # Simple query: "company" "conference" — EFTS handles implicit AND
    query = f'{search_entity} "conference"'
    url = (
        f"https://efts.sec.gov/LATEST/search-index?"
        f"q={requests.utils.quote(query)}"
        f"&forms=8-K,6-K"
        f"&dateRange=custom&startdt=2020-01-01&enddt=2026-12-31"
    )
    try:
        r = requests.get(url, headers=headers, timeout=15)
        if r.status_code != 200:
            print(f"    [EFTS] HTTP {r.status_code}")
            return events
        data = r.json()
        hits = data.get("hits", {}).get("hits", [])
        for hit in hits:
            src = hit.get("_source", {})
            fdate = src.get("file_date", "")
            d = _parse_date_str(fdate)
            if d is None:
                continue
            # Build title from filing metadata
            desc = src.get("file_description", "") or src.get("file_type", "")
            display = src.get("display_names", [""])[0]
            display_name = display.split("(")[0].strip() if display else ticker
            items_raw = src.get("items", [])
            item_labels = []
            for item_code in items_raw:
                label = _ITEM_LABELS.get(item_code)
                if label and label != "Financial Exhibits":
                    item_labels.append(label)
            label_str = "; ".join(item_labels[:2]) if item_labels else desc
            title = f"{display_name}: {label_str}" if label_str else f"{display_name}: Conference/Presentation Filing"
            events.append((d, title))
        print(f"    SEC EDGAR EFTS conferences: {len(events)} hits")
    except Exception as e:
        print(f"    [EFTS] {e}")
    return events


def fetch_edgar_efts_conference_names(ticker: str, company_name: str = None) -> list:
    """
    Search EDGAR EFTS for specific conference names in 8-K filings.
    Runs batched queries for major conference acronyms to catch filings
    that mention specific conferences (e.g. "AACR", "ASCO") but not
    the generic word "conference".
    """
    events: list[tuple[date, str]] = []
    headers = {**_REQ_HEADERS, "User-Agent": "AutoEvtResearch/1.0 noreply@example.com"}
    cik = _get_cik(ticker)
    if cik is None:
        return events

    if company_name:
        short_name = company_name.split(",")[0].strip()
        entity = f'"{short_name}"'
    else:
        entity = f'"{ticker}"'

    # Batch conference acronyms into groups to reduce API calls
    conf_batches = [
        '"AACR" OR "ASCO" OR "ESMO" OR "SITC"',
        '"ASH" OR "SABCS" OR "EHA" OR "SNO"',
        '"AAN" OR "CTAD" OR "EULAR" OR "ACR"',
        '"AASLD" OR "DDW" OR "JPM" OR "J.P. Morgan"',
        '"poster" OR "oral presentation" OR "late-breaking"',
    ]
    seen_dates: set[date] = set()
    for batch in conf_batches:
        query = f'{entity} {batch}'
        url = (
            f"https://efts.sec.gov/LATEST/search-index?"
            f"q={requests.utils.quote(query)}"
            f"&forms=8-K,6-K&dateRange=custom&startdt=2020-01-01&enddt=2026-12-31"
        )
        try:
            r = requests.get(url, headers=headers, timeout=15)
            if r.status_code != 200:
                continue
            data = r.json()
            for hit in data.get("hits", {}).get("hits", []):
                src = hit.get("_source", {})
                fdate = src.get("file_date", "")
                d = _parse_date_str(fdate)
                if d is None or d in seen_dates:
                    continue
                seen_dates.add(d)
                display = src.get("display_names", [""])[0]
                display_name = display.split("(")[0].strip() if display else ticker
                title = f"{display_name}: Conference/Presentation Filing"
                events.append((d, title))
        except Exception:
            continue
    print(f"    SEC EDGAR EFTS conference names: {len(events)} hits")
    return events


# ── Merge and deduplicate events ──────────────────────────────────────────────

def build_event_map(raw_events: list) -> dict:
    """
    Collapse all (date, title) pairs into {date: best_title}.
    Prefer longer, more informative titles.
    """
    event_map: dict[date, str] = {}
    for d, title in raw_events:
        title = title.strip()
        if not title:
            continue
        if d not in event_map or len(title) > len(event_map[d]):
            event_map[d] = title
    return event_map


# ── Conference keyword trigger ────────────────────────────────────────────────

_CONF_KEYWORDS = {
    # Oncology
    "AACR", "ASCO", "ESMO", "SITC", "ASH", "WCLC", "SABCS", "SNO", "EHA",
    "ASCO GI", "ASCO GU", "ESMO IO",
    # Neurology / Neuroscience
    "AAN", "CTAD", "AES", "AANEM", "AD/PD", "ECTRIMS", "ACTRIMS",
    # Immunology / Rheumatology
    "ACR", "EULAR", "AAD",
    # Hepatology / GI
    "AASLD", "DDW", "UEG",
    # Rare / Genetics
    "WORLDSymposium", "WORLD SYMPOSIUM", "ASHG", "ACMG",
    # Respiratory / Allergy
    "ATS", "ERS", "AAAAI",
    # Nephrology / Cardiology
    "ASN", "AHA", "ACC", "ESC",
    # Ophthalmology
    "AAO", "ARVO",
    # Infectious Disease
    "IDWeek", "CROI", "ECCMID",
    # General / Cross-specialty / Investor
    "JPM", "J.P. MORGAN", "BIO INTERNATIONAL", "ISTH", "DIA", "WCBP",
    "COWEN", "JEFFERIES", "GOLDMAN SACHS", "PIPER SANDLER",
    "LEERINK", "STIFEL", "NEEDHAM", "WELLS FARGO",
    "INVESTOR DAY", "ANALYST DAY", "R&D DAY", "SCIENCE DAY",
    # Presentation format keywords
    "ORAL PRESENTATION", "POSTER PRESENTATION", "LATE-BREAKING",
    "PRESENTED AT", "PRESENTED DATA", "DATA AT",
    "SCIENTIFIC SESSION", "PLENARY SESSION",
    # Generic
    "CONFERENCE", "SYMPOSIUM", "ANNUAL MEETING", "CONGRESS",
}

# ── Conference calendar for date remapping ──────────────────────────────────
# Static dict mapping (acronym, year) → (start_date, end_date).
# Used to remap retroactive conference mentions (e.g. in 6-K quarterly updates)
# back to actual conference dates where abstracts were released.
_CONFERENCE_CALENDAR = {
    # AACR (American Association for Cancer Research) — April
    ("AACR", 2020): (date(2020, 4, 27), date(2020, 4, 28)),
    ("AACR", 2021): (date(2021, 4, 10), date(2021, 4, 15)),
    ("AACR", 2022): (date(2022, 4, 8), date(2022, 4, 13)),
    ("AACR", 2023): (date(2023, 4, 14), date(2023, 4, 19)),
    ("AACR", 2024): (date(2024, 4, 5), date(2024, 4, 10)),
    ("AACR", 2025): (date(2025, 4, 25), date(2025, 4, 30)),
    # ASCO (American Society of Clinical Oncology) — late May/early June
    ("ASCO", 2020): (date(2020, 5, 29), date(2020, 5, 31)),
    ("ASCO", 2021): (date(2021, 6, 4), date(2021, 6, 8)),
    ("ASCO", 2022): (date(2022, 6, 3), date(2022, 6, 7)),
    ("ASCO", 2023): (date(2023, 6, 2), date(2023, 6, 6)),
    ("ASCO", 2024): (date(2024, 5, 31), date(2024, 6, 4)),
    ("ASCO", 2025): (date(2025, 5, 30), date(2025, 6, 3)),
    # ESMO (European Society for Medical Oncology) — Sep/Oct
    ("ESMO", 2020): (date(2020, 9, 19), date(2020, 9, 21)),
    ("ESMO", 2021): (date(2021, 9, 16), date(2021, 9, 21)),
    ("ESMO", 2022): (date(2022, 9, 9), date(2022, 9, 13)),
    ("ESMO", 2023): (date(2023, 10, 20), date(2023, 10, 24)),
    ("ESMO", 2024): (date(2024, 9, 13), date(2024, 9, 17)),
    ("ESMO", 2025): (date(2025, 9, 12), date(2025, 9, 16)),
    # SITC (Society for Immunotherapy of Cancer) — November
    ("SITC", 2020): (date(2020, 11, 9), date(2020, 11, 14)),
    ("SITC", 2021): (date(2021, 11, 10), date(2021, 11, 14)),
    ("SITC", 2022): (date(2022, 11, 8), date(2022, 11, 12)),
    ("SITC", 2023): (date(2023, 11, 1), date(2023, 11, 5)),
    ("SITC", 2024): (date(2024, 11, 6), date(2024, 11, 10)),
    ("SITC", 2025): (date(2025, 11, 5), date(2025, 11, 9)),
    # ASH (American Society of Hematology) — December
    ("ASH", 2020): (date(2020, 12, 5), date(2020, 12, 8)),
    ("ASH", 2021): (date(2021, 12, 11), date(2021, 12, 14)),
    ("ASH", 2022): (date(2022, 12, 10), date(2022, 12, 13)),
    ("ASH", 2023): (date(2023, 12, 9), date(2023, 12, 12)),
    ("ASH", 2024): (date(2024, 12, 7), date(2024, 12, 10)),
    ("ASH", 2025): (date(2025, 12, 6), date(2025, 12, 9)),
    # WCLC (World Conference on Lung Cancer) — Aug/Sep
    ("WCLC", 2020): (date(2020, 8, 8), date(2020, 8, 10)),
    ("WCLC", 2021): (date(2021, 9, 8), date(2021, 9, 14)),
    ("WCLC", 2022): (date(2022, 8, 6), date(2022, 8, 9)),
    ("WCLC", 2023): (date(2023, 9, 9), date(2023, 9, 12)),
    ("WCLC", 2024): (date(2024, 9, 7), date(2024, 9, 10)),
    ("WCLC", 2025): (date(2025, 9, 6), date(2025, 9, 9)),
    # SABCS (San Antonio Breast Cancer Symposium) — December
    ("SABCS", 2020): (date(2020, 12, 8), date(2020, 12, 11)),
    ("SABCS", 2021): (date(2021, 12, 7), date(2021, 12, 10)),
    ("SABCS", 2022): (date(2022, 12, 6), date(2022, 12, 9)),
    ("SABCS", 2023): (date(2023, 12, 5), date(2023, 12, 8)),
    ("SABCS", 2024): (date(2024, 12, 10), date(2024, 12, 13)),
    ("SABCS", 2025): (date(2025, 12, 9), date(2025, 12, 12)),
    # SNO (Society for Neuro-Oncology) — November
    ("SNO", 2020): (date(2020, 11, 19), date(2020, 11, 22)),
    ("SNO", 2021): (date(2021, 11, 18), date(2021, 11, 21)),
    ("SNO", 2022): (date(2022, 11, 16), date(2022, 11, 20)),
    ("SNO", 2023): (date(2023, 11, 15), date(2023, 11, 19)),
    ("SNO", 2024): (date(2024, 11, 21), date(2024, 11, 24)),
    ("SNO", 2025): (date(2025, 11, 20), date(2025, 11, 23)),
    # EHA (European Hematology Association) — June
    ("EHA", 2020): (date(2020, 6, 11), date(2020, 6, 14)),
    ("EHA", 2021): (date(2021, 6, 9), date(2021, 6, 17)),
    ("EHA", 2022): (date(2022, 6, 9), date(2022, 6, 12)),
    ("EHA", 2023): (date(2023, 6, 8), date(2023, 6, 11)),
    ("EHA", 2024): (date(2024, 6, 13), date(2024, 6, 16)),
    ("EHA", 2025): (date(2025, 6, 12), date(2025, 6, 15)),
    # AAN (American Academy of Neurology) — April
    ("AAN", 2020): (date(2020, 4, 25), date(2020, 5, 1)),
    ("AAN", 2021): (date(2021, 4, 17), date(2021, 4, 22)),
    ("AAN", 2022): (date(2022, 4, 2), date(2022, 4, 7)),
    ("AAN", 2023): (date(2023, 4, 22), date(2023, 4, 27)),
    ("AAN", 2024): (date(2024, 4, 13), date(2024, 4, 18)),
    ("AAN", 2025): (date(2025, 4, 5), date(2025, 4, 10)),
    # ACR (American College of Rheumatology) — November
    ("ACR", 2020): (date(2020, 11, 5), date(2020, 11, 9)),
    ("ACR", 2021): (date(2021, 11, 3), date(2021, 11, 9)),
    ("ACR", 2022): (date(2022, 11, 10), date(2022, 11, 14)),
    ("ACR", 2023): (date(2023, 11, 10), date(2023, 11, 15)),
    ("ACR", 2024): (date(2024, 11, 14), date(2024, 11, 19)),
    ("ACR", 2025): (date(2025, 11, 13), date(2025, 11, 18)),
    # EULAR (European Alliance of Associations for Rheumatology) — June
    ("EULAR", 2020): (date(2020, 6, 3), date(2020, 6, 6)),
    ("EULAR", 2021): (date(2021, 6, 2), date(2021, 6, 5)),
    ("EULAR", 2022): (date(2022, 6, 1), date(2022, 6, 4)),
    ("EULAR", 2023): (date(2023, 5, 31), date(2023, 6, 3)),
    ("EULAR", 2024): (date(2024, 6, 12), date(2024, 6, 15)),
    ("EULAR", 2025): (date(2025, 6, 4), date(2025, 6, 7)),
    # AASLD (The Liver Meeting) — November
    ("AASLD", 2020): (date(2020, 11, 13), date(2020, 11, 16)),
    ("AASLD", 2021): (date(2021, 11, 12), date(2021, 11, 15)),
    ("AASLD", 2022): (date(2022, 11, 4), date(2022, 11, 8)),
    ("AASLD", 2023): (date(2023, 11, 10), date(2023, 11, 14)),
    ("AASLD", 2024): (date(2024, 11, 15), date(2024, 11, 19)),
    ("AASLD", 2025): (date(2025, 11, 14), date(2025, 11, 18)),
    # DDW (Digestive Disease Week) — May
    ("DDW", 2020): (date(2020, 5, 2), date(2020, 5, 5)),
    ("DDW", 2021): (date(2021, 5, 21), date(2021, 5, 23)),
    ("DDW", 2022): (date(2022, 5, 21), date(2022, 5, 24)),
    ("DDW", 2023): (date(2023, 5, 6), date(2023, 5, 9)),
    ("DDW", 2024): (date(2024, 5, 18), date(2024, 5, 21)),
    ("DDW", 2025): (date(2025, 5, 3), date(2025, 5, 6)),
    # ASCO GI (Gastrointestinal Cancers Symposium) — January
    ("ASCO GI", 2020): (date(2020, 1, 23), date(2020, 1, 25)),
    ("ASCO GI", 2021): (date(2021, 1, 15), date(2021, 1, 17)),
    ("ASCO GI", 2022): (date(2022, 1, 20), date(2022, 1, 22)),
    ("ASCO GI", 2023): (date(2023, 1, 19), date(2023, 1, 21)),
    ("ASCO GI", 2024): (date(2024, 1, 18), date(2024, 1, 20)),
    ("ASCO GI", 2025): (date(2025, 1, 23), date(2025, 1, 25)),
    # ASCO GU (Genitourinary Cancers Symposium) — February
    ("ASCO GU", 2020): (date(2020, 2, 13), date(2020, 2, 15)),
    ("ASCO GU", 2021): (date(2021, 2, 11), date(2021, 2, 13)),
    ("ASCO GU", 2022): (date(2022, 2, 17), date(2022, 2, 19)),
    ("ASCO GU", 2023): (date(2023, 2, 16), date(2023, 2, 18)),
    ("ASCO GU", 2024): (date(2024, 1, 25), date(2024, 1, 27)),
    ("ASCO GU", 2025): (date(2025, 2, 13), date(2025, 2, 15)),
    # JPM (J.P. Morgan Healthcare Conference) — January
    ("JPM", 2020): (date(2020, 1, 13), date(2020, 1, 16)),
    ("JPM", 2021): (date(2021, 1, 11), date(2021, 1, 14)),
    ("JPM", 2022): (date(2022, 1, 10), date(2022, 1, 13)),
    ("JPM", 2023): (date(2023, 1, 9), date(2023, 1, 12)),
    ("JPM", 2024): (date(2024, 1, 8), date(2024, 1, 11)),
    ("JPM", 2025): (date(2025, 1, 13), date(2025, 1, 16)),
}

# Normalize text mentions to canonical calendar keys
_CONF_ALIASES = {
    "J.P. MORGAN": "JPM",
    "JP MORGAN": "JPM",
    "J.P.MORGAN": "JPM",
}

# Ordered list for conference name extraction (multi-word first for greedy match)
_CONF_ACRONYMS = [
    "ASCO GI", "ASCO GU", "ESMO IO",
    "AACR", "ASCO", "ESMO", "SITC", "ASH", "WCLC", "SABCS", "SNO", "EHA",
    "AAN", "CTAD", "AES", "EULAR", "ACR", "AASLD", "DDW",
    "JPM", "J.P. MORGAN", "AHA", "ACC", "ESC", "AAO", "ARVO",
    "ATS", "ERS", "ASHG", "ACMG", "ASN", "IDWeek", "CROI",
]

# Short acronyms (<=3 chars) that match inside common words — need word boundary
_SHORT_CONF = {"ACC", "ACR", "AAN", "AAD", "AES", "ASH", "ASN", "AHA", "AAO",
               "ATS", "ERS", "ESC", "EHA", "DDW", "DIA", "SNO", "UEG"}


def _has_conf_match(text: str) -> bool:
    """Conference keyword matching with word boundaries for short acronyms
    and 'conference call' exclusion."""
    upper = text.upper()
    # Remove "conference call" (earnings calls, not academic conferences)
    cleaned = re.sub(r'CONFERENCE\s+CALL', '', upper)
    for kw in _CONF_KEYWORDS:
        if kw in _SHORT_CONF:
            if re.search(r'\b' + kw + r'\b', cleaned):
                return True
        else:
            if kw in cleaned:
                return True
    return False


def _has_conference_keyword(d: date, event_map: dict) -> bool:
    """
    Return True if any event title on T, T-1, T-2, or T-3 contains a
    conference keyword.  T-3 is always checked (covers Friday filings for
    Monday trading dates, and abstract-release days earlier in the week).
    """
    for offset in range(4):
        title = event_map.get(d - timedelta(days=offset), "")
        if _has_conf_match(title):
            return True
    return False


# ── Event categorization ─────────────────────────────────────────────────────

_CATEGORY_RULES = [
    ("Regulatory",  {"FDA", "EMA", "PDUFA", "NDA", "BLA", "IND ", "BREAKTHROUGH",
                     "FAST TRACK", "PRIORITY REVIEW", "APPROVAL", "CRL",
                     "COMPLETE RESPONSE", "ADVISORY COMMITTEE", "ADCOM"}),
    ("Clinical",    {"PHASE 1", "PHASE 2", "PHASE 3", "PHASE I", "PHASE II", "PHASE III",
                     "CLINICAL TRIAL", "CLINICAL DATA", "TOPLINE", "TOP-LINE",
                     "PRIMARY ENDPOINT", "ORR", "PFS", "OVERALL SURVIVAL",
                     "EFFICACY", "SAFETY DATA", "DOSE ESCALATION", "INTERIM ANALYSIS",
                     "PIVOTAL", "ENROLLED", "ENROLLMENT"}),
    ("Conference",  _CONF_KEYWORDS),
    ("Earnings",    {"QUARTERLY RESULTS", "ANNUAL RESULTS", "FINANCIAL RESULTS",
                     "EARNINGS", "Q1 ", "Q2 ", "Q3 ", "Q4 ",
                     "10-K", "10-Q", "FISCAL YEAR"}),
    ("Financial",   {"OFFERING", "IPO", "ATM", "SHELF REGISTRATION",
                     "SECURITIES SALE", "WARRANT", "EQUITY",
                     "FINANCING", "CAPITAL RAISE"}),
    ("Corporate",   {"APPOINT", "RESIGN", "HIRE", "OFFICER", "DIRECTOR", "CEO",
                     "CFO", "CMO", "CSO", "BOARD", "MANAGEMENT CHANGE",
                     "COLLABORATION", "PARTNERSHIP", "LICENSE AGREEMENT",
                     "ACQUISITION", "MERGER", "RESTRUCTURING", "LAYOFF"}),
    ("Sector",      set()),   # catch-all for sector-driven moves
]


def classify_event(summary: str, d: date, event_map: dict,
                   dod_changes: dict, xbi_dod: dict) -> str:
    """Classify an event into a category based on actual event content.
    Uses raw event text (factual) over AI-generated analysis (speculative)."""
    raw = event_map.get(d, "")
    # Use raw event text for classification (factual content from news/filings)
    # Only fall back to summary if raw event exists and is short (i.e. summary IS the content)
    if raw:
        classify_text = raw.upper()
    else:
        # No raw event — volatile day with AI-generated analysis only.
        upper_s = summary.upper()
        if "UNEXPLAINED" in upper_s:
            return "Other"
        # Only classify as Sector if Claude explicitly attributes to sector
        if upper_s.startswith("BROAD BIOTECH"):
            return "Sector"
        # Fall through to keyword matching on analysis text
        classify_text = upper_s

    for category, keywords in _CATEGORY_RULES:
        if category == "Sector":
            continue
        if category == "Conference":
            if _has_conf_match(classify_text):
                return category
            continue
        if any(kw in classify_text for kw in keywords):
            return category

    # Sector: no company catalyst but significant XBI move
    xbi = xbi_dod.get(d)
    stock = dod_changes.get(d)
    if stock is not None and xbi is not None and abs(xbi) > 0.03:
        return "Sector"

    return "Other"


def _extract_conference_name(text: str, year: int) -> Optional[str]:
    """Extract specific conference acronym from text. Returns e.g. '2023 AACR'."""
    upper = text.upper()
    # Remove "conference call" noise
    cleaned = re.sub(r'CONFERENCE\s+CALL', '', upper)
    for acro in _CONF_ACRONYMS:
        if acro.upper() in _SHORT_CONF:
            if re.search(r'\b' + acro.upper() + r'\b', cleaned):
                return f"{year} {acro}"
        else:
            if acro.upper() in cleaned:
                return f"{year} {acro}"
    for kw in ("CONFERENCE", "SYMPOSIUM", "CONGRESS", "ANNUAL MEETING"):
        if kw in cleaned:
            return f"{year} Conference"
    return None


# ── Conference date remapping helpers ────────────────────────────────────────

def _extract_all_conference_acronyms(text: str) -> list:
    """Extract ALL distinct conference acronyms found in text.
    Returns canonical acronyms matching _CONFERENCE_CALENDAR keys.
    Multi-word acronyms checked first to avoid substring double-counts
    (e.g. 'ASCO GI' consumed before 'ASCO' can match the same region).
    """
    upper = text.upper()
    cleaned = re.sub(r'CONFERENCE\s+CALL', '', upper)
    found: list[str] = []

    # Multi-word first (ASCO GI before ASCO), then single-word
    ordered = sorted(
        [a for a in _CONF_ACRONYMS if " " in a], key=len, reverse=True
    ) + [a for a in _CONF_ACRONYMS if " " not in a]

    for acro in ordered:
        acro_upper = acro.upper()
        canonical = _CONF_ALIASES.get(acro_upper, acro_upper)
        if canonical in found:
            continue

        if acro_upper in _SHORT_CONF:
            if re.search(r'\b' + acro_upper + r'\b', cleaned):
                found.append(canonical)
                # Mask matched region to prevent substring matches
                cleaned = re.sub(r'\b' + acro_upper + r'\b', ' ' * len(acro_upper), cleaned, count=1)
        else:
            if acro_upper in cleaned:
                found.append(canonical)
                cleaned = cleaned.replace(acro_upper, ' ' * len(acro_upper), 1)

    # Check aliases directly (e.g. "J.P. MORGAN" → "JPM")
    for alias, canonical in _CONF_ALIASES.items():
        if canonical not in found and alias in upper:
            found.append(canonical)

    return found


def _find_conference_for_filing(acronym: str, filing_date: date) -> Optional[tuple]:
    """Find the most recent conference whose end_date < filing_date.
    Searches current year and 2 years back. Returns (start, end) or None.
    """
    best = None
    for year_offset in range(3):   # current year, -1, -2
        year = filing_date.year - year_offset
        key = (acronym, year)
        if key not in _CONFERENCE_CALENDAR:
            continue
        start, end = _CONFERENCE_CALENDAR[key]
        if end < filing_date:
            if best is None or end > best[1]:   # most recent
                best = (start, end)
    return best


def _extract_snippet_around_keyword(text: str, keyword: str,
                                    context_chars: int = 150) -> str:
    """Extract ~300 chars centered on keyword occurrence, trimmed to sentence boundaries."""
    upper = text.upper()
    kw_upper = keyword.upper()
    idx = upper.find(kw_upper)
    if idx < 0:
        return ""

    start = max(0, idx - context_chars)
    end = min(len(text), idx + len(keyword) + context_chars)
    snippet = text[start:end]

    # Trim to sentence boundaries
    if start > 0:
        dot_pos = snippet.find(". ")
        if 0 < dot_pos < context_chars // 2:
            snippet = snippet[dot_pos + 2:]
    if end < len(text):
        last_dot = snippet.rfind(". ")
        if last_dot > len(snippet) // 2:
            snippet = snippet[:last_dot + 1]

    return snippet.strip()


def _remap_conference_dates(event_map: dict, date_map: dict) -> dict:
    """Remap retroactive conference mentions to actual conference dates.

    Foreign private issuers mention conferences in quarterly 6-K updates filed
    weeks after the actual conference. This creates synthetic events at the
    conference start date so they appear in the correct time window for
    volatility analysis and conference keyword detection.

    Only processes EDGAR filings (identified by '|' separator in title).
    Original filing-date events are always preserved.
    """
    remapped = 0
    for filing_date in sorted(event_map):
        title = event_map[filing_date]
        # Only process EDGAR filings with body content
        if "|" not in title:
            continue

        acronyms = _extract_all_conference_acronyms(title)
        if not acronyms:
            continue

        for acro in acronyms:
            conf = _find_conference_for_filing(acro, filing_date)
            if conf is None:
                continue
            conf_start, conf_end = conf

            # Only remap if filing is retroactive (after conference ended)
            if filing_date <= conf_end:
                continue

            # Conference start must be within the sheet's date range
            if conf_start not in date_map:
                continue

            # Extract snippet around the conference keyword for context
            snippet = _extract_snippet_around_keyword(title, acro)
            year = conf_start.year
            if snippet:
                synthetic_title = f"{year} {acro}: {snippet}"
            else:
                synthetic_title = (
                    f"{year} {acro} presentation mentioned in quarterly update"
                )

            # Keep longer of existing vs synthetic (never overwrite a better event)
            existing = event_map.get(conf_start, "")
            if len(synthetic_title) > len(existing):
                event_map[conf_start] = synthetic_title
                remapped += 1

    if remapped:
        print(f"    Conference remapping: {remapped} events remapped to actual conference dates")
    return event_map


# ── Claude Haiku summariser ───────────────────────────────────────────────────

def _complete_sentence(text: str, max_words: int) -> str:
    """Truncate to max_words but ensure the result ends at a sentence boundary."""
    words = text.split()
    if len(words) <= max_words:
        return text
    truncated = " ".join(words[:max_words])
    # Common abbreviations that end with period but aren't sentence endings
    _ABBREVS = {"Ltd.", "Inc.", "Corp.", "Co.", "Dr.", "Mr.", "Mrs.", "Ms.",
                "Jr.", "Sr.", "vs.", "etc.", "approx.", "U.S.", "Ph.D.",
                "M.D.", "St.", "No.", "Vol.", "e.g.", "i.e.", "al."}
    # Find last sentence-ending punctuation (skip abbreviation periods)
    for i in range(len(truncated) - 1, max(len(truncated) // 3, 0), -1):
        if truncated[i] in ".;!?":
            candidate = truncated[:i + 1]
            # Check if this period belongs to an abbreviation
            last_word = candidate.rsplit(None, 1)[-1] if candidate else ""
            if last_word in _ABBREVS:
                continue
            return candidate
    # No good sentence boundary found — return full truncation
    return truncated


def _is_generic_label(title: str) -> bool:
    """Detect generic fallback labels that Claude cannot meaningfully summarize."""
    if "|" in title:
        return False   # Has body content from press release extraction
    if len(title) < 60:
        for label in _ITEM_LABELS.values():
            if label in title:
                return True
        if title.endswith("8-K Filing") or title.endswith("6-K Report"):
            return True
    return False


def summarize(title: str, client: Optional[anthropic.Anthropic]) -> str:
    """Summarize a press-release title/content in ≤30 words using Claude Haiku."""
    # Skip Claude for generic labels — they contain no useful content
    if _is_generic_label(title):
        return _complete_sentence(title, 30)

    # Detect conference content for structured format
    has_conf = _has_conf_match(title)

    if has_conf:
        prompt = (
            "You are a biotech analyst. Summarize this conference/presentation news "
            "in 30 words or fewer using this format when applicable:\n"
            '"[Conference] [Format] ([Stage]): [Key Metrics]."\n'
            "Format: Oral/Poster/Abstract/Late-breaking (LBA). "
            "Stage: Preclinical/Initial/Primary/Updated. "
            "Include ORR/PFS/OS/safety if present. "
            "Write complete sentences. Return only the summary, no quotes:\n\n"
            + title
        )
    elif "|" in title:
        prompt = (
            "You are a biotech analyst. Summarize the following press release content "
            "in 30 words or fewer. Be factual and specific — include drug name, "
            "indication, and outcome if present. Write complete sentences. "
            "Return only the summary, no quotes:\n\n"
            + title
        )
    else:
        prompt = (
            "You are a biotech analyst. Summarize the following press release headline "
            "in 30 words or fewer. Be factual and specific — include drug name, "
            "indication, and outcome if present. Write complete sentences. "
            "Return only the summary, no quotes:\n\n"
            + title
        )

    if client is None:
        words = title.split("|")[0].split() if "|" in title else title.split()
        return _complete_sentence(" ".join(words), 30)

    try:
        msg = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=120,
            messages=[{"role": "user", "content": prompt}],
        )
        summary = msg.content[0].text.strip().strip('"').strip("'")
        # Reject unhelpful Claude responses (e.g. "I don't see a headline")
        _REJECT = {"I DON'T SEE", "I CANNOT", "I DON'T HAVE", "NO PRESS RELEASE",
                    "NOT ENOUGH INFORMATION", "APPEARS TO BE CORRUPTED"}
        if any(r in summary.upper() for r in _REJECT):
            words = title.split("|")[0].split() if "|" in title else title.split()
            return _complete_sentence(" ".join(words), 30)
        return _complete_sentence(summary, 30)
    except Exception:
        # Graceful fallback: truncate headline part only
        words = title.split("|")[0].split() if "|" in title else title.split()
        return _complete_sentence(" ".join(words), 30)


# ── Volatility Catalyst Analysis ─────────────────────────────────────────────

def compute_dod_changes(trading_prices: dict) -> dict:
    """
    Return {date: pct_change_as_fraction} for every actual trading day.
    e.g. +0.12 means +12%, -0.15 means -15%.
    Gap > 4 calendar days (e.g. trading halts) is excluded.
    """
    dod: dict[date, float] = {}
    dates = sorted(trading_prices)
    for i in range(1, len(dates)):
        t0, t1 = dates[i - 1], dates[i]
        p0 = trading_prices[t0]
        if (t1 - t0).days <= 4 and p0 > 0:
            dod[t1] = (trading_prices[t1] - p0) / p0
    return dod


def fetch_xbi_changes(dates: list) -> dict:
    """Fetch XBI ETF daily % changes (as fractions) for a list of dates."""
    if not dates:
        return {}
    start = min(dates) - timedelta(days=7)
    end   = max(dates) + timedelta(days=2)
    try:
        hist = yf.Ticker("XBI").history(
            start=start.isoformat(), end=end.isoformat(), auto_adjust=True
        )
        xbi_raw: dict[date, float] = {ts.date(): float(row["Close"]) for ts, row in hist.iterrows()}
        xbi_dates = sorted(xbi_raw)
        xbi_dod: dict[date, float] = {}
        for i in range(1, len(xbi_dates)):
            t0, t1 = xbi_dates[i - 1], xbi_dates[i]
            p0 = xbi_raw[t0]
            if (t1 - t0).days <= 4 and p0 > 0:
                xbi_dod[t1] = (xbi_raw[t1] - p0) / p0
        return xbi_dod
    except Exception as e:
        print(f"    [XBI fetch] {e}")
        return {}


def analyze_volatile_catalyst(
    d: date,
    stock_pct: float,
    event_map: dict,
    xbi_dod: dict,
    client,          # anthropic.Anthropic | None
) -> str:
    """
    Call Claude to explain the catalyst for a high-volatility day (>10% move).
    Uses dates T and T-1 (also T-2 to cover weekends before Monday moves).
    Returns a ≤65-word explanation with structured conference data when applicable.
    """
    # Company news for T, T-1, T-2, and T-3 on Mondays (catches Friday conference filings)
    lookback = 4 if d.weekday() == 0 else 3
    news_items = []
    for offset in range(lookback):
        check = d - timedelta(days=offset)
        if check in event_map:
            news_items.append(f"{check}: {event_map[check]}")
    news_context = "; ".join(news_items) if news_items else "None"

    xbi_val = xbi_dod.get(d)
    xbi_str = f"{xbi_val * 100:+.1f}" if xbi_val is not None else "N/A"

    prompt = (
        f"You are an expert Biotech Equity Research Analyst. On {d}, the stock moved by {stock_pct:+.1f}%.\n\n"
        f"Context for this date and the day prior:\n\n"
        f"Company News/Filings: {news_context}\n\n"
        f"Biotech Sector Trend (XBI ETF change): {xbi_str}%\n\n"
        f"Your Task: Explain the core catalyst for this price movement in under 65 words. "
        f"Write complete sentences.\n\n"
        f"Rules:\n\n"
        f"CONFERENCE DATA EXTRACTION (CRITICAL): If the news mentions major medical conferences "
        f"(e.g., AACR, ASCO, ESMO, SITC, ASH, WCLC, AAN, JPM, SABCS, SNO, EHA, ACR, EULAR, AASLD, DDW), "
        f"you MUST structure your response using this dense format:\n"
        f'"[Conference] [Format] ([Stage]): [Key Metrics]."\n\n'
        f"Format: Specify if it is an Oral Presentation, Poster, Abstract, or Late-breaking (LBA).\n\n"
        f"Stage: Specify if data is Preclinical, Initial, Primary, or Updated.\n\n"
        f"Key Metrics: Extract specific numbers for efficacy (e.g., ORR, CR, PR, PFS, OS) and briefly note "
        f"side effects/safety if mentioned.\n"
        f'(Example: "ESMO Late-breaking Oral (Primary): 45% ORR, 8.2m mPFS in solid tumors; severe neutropenia '
        f'in 10%. Drove +15% stock surge.")\n\n'
        f"If it's a regulatory event (PDUFA, IND) or financial event (Public Offering), state it clearly.\n\n"
        f"If there is NO company-specific news, attribute the movement to the sector trend: "
        f"'Broad biotech sector movement ({xbi_str}% XBI).'\n\n"
        f"DO NOT hallucinate. If no news and no sector correlation exist, output exactly: "
        f"'Unexplained volatility; no material company catalyst found.'"
    )
    if client is None:
        # Rule-based fallback (no API key): use available context directly
        if news_items:
            raw = news_items[0].split(": ", 1)[-1]   # strip the date prefix
            return _complete_sentence(raw, 30)
        elif xbi_val is not None and abs(xbi_val) > 0.03:
            direction = "up" if xbi_val > 0 else "down"
            return f"Broad biotech sector move; XBI {direction} {abs(xbi_val)*100:.1f}% same day."
        return "Unexplained volatility; no material company catalyst found."
    try:
        msg = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=220,
            messages=[{"role": "user", "content": prompt}],
        )
        text = msg.content[0].text.strip().strip('"').strip("'")
        return _complete_sentence(text, 65)
    except Exception as e:
        print(f"    [Claude API error] {e}")
        return "Unexplained volatility; no material company catalyst found."


# ── Main orchestrator ─────────────────────────────────────────────────────────

def fill_historical_events(ticker: str, force: bool = False) -> None:
    """
    force=True  → overwrite EVT cells even if they already contain text
                  (useful to refresh with better summaries)
    """
    ticker = ticker.upper()
    sep = "=" * 60
    print(f"\n{sep}")
    print(f"  Fill for ticker: {ticker}" + (" [FORCE mode]" if force else ""))
    print(f"{sep}\n")

    # ── 1. Find the DCF file ──────────────────────────────────────────────────
    print("[1] Locating DCF file…")
    dcf_path = find_dcf_file(ticker)
    print(f"    {dcf_path}\n")

    # ── 2. Load workbook (two passes: read-only for value inspection, full for write) ──
    print("[2] Loading workbook…")
    wb_read  = openpyxl.load_workbook(dcf_path, data_only=True)   # cached values
    wb_write = openpyxl.load_workbook(dcf_path, data_only=False)  # formulas preserved
    ws_read  = wb_read["Historical Events"]
    ws_write = wb_write["Historical Events"]
    # Ensure no gridlines on all sheets
    for ws_name in wb_write.sheetnames:
        wb_write[ws_name].sheet_view.showGridLines = False
    print(f"    Sheet 'Historical Events' loaded (gridlines off).\n")

    # ── 3. Parse all dates in the sheet ──────────────────────────────────────
    print("[3] Parsing dates from sheet…")
    date_map   = load_date_map(ws_read)   # {date: (row, block_idx)}
    all_dates  = sorted(date_map.keys())
    print(f"    {len(all_dates)} dates found: {all_dates[0]} → {all_dates[-1]}\n")

    # ── 4. Fetch stock prices ─────────────────────────────────────────────────
    print("[4] Fetching historical stock prices (Yahoo Finance)…")
    prices, trading_prices = fetch_prices(ticker, all_dates)
    print(f"    {len(prices)} price data points retrieved.\n")

    # ── 5. Fetch press releases / news events ─────────────────────────────────
    print("[5] Searching for press releases and events…")
    company_name = _get_company_name(ticker)
    print(f"    Company name: {company_name or '(unknown)'}")
    raw_events: list[tuple[date, str]] = []
    if company_name:
        raw_events += fetch_ir_rss(ticker, company_name=company_name)
    raw_events += fetch_yahoo_news(ticker)
    raw_events += fetch_globenewswire(ticker)
    raw_events += fetch_edgar_8k(ticker)
    raw_events += fetch_google_news(ticker, company_name=company_name)
    raw_events += fetch_edgar_efts_conferences(ticker, company_name=company_name)
    raw_events += fetch_edgar_efts_conference_names(ticker, company_name=company_name)
    if company_name:
        raw_events += fetch_google_news_conferences(ticker, company_name)

    event_map = build_event_map(raw_events)
    print(f"    {len(event_map)} unique event dates after deduplication.\n")

    # ── 5a. Remap retroactive conference mentions to actual dates ─────────
    print("[5a] Conference date remapping…")
    event_map = _remap_conference_dates(event_map, date_map)
    print(f"    {len(event_map)} unique event dates after remapping.\n")

    # ── 5b. Identify deep-analysis dates: |DoD %| > 10% OR conference keyword ──
    print("[5b] Identifying high-volatility and conference days…")
    dod_changes = compute_dod_changes(trading_prices)
    volatile_dates = sorted(
        d for d in date_map
        if (d in dod_changes and abs(dod_changes[d]) > 0.10)
        or _has_conference_keyword(d, event_map)
    )
    n_vol  = sum(1 for d in volatile_dates if d in dod_changes and abs(dod_changes[d]) > 0.10)
    n_conf = sum(1 for d in volatile_dates if _has_conference_keyword(d, event_map))
    print(f"    {len(volatile_dates)} dates total: {n_vol} high-volatility, {n_conf} conference-keyword")

    # ── 5c. Fetch XBI sector benchmark for volatile dates ────────────────────
    print("[5c] Fetching XBI ETF changes for volatile dates…")
    xbi_dod = fetch_xbi_changes(volatile_dates)
    print(f"    XBI data for {len(xbi_dod)} dates\n")

    # ── 6. Summarize routine event dates with Claude Haiku ────────────────────
    print("[6] Summarizing press-release dates with Claude Haiku…")
    api_key = os.environ.get("MY_PYTHON_SCRIPT_KEY")
    if api_key:
        client = anthropic.Anthropic(api_key=api_key)
        print("    API key found — AI summaries enabled.")
    else:
        client = None
        print("    [WARNING] MY_PYTHON_SCRIPT_KEY not set.")
        print("    Add MY_PYTHON_SCRIPT_KEY=sk-ant-... to your .env file for AI summaries.")
        print("    Falling back to title truncation.\n")
    summaries: dict[date, str] = {}
    relevant = {d: t for d, t in event_map.items() if d in date_map}
    for d, title in sorted(relevant.items()):
        s = summarize(title, client)
        summaries[d] = s
        print(f"    {d}: {s}")
    print(f"    {len(summaries)} summaries generated.\n")

    # ── 6b. Deep-retrieve & AI-analyze volatile + conference days ─────────────
    print("[6b] Analyzing volatile/conference days with Claude…")
    volatile_summaries: dict[date, str] = {}
    for d in volatile_dates:
        stock_pct = dod_changes.get(d, 0.0) * 100   # may be <10% for conf-only dates
        analysis  = analyze_volatile_catalyst(d, stock_pct, event_map, xbi_dod, client)
        volatile_summaries[d] = analysis
        print(f"    {d} ({stock_pct:+.1f}%): {analysis}")
    print(f"    {len(volatile_summaries)} analyses generated.\n")

    # Volatile analyses take priority over simple summaries for their dates
    all_summaries = {**summaries, **volatile_summaries}

    # ── 7. Write values to sheet ──────────────────────────────────────────────
    print("[7] Writing data to 'Historical Events' sheet…")
    filled_price = 0
    filled_evt   = 0

    filled_cat   = 0

    for d, (row, bi) in date_map.items():
        _label, dcol, pcol, _dod, ecol, cat_col = BLOCKS[bi]

        # Share Price — always write from Yahoo Finance (Bloomberg values can be wrong/stale)
        if d in prices:
            ws_write.cell(row=row, column=pcol).value = prices[d]
            filled_price += 1

        # EVT — overwrite blank cells (or all cells if --force)
        cur_evt = ws_read.cell(row=row, column=ecol).value
        if (is_blank(cur_evt) or force) and d in all_summaries:
            ws_write.cell(row=row, column=ecol).value = all_summaries[d]
            filled_evt += 1

        # Category — write for dates that have events
        if d in all_summaries:
            cat = classify_event(all_summaries[d], d, event_map, dod_changes, xbi_dod)
            ws_write.cell(row=row, column=cat_col).value = cat
            filled_cat += 1

    print(f"    Share Price cells filled : {filled_price}")
    print(f"    EVT cells filled         : {filled_evt}")
    print(f"    Category cells filled    : {filled_cat}\n")

    # ── 8. Save ───────────────────────────────────────────────────────────────
    print("[8] Saving workbook…")
    wb_write.save(dcf_path)
    wb_read.close()
    wb_write.close()
    print(f"    Saved → {dcf_path}")

    # ── 9. Print categorized summary ──────────────────────────────────────
    cat_counts: dict[str, int] = {}
    conf_details: list[str] = []

    for d, (row, bi) in date_map.items():
        if d not in all_summaries:
            continue
        cat = classify_event(all_summaries[d], d, event_map, dod_changes, xbi_dod)
        cat_counts[cat] = cat_counts.get(cat, 0) + 1
        if cat == "Conference":
            year_label = BLOCKS[bi][0]   # "FA 2023"
            year_int = int(year_label.split()[-1])
            detail = _extract_conference_name(
                all_summaries[d] + " " + event_map.get(d, ""), year_int
            )
            if detail and detail not in conf_details:
                conf_details.append(detail)

    print(f"\n{'═' * 70}")
    print(f"  {ticker} Historical Events — Categorized Summary")
    print(f"{'═' * 70}")
    print(f"  {'Category':<20} {'Count':>8}  Details")
    print(f"  {'-' * 20} {'-' * 8}  {'-' * 38}")

    for cat_name in ["Regulatory", "Clinical", "Conference", "Earnings",
                     "Financial", "Corporate", "Sector", "Other"]:
        count = cat_counts.get(cat_name, 0)
        if count == 0:
            continue
        detail_str = ""
        if cat_name == "Conference" and conf_details:
            detail_str = ", ".join(sorted(set(conf_details)))
        print(f"  {cat_name:<20} {count:>8}  {detail_str}")

    total_events = sum(cat_counts.values())
    print(f"  {'-' * 20} {'-' * 8}")
    print(f"  {'TOTAL':<20} {total_events:>8}")

    n_prices = sum(1 for d in date_map if d in prices)
    n_volatile = sum(1 for d in volatile_dates if d in dod_changes and abs(dod_changes[d]) > 0.10)
    n_trading = sum(1 for d in date_map if d in trading_prices)
    print(f"\n  Price data points: {n_prices}")
    print(f"  Volatile moves (>10%): {n_volatile}")
    print(f"  Trading days: {n_trading}")
    print(f"{'═' * 70}")
    print(f"\n{sep}")
    print(f"  Done! {ticker} Historical Events updated successfully.")
    print(f"{sep}\n")


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    args = sys.argv[1:]
    force = "--force" in args
    args  = [a for a in args if a != "--force"]
    raw   = " ".join(args)

    # Accept both   "Fill for ticker: CMPX"   and   "CMPX"
    m = re.search(r"Fill for ticker[:\s]+([A-Z]{1,6})", raw, re.IGNORECASE)
    ticker = (m.group(1) if m else raw.strip()).upper()

    fill_historical_events(ticker, force=force)


if __name__ == "__main__":
    main()
