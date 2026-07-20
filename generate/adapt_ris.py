#!/usr/bin/env python3
"""
adapt_ris.py — adapt the RIS (Restated Income Statement) to the ticker's pipeline.

The DCF template's RIS is hard-wired to the template company's (BCYC) drugs and
R&D/G&A note decomposition. For any other ticker that leaves:
  • revenue rows SUMIFS-ing on BCYC drug names  → 0 (no drug revenue flows), and
  • the R&D total (=SUMIF of BCYC-labeled ISN note rows) → 0, so ratio cells like
    `=DefRev/R&D` throw #DIV/0! and cascade through RCFS → VALUATION → Catalyst.

This step (previously a manual edit — see CMPX) automates the adaptation:
  1. Relabel RIS drug rows (revenue timeline R8-11, sales R26-29, COGS R37-40)
     to THIS ticker's assets, read live from the Scenarios sheet; unused slots
     become "Reserved" (their SUMIFS then match nothing → 0).
  2. Override the HISTORICAL columns of the R&D total (R46) and G&A total (R61)
     to pull the IS-level figure straight from FY DATA, so both are correct and
     non-zero (killing the div/0). Forecast note rows are handled later by
     tools/fix_statement_logic.py, which links R&D/G&A to Pipeline Timeline.

    python generate/adapt_ris.py --ticker MOLN

Idempotent. Surgical XML patching (never openpyxl.save) — styles/formulas kept.
"""
import argparse
import re
import shutil
import time
import warnings
import zipfile
from pathlib import Path

warnings.filterwarnings("ignore")
import openpyxl.drawing.text as _t  # noqa: E402
_t.Font.pitchFamily.max = 127
import openpyxl  # noqa: E402
from openpyxl.utils import get_column_letter as cl  # noqa: E402
from openpyxl.utils import column_index_from_string as ci  # noqa: E402

# RIS row layout (drug slots) and the historical year columns to fix.
TIMELINE_ROWS = [8, 9, 10, 11]
REVENUE_ROWS = [26, 27, 28, 29]
COGS_ROWS = [37, 38, 39, 40]
RD_TOTAL_ROW = 46
GA_TOTAL_ROW = 61
TOTAL_OPEX_ROW = 71     # RIS Total Operating Expenses
REVENUE_ROW = 30        # RIS Total Revenue From Sales (product)
YEAR_COLS = 18                           # RBS grid width (cols F..W)
RIS_LAST_COL = "X"                      # RIS F=2020 through X=2038
# NB: the historical FY columns (hist_cols), the last actual fiscal year
# (last_actual) and the model horizon (TERMINAL_YEAR = rbs_base + YEAR_COLS - 1)
# are all derived per-ticker in main() from the FY DATA base year and the
# populated year columns, not hardcoded — see the FY DATA R&D-row scan below.
FYD = "'FY DATA'"


def _cash_plug_formula(col: str, cash_row: int,
                       financial_asset_rows: list[int]) -> str:
    """Balance cash after preserving any non-cash financial assets.

    RBS row 66 is total equity and row 63 is NOA.  Older logic assumed cash
    was the only financial asset, which double-counted marketable securities
    for issuers such as TARA.  The row inputs are discovered from the workbook
    rather than template row numbers.
    """
    other_refs = [f"{col}{row}" for row in financial_asset_rows if row != cash_row]
    formula = f"{col}66-{col}63"
    if other_refs:
        formula += f"-SUM({','.join(other_refs)})"
    return formula


def xesc(s):
    return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def sheet_zip_path(zf, name):
    wbx = zf.read("xl/workbook.xml").decode("utf8", "ignore")
    m = re.search(rf'<sheet name="{re.escape(name)}"[^>]*r:id="(rId\d+)"', wbx)
    rid = m.group(1)
    rels = zf.read("xl/_rels/workbook.xml.rels").decode("utf8", "ignore")
    tgt = re.search(rf'Id="{rid}"[^>]*Target="([^"]+)"', rels).group(1).lstrip("/")
    return tgt if tgt.startswith("xl/") else "xl/" + tgt


def _is_absolute_asset_row(col_a, col_b) -> bool:
    """True only for structural asset rows in the Scenarios ' Absolute' module.

    Those rows carry col A == 4 (scenario 4) and col B == ' Absolute'
    (generate_scenarios.py).  Restricting to them keeps catalyst scenario
    headers ('{event} (+)'/'(-)') and paren-containing breakdown names
    ('{drug} {ind} Cumulative') — which live on scenario-header rows whose
    col A is empty and col B is a numeric scenario id — out of the asset list.
    """
    if col_a == 4:
        return True
    return isinstance(col_b, str) and col_b.strip() == "Absolute"


def read_assets(dcf: Path):
    """Distinct drug-asset labels from the Scenarios ' Absolute' module, in
    first-seen order."""
    wb = openpyxl.load_workbook(dcf, data_only=False, read_only=True)
    sc = wb["Scenarios"]
    seen, assets = set(), []
    for row in sc.iter_rows(min_col=1, max_col=3):
        col_a, col_b, v = row[0].value, row[1].value, row[2].value
        if not _is_absolute_asset_row(col_a, col_b):
            continue
        if isinstance(v, str) and "(" in v and not v.startswith("=") \
                and "Market Share" not in v and "TAM" not in v:
            if v not in seen:
                seen.add(v)
                assets.append(v)
    wb.close()
    return assets


def patch_text(xml, addr, text):
    style = ""
    m = re.search(rf'<c r="{addr}"([^>]*?)(?:/>|>.*?</c>)', xml, re.S)
    if m:
        sm = re.search(r'\ss="(\d+)"', m.group(1))
        style = f' s="{sm.group(1)}"' if sm else ""
    cell = f'<c r="{addr}"{style} t="inlineStr"><is><t xml:space="preserve">{xesc(text)}</t></is></c>'
    new, n = re.subn(rf'<c r="{addr}"[^>]*?(?:/>|>.*?</c>)', cell, xml, flags=re.S)
    return new if n else xml


def patch_formula(xml, addr, formula):
    style = ""
    m = re.search(rf'<c r="{addr}"([^>]*?)(?:/>|>.*?</c>)', xml, re.S)
    if m:
        sm = re.search(r'\ss="(\d+)"', m.group(1))
        style = f' s="{sm.group(1)}"' if sm else ""
    cell = f'<c r="{addr}"{style}><f>{xesc(formula)}</f></c>'
    new, n = re.subn(rf'<c r="{addr}"[^>]*?(?:/>|>.*?</c>)', cell, xml, flags=re.S)
    return new if n else xml


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--ticker")
    ap.add_argument("--path")
    a = ap.parse_args()
    dcf = Path(a.path) if a.path else Path(
        f"/mnt/c/Users/yzsun/Desktop/DD/{a.ticker}/DCF {a.ticker}.xlsx")

    assets = read_assets(dcf)
    if not assets:
        raise SystemExit("No drug assets found in Scenarios — run generate_scenarios first.")
    print(f"Assets from Scenarios ({len(assets)}):")
    for x in assets:
        print(f"  • {x[:70]}")

    # The RIS template has four visible drug slots.  For companies with more
    # than four clinical assets, keep the first three as visible rows and use
    # the fourth row as an "Other Pipeline" dynamic aggregate so no Pipeline
    # revenue/COGS is dropped from valuation.
    n_slots = len(TIMELINE_ROWS)
    aggregate_tail = len(assets) > n_slots
    visible_assets = assets[: n_slots - 1] if aggregate_tail else assets[:n_slots]
    tail_assets = assets[n_slots - 1:] if aggregate_tail else []

    if aggregate_tail:
        print(
            f"  Aggregating {len(tail_assets)} assets into RIS row {TIMELINE_ROWS[-1]} "
            "so all Pipeline revenue/COGS remains modeled: "
            + ", ".join(a[:30] for a in tail_assets)
        )

    # Build the label patches (fill drug slots, pad with Reserved/aggregate).
    labels = {}
    for slots, suffix in [(TIMELINE_ROWS, ""), (REVENUE_ROWS, " Revenue"), (COGS_ROWS, " COGS")]:
        for i, r in enumerate(slots):
            if aggregate_tail and i == n_slots - 1:
                if suffix == "":
                    labels[f"D{r}"] = "Other Pipeline Active Stage Load"
                else:
                    labels[f"D{r}"] = "Other Pipeline" + suffix
            elif i < len(visible_assets):
                labels[f"D{r}"] = visible_assets[i] + suffix
            else:
                labels[f"D{r}"] = "Reserved"

    # RBS: the restated PP&E (R11) is pulled from the Schedules rollforward,
    # whose ratio cells 0/0 for a pre-revenue company; repoint the HISTORICAL
    # columns straight to FY DATA so Net Operating Assets — and the accounting-
    # identity check (R47/R67) — balance.
    #   NB: RBS and FY DATA use DIFFERENT base years (RBS starts one year later),
    #   so RBS column X must read FY DATA column X+offset, not the same letter.
    wb = openpyxl.load_workbook(dcf, data_only=False, read_only=True)
    rbs_base = int(wb["RBS"]["F4"].value)
    fyd_base = int(wb["FY DATA K USD"]["F4"].value)
    # Derive the last ACTUAL fiscal year — the last FY DATA year column that
    # actually carries financials — instead of hardcoding it. Scan the IS R&D row
    # (present for every biotech) across the year columns F..K.
    kws = wb["FY DATA K USD"]
    rd_row = next((r for r in range(1, kws.max_row + 1)
                   if str(kws.cell(r, 3).value) == "IS"
                   and "research and development" in str(kws.cell(r, 4).value or "").lower()), None)
    last_actual = fyd_base
    if rd_row:
        for i in range(6):                       # FY DATA year columns F..K
            v = kws.cell(rd_row, 6 + i).value
            if isinstance(v, (int, float)) and v:
                last_actual = fyd_base + i
    # Locate the RBS "Cash And Cash Equivalents" row by label rather than
    # assuming row 33: when the balance sheet carries marketable securities
    # (classified as a Financial Asset, listed above cash) the cash row shifts
    # up, and the cash plug / VALUATION bridge must follow it.  Falls back to 33.
    rbs_ws = wb["RBS"]
    cash_row = 33
    financial_assets_header = None
    total_financial_assets_row = None
    for r in range(1, (rbs_ws.max_row or 0) + 1):
        label = str(rbs_ws.cell(r, 4).value or "").strip()
        if label == "Financial Assets":
            financial_assets_header = r
        elif (label == "Total Financial Assets"
              and financial_assets_header is not None
              and total_financial_assets_row is None):
            total_financial_assets_row = r
        elif label == "Cash And Cash Equivalents":
            cash_row = r
    if financial_assets_header is not None and total_financial_assets_row is not None:
        financial_asset_rows = [
            r for r in range(financial_assets_header + 1, total_financial_assets_row)
            if str(rbs_ws.cell(r, 3).value or "").strip() == "BS"
            and str(rbs_ws.cell(r, 4).value or "").strip()
        ]
    else:
        financial_asset_rows = [cash_row]
    if cash_row not in financial_asset_rows:
        financial_asset_rows.append(cash_row)
    wb.close()
    ris_base = fyd_base          # RIS column F is the same base year as FY DATA
    offset = rbs_base - fyd_base
    TERMINAL_YEAR = rbs_base + YEAR_COLS - 1   # model horizon = last grid col (W)

    # RIS opex overrides (historical columns only): pull the IS-level total
    # straight from FY DATA so R&D/G&A are correct and non-zero (kills div/0).
    # hist_cols = the populated FY DATA year columns (F.. up to last_actual),
    # derived from the data rather than a fixed 5-column assumption so tickers
    # with fewer/more reported years don't clobber forecast cols or miss cols.
    hist_cols = [cl(6 + i) for i in range(last_actual - fyd_base + 1)]
    ris_formulas = {}
    # Pipeline uses F=2010 while RIS/Scenarios use the fiscal/model columns
    # directly.  Existing template formulas map RIS F -> Pipeline P, i.e. +10
    # columns.  Keep that convention for dynamic revenue/COGS formulas.
    def pipeline_col_for_ris_col(col: str) -> str:
        return cl(ci(col) + 10)

    def scen_sumifs(asset_label: str, col: str) -> str:
        return (
            f'SUMIFS(Scenarios!{col}:{col},Scenarios!$C:$C,'
            f'"{asset_label}",Scenarios!$A:$A,$E$2)'
        )

    # RIS starts one year earlier than RBS, so its 2038 terminal column is X.
    # Stopping at W left X carrying stale template revenue/COGS formulas and
    # made terminal free cash flow collapse despite a valid 2038 Pipeline.
    for col_idx in range(ci("F"), ci(RIS_LAST_COL) + 1):
        col = cl(col_idx)
        pcol = pipeline_col_for_ris_col(col)
        # Stage timeline rows.
        for i, r in enumerate(TIMELINE_ROWS):
            if aggregate_tail and i == n_slots - 1:
                terms = [
                    f'IF({scen_sumifs(asset, col)}<5,{scen_sumifs(asset, col)},0)'
                    for asset in tail_assets
                ]
                ris_formulas[f"{col}{r}"] = "SUM(" + ",".join(terms) + ")" if terms else "0"
            elif i < len(visible_assets):
                ris_formulas[f"{col}{r}"] = (
                    f'SUMIFS(Scenarios!{col}:{col},Scenarios!$C:$C,$D{r},'
                    f'Scenarios!$A:$A,$E$2)'
                )
            else:
                ris_formulas[f"{col}{r}"] = (
                    f'SUMIFS(Scenarios!{col}:{col},Scenarios!$C:$C,$D{r},'
                    f'Scenarios!$A:$A,$E$2)'
                )

        # Revenue and COGS rows.
        for slots, suffix in [(REVENUE_ROWS, " Revenue"), (COGS_ROWS, " COGS")]:
            for i, r in enumerate(slots):
                if aggregate_tail and i == n_slots - 1:
                    prev_refs = ",".join(f"{col}{rr}" for rr in slots[: n_slots - 1])
                    ris_formulas[f"{col}{r}"] = (
                        f'SUMIFS(Pipeline!{pcol}:{pcol},Pipeline!$D:$D,"*{suffix}")'
                        f'-SUM({prev_refs})'
                    )
                else:
                    ris_formulas[f"{col}{r}"] = (
                        f'SUMIFS(Pipeline!{pcol}:{pcol},Pipeline!$D:$D,$D{r})'
                    )

    for col in hist_cols:
        ris_formulas[f"{col}{RD_TOTAL_ROW}"] = (
            f'SUMIFS({FYD}!{col}:{col},{FYD}!$D:$D,"Research And Development",{FYD}!$C:$C,"IS")')
        ris_formulas[f"{col}{GA_TOTAL_ROW}"] = (
            f'SUMIFS({FYD}!{col}:{col},{FYD}!$D:$D,"General And Administrative",{FYD}!$C:$C,"IS")')

    def rcol(base, year):
        return cl(ci("F") + (year - base))

    # Forecast R&D/G&A is intentionally not set here.  The later statementlogic
    # step writes visible note-level rows and stage-linked forecast formulas.

    # RBS PP&E (R11): historical → FY DATA (offset-corrected column); forecast →
    # roll flat from the prior column (the Schedules rollforward 0/0's for a
    # pre-revenue company). Together this balances the accounting-identity check
    # (R47/R67) across every year.
    rbs_formulas = {}
    for y in range(rbs_base, last_actual + 1):    # historical
        c = rcol(rbs_base, y)
        fcol = cl(ci(c) + offset)
        rbs_formulas[f"{c}11"] = (
            f'SUMIFS({FYD}!{fcol}:{fcol},{FYD}!$D:$D,"Property And Equipment, Net",{FYD}!$C:$C,"BS")')
    for y in range(last_actual + 1, TERMINAL_YEAR + 1):   # forecast: roll flat
        c = ci(rcol(rbs_base, y))
        rbs_formulas[f"{cl(c)}11"] = f"{cl(c - 1)}11"

    # RBS forecast cash plug: RCFS ending cash can drift from the independently
    # forecast operating assets/liabilities and equity rows.  Use cash as the
    # balance-sheet plug AFTER preserving other financial assets (e.g.
    # marketable securities), so RBS R47/R67 remain true in every forecast year.
    for y in range(last_actual + 1, TERMINAL_YEAR + 1):
        c = rcol(rbs_base, y)
        rbs_formulas[f"{c}{cash_row}"] = _cash_plug_formula(
            c, cash_row, financial_asset_rows
        )

    bak = dcf.with_name(f"{dcf.stem}_pre_adaptris_{time.strftime('%Y%m%d_%H%M%S')}.xlsx")
    shutil.copy2(dcf, bak)

    # VALUATION cash bridge: the template pulls RBS cash for the valuation-date
    # YEAR, which for a pre-revenue name is a projected (often negative) cash
    # balance — double-counting the burn already in the discounted FCF. Use the
    # last REPORTED net cash instead (standard: EV + current net cash).
    cash_col = cl(ci("F") + (last_actual - rbs_base))     # RBS column for 2024
    val_formulas = {
        "C45": f"RBS!{cash_col}{cash_row}",
        # Shares Outstanding: the template's per-year share cell still holds the
        # template company's count; use the live figure filled into BBG DAPI
        # (row 12 = "Shares Outstanding", col I = FINAL).
        "C47": "'BBG DAPI'!$I$12",
    }

    with zipfile.ZipFile(dcf) as zf:
        ris_path = sheet_zip_path(zf, "RIS")
        rbs_path = sheet_zip_path(zf, "RBS")
        val_path = sheet_zip_path(zf, "VALUATION")
        blobs = {n: zf.read(n) for n in zf.namelist()}

    xml = blobs[ris_path].decode("utf8", "ignore")
    for addr, text in labels.items():
        xml = patch_text(xml, addr, text)
    for addr, f in ris_formulas.items():
        xml = patch_formula(xml, addr, f)
    blobs[ris_path] = xml.encode("utf8")

    rxml = blobs[rbs_path].decode("utf8", "ignore")
    for addr, f in rbs_formulas.items():
        rxml = patch_formula(rxml, addr, f)
    blobs[rbs_path] = rxml.encode("utf8")

    vxml = blobs[val_path].decode("utf8", "ignore")
    for addr, f in val_formulas.items():
        vxml = patch_formula(vxml, addr, f)
    blobs[val_path] = vxml.encode("utf8")

    with zipfile.ZipFile(dcf, "w", zipfile.ZIP_DEFLATED) as zo:
        for n, b in blobs.items():
            zo.writestr(n, b)

    formulas = ris_formulas
    print(f"\nRIS adapted ({len(labels)} labels, {len(formulas)} formula cells; +{len(rbs_formulas)} RBS PP&E cells):")
    for r, base in [(8, "timeline"), (26, "revenue"), (37, "COGS")]:
        print(f"  R{r}+ {base}: " + " | ".join(labels[f'D{r+i}'][:22] for i in range(len(TIMELINE_ROWS))))
    print(f"  R{RD_TOTAL_ROW} R&D total + R{GA_TOTAL_ROW} G&A total → direct FY DATA IS pull (hist cols {hist_cols[0]}-{hist_cols[-1]})")
    print(f"backup: {bak.name}")


if __name__ == "__main__":
    main()
