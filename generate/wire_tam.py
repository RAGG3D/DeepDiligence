#!/usr/bin/env python3
"""
wire_tam.py — write data-center indication TAM ($MM) into Pipeline TAM rows.

Pipeline revenue = TAM × market-share × maturity. The template's TAM rows SUMIF
the big TAM Solid/Blood tables — which have no rows for this ticker's indications
(e.g. AML, SCLC, NEC), so TAM = 0 and revenue is 0 regardless of market share.

This step first upserts any `#### 3.N.1 TAM … Addressable Market ($MM)` report
anchors into the DD data center, then reads `datastore/export/tam_by_indication_year.csv`
and overwrites the matching Pipeline "… TAM" row with those database values.
The Pipeline workbook is never allowed to use report-only TAM data.

    python generate/wire_tam.py --ticker MOLN --report-dir .../pipeline_base4

Surgical XML patching; idempotent.
"""
import argparse
import glob
import re
import shutil
import subprocess
import time
import warnings
import zipfile
from pathlib import Path

import csv

warnings.filterwarnings("ignore")
import openpyxl.drawing.text as _t  # noqa: E402
_t.Font.pitchFamily.max = 127
import openpyxl  # noqa: E402
from openpyxl.utils import get_column_letter as cl  # noqa: E402

import sys  # noqa: E402

REPO = Path(__file__).resolve().parents[1]
if str(REPO) not in sys.path:
    sys.path.insert(0, str(REPO))

from datastore.tam_overrides import (  # noqa: E402
    EXPORT_DIR,
    load_overrides,
    parse_report_tam,
    save_overrides,
    upsert_overrides,
)

# Standard maturity ramp (years-since-launch → factor), tiers AVG/BIC/T1, plateau
# 1.056. Rows 551/552/553 of the TAM sheet, cols F..AH. The template ships these
# EMPTY, so the Pipeline revenue formula's INDEX into them yields 0 (→ zero
# revenue) until they are filled. Values match the datastore param_maturity /
# a completed build (CMPX).
_PLATEAU = [1.056] * 21
MATURITY = {
    551: [0.193, 0.332, 0.423, 0.541, 0.669, 0.801, 0.913, 1.0] + _PLATEAU,   # AVG
    552: [0.003, 0.078, 0.175, 0.342, 0.586, 0.795, 0.921, 1.0] + _PLATEAU,   # BIC
    553: [0.046, 0.129, 0.233, 0.410, 0.554, 0.777, 0.930, 1.0] + _PLATEAU,   # T1
}


def parse_reports(report_dir, ticker):
    """{drug_short_upper: {ABBR: {year: $MM}}} from the TAM tables."""
    out = {}
    paths = [Path(p) for p in sorted(glob.glob(f"{report_dir}/{ticker}_*_research_*.md"))]
    latest = {}
    prefix = f"{ticker}_"
    marker = "_research_"
    for path in paths:
        if not path.stem.startswith(prefix) or marker not in path.stem:
            continue
        drug, version = path.stem[len(prefix):].split(marker, 1)
        current = latest.get(drug.upper())
        current_version = current.stem.split(marker, 1)[1] if current else ""
        if current is None or (version, path.stat().st_mtime) > (
            current_version, current.stat().st_mtime
        ):
            latest[drug.upper()] = path
    for path in sorted(latest.values()):
        name = Path(path).name
        drug = name.split("_")[1].upper()
        text = Path(path).read_text(encoding="utf-8", errors="ignore")
        # per-indication sections: ### 3.N <name> (ABBR)
        for m in re.finditer(r'(?m)^##+\s*3\.\d+\s+.*?\(([A-Za-z0-9/+\-]+)\)(.*?)(?=^##+\s*3\.\d+\s|\Z)',
                             text, re.S):
            abbr = m.group(1).strip().upper()
            body = m.group(2)
            # find the TAM table (near "Addressable Market ($MM)")
            tm = re.search(r'Addressable Market \(\$MM\)(.*?)(?=\n#|\Z)', body, re.S)
            if not tm:
                continue
            scope = tm.group(1)
            anchors = {}
            for row in re.finditer(r'\|\s*(\d{4})\s*\|\s*\$?\s*([\d,]+(?:\.\d+)?)', scope):
                anchors[int(row.group(1))] = float(row.group(2).replace(",", ""))
            if anchors:
                out.setdefault(drug, {})[abbr] = anchors
    return out


def upsert_report_tam_to_datastore(report_dir: Path, ticker: str) -> int:
    """Persist report TAM anchors and rebuild DuckDB before workbook reads it.

    Publishing only the override CSVs left ``dd.duckdb`` on the pre-Pipeline
    snapshot produced by the peer-database step.  The seed JSON is the system
    of record, so use the normal datastore builder to atomically refresh both
    Layer 1/2 DuckDB and every export after a new ticker's TAM is discovered.
    """
    entries = parse_report_tam(report_dir, ticker.upper())
    if not entries:
        return 0
    # A ticker refresh replaces its complete curated TAM namespace.  Otherwise
    # old indications survive after a program or protocol changes and can be
    # selected by later builds despite no longer appearing in current research.
    retained = [
        item for item in load_overrides()
        if str(item.get("source_ticker", "")).upper() != ticker.upper()
    ]
    save_overrides(retained)
    upsert_overrides(entries)
    subprocess.run(
        [sys.executable, str(REPO / "datastore" / "build_datastore.py")],
        cwd=str(REPO),
        check=True,
    )
    return len(entries)


def _base_tam_from_revenue(export_dir: Path) -> dict:
    """Reconstruct the base (marketed-sales) TAM = SUM(revenue) by indication/year
    from the published drug_indication_revenue.csv. Used as a fallback when a
    foreign ticker's override is stripped from a code that DOES have real marketed
    sales, so the code falls back to shared industry TAM instead of dropping to 0."""
    path = Path(export_dir) / "drug_indication_revenue.csv"
    base: dict = {}
    if not path.exists():
        return base
    with path.open(newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            code = (row.get("indication_code") or "").strip().upper()
            try:
                year = int(row.get("year", ""))
                rev = float(row.get("revenue_usd_m", ""))
            except (TypeError, ValueError):
                continue
            series = base.setdefault(code, {})
            series[year] = series.get(year, 0.0) + rev
    return base


def load_datastore_tam(active_ticker: str | None = None,
                       export_dir: Path = EXPORT_DIR) -> dict:
    """Return {INDICATION: {year: tam_usd_m}} from the published data center.

    When ``active_ticker`` is given, tam_override rows curated for a DIFFERENT
    ticker are not allowed to bleed into this build (G2): a code whose only
    override belongs to another ticker falls back to its real marketed-sales TAM
    (or is dropped if it has none), while the active ticker's own curated
    overrides always win. Passing None keeps the legacy unscoped behavior.
    """
    export_dir = Path(export_dir)
    out: dict = {}
    tam_path = export_dir / "tam_by_indication_year.csv"
    if tam_path.exists():
        with tam_path.open(newline="", encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                code = (row.get("indication_code") or "").strip().upper()
                try:
                    year = int(row.get("year", ""))
                    tam = float(row.get("tam_usd_m", ""))
                except (TypeError, ValueError):
                    continue
                out.setdefault(code, {})[year] = tam
    if not active_ticker:
        return out

    active = str(active_ticker).strip().upper()
    active_ov: dict = {}      # {code: {year: tam}} owned by the active ticker
    active_codes: set = set()
    foreign_codes: set = set()
    ov_path = export_dir / "tam_override.csv"
    if ov_path.exists():
        with ov_path.open(newline="", encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                code = (row.get("indication_code") or "").strip().upper()
                src = (row.get("source_ticker") or "").strip().upper()
                try:
                    year = int(row.get("year", ""))
                    tam = float(row.get("tam_usd_m", ""))
                except (TypeError, ValueError):
                    continue
                if src == active:
                    active_codes.add(code)
                    active_ov.setdefault(code, {})[year] = tam
                else:
                    foreign_codes.add(code)

    # Codes overridden ONLY by another ticker must not leak in: fall back to real
    # marketed-sales TAM if the code has any, otherwise drop it entirely.
    foreign_only = foreign_codes - active_codes
    if foreign_only:
        base = _base_tam_from_revenue(export_dir)
        for code in foreign_only:
            if code in base:
                out[code] = dict(base[code])
            else:
                out.pop(code, None)

    # The active ticker's own curated overrides always win.
    for code, series in active_ov.items():
        out.setdefault(code, {}).update(series)
    return out


def interp(anchors, year):
    """CAGR interpolation across anchor years; hold ends."""
    ys = sorted(anchors)
    if not ys:
        return 0.0
    if year <= ys[0]:
        if len(ys) >= 2 and anchors[ys[0]] > 0 and anchors[ys[1]] > 0:
            g = (anchors[ys[1]] / anchors[ys[0]]) ** (1.0 / (ys[1] - ys[0]))
            return anchors[ys[0]] * g ** (year - ys[0])
        return anchors[ys[0]]
    if year >= ys[-1]:
        return anchors[ys[-1]]
    for a, b in zip(ys, ys[1:]):
        if a <= year <= b:
            if anchors[a] > 0 and anchors[b] > 0:
                g = (anchors[b] / anchors[a]) ** (1.0 / (b - a))
                return anchors[a] * g ** (year - a)
            return anchors[a] + (anchors[b] - anchors[a]) * (year - a) / (b - a)
    return anchors[ys[-1]]


def patch_num(xml, addr, value):
    m = re.search(rf'<c r="{addr}"([^>]*?)(?:/>|>.*?</c>)', xml, re.S)
    if not m:
        return xml
    sm = re.search(r'\ss="(\d+)"', m.group(1))
    style = f' s="{sm.group(1)}"' if sm else ""
    cell = f'<c r="{addr}"{style}><v>{value:.4f}</v></c>'
    return re.sub(rf'<c r="{addr}"[^>]*?(?:/>|>.*?</c>)', cell, xml, count=1, flags=re.S)


def sheet_zip_path(zf, name):
    rid = re.search(rf'<sheet name="{re.escape(name)}"[^>]*r:id="(rId\d+)"',
                    zf.read("xl/workbook.xml").decode("utf8", "ignore")).group(1)
    rels = zf.read("xl/_rels/workbook.xml.rels").decode("utf8", "ignore")
    tgt = re.search(rf'Id="{rid}"[^>]*Target="([^"]+)"', rels).group(1).lstrip("/")
    return tgt if tgt.startswith("xl/") else "xl/" + tgt


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--ticker", required=True)
    ap.add_argument("--report-dir")
    ap.add_argument("--path")
    a = ap.parse_args()
    dcf = Path(a.path) if a.path else Path(
        f"/mnt/c/Users/yzsun/Desktop/DD/{a.ticker}/DCF {a.ticker}.xlsx")
    rdir = a.report_dir or f"/mnt/c/Users/yzsun/Desktop/DD/{a.ticker}/pipeline_base4"

    # discover Pipeline TAM rows + year columns
    wb = openpyxl.load_workbook(dcf, data_only=False, read_only=True)
    pl = wb["Pipeline"]
    # Year header row 7 is F=<base> then =F7+1 formulas, so only the first cell is
    # a literal int. Derive the columns POSITIONALLY from that base (through 2038).
    year_col = {}
    base_col = base_year = None
    for c in range(6, 60):
        v = pl.cell(7, c).value
        if isinstance(v, int) and 2000 <= v <= 2100:
            base_col, base_year = c, v
            break
    if base_year is not None:
        for c in range(base_col, base_col + (2038 - base_year) + 1):
            year_col[base_year + (c - base_col)] = c
    tam_rows = []   # (row, drug_ref_row, abbr)
    drug_at = {}    # drug_row -> short upper (from D-cell value)
    for r in range(8, pl.max_row + 1):
        d = pl.cell(r, 4).value
        if isinstance(d, str):
            m = re.match(r'=D(\d+)&"\s*(.+?)\s+TAM"\s*$', d)
            if m:
                tam_rows.append((r, int(m.group(1)), m.group(2).strip().upper()))
            elif not d.startswith("="):
                drug_at[r] = d.split(" (")[0].strip().upper()
    wb.close()

    n_upserted = upsert_report_tam_to_datastore(Path(rdir), a.ticker)
    if n_upserted:
        print(f"Upserted {n_upserted} report TAM tables into datastore overrides.")

    tam_db = load_datastore_tam(a.ticker)
    if not tam_db:
        raise SystemExit("No datastore TAM export found at datastore/export/tam_by_indication_year.csv")
    print("Using datastore/export/tam_by_indication_year.csv for Pipeline TAM.")

    report_tam = parse_reports(rdir, a.ticker)

    tam_sheet = next((s for s in ("TAM Solid+MM", "TAM Solid")
                      if s in openpyxl.load_workbook(dcf, read_only=True).sheetnames), None)
    with zipfile.ZipFile(dcf) as zf:
        sp = sheet_zip_path(zf, "Pipeline")
        tp = sheet_zip_path(zf, tam_sheet) if tam_sheet else None
        blobs = {n: zf.read(n) for n in zf.namelist()}
    xml = blobs[sp].decode("utf8", "ignore")

    written = 0
    for row, dref, abbr in tam_rows:
        drug = drug_at.get(dref)
        series = tam_db.get(abbr)
        if not series:
            # Last-resort self-healing: if the report has anchors, publish them
            # first and reload, then still write only from the datastore export.
            anchors = (report_tam.get(drug, {}) or {}).get(abbr)
            if anchors:
                n_upserted = upsert_report_tam_to_datastore(Path(rdir), a.ticker)
                tam_db = load_datastore_tam(a.ticker)
                series = tam_db.get(abbr)
        if not series:
            print(f"  ! no datastore TAM for {drug}/{abbr} (Pipeline R{row})")
            continue
        for y, c in year_col.items():
            value = series[y] if y in series else interp(series, y)
            xml = patch_num(xml, f"{cl(c)}{row}", value)
        written += 1
        print(f"  wrote R{row} {drug}/{abbr} from datastore: 2024≈{interp(series,2024):.0f} "
              f"2030≈{interp(series,2030):.0f} 2038≈{interp(series,2038):.0f} $MM")

    bak = dcf.with_name(f"{dcf.stem}_pre_wiretam_{time.strftime('%Y%m%d_%H%M%S')}.xlsx")
    shutil.copy2(dcf, bak)
    blobs[sp] = xml.encode("utf8")

    # Maturity ramp curve + COGS param — the Pipeline revenue/COGS formulas INDEX
    # into rows 551-553 ($F..$AH) and $P$562 of the TAM sheet, but the template's
    # sheet ends ~row 415: those rows are ABSENT (→ INDEX yields 0 → zero revenue).
    # Insert them (as new <row> elements appended in ascending order).
    mat_done = 0
    if tp:
        txml = blobs[tp].decode("utf8", "ignore")
        if 'r="551"' not in txml:
            rows_xml = ""
            for row, curve in MATURITY.items():         # 551 AVG / 552 BIC / 553 T1
                cells = "".join(f'<c r="{cl(6 + i)}{row}"><v>{f}</v></c>'
                                for i, f in enumerate(curve))
                rows_xml += f'<row r="{row}">{cells}</row>'
                mat_done += 1
            rows_xml += '<row r="562"><c r="P562"><v>0.3</v></c></row>'  # COGS/Price
            txml = txml.replace("</sheetData>", rows_xml + "</sheetData>", 1)
            # extend the sheet dimension so Excel/LibreOffice see the new rows
            txml = re.sub(r'(<dimension ref="[A-Z]+)\d+(:[A-Z]+)\d+"',
                          r'\g<1>1\g<2>562"', txml, count=1)
            blobs[tp] = txml.encode("utf8")

    with zipfile.ZipFile(dcf, "w", zipfile.ZIP_DEFLATED) as zo:
        for n, b in blobs.items():
            zo.writestr(n, b)
    print(f"\nWired TAM into {written} Pipeline rows + maturity curve into "
          f"{mat_done} rows of '{tam_sheet}'. backup: {bak.name}")


if __name__ == "__main__":
    main()
