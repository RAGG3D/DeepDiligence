#!/usr/bin/env python3
"""
adapt_catalyst.py — repoint the Catalyst sum-of-parts breakdown to this ticker.

The Catalyst tab is a per-drug price/share breakdown: each clinical asset's DCF
value/share (pulled from the VALUATION scenario waterfall) × its likelihood of
approval (LOA), summed to Final Market Price. Framework v7 embeds the scenario
What-If results in Catalyst B:C and scales target groups dynamically.

The value cells already reference the VALUATION waterfall, which — after
adapt_ris + the Scenarios "Break Down" modules — computes THIS ticker's per-drug
values (e.g. G26="MP0533 Only", G27="+MP0712"). What stays BCYC is cosmetic:
the drug labels in row 7 (BT8009…) and the LOAs (tuned for BCYC's late-stage
assets). This step fixes those:
  • row 7 drug labels  → the ticker's assets (short code), unused slots "—"
  • LOAs               → a stage-appropriate default (Phase-1 ≈ 0.10) PLACEHOLDER
  • B5 event title      → "{ticker} pipeline catalyst"

LOA is analyst judgement (like market share) — the 0.10 default is a placeholder
to be overwritten by real per-drug conviction; it is NOT a researched value.

    python generate/adapt_catalyst.py --ticker MOLN
"""
import argparse
import json
import re
import shutil
import time
import warnings
import zipfile
from pathlib import Path

warnings.filterwarnings("ignore")
import openpyxl.drawing.text as _t  # noqa: E402
_t.Font.pitchFamily.max = 127
import openpyxl  # noqa: E402

# Catalyst drug columns (row 7 label, row 9 LOA) in waterfall order.
DRUG_SLOTS = [("G7", "J9"), ("K7", "N9"), ("O7", "R9"), ("S7", "V9")]
# Left-side Gap-analysis labels (col F), one per DRUG_SLOTS entry in the same
# waterfall order.  The template ships BCYC text ("BT8009, 80% LOA" …); this
# step overwrites them so no stale drug/LOA is left behind on a fresh run.
GAP_SLOTS = ["F51", "F52", "F53", "F54"]
# Row-9 value cell for each slot (same column as the row-7 label).  The template
# wires these to the VALUATION waterfall breakdown; adapt_catalyst leaves them
# untouched EXCEPT for the aggregate "Other Pipeline" tail slot (see main()).
VALUE_ADDRS = ["G9", "K9", "O9", "S9"]
DEFAULT_LOA = 0.10   # Phase-1 placeholder


def xesc(s):
    return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def sheet_zip_path(zf, name):
    wbx = zf.read("xl/workbook.xml").decode("utf8", "ignore")
    rid = re.search(rf'<sheet name="{re.escape(name)}"[^>]*r:id="(rId\d+)"', wbx).group(1)
    rels = zf.read("xl/_rels/workbook.xml.rels").decode("utf8", "ignore")
    tgt = re.search(rf'Id="{rid}"[^>]*Target="([^"]+)"', rels).group(1).lstrip("/")
    return tgt if tgt.startswith("xl/") else "xl/" + tgt


def _is_absolute_asset_row(col_a, col_b) -> bool:
    """True only for structural asset rows in the Scenarios ' Absolute' module
    (col A == 4, col B == ' Absolute').  Keeps catalyst scenario headers
    ('{event} (+)'/'(-)') and paren-containing breakdown names ('... Cumulative')
    — which sit on scenario-header rows (col A empty, col B a numeric scenario
    id) — from being mistaken for drug assets."""
    if col_a == 4:
        return True
    return isinstance(col_b, str) and col_b.strip() == "Absolute"


def _indication_from_formula(value):
    text = str(value or "")
    match = re.search(r'&"\s*(.*?)\s+Market Share"', text)
    if match:
        return match.group(1).strip()
    match = re.search(r'\)\s+(.*?)\s+Market Share$', text)
    return match.group(1).strip() if match else "All"


def read_assets(dcf):
    """Return every drug×indication target from Scenarios Absolute.

    Catalyst outcomes are indication-specific. Treating one drug with multiple
    indications as a single target silently loses scenarios, so market-share
    rows, not only parent drug rows, define the framework universe.
    """
    wb = openpyxl.load_workbook(dcf, data_only=False, read_only=True)
    seen, targets = set(), []
    current_asset = None
    for row in wb["Scenarios"].iter_rows(min_col=1, max_col=4):
        col_a, col_b, v, unit = row[0].value, row[1].value, row[2].value, row[3].value
        if col_a != 4 or not isinstance(col_b, str) or col_b.strip() != "Absolute":
            continue
        if unit == "[%]" and current_asset:
            indication = _indication_from_formula(v)
            label = current_asset if indication in ("", "All") else f"{current_asset} - {indication}"
            if label not in seen:
                seen.add(label); targets.append(label)
        elif isinstance(v, str) and not v.startswith("=") and "TAM" not in v:
            current_asset = v.split(" (")[0].strip()
    wb.close()
    return targets


def _write_manifest(
    ticker,
    dcf,
    targets,
    *,
    active_targets=None,
    scenario_count=None,
    conviction_threshold=0.10,
):
    artifact_dir = Path(__file__).resolve().parents[1] / "artifacts" / ticker.upper()
    artifact_dir.mkdir(parents=True, exist_ok=True)
    from openpyxl.utils import get_column_letter

    if active_targets:
        active_targets = list(dict.fromkeys(active_targets))
        missing = [name for name in active_targets if name not in targets]
        if missing:
            raise ValueError(f"active Catalyst targets absent from Scenarios Absolute: {missing}")
        if not scenario_count or scenario_count < 1:
            raise ValueError("active Catalyst manifest requires a positive scenario_count")
        ordered_targets = active_targets + [name for name in targets if name not in active_targets]
        scenario_last = 9 + int(scenario_count)
        table_title_row = scenario_last + 2
        table_target_row = table_title_row + 1
        table_header_row = table_title_row + 2
        table_input_first = table_title_row + 3
        table_input_last = table_input_first + 3
        base_col = 3
        final_col = 4
        upside_col = 5
        rjconv_col = 6
        outcome_first_col = 7
        group_first_col = outcome_first_col + len(active_targets)
        manifest = {
            "ticker": ticker.upper(),
            "framework_version": 7,
            "mode": "active_cartesian",
            "conviction_threshold": conviction_threshold,
            "active_targets": active_targets,
            "targets": [],
            "event_metadata": {"name": "C2", "disclosure": "C3", "source": "C4"},
            "manual_cells": [],
            "neutral_defaults": {},
            "layout": {
                "main_header_row": 7,
                "main_base_row": 9,
                "main_scenario_first": 10,
                "main_scenario_last": scenario_last,
                "scenario_count": int(scenario_count),
                "outcome_first_col": outcome_first_col,
                "outcome_last_col": outcome_first_col + len(active_targets) - 1,
                "base_col": base_col,
                "final_market_col": final_col,
                "upside_col": upside_col,
                "rjconv_col": rjconv_col,
                "embedded_data_table_ref": f"B8:C{scenario_last}",
                "embedded_data_table_input": "C8",
                "scenario_input_bridge": "B6",
                "terminal_growth_bridge": "C6",
                "target_group_first_col": group_first_col,
                "table3_title_row": table_title_row,
                "table3_target_row": table_target_row,
                "table3_header_row": table_header_row,
                "table3_input_first": table_input_first,
                "table3_input_last": table_input_last,
            },
        }
        outcomes = ("increase", "remain", "decrease", "suspension")
        wb = openpyxl.load_workbook(dcf, data_only=False, read_only=True)
        ws = wb["Catalyst"]
        try:
            for i, target_name in enumerate(ordered_targets):
                group_start = group_first_col + 4 * i
                main_range = (
                    f"{get_column_letter(group_start)}6:"
                    f"{get_column_letter(group_start + 3)}{scenario_last}"
                )
                table3_range = (
                    f"{get_column_letter(group_start)}{table_target_row}:"
                    f"{get_column_letter(group_start + 2)}{table_input_last}"
                )
                allowed = []
                for offset, outcome in enumerate(outcomes):
                    row = table_input_first + offset
                    conv = ws.cell(row, group_start + 2).value
                    if isinstance(conv, (int, float)) and conv >= conviction_threshold:
                        allowed.append(outcome.title())
                manifest["targets"].append({
                    "name": target_name,
                    "active": target_name in active_targets,
                    "hidden": False,
                    "masked": target_name not in active_targets,
                    "allowed_outcomes": allowed if target_name in active_targets else [],
                    "main_range": main_range,
                    "display_ranges": [main_range, table3_range],
                    "market_share_change_col": get_column_letter(group_start),
                    "loa_change_col": get_column_letter(group_start + 1),
                    "conv_col": get_column_letter(group_start + 2),
                })
                for offset, outcome in enumerate(outcomes):
                    row = table_input_first + offset
                    for col, kind in (
                        (group_start, "ms"),
                        (group_start + 1, "loa"),
                        (group_start + 2, "conv"),
                    ):
                        addr = f"{get_column_letter(col)}{row}"
                        manifest["manual_cells"].append(addr)
                        manifest["neutral_defaults"][addr] = (
                            1 if kind == "conv" and outcome == "remain" else 0
                        )
        finally:
            wb.close()
        path = artifact_dir / f"{ticker.upper()}_catalyst_manifest.json"
        path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")
        return path

    scenario_last = 9 + 4 * len(targets)
    table_title_row = scenario_last + 2
    table_target_row = table_title_row + 1
    table_header_row = table_title_row + 2
    table_input_first = table_title_row + 3
    table_input_last = table_input_first + 3
    manifest = {
        "ticker": ticker.upper(), "framework_version": 3,
        "targets": [],
        "event_metadata": {"name": "C2", "disclosure": "C3", "source": "C4"},
        "manual_cells": [], "neutral_defaults": {},
        "layout": {
            "main_header_row": 7,
            "main_base_row": 9,
            "main_scenario_first": 10,
            "main_scenario_last": scenario_last,
            "table3_title_row": table_title_row,
            "table3_target_row": table_target_row,
            "table3_header_row": table_header_row,
            "table3_input_first": table_input_first,
            "table3_input_last": table_input_last,
        },
    }
    outcomes = ("increase", "remain", "decrease", "suspension")
    for i, target_name in enumerate(targets):
        group_start = 7 + 4 * i
        market_share_col = group_start + 2
        loa_col = group_start + 3
        conv_col = group_start + 4
        main_range = (
            f"{get_column_letter(group_start)}6:"
            f"{get_column_letter(group_start + 3)}{scenario_last}"
        )
        table3_range = (
            f"{get_column_letter(market_share_col)}{table_target_row}:"
            f"{get_column_letter(conv_col)}{table_input_last}"
        )
        manifest["targets"].append({
            "name": target_name,
            "main_range": main_range,
            "display_ranges": [main_range, table3_range],
            "market_share_change_col": get_column_letter(market_share_col),
            "loa_change_col": get_column_letter(loa_col),
            "conv_col": get_column_letter(conv_col),
        })
        for offset, outcome in enumerate(outcomes):
            row = table_input_first + offset
            for col, kind in (
                (market_share_col, "ms"), (loa_col, "loa"), (conv_col, "conv")
            ):
                addr = f"{get_column_letter(col)}{row}"
                manifest["manual_cells"].append(addr)
                manifest["neutral_defaults"][addr] = (
                    1 if kind == "conv" and outcome == "remain" else 0
                )
    path = artifact_dir / f"{ticker.upper()}_catalyst_manifest.json"
    path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    return path


def _style(xml, addr):
    m = re.search(rf'<c r="{addr}"([^>]*?)(?:/>|>.*?</c>)', xml, re.S)
    if not m:
        return None
    sm = re.search(r'\ss="(\d+)"', m.group(1))
    return f' s="{sm.group(1)}"' if sm else ""


def patch_text(xml, addr, text):
    s = _style(xml, addr)
    if s is None:
        return xml
    cell = f'<c r="{addr}"{s} t="inlineStr"><is><t xml:space="preserve">{xesc(text)}</t></is></c>'
    return re.sub(rf'<c r="{addr}"[^>]*?(?:/>|>.*?</c>)', cell, xml, count=1, flags=re.S)


def patch_num(xml, addr, value):
    s = _style(xml, addr)
    if s is None:
        return xml
    cell = f'<c r="{addr}"{s}><v>{value}</v></c>'
    return re.sub(rf'<c r="{addr}"[^>]*?(?:/>|>.*?</c>)', cell, xml, count=1, flags=re.S)


def patch_formula(xml, addr, formula):
    s = _style(xml, addr)
    if s is None:
        return xml
    cell = f'<c r="{addr}"{s}><f>{xesc(formula)}</f></c>'
    return re.sub(rf'<c r="{addr}"[^>]*?(?:/>|>.*?</c>)', cell, xml, count=1, flags=re.S)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--ticker")
    ap.add_argument("--path")
    ap.add_argument("--loa", type=float, default=DEFAULT_LOA)
    a = ap.parse_args()
    dcf = Path(a.path) if a.path else Path(
        f"/mnt/c/Users/yzsun/Desktop/DD/{a.ticker}/DCF {a.ticker}.xlsx")

    assets = read_assets(dcf)
    shorts = [x.split(" (")[0].strip() for x in assets]
    print(f"Assets: {shorts}")

    n_slots = len(DRUG_SLOTS)
    aggregate_tail = False
    if aggregate_tail:
        tail = shorts[n_slots - 1:]
        print(f"Aggregating {len(tail)} assets beyond the first {n_slots - 1} "
              f"into an 'Other Pipeline' Catalyst slot ({tail}) so no program "
              f"is dropped from the sum-of-parts / Final Market Price W9.")

    labels, loas, value_formulas = {}, {}, {}
    for i, (lab_addr, loa_addr) in enumerate(DRUG_SLOTS):
        gap_addr = GAP_SLOTS[i]
        if aggregate_tail and i == n_slots - 1:
            # Mirror adapt_ris's 'Other Pipeline' aggregate: the last slot
            # absorbs every asset beyond the first three.  Its value = the full
            # per-share valuation (VALUATION!C48, the waterfall total) minus the
            # three visible per-drug values, so the four value cells still sum to
            # the total and no tail program is lost from W9.
            labels[lab_addr] = "Other Pipeline"
            loas[loa_addr] = a.loa           # aggregate LOA (placeholder)
            visible = "-".join(VALUE_ADDRS[: n_slots - 1])
            value_formulas[VALUE_ADDRS[i]] = f"VALUATION!C48-{visible}"
            labels[gap_addr] = f"{labels[lab_addr]}, {round(loas[loa_addr] * 100)}% LOA"
        elif i < min(len(shorts), n_slots):
            labels[lab_addr] = shorts[i]
            loas[loa_addr] = a.loa
            labels[gap_addr] = f"{labels[lab_addr]}, {round(loas[loa_addr] * 100)}% LOA"
        else:
            labels[lab_addr] = "—"
            loas[loa_addr] = 0
            # Unused slot: overwrite the stale template drug/LOA with a neutral
            # placeholder so nothing BCYC-specific survives.
            labels[gap_addr] = "Reserved, 0% LOA"
    labels["B5"] = f"{a.ticker} pipeline catalyst"

    bak = dcf.with_name(f"{dcf.stem}_pre_catalyst_{time.strftime('%Y%m%d_%H%M%S')}.xlsx")
    shutil.copy2(dcf, bak)
    with zipfile.ZipFile(dcf) as zf:
        sp = sheet_zip_path(zf, "Catalyst")
        blobs = {n: zf.read(n) for n in zf.namelist()}
    xml = blobs[sp].decode("utf8", "ignore")
    for addr, text in labels.items():
        xml = patch_text(xml, addr, text)
    for addr, val in loas.items():
        xml = patch_num(xml, addr, val)
    for addr, formula in value_formulas.items():
        xml = patch_formula(xml, addr, formula)
    blobs[sp] = xml.encode("utf8")
    with zipfile.ZipFile(dcf, "w", zipfile.ZIP_DEFLATED) as zo:
        for n, b in blobs.items():
            zo.writestr(n, b)

    manifest_path = _write_manifest(a.ticker or dcf.stem.replace("DCF ", ""), dcf, shorts)

    print(f"Catalyst repointed: row7 {[labels[s[0]] for s in DRUG_SLOTS]}, "
          f"LOA={a.loa} (placeholder). backup: {bak.name}")
    print(f"Catalyst lifecycle manifest: {manifest_path}")


if __name__ == "__main__":
    main()
