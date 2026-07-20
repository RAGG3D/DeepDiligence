#!/usr/bin/env python3
"""Build the embedded full drug×indication Catalyst framework and scenarios."""
from __future__ import annotations

import argparse
import json
import shutil
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path
from xml.etree import ElementTree as ET

REPO = Path(__file__).resolve().parents[1]
if str(REPO) not in sys.path:
    sys.path.insert(0, str(REPO))

from generate.adapt_catalyst import _write_manifest, read_assets
from tools import catalyst_workflow as lifecycle


def _default_path(ticker: str) -> Path:
    return Path(f"/mnt/c/Users/yzsun/Desktop/DD/{ticker}/DCF {ticker}.xlsx")


def _windows_path(path: Path) -> str:
    result = subprocess.run(["wslpath", "-w", str(path)], check=True, capture_output=True, text=True)
    return result.stdout.strip()


def _restore_active_fills_only(path: Path, state: dict) -> None:
    """Return an active run to neutral appearance before a structural rebuild.

    Inputs and event metadata are intentionally preserved.  After the new
    framework is built, the same active run is re-masked with fresh grey fills
    and fonts without opening a duplicate database run.
    """
    original = state.get("original_styles") or {}
    backup = Path(state.get("backup") or "")
    if not original or not backup.exists():
        return
    parts, order = lifecycle._read_parts(path)
    sheet_path = lifecycle._sheet_path(parts, "Catalyst")
    sheet = ET.fromstring(parts[sheet_path])
    lifecycle._restore_original_fills(parts, sheet, original, backup)
    parts[sheet_path] = lifecycle._excel_xml(sheet)
    lifecycle._write_parts(path, parts, order)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--ticker", required=True)
    ap.add_argument("--path")
    ap.add_argument(
        "--research",
        help="Catalyst research JSON. When relevant_targets is present, build the "
             "conviction-filtered Cartesian active-catalyst layout.",
    )
    ap.add_argument("--conviction-threshold", type=float, default=0.10)
    ap.add_argument("--no-backup", action="store_true")
    args = ap.parse_args()
    path = Path(args.path) if args.path else _default_path(args.ticker.upper())
    targets = read_assets(path)
    if not targets:
        raise SystemExit("No drug×indication targets found in Scenarios Absolute")
    powershell = Path("/mnt/c/Windows/System32/WindowsPowerShell/v1.0/powershell.exe")
    if not powershell.exists():
        raise SystemExit("Excel-native Catalyst build requires Windows PowerShell/Excel COM")
    if not args.no_backup:
        backup = path.with_name(
            f"{path.stem}_pre_catalyst_framework_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        shutil.copy2(path, backup)
        print(f"Backup: {backup}")

    artifact_dir = REPO / "artifacts" / args.ticker.upper()
    state_path = artifact_dir / f"{args.ticker.upper()}_catalyst_active_state.json"
    active_state = None
    if state_path.exists():
        active_state = json.loads(state_path.read_text(encoding="utf-8"))
        _restore_active_fills_only(path, active_state)

    research = None
    if args.research:
        research = json.loads(Path(args.research).read_text(encoding="utf-8"))
    elif active_state is not None:
        research = active_state.get("research") or None
    relevant_targets = list(dict.fromkeys((research or {}).get("relevant_targets") or []))
    missing = [name for name in relevant_targets if name not in targets]
    if missing:
        raise SystemExit(f"Catalyst research targets absent from Scenarios Absolute: {missing}")

    # One builder owns both active and neutral layouts. Without event research,
    # every target is included; neutral defaults leave only Remain above the
    # conviction threshold. This prevents the retired VALUATION O:P/v3 layout
    # from reappearing through a clean or sync entry point.
    builder_targets = relevant_targets or targets
    script = REPO / "tools" / "build_active_catalyst_combinations.ps1"
    cmd = [
        str(powershell), "-NoProfile", "-ExecutionPolicy", "Bypass", "-File",
        _windows_path(script), "-Path", _windows_path(path),
        "-Ticker", args.ticker.upper(),
        "-RelevantTargetsJson", json.dumps(builder_targets),
        "-ConvictionThreshold", str(args.conviction_threshold),
    ]
    subprocess.run(cmd, check=True)
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=False, read_only=True)
    try:
        ws = wb["Catalyst"]
        title_row = next(
            row for row in range(1, ws.max_row + 1)
            if ws.cell(row, 2).value == "Catalyst Input Changes (Table 3)"
        )
    finally:
        wb.close()
    scenario_count = title_row - 11
    manifest = _write_manifest(
        args.ticker.upper(), path, targets,
        active_targets=builder_targets,
        scenario_count=scenario_count,
        conviction_threshold=args.conviction_threshold,
    )
    if active_state is not None:
        manifest_data = json.loads(manifest.read_text(encoding="utf-8"))
        neutral_backup = path.with_name(
            f"{path.stem}_pre_active_two_table_"
            f"{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        )
        shutil.copy2(path, neutral_backup)
        active_state["original_styles"] = lifecycle._patch_run(
            path, manifest_data, active_state.get("research") or {}
        )
        active_state["backup"] = str(neutral_backup)
        active_state["workbook"] = str(path)
        active_state["framework_version"] = manifest_data.get("framework_version", 3)
        state_path.write_text(
            json.dumps(active_state, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )
        print(
            f"Reapplied existing active Catalyst run {active_state.get('run_id')} "
            "to the two-table layout"
        )
    mode = "active Cartesian" if relevant_targets else "full-universe neutral Cartesian"
    # Excel-native builders calculate in Manual mode for deterministic What-If
    # table construction. Normalize the package after every standalone build so
    # modern Excel cannot display cached formulas as stale strikethrough values.
    subprocess.run([
        sys.executable,
        str(REPO / "tools" / "normalize_calc_state.py"),
        "--path", str(path),
        "--no-backup",
    ], check=True)
    print(f"Catalyst {mode} framework verified for {len(targets)} targets; manifest → {manifest}")


if __name__ == "__main__":
    main()
