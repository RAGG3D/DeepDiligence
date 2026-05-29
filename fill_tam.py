#!/usr/bin/env python3
"""
fill_tam.py -- Fill TAM Solid and TAM Blood sheets with FY2024/FY2025 drug revenue data.

Uses surgical zip-patching (NEVER openpyxl .save()) to preserve sharedStrings.xml,
calcChain.xml, styles, and all other workbook parts.

Usage:
    python fill_tam.py [--dry-run] [--file PATH]
"""

import argparse
import logging
import re
import shutil
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


_EMPTY_CALC_CHAIN = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
    '<calcChain xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"></calcChain>'
)
logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

# ── File path ────────────────────────────────────────────────────────────────
DEFAULT_FILE = Path("/mnt/c/Users/yzsun/Desktop/DD/base/DCF Template 2020.xlsx")

# ── Sheet names ──────────────────────────────────────────────────────────────
SHEET_SOLID = "TAM Solid"
SHEET_BLOOD = "TAM Blood"

# ── xlsx XML namespaces ──────────────────────────────────────────────────────
_NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

# ══════════════════════════════════════════════════════════════════════════════
#  REVENUE DATA -- TAM Solid  (copied from DCF CMPX.xlsx 2026-03-10)
#  Key: row_number → {2024: T_col_value, 2025: U_col_value}
#  Values: number → numeric cell, string '=...' → formula cell
#  Covers: Oncology Drug Market (R9-R401), TAM/Summary (R402-R505),
#          Parameters (R510-R562)
# ══════════════════════════════════════════════════════════════════════════════

SOLID_CELLS: Dict[int, Dict[int, Any]] = {
    # ══════════════════════════════════════════════════════════════════════
    #  ONCOLOGY DRUG MARKET (R9-R343)
    # ══════════════════════════════════════════════════════════════════════
    9: {2024: 467.64, 2025: 467.64},  # Temodar
    10: {2024: '=T9/3/(T526/T528)'=, 2025: '=U9/3/(U526/U528)'=},  # GBM
    11: {2024: '=S11*S11/R11'=, 2025: '=T11*T11/S11'=},  # Tafinlar
    12: {2024: 0, 2025: 0},  # Melanoma
    13: {2024: '=S13/M551*N551'=, 2025: '=T13^2/S13'=},  # Throid Cancer
    14: {2024: '=S14/L546*M546'=, 2025: '=T14^2/S14'=},  # NSCLC
    15: {2024: '=T11-T13-T14'=, 2025: '=U11-U13-U14'=},  # GBM
    16: {2024: 26.33, 2025: 26.33},  # Valstar Intravesical Solution
    17: {2024: '=T16'=, 2025: '=U16'=},  # BLCA
    18: {2024: 712, 2025: 712},  # Padcev
    19: {2024: '=T18'=, 2025: '=U18'=},  # BLCA
    20: {2024: 107, 2025: 107},  # Balversa
    21: {2024: '=T20'=, 2025: '=U20'=},  # BLCA
    22: {2024: 71, 2025: 277},  # Ankita
    23: {2024: '=T22'=, 2025: '=U22'=},  # BLCA
    24: {2024: 59, 2025: 59},  # Jelmyto
    25: {2024: '=T24'=, 2025: '=U24'=},  # BLCA
    26: {2024: '=AVERAGE(Q26:S26)'=, 2025: '=T26'=},  # Bavencio
    27: {2024: '=S27/L542*M542'=, 2025: '=T27'=},  # BLCA
    28: {2024: '=T26-T27'=, 2025: '=T28'=},  # RCC
    29: {2024: '=S29*1.1'=, 2025: '=T29*1.1'=},  # Trodelvy
    30: {2024: '=S30/H550*I550'=, 2025: '=T30*1.1'=},  # TNBC
    31: {2024: '=T29-T32-T30'=, 2025: '=U29-U32-U30'=},  # BLCA
    32: {2024: '=G551*S32/F551'=, 2025: '=T32*1.1'=},  # BRCA
    33: {2024: 69, 2025: 68},  # Taxotere
    34: {2024: '=T$33*(T$489/(T$489+T$490+T$491))'=, 2025: '=U$33*(U$489/(U$489+U$490+U$491))'=},  # TNBC
    35: {2024: '=T$33*(T522/(T$489+T$490+T$491))'=, 2025: '=U$33*(U522/(U$489+U$490+U$491))'=},  # BRCA
    36: {2024: '=T$33*(T523/(T$489+T$490+T$491))'=, 2025: '=U$33*(U523/(U$489+U$490+U$491))'=},  # BLCA
    37: {2024: 46, 2025: 45},  # Evista
    38: {2024: '=T$37'=, 2025: '=U$37'=},  # BRCA
    39: {2024: '=AVERAGE(F39:R39)'=, 2025: 384.58},  # Femara
    40: {2024: '=T$39'=, 2025: '=U$39'=},  # BRCA
    41: {2024: 939, 2025: 939},  # Zoladex
    42: {2024: '=T41-T43-T44'=, 2025: '=U41-U43-U44'=},  # PRAD
    43: {2024: '=(T41/2)*0.17'=, 2025: '=(U41/2)*0.17'=},  # TNBC
    44: {2024: '=T41/2*0.83'=, 2025: '=U41/2*0.83'=},  # BRCA
    45: {2024: 319.54, 2025: 319.54},  # Halaven
    46: {2024: '=(T45/2)*0.17'=, 2025: '=(U45/2)*0.17'=},  # TNBC
    47: {2024: '=T45-T46'=, 2025: '=U45-U46'=},  # BRCA
    48: {2024: 43, 2025: 43},  # Aybintio
    49: {2024: '=T48*0.17'=, 2025: '=U48*0.17'=},  # TNBC
    50: {2024: '=T48*0.83'=, 2025: '=U48*0.83'=},  # BRCA
    51: {2024: 5909.93, 2025: 5909.93},  # Herceptin
    52: {2024: '=S52/J546*M546'=, 2025: '=T52'=},  # GC
    53: {2024: '=T51-T52'=, 2025: '=U51-U52'=},  # BRCA
    54: {2024: 65, 2025: 65},  # Herzuma
    55: {2024: '=S55'=, 2025: '=T55'=},  # GC
    56: {2024: '=S56'=, 2025: '=T56'=},  # BRCA
    57: {2024: 161, 2025: 161},  # Ogivri
    58: {2024: '=R58'=, 2025: '=T58'=},  # GC
    59: {2024: '=R59'=, 2025: '=T59'=},  # BRCA
    60: {2024: 368, 2025: 368},  # Kanjinti
    61: {2024: '=T60*(T$459/(T$459+T$457))'=, 2025: '=T61'=},  # GC
    62: {2024: '=T60*(T$457/(T$459+T$457))'=, 2025: '=T62'=},  # BRCA
    63: {2024: 90, 2025: 90},  # Trazimera
    64: {2024: '=S64'=, 2025: '=T64'=},  # GC
    65: {2024: '=S65'=, 2025: '=T65'=},  # BRCA
    66: {2024: 156, 2025: 156},  # Ontruzant
    67: {2024: '=S67'=, 2025: '=T67'=},  # GC
    68: {2024: '=S68'=, 2025: '=T68'=},  # BRCA
    69: {2024: 186, 2025: 186},  # Doxil
    70: {2024: '=R70'=, 2025: '=T70'=},  # OV
    71: {2024: '=R71'=, 2025: '=T71'=},  # BRCA
    72: {2024: 326, 2025: 326},  # Elahere
    73: {2024: '=T72'=, 2025: '=U72'=},  # OV
    74: {2024: 124.5, 2025: 124.5},  # Rubraca
    75: {2024: '=T74'=, 2025: '=U74'=},  # OV
    76: {2024: 537, 2025: 537},  # Tevimbra
    77: {2024: '=T76'=, 2025: '=U76'=},  # ESCA
    78: {2024: 427.64, 2025: 427.64},  # Eloxatin
    79: {2024: '=T78'=, 2025: '=U78'=},  # CRC
    80: {2024: 1170, 2025: 1170},  # Vectibix
    81: {2024: '=T80'=, 2025: '=U80'=},  # CRC
    82: {2024: 109, 2025: 109},  # Fruzaqla
    83: {2024: '=T82'=, 2025: '=U82'=},  # CRC
    84: {2024: 2, 2025: 2},  # Fusilev
    85: {2024: '=T84'=, 2025: '=U84'=},  # CRC
    86: {2024: 29.15, 2025: 29.15},  # Camptosar
    87: {2024: '=T86/3'=, 2025: '=U86/3'=},  # CRC
    88: {2024: 848.85, 2025: 848.85},  # Xeloda
    89: {2024: '=T$88/3'=, 2025: '=U$88/3'=},  # CRC
    90: {2024: '=(T$88-T$89)*(T$490/(T$490+T$489+T$492))'=, 2025: '=(U$88-U$89)*(U$490/(U$490+U$489+U$492))'=},  # BRCA
    91: {2024: '=(T$88-T$89)*(T$489/(T$490+T$489+T$492))'=, 2025: '=(U$88-U$89)*(U$489/(U$490+U$489+U$492))'=},  # TNBC
    92: {2024: '=T88-T91-T90-T89'=, 2025: '=U88-U91-U90-U89'=},  # GC
    93: {2024: 566, 2025: 566},  # Stivarga
    94: {2024: '=S94*T93/S93'=, 2025: '=T94*U93/T93'=},  # CRC
    95: {2024: '=T93-T94'=, 2025: '=U93-U94'=},  # HCC
    96: {2024: 490, 2025: 490},  # Lonsurf
    97: {2024: '=S97'=, 2025: '=T97'=},  # GC
    98: {2024: '=S98'=, 2025: '=T98'=},  # CRC
    99: {2024: 9, 2025: 9},  # Khapzory
    100: {2024: '=T99/2'=, 2025: '=U99/2'=},  # CRC
    101: {2024: 69.5, 2025: 69.5},  # Vegzelma
    102: {2024: '=T101/4'=, 2025: '=U101/4'=},  # OV
    103: {2024: '=T101/4'=, 2025: '=U101/4'=},  # CRC
    104: {2024: '=T101-T103-T102'=, 2025: '=U101-U103-U102'=},  # NSCLC
    105: {2024: '=T106+T107'=, 2025: '=U106+U107'=},  # Tukysa
    106: {2024: '=S106^2/R106'=, 2025: '=T106^2/S106'=},  # BRCA
    107: {2024: '=S107/F551*G551'=, 2025: '=T107/G551*H551'=},  # CRC
    108: {2024: 2811, 2025: 2811},  # Lynparza
    109: {2024: '=S109/N$516*O$516'=, 2025: '=T109/O$516*P$516'=},  # OV
    110: {2024: '=S110'=, 2025: '=T110'=},  # BRCA
    111: {2024: '=S111'=, 2025: '=T111'=},  # TNBC
    112: {2024: '=T108-T111-T110-T109'=, 2025: '=U108-U111-U110-U109'=},  # PRAD
    113: {2024: 43, 2025: 43},  # Talzenna
    114: {2024: '=S114'=, 2025: '=T114'=},  # BRCA
    115: {2024: '=S115'=, 2025: '=T115'=},  # TNBC
    116: {2024: '=S116'=, 2025: '=T116'=},  # PRAD
    117: {2024: 921, 2025: 921},  # Zejula
    118: {2024: '=S118'=, 2025: '=T118'=},  # OV
    119: {2024: '=S119'=, 2025: '=T119'=},  # PRAD
    120: {2024: 980, 2025: 980},  # Pluvicto
    121: {2024: '=T120'=, 2025: '=U120'=},  # mCRPC
    122: {2024: 2010.85, 2025: 2010.85},  # Zytiga
    123: {2024: '=T122'=, 2025: '=U122'=},  # mCRPC
    124: {2024: 5038, 2025: 5038},  # Xtandi
    125: {2024: '=T124'=, 2025: '=U124'=},  # mCRPC
    126: {2024: 325, 2025: 325},  # Rybrevant
    127: {2024: '=T126'=, 2025: '=U126'=},  # NSCLC
    128: {2024: 4, 2025: 4},  # Lazcluze
    129: {2024: '=T128'=, 2025: '=U128'=},  # NSCLC
    130: {2024: 218, 2025: 218},  # Imjudo
    131: {2024: '=T130-T132'=, 2025: '=U130-U132'=},  # NSCLC
    132: {2024: 150, 2025: 150},  # HCC
    133: {2024: 1652, 2025: 1652},  # Alecensa
    134: {2024: '=T133'=, 2025: '=U133'=},  # NSCLC
    135: {2024: 63, 2025: 63},  # Zykadia
    136: {2024: '=T135'=, 2025: '=U135'=},  # NSCLC
    137: {2024: 184, 2025: 184},  # Alunbrig
    138: {2024: '=T137'=, 2025: '=U137'=},  # NSCLC
    139: {2024: '=S139/F544*G544'=, 2025: '=T139/G544*H544'=},  # Augtyro (Forcasting)
    140: {2024: '=T139'=, 2025: '=U139'=},  # NSCLC
    141: {2024: 455.23, 2025: 455.23},  # Iressa
    142: {2024: '=T141'=, 2025: '=U141'=},  # NSCLC
    143: {2024: 5799, 2025: 5799},  # Tagrisso
    144: {2024: '=T143'=, 2025: '=U143'=},  # NSCLC
    145: {2024: 153, 2025: 511},  # Krazati
    146: {2024: '=T145'=, 2025: '=U145'=},  # NSCLC
    147: {2024: 6, 2025: 6},  # Vizimpro
    148: {2024: '=T147'=, 2025: '=U147'=},  # NSCLC
    149: {2024: 154, 2025: 154},  # Tebrecta
    150: {2024: '=T149'=, 2025: '=U149'=},  # NSCLC
    151: {2024: 538, 2025: 538},  # Lorbrena
    152: {2024: '=T151'=, 2025: '=U151'=},  # NSCLC
    153: {2024: 54, 2025: 54},  # Tepmetko
    154: {2024: '=T153'=, 2025: '=U153'=},  # NSCLC
    155: {2024: 280, 2025: 280},  # Lumakras
    156: {2024: '=T155'=, 2025: '=U155'=},  # NSCLC
    157: {2024: 87, 2025: 87},  # Rozlytrek
    158: {2024: '=T157'=, 2025: '=U157'=},  # NSCLC
    159: {2024: 54.5, 2025: 54.5},  # Pemfexy
    160: {2024: '=T159'=, 2025: '=U159'=},  # NSCLC
    161: {2024: '=S161'=, 2025: '=T161'=},  # Avastin
    162: {2024: '=S162'=, 2025: '=T162'=},  # OV
    163: {2024: '=S163'=, 2025: '=T163'=},  # BRCA
    164: {2024: '=S164'=, 2025: '=T164'=},  # RCC
    165: {2024: '=S165'=, 2025: '=T165'=},  # TNBC
    166: {2024: '=S166'=, 2025: '=T166'=},  # GBM
    167: {2024: '=S167'=, 2025: '=T167'=},  # NSCLC
    168: {2024: 2238, 2025: 2238},  # Yervoy, Mono Antibody
    169: {2024: '=S169'=, 2025: '=T169'=},  # ESCA
    170: {2024: '=S170'=, 2025: '=T170'=},  # CRC
    171: {2024: '=S171'=, 2025: '=T171'=},  # NSCLC
    172: {2024: '=S172'=, 2025: '=T172'=},  # Melanoma NCAM+
    173: {2024: '=S173'=, 2025: '=T173'=},  # MPM
    174: {2024: '=S174'=, 2025: '=T174'=},  # RCC
    175: {2024: 451.33, 2025: 451.33},  # Xalkori
    176: {2024: '=T175'=, 2025: '=U175'=},  # NSCLC
    177: {2024: 836.56, 2025: 836.56},  # Cyramza, Mono Antibody
    178: {2024: '=T177/3.2'=, 2025: '=T178'=},  # GC
    179: {2024: '=T177/3.2'=, 2025: '=T179'=},  # CRC
    180: {2024: 26, 2025: '=T180'=},  # HCC
    181: {2024: '=T177/3.2'=, 2025: '=T181'=},  # NSCLC
    182: {2024: 0, 2025: 0},  # Mekinist
    183: {2024: 0, 2025: 0},  # Melanoma
    184: {2024: 0, 2025: 0},  # NSCLC
    185: {2024: 1922, 2025: 1922},  # Tafinlar
    186: {2024: 0, 2025: 0},  # Melanoma
    187: {2024: '=S187'=, 2025: '=T187'=},  # NSCLC
    188: {2024: '=S188'=, 2025: '=T188'=},  # BTC
    189: {2024: '=S189'=, 2025: '=T189'=},  # Melanoma NCAM+
    190: {2024: 4195, 2025: 4195},  # Tecentriq, Mono Antibody
    191: {2024: '=S191'=, 2025: '=T191'=},  # Melanoma
    192: {2024: '=S192'=, 2025: '=T192'=},  # NSCLC
    193: {2024: '=S193'=, 2025: '=T193'=},  # TNBC
    194: {2024: 218, 2025: 183.83},  # ES-SCLC
    195: {2024: 1900, 2025: 1900},  # HCC
    196: {2024: 254, 2025: 254},  # Retevmo
    197: {2024: '=T196/2'=, 2025: '=U196/2'=},  # NSCLC
    198: {2024: 2696, 2025: 2696},  # Enhertu
    199: {2024: '=S199'=, 2025: '=T199'=},  # BRCA
    200: {2024: '=S200'=, 2025: '=T200'=},  # TNBC
    201: {2024: '=S201'=, 2025: '=T201'=},  # GC
    202: {2024: '=S202'=, 2025: '=T202'=},  # NSCLC
    203: {2024: 4019, 2025: 4019},  # Imfinzi
    204: {2024: '=S204'=, 2025: '=T204'=},  # NSCLC
    205: {2024: '=S205'=, 2025: '=T205'=},  # BTC
    206: {2024: '=S206'=, 2025: '=T206'=},  # ES-SCLC
    207: {2024: '=S207'=, 2025: '=T207'=},  # HCC
    208: {2024: 863, 2025: 863},  # Libtayo
    209: {2024: '=S209'=, 2025: '=T209'=},  # NSCLC
    210: {2024: 193, 2025: 193},  # Mektovi
    211: {2024: '=S211'=, 2025: '=T211'=},  # Melanoma
    212: {2024: '=S212'=, 2025: '=T212'=},  # NSCLC
    213: {2024: 236, 2025: 236},  # Braftovi
    214: {2024: '=S214'=, 2025: '=T214'=},  # CRC
    215: {2024: '=S215'=, 2025: '=T215'=},  # Melanoma
    216: {2024: '=S216'=, 2025: '=T216'=},  # Melanoma NCAM+
    217: {2024: '=S217'=, 2025: '=T217'=},  # NSCLC
    218: {2024: 423, 2025: 423},  # Zirabev
    219: {2024: '=S219'=, 2025: '=T219'=},  # GBM
    220: {2024: '=S220'=, 2025: '=T220'=},  # OV
    221: {2024: '=S221'=, 2025: '=T221'=},  # BRCA
    222: {2024: '=S222'=, 2025: '=T222'=},  # TNBC
    223: {2024: '=S223'=, 2025: '=T223'=},  # CRC
    224: {2024: '=S224'=, 2025: '=T224'=},  # NSCLC
    225: {2024: 543.2, 2025: 543.2},  # Mvasi
    226: {2024: '=T225/4/(T527/T526)'=, 2025: '=T226'=},  # GBM
    227: {2024: '=T$225/4'=, 2025: '=T227'=},  # OV
    228: {2024: '=T$225/4*0.83'=, 2025: '=T228'=},  # BRCA
    229: {2024: '=T$225/4*0.17'=, 2025: '=T229'=},  # TNBC
    230: {2024: '=T$225/4'=, 2025: '=T230'=},  # CRC
    231: {2024: '=T$225/4'=, 2025: '=T231'=},  # NSCLC
    232: {2024: 101, 2025: 101},  # Imlygic
    233: {2024: 2, 2025: 7.9},  # Melanoma NCAM+
    234: {2024: 627, 2025: 627},  # Opdualag IV, Mono Antibody
    235: {2024: '=T234'=, 2025: '=U234'=},  # Melanoma
    236: {2024: 200, 2025: 200},  # Pemazyre (Incyte, pemigatinib)
    237: {2024: 200, 2025: 200},  # BTC
    238: {2024: 25, 2025: 25},  # Tibsovo (Servier, ivosidenib)
    239: {2024: 25, 2025: 25},  # BTC
    240: {2024: 50, 2025: 50},  # Lytgobi (Taiho, futibatinib)
    241: {2024: 50, 2025: 50},  # BTC
    242: {2025: 1.5},  # Truseltiq (Helsinn, infigratinib)
    243: {2025: 1.5},  # BTC
    244: {2024: 2200, 2025: 2200},  # Lenvima (Eisai/Merck, lenvatinib)
    245: {2024: 763, 2025: 763},  # EC
    246: {2024: 1650, 2025: 1650},  # HCC
    247: {2024: 1155, 2025: 1155},  # RCC
    248: {2024: 371, 2025: 371},  # Jemperli (GSK, dostarlimab)
    249: {2024: 371, 2025: 371},  # EC
    250: {2024: 245, 2025: 228.6},  # Zepzelca (Jazz, lurbinectedin)
    251: {2024: 245, 2025: 228.6},  # ES-SCLC
    252: {2024: 75, 2025: 75},  # Imdelltra (Amgen, tarlatamab)
    253: {2024: 75, 2025: 75},  # ES-SCLC
    254: {2024: 100, 2025: 100},  # Nexavar (Bayer, sorafenib)
    255: {2024: 50, 2025: 277},  # HCC
    256: {2024: 5, 2025: 67.8},  # RCC
    257: {2024: 1809, 2025: 1809},  # Cabometyx (Exelixis, cabozantinib)
    258: {2024: 480, 2025: 480},  # HCC
    259: {2024: 1230, 2025: 1230},  # RCC
    260: {2024: 8, 2025: 8},  # Amtagvi (Iovance, lifileucel)
    261: {2024: 8, 2025: 8},  # Melanoma NCAM+
    262: {2024: 90, 2025: 427.1},  # Sutent (Pfizer, sunitinib)
    263: {2024: 90, 2025: 427.1},  # RCC
    264: {2024: 180, 2025: 436.7},  # Votrient (Novartis, pazopanib)
    265: {2024: 180, 2025: 436.7},  # RCC
    266: {2024: 784, 2025: 742.7},  # Inlyta (Pfizer, axitinib)
    267: {2024: 784, 2025: 742.7},  # RCC
    268: {2024: 130, 2025: 130},  # Fotivda (EUSA/LG Chem, tivozanib)
    269: {2024: 130, 2025: 130},  # RCC
    270: {2024: 18, 2025: 80.6},  # Afinitor (Novartis, everolimus)
    271: {2024: 18, 2025: 80.6},  # RCC
    272: {2024: 2, 2025: 32.4},  # Torisel (Pfizer, temsirolimus)
    273: {2024: 2, 2025: 32.4},  # RCC
    274: {2024: 400, 2025: 400},  # Welireg (Merck, belzutifan)
    275: {2024: 400, 2025: 400},  # RCC
    276: {2024: 256.46, 2025: 256.46},  # Peg-Intron
    277: {2024: '=T276/2'=, 2025: '=U276/2'=},  # Melanoma
    278: {2024: 99.5, 2025: 99.5},  # INTRON A
    279: {2024: '=T278/4'=, 2025: '=U278/4'=},  # Melanoma
    280: {2024: 82, 2025: 82},  # Zelboraf
    282: {2024: 7, 2025: 19.4},  # Melanoma NCAM+
    283: {2024: 56, 2025: 56},  # Cotellic
    284: {2024: '=T283/2'=, 2025: '=U283/2'=},  # Melanoma
    285: {2024: 1638, 2025: 1638},  # Erbitux
    286: {2024: '=T285/2'=, 2025: '=U285/2'=},  # CRC
    287: {2024: '=T285/2'=, 2025: '=U285/2'=},  # HNSCC
    289: {2024: '=S289'=, 2025: '=T289'=},  # Alimta
    290: {2024: '=S290'=, 2025: '=T290'=},  # Melanoma
    291: {2024: '=S291'=, 2025: '=T291'=},  # CRC
    292: {2024: '=S292'=, 2025: '=T292'=},  # MPM
    293: {2024: '=S293'=, 2025: '=T293'=},  # HNSCC
    295: {2024: 20, 2025: 20},  # Gemzar
    296: {2024: '=S296'=, 2025: '=T296'=},  # Melanoma
    297: {2024: '=S297'=, 2025: '=T297'=},  # CRC
    298: {2024: '=S298'=, 2025: '=T298'=},  # HNSCC
    299: {2024: '=S299'=, 2025: '=T299'=},  # BRCA
    300: {2024: '=S300'=, 2025: '=T300'=},  # TNBC
    301: {2024: '=S301'=, 2025: '=T301'=},  # BLCA
    303: {2024: 1231, 2025: 1231},  # Abraxane
    304: {2024: '=S304'=, 2025: '=T304'=},  # Melanoma
    305: {2024: '=S305'=, 2025: '=T305'=},  # CRC
    306: {2024: '=S306'=, 2025: '=T306'=},  # HNSCC
    307: {2024: '=S307'=, 2025: '=T307'=},  # BRCA
    308: {2024: '=S308'=, 2025: '=T308'=},  # TNBC
    309: {2024: '=S309'=, 2025: '=T309'=},  # GC
    311: {2024: 10064, 2025: 10064},  # Opdivo
    312: {2024: '=S312'=, 2025: '=T312'=},  # ESCA
    313: {2024: '=S313'=, 2025: '=T313'=},  # GC
    314: {2024: '=S314'=, 2025: '=T314'=},  # BLCA
    315: {2024: '=S315'=, 2025: '=T315'=},  # CRC
    316: {2024: '=S316'=, 2025: '=T316'=},  # HNSCC
    317: {2024: '=S317'=, 2025: '=T317'=},  # Melanoma
    318: {2024: '=S318'=, 2025: '=T318'=},  # NSCLC
    319: {2024: 12, 2025: 12},  # BTC
    320: {2024: 10, 2025: 52.5},  # HCC
    321: {2024: 114, 2025: 114},  # Melanoma NCAM+
    322: {2024: 133, 2025: 133},  # MPM
    323: {2024: 1152, 2025: 1152},  # RCC
    325: {2024: 25012, 2025: 25012},  # Keytruda
    326: {2024: '=S326'=, 2025: '=T326'=},  # BLCA
    327: {2024: '=S327'=, 2025: '=T327'=},  # ESCA
    328: {2024: '=S328'=, 2025: '=T328'=},  # GC
    329: {2024: '=S329'=, 2025: '=T329'=},  # TNBC
    330: {2024: '=S330'=, 2025: '=T330'=},  # CRC
    331: {2024: '=S331*T332'=, 2025: '=T331*U332'=},  # HNSCC
    333: {2024: '=S333*T334'=, 2025: '=T333*U334'=},  # Melanoma
    335: {2024: '=S335*T400'=, 2025: '=T335*U400'=},  # NSCLC
    336: {2024: 30, 2025: 30},  # BTC
    337: {2024: 590, 2025: 590},  # EC
    338: {2024: 20, 2025: 71.43},  # HCC
    339: {2024: 342, 2025: 342},  # Melanoma NCAM+
    340: {2024: 28, 2025: 28},  # MPM
    341: {2024: 2430, 2025: 2430},  # RCC
    401: {2024: '=S401'=, 2025: '=T401'=},  # List Price Per Year

    # ══════════════════════════════════════════════════════════════════════
    #  TAM / SUMMARY SECTION (R402-R505)
    # ══════════════════════════════════════════════════════════════════════
    406: {2024: '=SUMIF($D$9:$D$309, $D406, T$9:T$309)'=, 2025: '=SUMIF($D$9:$D$309, $D406, U$9:U$309)'=},  # BTC
    408: {2024: '=S408*S409'=, 2025: '=T408*T409'=},  # CRC
    410: {2024: '=SUMIF($D$9:$D$309, $D410, T$9:T$309)'=, 2025: '=SUMIF($D$9:$D$309, $D410, U$9:U$309)'=},  # NSCLC
    412: {2024: '=SUMIF($D$9:$D$309, $D412, T$9:T$309)'=, 2025: '=SUMIF($D$9:$D$309, $D412, U$9:U$309)'=},  # HNSCC
    414: {2024: '=SUM(T412,T410,T408)'=, 2025: '=SUM(U412,U410,U408)'=},  # CRC/NSCLC/HNSCC
    415: {2024: '=SUMIF($D$9:$D$309, $D415, T$9:T$309)'=, 2025: '=SUMIF($D$9:$D$309, $D415, U$9:U$309)'=},  # Melanoma
    417: {2024: '=(S417/S415)*T415'=, 2025: '=(S417/S415)*U415'=},  # AML Drug TAM Estimation
    421: {2024: '=T145/T410'=, 2025: '=U145/U410'=},  # Krazati Market Share
    427: {2024: 2811, 2025: 2811},  # Lynparza
    433: {2024: 43, 2025: 43},  # Talzenna
    438: {2024: 921, 2025: 921},  # Zejula
    442: {2024: 980, 2025: 980},  # Pluvicto
    444: {2024: 2010.85, 2025: 2010.85},  # Zytiga
    446: {2024: 5038, 2025: 5038},  # Xtandi
    448: {2024: '=S448*O515'=, 2025: '=T448*P515'=},  # mCRPC TAM
    453: {2024: 1652, 2025: 1652},  # Alecensa
    455: {2024: 63, 2025: 63},  # Zykadia
    457: {2024: 538, 2025: 538},  # Lorbrena
    459: {2024: '=S459/F544*G544'=, 2025: '=T459/G544*H544'=},  # Augtyro (Forcasting)
    461: {2024: 87, 2025: 87},  # Rozlytrek
    463: {2024: 451.33, 2025: 451.33},  # Xalkori
    465: {2024: '=S465*(1+T466)'=, 2025: '=T465*(1+U466)'=},  # Total ROS1+ NSCLC
    471: {2024: 1652, 2025: 1652},  # Alecensa
    472: {2024: 63, 2025: 63},  # Zykadia
    473: {2024: 184, 2025: 184},  # Alunbrig
    474: {2024: 455.23, 2025: 455.23},  # Iressa
    475: {2024: 5799, 2025: 5799},  # Tagrisso
    476: {2024: 153, 2025: 511},  # Krazati
    477: {2024: 6, 2025: 6},  # Vizimpo
    478: {2024: 154, 2025: 154},  # Tebrecta
    479: {2024: 538, 2025: 538},  # Lorbrena
    480: {2024: 54, 2025: 54},  # Tepmetko
    481: {2024: 280, 2025: 280},  # Lumakras
    482: {2024: 87, 2025: 87},  # Rozlytrek
    483: {2024: 54.5, 2025: 54.5},  # Pemfexy
    484: {2024: 6069.43, 2025: 6069.43},  # Avastin
    485: {2024: 2238, 2025: 2238},  # Yervoy, Mono Antibody
    486: {2024: 451.33, 2025: 451.33},  # Xalkori
    487: {2024: 836.56, 2025: 836.56},  # Cyramza, Mono Antibody
    488: {2024: 873, 2025: 873},  # Mekinist
    489: {2024: 1922, 2025: 1922},  # Tafinlar
    490: {2024: 4195, 2025: 4195},  # Tecentriq, Mono Antibody
    491: {2024: 254, 2025: 254},  # Retevmo
    492: {2024: 2696, 2025: 2696},  # Enhertu
    493: {2024: 4019, 2025: 4019},  # Imfinzi
    494: {2024: 863, 2025: 863},  # Libtayo
    495: {2024: 193, 2025: 193},  # Mektovi
    496: {2024: 236, 2025: 236},  # Braftovi
    497: {2024: 423, 2025: 423},  # Zirabev
    498: {2024: 543.2, 2025: 543.2},  # Mvasi
    499: {2024: 1638, 2025: 1638},  # Erbitux
    500: {2024: 2098.86, 2025: 2098.86},  # Alimta
    501: {2024: 20, 2025: 20},  # Gemzar
    502: {2024: 1231, 2025: 1231},  # Abraxane
    503: {2024: 10064, 2025: 10064},  # Opdivo
    504: {2024: 25012, 2025: 25012},  # Keytruda
    505: {2024: '=SUM(T471:T504)'=, 2025: '=SUM(U471:U504)'=},  # Total Solid Tumor

    # ══════════════════════════════════════════════════════════════════════
    #  PARAMETERS SECTION (R510-R562) — uses backup row numbers
    # ══════════════════════════════════════════════════════════════════════
    515: {2024: '=S515'=, 2025: '=T515'=},  # Average
    517: {2024: 8116.97, 2025: 8189.27},  # World Population
    518: {2024: '=AVERAGE(P518:S518)'=, 2025: '=T518'=},  # Growth Rate
    520: {2024: '=S520'=, 2025: '=T520'=},  # OV Incidence
    521: {2024: '=S521'=, 2025: '=T521'=},  # TNBC Incidence
    522: {2024: '=S522'=, 2025: '=T522'=},  # BRCA Incidence
    523: {2024: '=S523'=, 2025: '=T523'=},  # BLCA Incidence
    524: {2024: '=S524'=, 2025: '=T524'=},  # GC Incidence
    525: {2024: '=S525'=, 2025: '=T525'=},  # MCC Incidence
    526: {2024: '=S526'=, 2025: '=T526'=},  # GBM Incidence
    527: {2024: '=S527'=, 2025: '=T527'=},  # NSCLC Incidence
    528: {2024: '=S528'=, 2025: '=T528'=},  # Melanoma Incidence
    529: {2024: '=S529'=, 2025: '=T529'=},  # BTC Incidence
    530: {2024: '=S530'=, 2025: '=T530'=},  # EC Incidence
    531: {2024: '=S531'=, 2025: '=T531'=},  # ES-SCLC Incidence
    532: {2024: '=S532'=, 2025: '=T532'=},  # HCC Incidence
    533: {2024: '=S533'=, 2025: '=T533'=},  # Melanoma NCAM+ Incidence
    534: {2024: '=S534'=, 2025: '=T534'=},  # MPM Incidence
    535: {2024: '=S535'=, 2025: '=T535'=},  # RCC Incidence
    551: {2024: '=S551'=, 2025: '=T551'=},  # Average Growth
    552: {2024: '=T551'=, 2025: '=U551'=},  # Best-In-Class Growth
    553: {2024: '=T552'=, 2025: '=U552'=},  # Tier One Growth
    562: {2024: '=S562'=, 2025: '=T562'=},  # COGS/Price
}

# Backward-compatible alias
SOLID_DATA = {r: {y: v for y, v in d.items() if not isinstance(v, str)}
              for r, d in SOLID_CELLS.items()
              if any(not isinstance(v, str) for v in d.values())}


# ══════════════════════════════════════════════════════════════════════════════
#  REVENUE DATA -- TAM Blood
#  All values in MM USD unless noted (CHF/EUR drugs have separate handling).
#  Key: row_number → {2024: value, 2025: value}
#  Blood sheet currently ends at T=2023; need to ADD columns U=2024, V=2025.
# ══════════════════════════════════════════════════════════════════════════════

BLOOD_DATA: Dict[int, Dict[int, float]] = {
    # ── CAR-T ────────────────────────────────────────────────────────────
    16: {2024: 1570, 2025: 1495},  # Yescarta -- Gilead
    20: {2024: 403, 2025: 344},    # Tecartus -- Gilead
    24: {2024: 443, 2025: 381},    # Kymriah -- Novartis
    28: {2024: 747, 2025: 1358},   # Breyanzi -- BMS

    # ── BTK Inhibitors ───────────────────────────────────────────────────
    35: {2024: 3347, 2025: 2869},  # Imbruvica -- AbbVie
    39: {2024: 3139, 2025: 3518},  # Calquence -- AZ
    43: {2024: 2600, 2025: 3900},  # BRUKINSA -- BeiGene

    # ── Other Heme / Oncology ────────────────────────────────────────────
    47: {2024: 1000, 2025: 1050},  # Other Oncology (Eli Lilly) -- est Verzenio+Jaypirca heme
    55: {2024: 2583, 2025: 2792},  # Venclexta -- AbbVie
    59: {2024: 45, 2025: 50},      # Tazverik -- Ipsen (EZH2 inhibitor, growing)
    65: {2024: 100, 2025: 130},    # Monjuvi -- Incyte/MorphoSys (FL approval Jun 2025)
    71: {2024: 300, 2025: 250},    # Xalkori -- Pfizer (same declining trend as Solid R158)
    75: {2024: 55, 2025: 50},      # Zynlonta -- ADC Therapeutics (small, declining)

    # ── IO in Heme ───────────────────────────────────────────────────────
    79: {2024: 9304, 2025: 10049}, # Opdivo -- BMS (total worldwide)
    83: {2024: 910, 2025: 986},    # Gazyva -- Roche (CHF)
    89: {2024: 29482, 2025: 31680},# Keytruda -- Merck (total worldwide)
    93: {2024: 90, 2025: 95},      # Pemazyre -- Incyte (FGFR2 inhibitor, modest growth)
    97: {2024: 1121, 2025: 1470},  # Polivy -- Roche (CHF)
    103: {2024: 1379, 2025: 1251}, # Rituxan -- Roche (CHF)
    109: {2024: 145, 2025: 146},   # Xpovio -- Karyopharm
}


# ══════════════════════════════════════════════════════════════════════════════
#  XML HELPERS (adapted from excel_writer.py)
# ══════════════════════════════════════════════════════════════════════════════

def _cell_addr(col_idx: int, row: int) -> str:
    return f"{get_column_letter(col_idx)}{row}"


def _col_of(addr: str) -> int:
    col_str = "".join(c for c in addr if c.isalpha())
    return column_index_from_string(col_str)


def _num_str(value: Any) -> str:
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value)


def _get_sheet_zip_paths(xlsx_path: Path) -> Dict[str, str]:
    """Return {sheet_name: zip_entry_path} for all worksheets."""
    import xml.etree.ElementTree as ET
    with zipfile.ZipFile(xlsx_path) as zf:
        wb_xml = ET.fromstring(zf.read("xl/workbook.xml"))
        rels_xml = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))

    rid_to_path: Dict[str, str] = {}
    for rel in rels_xml:
        if "worksheet" in rel.get("Type", ""):
            rid = rel.get("Id", "")
            target = rel.get("Target", "")
            rid_to_path[rid] = (
                f"xl/{target}" if not target.startswith("/") else target.lstrip("/")
            )

    sheet_map: Dict[str, str] = {}
    for sheet_elem in wb_xml.findall(f".//{{{_NS_MAIN}}}sheet"):
        name = sheet_elem.get("name", "")
        rid = sheet_elem.get(f"{{{_NS_R}}}id", "")
        if rid in rid_to_path:
            sheet_map[name] = rid_to_path[rid]
    return sheet_map


def _find_cell(xml: str, addr: str) -> Optional[Tuple[int, int, str, str]]:
    """Find a cell by address. Returns (lt, end_pos, open_tag, inner) or None."""
    search = f'r="{addr}"'
    start = 0
    while True:
        pos = xml.find(search, start)
        if pos == -1:
            return None
        lt = xml.rfind("<", 0, pos)
        if lt == -1 or xml[lt + 1] != "c" or xml[lt + 2] not in (" ", "\t", "\n", "/", ">"):
            start = pos + 1
            continue
        tag_end = xml.index(">", lt) + 1
        open_tag = xml[lt:tag_end]
        if xml[tag_end - 2: tag_end] == "/>":
            return (lt, tag_end, open_tag, "")
        c_end = xml.index("</c>", tag_end) + 4
        inner = xml[tag_end:c_end - 4]
        return (lt, c_end, open_tag, inner)


def _get_cell_style(xml: str, addr: str) -> Optional[str]:
    """Extract s="N" style attribute from a cell."""
    found = _find_cell(xml, addr)
    if not found:
        return None
    _, _, open_tag, _ = found
    m = re.search(r's="(\d+)"', open_tag)
    return m.group(1) if m else None


def _get_ref_col_style(xml: str, row: int, ref_col: str) -> Optional[str]:
    """Get the style of the cell at (ref_col, row) to use as reference."""
    return _get_cell_style(xml, f"{ref_col}{row}")


def _set_style_in_tag(open_tag: str, style: Optional[str]) -> str:
    """Ensure open_tag has s="style" attribute (add/replace/remove)."""
    if style is None:
        return open_tag
    if re.search(r's="[^"]*"', open_tag):
        return re.sub(r's="[^"]*"', f's="{style}"', open_tag)
    # Insert s= after r= attribute
    return re.sub(r'(r="[^"]*")', rf'\1 s="{style}"', open_tag)


def _build_cell_xml(addr: str, style: Optional[str], content: str) -> str:
    """Build a complete <c ...>content</c> element."""
    s_attr = f' s="{style}"' if style else ""
    return f'<c r="{addr}"{s_attr}>{content}</c>'


def _patch_numeric_cell(xml: str, addr: str, val_str: str,
                        style: Optional[str] = None) -> str:
    """Find cell r="ADDR" and update its <v> content, or insert new cell.
    If style is provided, set/replace the s= attribute."""
    found = _find_cell(xml, addr)
    if found:
        lt, end_pos, open_tag, inner = found
        # Clean open_tag: remove t="s" etc.
        open_tag_clean = re.sub(r'\s+t="[^"]*"', '', open_tag)
        if style:
            open_tag_clean = _set_style_in_tag(open_tag_clean, style)
        # Self-closing tag
        if open_tag_clean.endswith("/>"):
            new_cell = open_tag_clean[:-2] + f"><v>{val_str}</v></c>"
            return xml[:lt] + new_cell + xml[end_pos:]
        # Has inner content -- replace/add <v>
        if "<v>" in inner:
            v_m = re.search(r'<v>.*?</v>', inner)
            new_inner = inner[:v_m.start()] + f"<v>{val_str}</v>" + inner[v_m.end():]
        else:
            new_inner = inner + f"<v>{val_str}</v>"
        # Remove any <f> tags (we're writing a value, not formula)
        new_inner = re.sub(r'<f[^>]*>.*?</f>', '', new_inner)
        new_inner = re.sub(r'<f[^>]*/>', '', new_inner)
        return xml[:lt] + open_tag_clean + new_inner + "</c>" + xml[end_pos:]

    # Cell not found: insert into the row
    row_num = int("".join(c for c in addr if c.isdigit()))
    return _insert_cell_into_row(xml, row_num, addr, val_str, style)


def _patch_formula_cell(xml: str, addr: str, formula: str,
                        style: Optional[str] = None) -> str:
    """Write a formula cell <c r="ADDR" s="S"><f>FORMULA</f></c>.
    Replaces existing cell content if present, or inserts new cell."""
    found = _find_cell(xml, addr)
    if found:
        lt, end_pos, open_tag, inner = found
        # Clean open_tag: remove t="s", t="shared" etc.
        open_tag_clean = re.sub(r'\s+t="[^"]*"', '', open_tag)
        if style:
            open_tag_clean = _set_style_in_tag(open_tag_clean, style)
        if open_tag_clean.endswith("/>"):
            new_cell = open_tag_clean[:-2] + f"><f>{formula}</f></c>"
        else:
            new_cell = open_tag_clean + f"<f>{formula}</f></c>"
        return xml[:lt] + new_cell + xml[end_pos:]

    # Insert new formula cell
    row_num = int("".join(c for c in addr if c.isdigit()))
    return _insert_cell_into_row(xml, row_num, addr, None, style, formula)


def _insert_cell_into_row(xml: str, row_num: int, addr: str,
                          val_str: Optional[str], style: Optional[str] = None,
                          formula: Optional[str] = None) -> str:
    """Insert a new cell into an existing row."""
    row_search = f'r="{row_num}"'
    row_pos = 0
    while True:
        rp = xml.find(row_search, row_pos)
        if rp == -1:
            logger.warning(f"  Row {row_num} not found in XML; cannot insert {addr}")
            return xml
        lt = xml.rfind("<", 0, rp)
        if lt != -1 and xml[lt + 1: lt + 4] == "row":
            break
        row_pos = rp + 1

    row_tag_end = xml.index(">", lt) + 1

    # Build cell content
    if formula:
        content = f"<f>{formula}</f>"
    elif val_str is not None:
        content = f"<v>{val_str}</v>"
    else:
        content = ""
    new_cell = _build_cell_xml(addr, style, content)

    # Self-closing row? <row r="16" ... />
    if xml[row_tag_end - 2: row_tag_end] == "/>":
        return (
            xml[:row_tag_end - 2]
            + f">{new_cell}</row>"
            + xml[row_tag_end:]
        )

    row_end = xml.index("</row>", row_tag_end)
    row_body = xml[row_tag_end:row_end]
    col_idx = _col_of(addr)

    # Insert in column order among existing cells
    cells = list(re.finditer(r'<c\b[^>]*\br="([A-Z]+\d+)"', row_body))
    insert_at = len(row_body)  # default: append
    for m in cells:
        if _col_of(m.group(1)) > col_idx:
            insert_at = m.start()
            break

    new_body = row_body[:insert_at] + new_cell + row_body[insert_at:]
    return xml[:row_tag_end] + new_body + xml[row_end:]


def _update_dimension(xml: str, new_max_col: str) -> str:
    """Update <dimension ref="A1:T179"/> to extend to new_max_col."""
    match = re.search(r'<dimension\s+ref="([A-Z]+\d+):([A-Z]+)(\d+)"', xml)
    if match:
        start_ref = match.group(1)
        old_col = match.group(2)
        max_row = match.group(3)
        old_col_idx = column_index_from_string(old_col)
        new_col_idx = column_index_from_string(new_max_col)
        if new_col_idx > old_col_idx:
            new_ref = f'<dimension ref="{start_ref}:{new_max_col}{max_row}"'
            xml = xml[:match.start()] + new_ref + xml[match.end():]
            logger.info(f"  Updated dimension: {old_col}{max_row} -> {new_max_col}{max_row}")
    return xml


# ══════════════════════════════════════════════════════════════════════════════
#  FORMULA SHIFTING
# ══════════════════════════════════════════════════════════════════════════════

_COL_REF_RE = re.compile(r'(?<![A-Za-z])(\$?)([A-Z]{1,3})(\$?)(\d+)(?![A-Za-z(])')


def _shift_col(col_letter: str, delta: int) -> str:
    """Shift a column letter by delta positions."""
    idx = column_index_from_string(col_letter) + delta
    if idx < 1:
        idx = 1
    return get_column_letter(idx)


def _shift_formula(formula: str, col_delta: int) -> str:
    """Shift non-absolute column references in a formula by col_delta.
    $-prefixed columns stay unchanged. Rows always unchanged."""
    def replace_ref(m):
        dollar_col = m.group(1)
        col = m.group(2)
        dollar_row = m.group(3)
        row = m.group(4)
        if dollar_col == "$":
            return m.group(0)  # Absolute column -- unchanged
        new_col = _shift_col(col, col_delta)
        return f"{dollar_col}{new_col}{dollar_row}{row}"
    return _COL_REF_RE.sub(replace_ref, formula)


def _resolve_shared_formulas(xml: str, col_letter: str) -> Dict[int, str]:
    """Resolve all formula cells in a column, expanding shared formula refs.
    Returns {row_number: formula_text}."""
    cell_pat = re.compile(
        r'<c\s([^>]*?)/>|<c\s([^>]*?)>(.*?)</c>', re.DOTALL
    )

    # Step 1: Find ALL shared formula masters across ALL columns
    # A master has: t="shared", ref="...", si="N", and formula text
    si_masters: Dict[str, Tuple[str, str]] = {}  # si -> (master_addr, formula)
    for m in cell_pat.finditer(xml):
        if m.group(1) is not None:
            continue  # Self-closing, no formula
        attrs, inner = m.group(2), m.group(3) or ""
        if '<f' not in inner:
            continue
        # Check if this is a shared formula master (has ref= attribute)
        f_m = re.search(r'<f\s[^>]*?ref="[^"]*"[^>]*>(.*?)</f>', inner)
        if not f_m:
            continue
        formula = f_m.group(1).strip()
        if not formula:
            continue
        si_m = re.search(r'si="(\d+)"', inner)
        if not si_m:
            continue
        ref_m = re.search(r'r="([A-Z]+)(\d+)"', attrs)
        if not ref_m:
            continue
        si_masters[si_m.group(1)] = (f"{ref_m.group(1)}{ref_m.group(2)}", formula)

    # Step 2: Collect all cells in the target column
    result: Dict[int, str] = {}
    for m in cell_pat.finditer(xml):
        if m.group(1) is not None:
            attrs, inner = m.group(1), ""
        else:
            attrs, inner = m.group(2), m.group(3) or ""
        ref_m = re.search(r'r="([A-Z]+)(\d+)"', attrs)
        if not ref_m or ref_m.group(1) != col_letter:
            continue
        row = int(ref_m.group(2))
        if "<f" not in inner:
            continue

        # Check for explicit formula text (non-shared or shared master)
        f_m = re.search(r'<f[^>]*>(.*?)</f>', inner)
        if f_m and f_m.group(1).strip():
            result[row] = f_m.group(1)
            continue

        # Shared formula reference -- resolve from master
        si_m = re.search(r'si="(\d+)"', inner)
        if si_m and si_m.group(1) in si_masters:
            si = si_m.group(1)
            master_addr, master_formula = si_masters[si]
            master_col = "".join(c for c in master_addr if c.isalpha())
            col_delta = (column_index_from_string(col_letter)
                         - column_index_from_string(master_col))
            master_row = int("".join(c for c in master_addr if c.isdigit()))
            row_delta = row - master_row
            shifted = master_formula
            if col_delta != 0:
                shifted = _shift_formula(shifted, col_delta)
            if row_delta != 0:
                shifted = _shift_formula_rows(shifted, row_delta)
            result[row] = shifted

    return result


def _shift_formula_rows(formula: str, row_delta: int) -> str:
    """Shift non-absolute row references in a formula by row_delta."""
    def replace_ref(m):
        dollar_col = m.group(1)
        col = m.group(2)
        dollar_row = m.group(3)
        row = m.group(4)
        if dollar_row == "$":
            return m.group(0)
        new_row = int(row) + row_delta
        if new_row < 1:
            new_row = 1
        return f"{dollar_col}{col}{dollar_row}{new_row}"
    return _COL_REF_RE.sub(replace_ref, formula)


def _extend_formulas(xml: str, src_col: str, tgt_col: str,
                     dry_run: bool, sheet_name: str) -> Tuple[str, int]:
    """Copy formulas from src_col to tgt_col where tgt_col is missing them.
    Returns (updated_xml, count)."""
    col_delta = (column_index_from_string(tgt_col)
                 - column_index_from_string(src_col))
    src_formulas = _resolve_shared_formulas(xml, src_col)
    tgt_formulas = _resolve_shared_formulas(xml, tgt_col)

    count = 0
    for row, formula in sorted(src_formulas.items()):
        if row <= 6:
            continue  # Skip header rows
        if row in tgt_formulas:
            continue  # Target already has formula
        shifted = _shift_formula(formula, col_delta)
        src_style = _get_cell_style(xml, f"{src_col}{row}")
        tgt_addr = f"{tgt_col}{row}"
        if dry_run:
            logger.info(f"  [DRY-RUN] {sheet_name} {tgt_addr}: f={shifted[:60]}... s={src_style}")
        else:
            xml = _patch_formula_cell(xml, tgt_addr, shifted, src_style)
        count += 1

    return xml, count


# ══════════════════════════════════════════════════════════════════════════════
#  PARAMETER CARRY-FORWARD (copy 2023 values to 2024/2025 for non-drug rows)
# ══════════════════════════════════════════════════════════════════════════════

def _carry_forward_values(xml: str, ref_col: str, tgt_cols: List[str],
                          data_rows: Set[int], dry_run: bool,
                          sheet_name: str) -> Tuple[str, int]:
    """For rows that have a value in ref_col but NOT in any tgt_col,
    copy the value forward. Skip rows in data_rows (drug revenue rows)
    and formula rows."""
    cell_pat = re.compile(r'<c\s([^>]*?)/>|<c\s([^>]*?)>(.*?)</c>', re.DOTALL)

    # Collect ref_col value cells (non-formula, non-empty)
    ref_values: Dict[int, str] = {}
    ref_styles: Dict[int, Optional[str]] = {}
    for m in cell_pat.finditer(xml):
        if m.group(1) is not None:
            attrs, inner = m.group(1), ""
        else:
            attrs, inner = m.group(2), m.group(3) or ""
        ref_m = re.search(r'r="([A-Z]+)(\d+)"', attrs)
        if not ref_m or ref_m.group(1) != ref_col:
            continue
        row = int(ref_m.group(2))
        if row <= 6:
            continue
        if "<f" in inner:
            continue  # Skip formula cells
        v_m = re.search(r'<v>(.*?)</v>', inner)
        if not v_m:
            continue
        s_m = re.search(r's="(\d+)"', attrs)
        ref_values[row] = v_m.group(1)
        ref_styles[row] = s_m.group(1) if s_m else None

    # Check which target cols already have data
    tgt_has: Dict[str, Set[int]] = {col: set() for col in tgt_cols}
    for m in cell_pat.finditer(xml):
        if m.group(1) is not None:
            attrs, inner = m.group(1), ""
        else:
            attrs, inner = m.group(2), m.group(3) or ""
        ref_m = re.search(r'r="([A-Z]+)(\d+)"', attrs)
        if not ref_m:
            continue
        col = ref_m.group(1)
        if col not in tgt_has:
            continue
        row = int(ref_m.group(2))
        if "<v>" in inner or "<f" in inner:
            tgt_has[col].add(row)

    count = 0
    for row in sorted(ref_values.keys()):
        if row in data_rows:
            continue  # Drug revenue rows handled separately
        style = ref_styles[row]
        val = ref_values[row]
        for col in tgt_cols:
            if row in tgt_has[col]:
                continue
            addr = f"{col}{row}"
            if dry_run:
                logger.info(f"  [DRY-RUN] {sheet_name} {addr}: carry-forward v={val[:15]} s={style}")
            else:
                xml = _patch_numeric_cell(xml, addr, val, style)
            count += 1

    return xml, count


# ══════════════════════════════════════════════════════════════════════════════
#  STYLE FIX (retroactively fix cells patched in Pass 1/2 without styles)
# ══════════════════════════════════════════════════════════════════════════════

def _replicate_empty_cells(xml: str, ref_col: str, tgt_cols: List[str],
                           dry_run: bool, sheet_name: str) -> Tuple[str, int]:
    """Replicate empty styled cells from ref_col to tgt_cols where target has no cell.
    These are section headers, spacers, dividers with borders/backgrounds but no data."""
    cell_pat = re.compile(r'<c\s([^>]*?)/>|<c\s([^>]*?)>(.*?)</c>', re.DOTALL)

    # Find empty styled cells in ref_col (no <v> and no <f>)
    ref_empty: Dict[int, str] = {}  # row -> style
    for m in cell_pat.finditer(xml):
        if m.group(1) is not None:
            attrs, inner = m.group(1), ""
        else:
            attrs, inner = m.group(2), m.group(3) or ""
        ref_m = re.search(r'r="([A-Z]+)(\d+)"', attrs)
        if not ref_m or ref_m.group(1) != ref_col:
            continue
        row = int(ref_m.group(2))
        if "<v>" in inner or "<f" in inner:
            continue
        s_m = re.search(r's="(\d+)"', attrs)
        if s_m:
            ref_empty[row] = s_m.group(1)

    # Check which rows already have cells in target columns
    tgt_has: Dict[str, Set[int]] = {col: set() for col in tgt_cols}
    for m in cell_pat.finditer(xml):
        if m.group(1) is not None:
            attrs = m.group(1)
        else:
            attrs = m.group(2)
        ref_m = re.search(r'r="([A-Z]+)(\d+)"', attrs)
        if not ref_m:
            continue
        col = ref_m.group(1)
        if col in tgt_has:
            tgt_has[col].add(int(ref_m.group(2)))

    count = 0
    for row in sorted(ref_empty.keys()):
        style = ref_empty[row]
        for col in tgt_cols:
            if row in tgt_has[col]:
                continue
            addr = f"{col}{row}"
            if dry_run:
                logger.info(f"  [DRY-RUN] {sheet_name} {addr}: replicate empty cell s={style}")
            else:
                xml = _insert_cell_into_row(xml, row, addr, None, style)
            count += 1

    return xml, count


def _fix_existing_styles(xml: str, ref_col: str, fix_cols: List[str],
                         dry_run: bool, sheet_name: str) -> Tuple[str, int]:
    """For value cells in fix_cols that have wrong or missing style,
    set style to match the same row's ref_col cell."""
    cell_pat = re.compile(r'<c\s([^>]*?)/>|<c\s([^>]*?)>(.*?)</c>', re.DOTALL)

    # Collect ref_col styles
    ref_styles: Dict[int, str] = {}
    for m in cell_pat.finditer(xml):
        if m.group(1) is not None:
            attrs = m.group(1)
        else:
            attrs = m.group(2)
        ref_m = re.search(r'r="([A-Z]+)(\d+)"', attrs)
        if not ref_m or ref_m.group(1) != ref_col:
            continue
        row = int(ref_m.group(2))
        s_m = re.search(r's="(\d+)"', attrs)
        if s_m:
            ref_styles[row] = s_m.group(1)

    count = 0
    for m in cell_pat.finditer(xml):
        if m.group(1) is not None:
            attrs, inner = m.group(1), ""
        else:
            attrs, inner = m.group(2), m.group(3) or ""
        ref_m = re.search(r'r="([A-Z]+)(\d+)"', attrs)
        if not ref_m:
            continue
        col = ref_m.group(1)
        if col not in fix_cols:
            continue
        row = int(ref_m.group(2))
        if row <= 6 or row not in ref_styles:
            continue
        if "<f" in inner:
            continue  # Don't fix formula cells here
        if "<v>" not in inner and not attrs.strip().endswith("/"):
            continue  # Empty cell
        v_m = re.search(r'<v>(.*?)</v>', inner)
        if not v_m:
            continue

        expected_style = ref_styles[row]
        s_m = re.search(r's="(\d+)"', attrs)
        current_style = s_m.group(1) if s_m else None
        if current_style == expected_style:
            continue  # Already correct

        addr = f"{col}{row}"
        val = v_m.group(1)
        if dry_run:
            logger.info(f"  [DRY-RUN] {sheet_name} {addr}: style {current_style} -> {expected_style}")
        else:
            xml = _patch_numeric_cell(xml, addr, val, expected_style)
        count += 1

    return xml, count


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN PATCHING LOGIC
# ══════════════════════════════════════════════════════════════════════════════

def _discover_solid_year_cols(ws) -> Dict[int, int]:
    """Read TAM Solid row 6 to find {year: col_index}."""
    year_col: Dict[int, int] = {}
    for col_idx in range(1, 35):  # A through AH
        val = ws.cell(row=6, column=col_idx).value
        if isinstance(val, (int, float)) and 2000 <= int(val) <= 2050:
            year_col[int(val)] = col_idx
    return year_col


def _discover_blood_year_cols(ws) -> Dict[int, int]:
    """Read TAM Blood row 6 to find {year: col_index}."""
    year_col: Dict[int, int] = {}
    for col_idx in range(1, 30):
        val = ws.cell(row=6, column=col_idx).value
        if isinstance(val, (int, float)) and 2000 <= int(val) <= 2050:
            year_col[int(val)] = col_idx
    return year_col


def patch_solid_sheet(xml: str, year_col: Dict[int, int], dry_run: bool) -> str:
    """Patch TAM Solid sheet: data + style fix + carry-forward."""
    # Solid: S=2023 (col 19), T=2024 (col 20), U=2025 (col 21)
    ref_col = get_column_letter(year_col[2023])  # "S"
    col_2024 = get_column_letter(year_col[2024])  # "T"
    col_2025 = get_column_letter(year_col[2025])  # "U"
    data_rows = set(SOLID_CELLS.keys())

    # Step 1: Write data from SOLID_CELLS (both numeric and formula cells)
    n_num = 0
    n_fml = 0
    for row_num, year_vals in SOLID_CELLS.items():
        style = _get_ref_col_style(xml, row_num, ref_col)
        for year, value in year_vals.items():
            if year not in year_col:
                logger.warning(f"  Year {year} not in Solid+MM columns; skipping R{row_num}")
                continue
            col_idx = year_col[year]
            addr = _cell_addr(col_idx, row_num)
            if isinstance(value, str) and value.startswith('='):
                # Formula cell
                if dry_run:
                    logger.info(f"  [DRY-RUN] Solid R{row_num} {addr} f={value[:60]} s={style}")
                else:
                    xml = _patch_formula_cell(xml, addr, value[1:], style)
                n_fml += 1
            else:
                # Numeric cell
                val_str = _num_str(value)
                if dry_run:
                    logger.info(f"  [DRY-RUN] Solid R{row_num} {addr} = {value} s={style}")
                else:
                    xml = _patch_numeric_cell(xml, addr, val_str, style)
                n_num += 1
    logger.info(f"TAM Solid: {n_num} numeric + {n_fml} formula cells patched")

    # Step 2: Fix styles on previously-patched cells
    xml, n_style = _fix_existing_styles(xml, ref_col, [col_2024, col_2025],
                                        dry_run, "Solid")
    logger.info(f"TAM Solid: {n_style} styles fixed")

    # Step 3: Carry forward parameter values (non-drug, non-formula rows)
    # (SOLID_CELLS already provides all formulas for T/U, so _extend_formulas
    #  is no longer needed for these columns)
    xml, n_cf = _carry_forward_values(xml, ref_col, [col_2024, col_2025],
                                      data_rows, dry_run, "Solid")
    logger.info(f"TAM Solid: {n_cf} values carried forward")

    # Step 4: Replicate empty styled cells (section headers, spacers, dividers)
    xml, n_ec = _replicate_empty_cells(xml, ref_col, [col_2024, col_2025],
                                       dry_run, "Solid")
    logger.info(f"TAM Solid: {n_ec} empty cells replicated")

    return xml


def patch_blood_sheet(
    xml: str,
    year_col: Dict[int, int],
    max_data_row: int,
    dry_run: bool,
) -> str:
    """Patch TAM Blood sheet: data + style fix + formula extension + carry-forward."""
    # Blood: T=2023 (col 20), U=2024 (col 21), V=2025 (col 22)
    need_2024_col = 2024 not in year_col
    need_2025_col = 2025 not in year_col

    if need_2024_col:
        max_year = max(year_col.keys())
        max_col = year_col[max_year]
        col_2024 = max_col + 1
        year_col[2024] = col_2024
        logger.info(f"  Adding 2024 column at {get_column_letter(col_2024)} (col {col_2024})")
    else:
        col_2024 = year_col[2024]

    if need_2025_col:
        col_2025 = year_col[2024] + 1
        year_col[2025] = col_2025
        logger.info(f"  Adding 2025 column at {get_column_letter(col_2025)} (col {col_2025})")
    else:
        col_2025 = year_col[2025]

    ref_col = get_column_letter(year_col[2023])   # "T"
    col_2024_l = get_column_letter(col_2024)      # "U"
    col_2025_l = get_column_letter(col_2025)      # "V"
    data_rows = set(BLOOD_DATA.keys())

    # Add year header cells with style matching 2023 header
    hdr_style = _get_cell_style(xml, f"{ref_col}6")
    if need_2024_col:
        addr = _cell_addr(col_2024, 6)
        if not dry_run:
            xml = _patch_numeric_cell(xml, addr, "2024", hdr_style)
    if need_2025_col:
        addr = _cell_addr(col_2025, 6)
        if not dry_run:
            xml = _patch_numeric_cell(xml, addr, "2025", hdr_style)

    if not dry_run and (need_2024_col or need_2025_col):
        new_max_col = get_column_letter(max(year_col.values()))
        xml = _update_dimension(xml, new_max_col)

    # Step 1: Write drug revenue data with correct style
    changes = 0
    for row_num, year_vals in BLOOD_DATA.items():
        style = _get_ref_col_style(xml, row_num, ref_col)
        for year, value in year_vals.items():
            if year not in year_col:
                continue
            col_idx = year_col[year]
            addr = _cell_addr(col_idx, row_num)
            val_str = _num_str(value)
            if dry_run:
                logger.info(f"  [DRY-RUN] Blood R{row_num} {addr} = {value} s={style}")
            else:
                xml = _patch_numeric_cell(xml, addr, val_str, style)
            changes += 1
    logger.info(f"TAM Blood: {changes} data cells patched")

    # Step 2: Fix styles on previously-patched cells
    xml, n_style = _fix_existing_styles(xml, ref_col, [col_2024_l, col_2025_l],
                                        dry_run, "Blood")
    logger.info(f"TAM Blood: {n_style} styles fixed")

    # Step 3: Extend ALL formulas from T→U and T→V
    xml, n_tu = _extend_formulas(xml, ref_col, col_2024_l, dry_run, "Blood")
    logger.info(f"TAM Blood: {n_tu} formulas extended T->U")
    xml, n_tv = _extend_formulas(xml, ref_col, col_2025_l, dry_run, "Blood")
    logger.info(f"TAM Blood: {n_tv} formulas extended T->V")

    # Step 4: Carry forward parameter values
    xml, n_cf = _carry_forward_values(xml, ref_col, [col_2024_l, col_2025_l],
                                      data_rows, dry_run, "Blood")
    logger.info(f"TAM Blood: {n_cf} values carried forward")

    # Step 5: Replicate empty styled cells (section headers, spacers, dividers)
    xml, n_ec = _replicate_empty_cells(xml, ref_col, [col_2024_l, col_2025_l],
                                       dry_run, "Blood")
    logger.info(f"TAM Blood: {n_ec} empty cells replicated")

    return xml


def _apply_patches(
    xlsx_path: Path,
    solid_patcher,
    blood_patcher,
    dry_run: bool,
) -> None:
    """Apply patches to both TAM sheets via zip manipulation."""
    sheet_zip_paths = _get_sheet_zip_paths(xlsx_path)

    solid_zip = sheet_zip_paths.get(SHEET_SOLID)
    blood_zip = sheet_zip_paths.get(SHEET_BLOOD)

    if not solid_zip:
        logger.error(f"Sheet '{SHEET_SOLID}' not found in workbook")
        return
    if not blood_zip:
        logger.error(f"Sheet '{SHEET_BLOOD}' not found in workbook")
        return

    logger.info(f"  Solid+MM sheet: {solid_zip}")
    logger.info(f"  Blood sheet: {blood_zip}")

    if dry_run:
        # For dry run, just read and process without writing
        with zipfile.ZipFile(xlsx_path) as zf:
            solid_xml = zf.read(solid_zip).decode("utf-8")
            blood_xml = zf.read(blood_zip).decode("utf-8")
        solid_patcher(solid_xml)
        blood_patcher(blood_xml)
        logger.info("Dry run complete -- no changes written.")
        return

    # ── Read, patch, and rewrite ─────────────────────────────────────────
    modified: Dict[str, bytes] = {}

    with zipfile.ZipFile(xlsx_path) as zf:
        solid_xml = zf.read(solid_zip).decode("utf-8")
        blood_xml = zf.read(blood_zip).decode("utf-8")

    solid_xml = solid_patcher(solid_xml)
    blood_xml = blood_patcher(blood_xml)

    modified[solid_zip] = solid_xml.encode("utf-8")
    modified[blood_zip] = blood_xml.encode("utf-8")

    # ── Add fullCalcOnLoad to workbook.xml ───────────────────────────────
    with zipfile.ZipFile(xlsx_path) as zf:
        wb_xml_bytes = zf.read("xl/workbook.xml")
    wb_xml_str = wb_xml_bytes.decode("utf-8")
    if "fullCalcOnLoad" not in wb_xml_str:
        wb_xml_str = wb_xml_str.replace("<calcPr", '<calcPr fullCalcOnLoad="1"', 1)
        logger.info("Added fullCalcOnLoad='1' to workbook.xml")
    modified["xl/workbook.xml"] = wb_xml_str.encode("utf-8")

    # ── Write new zip ────────────────────────────────────────────────────
    tmp_path = xlsx_path.with_suffix(".~tam_patch.xlsx")
    with zipfile.ZipFile(xlsx_path, "r") as zin:
        with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                if item.filename == "xl/calcChain.xml":
                    continue  # removed — references stripped below
                if item.filename in modified:
                    zout.writestr(item, modified[item.filename])
                else:
                    zout.writestr(item, zin.read(item.filename))

    try:
        tmp_path.replace(xlsx_path)
    except PermissionError:
        import os
        os.remove(str(xlsx_path))
        tmp_path.rename(xlsx_path)
    logger.info(f"Patch applied -> {xlsx_path}")


# ══════════════════════════════════════════════════════════════════════════════
#  SUMMARY REPORT
# ══════════════════════════════════════════════════════════════════════════════

def generate_summary(dry_run: bool) -> str:
    """Generate a summary of all changes."""
    lines = [
        "# TAM Update Summary",
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"Mode: {'DRY RUN' if dry_run else 'APPLIED'}",
        "",
        "## TAM Solid Changes",
        "",
        "| Row | Drug | FY2024 ($M) | FY2025 ($M) | Source |",
        "|-----|------|-------------|-------------|--------|",
    ]

    DRUG_NAMES_SOLID = {
        11: "Tafinlar+Mek", 19: "Padcev", 25: "Jelmyto", 27: "Bavencio",
        29: "Trodelvy", 44: "Zoladex", 47: "Halaven", 50: "Aybintio",
        53: "Herceptin", 56: "Herzuma", 62: "Kanjinti", 65: "Trazimera",
        68: "Ontruzant", 75: "Elahere", 79: "Rubraca", 82: "Tevimbra",
        84: "Eloxatin", 86: "Vectibix", 88: "Fruzaqla", 99: "Stivarga",
        101: "Lonsurf", 106: "Vegzelma", 110: "Tukysa",
        112: "Rybrevant+Lazcluze", 116: "Imjudo", 118: "Alecensa",
        122: "Alunbrig", 128: "Tagrisso", 134: "Tebrecta", 136: "Lorbrena",
        140: "Lumakras", 143: "Rozlytrek", 145: "Pemfexy", 147: "Avastin",
        154: "Yervoy", 158: "Xalkori", 160: "Cyramza", 167: "Tafinlar(dup)",
        170: "Tecentriq", 174: "Retevmo", 176: "Enhertu", 181: "Imfinzi",
        183: "Libtayo", 185: "Mektovi", 188: "Braftovi", 192: "Zirabev",
        199: "Mvasi", 208: "Opdualag", 216: "Cotellic", 218: "Erbitux",
        222: "Alimta", 227: "Gemzar", 235: "Abraxane", 244: "Opdivo",
        253: "Keytruda",
        # Rows shifted +32 after HL/MM expansion (R288+)
        288: "Abecma", 290: "Adcetris", 295: "Carvykti", 297: "Darzalex",
        299: "Elrexfio", 301: "Kyprolis", 303: "Ninlaro", 305: "Pomalyst",
        308: "Revlimid", 314: "Sarclisa", 316: "Talvey", 318: "Tecvayli",
        # Original drugs shifted +32
        318: "Lynparza", 324: "Talzenna", 329: "Zejula",
        333: "Pluvicto", 335: "Zytiga", 337: "Xtandi",
    }
    # Summary rows (R403-R449, shifted +58 from R345-R391)
    for r in range(403, 450):
        if r in SOLID_DATA:
            DRUG_NAMES_SOLID[r] = f"[S] {DRUG_NAMES_SOLID.get(r, f'Row{r}')}"
    _SUMMARY_NAMES = {
        403: "Lorbrena", 406: "Rozlytrek", 408: "Xalkori", 416: "Alecensa",
        418: "Alunbrig", 420: "Tagrisso", 423: "Tebrecta", 424: "Lorbrena",
        426: "Lumakras", 427: "Rozlytrek", 428: "Pemfexy", 429: "Avastin",
        430: "Yervoy", 431: "Xalkori", 432: "Cyramza", 434: "Tafinlar",
        435: "Tecentriq", 436: "Retevmo", 437: "Enhertu", 438: "Imfinzi",
        439: "Libtayo", 440: "Mektovi", 441: "Braftovi", 442: "Zirabev",
        443: "Mvasi", 444: "Erbitux", 445: "Alimta", 446: "Gemzar",
        447: "Abraxane", 448: "Opdivo", 449: "Keytruda",
    }
    for r, n in _SUMMARY_NAMES.items():
        DRUG_NAMES_SOLID[r] = f"[S] {n}"
    SOURCES_SOLID: Dict[int, str] = {}

    for row_num in sorted(SOLID_DATA.keys()):
        vals = SOLID_DATA[row_num]
        name = DRUG_NAMES_SOLID.get(row_num, f"Row {row_num}")
        v24 = f"{vals.get(2024, 'N/A'):,}" if 2024 in vals else "N/A"
        v25 = f"{vals.get(2025, 'N/A'):,}" if 2025 in vals else "N/A"
        src = SOURCES_SOLID.get(row_num, "")
        lines.append(f"| {row_num} | {name} | {v24} | {v25} | {src} |")

    lines.extend([
        "",
        "## TAM Blood Changes",
        "",
        "| Row | Drug | FY2024 ($M) | FY2025 ($M) | Source |",
        "|-----|------|-------------|-------------|--------|",
    ])

    DRUG_NAMES_BLOOD = {
        16: "Yescarta", 20: "Tecartus", 24: "Kymriah", 28: "Breyanzi",
        35: "Imbruvica", 39: "Calquence", 43: "BRUKINSA",
        47: "Lilly Other Onc", 55: "Venclexta", 59: "Tazverik",
        65: "Monjuvi", 71: "Xalkori", 75: "Zynlonta",
        79: "Opdivo", 83: "Gazyva", 89: "Keytruda", 93: "Pemazyre",
        97: "Polivy", 103: "Rituxan", 109: "Xpovio",
    }
    SOURCES_BLOOD: Dict[int, str] = {}

    for row_num in sorted(BLOOD_DATA.keys()):
        vals = BLOOD_DATA[row_num]
        name = DRUG_NAMES_BLOOD.get(row_num, f"Row {row_num}")
        v24 = f"{vals.get(2024, 'N/A'):,}" if 2024 in vals else "N/A"
        v25 = f"{vals.get(2025, 'N/A'):,}" if 2025 in vals else "N/A"
        src = SOURCES_BLOOD.get(row_num, "")
        lines.append(f"| {row_num} | {name} | {v24} | {v25} | {src} |")

    lines.extend([
        "",
        "## Notes",
        "- All values in MM USD unless noted",
        "- Roche drugs reported in CHF (treated as ~USD for these purposes)",
        f"- Solid: {len(SOLID_DATA)} rows updated ({sum(len(v) for v in SOLID_DATA.values())} cells)",
        f"- Blood: {len(BLOOD_DATA)} rows updated ({sum(len(v) for v in BLOOD_DATA.values())} cells)",
        "- Summary rows R371-R417 mirror main drug rows",
        "- Biosimilar estimates based on market trends (Pfizer oncology biosims -35% ops in 2024)",
        "- Small drugs (<$100M) use conservative growth/decline estimates",
    ])

    return "\n".join(lines)


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Fill TAM Solid and TAM Blood with FY2024/FY2025 drug revenue data"
    )
    parser.add_argument("--dry-run", action="store_true", help="Preview changes without writing")
    parser.add_argument("--file", type=str, default=str(DEFAULT_FILE), help="Path to Excel file")
    args = parser.parse_args()

    xlsx_path = Path(args.file)
    if not xlsx_path.exists():
        logger.error(f"File not found: {xlsx_path}")
        return

    logger.info(f"Opening: {xlsx_path}")
    logger.info(f"Mode: {'DRY RUN' if args.dry_run else 'LIVE'}")

    # ── Backup ───────────────────────────────────────────────────────────
    if not args.dry_run:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = xlsx_path.with_name(f"{xlsx_path.stem}_pre_tam_update_{ts}.xlsx")
        shutil.copy2(xlsx_path, backup_path)
        logger.info(f"Backup: {backup_path}")

    # ── Read sheet structure with openpyxl (READ-ONLY, never .save()) ────
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True, read_only=True)

    ws_solid = wb[SHEET_SOLID]
    solid_year_col = _discover_solid_year_cols(ws_solid)
    logger.info(f"Solid+MM year columns: {solid_year_col}")

    ws_blood = wb[SHEET_BLOOD]
    blood_year_col = _discover_blood_year_cols(ws_blood)
    blood_max_row = ws_blood.max_row or 179
    logger.info(f"Blood year columns: {blood_year_col}")
    logger.info(f"Blood max row: {blood_max_row}")

    wb.close()

    # ── Create patchers ──────────────────────────────────────────────────
    def solid_patcher(xml: str) -> str:
        return patch_solid_sheet(xml, solid_year_col, args.dry_run)

    def blood_patcher(xml: str) -> str:
        return patch_blood_sheet(xml, blood_year_col, blood_max_row, args.dry_run)

    # ── Apply ────────────────────────────────────────────────────────────
    _apply_patches(xlsx_path, solid_patcher, blood_patcher, args.dry_run)

    # ── Summary report ───────────────────────────────────────────────────
    summary = generate_summary(args.dry_run)
    summary_path = Path("/home/nazdaq_44sun/Investment/auto_dcf/tam_update_summary.md")
    summary_path.write_text(summary)
    logger.info(f"Summary written: {summary_path}")

    # ── Print summary stats ──────────────────────────────────────────────
    n_solid = sum(len(v) for v in SOLID_DATA.values())
    n_blood = sum(len(v) for v in BLOOD_DATA.values())
    print(f"\n{'='*60}")
    print(f"TAM Update {'(DRY RUN)' if args.dry_run else 'COMPLETE'}")
    print(f"{'='*60}")
    print(f"  Solid+MM: {len(SOLID_DATA)} drugs, {n_solid} cells")
    print(f"  Blood:    {len(BLOOD_DATA)} drugs, {n_blood} cells")
    print(f"  Total:    {n_solid + n_blood} cells")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
