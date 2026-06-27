#!/usr/bin/env python3
"""
extract_tam_solid.py — pull the PARAMETER inputs out of the live TAM Solid sheet
into datastore/seed/params_seed.json.

The drug-market rows are sourced from the clean per-drug JSON files (tam_data/,
tam_hl_mm_data/) by build_datastore.py. This script only harvests the parameter
inputs that live in the sheet's "Parameters" section and are not yet in JSON:
incidence rates, world population, peer-drug growth ramps, and Xpovio COGS lines.

Excel's surgical XML patcher leaves cached formula results inconsistent, so we
read CONSTANT input cells (not formula results) wherever possible.

Usage:
    python datastore/extract_tam_solid.py \
        --xlsx "/mnt/c/Users/yzsun/Desktop/DD/CMPX/DCF CMPX.xlsx"
"""
import argparse
import json
import os

import openpyxl

# TAM Solid columns: F(6)=2010 ... AH(34)=2038
COL2YEAR = {c: 2010 + (c - 6) for c in range(6, 35)}

# Peer-drug growth ramps: sheet row -> drug name (R541..R549, "<Drug> Sale")
GROWTH_ROWS = {541: "Alimta", 543: "Alecensa", 545: "Iressa",
               547: "Alunbrig", 549: "Tagrisso"}

# Maturity curves the model actually uses: factor by years-since-launch, per tier
# (R551 Average / R552 Best-In-Class / R553 Tier-One, across cols F..AH).
MATURITY_ROWS = {551: "AVG", 552: "BIC", 553: "T1"}

# Xpovio cost lines (constants live in the K / 2015 column)
COGS_CELLS = {  # item -> (row, col)
    "xpovio_net_sale":       (556, 11),
    "xpovio_manufacturing":  (557, 11),
    "xpovio_selling_and_ga": (558, 11),
    "xpovio_ga":             (559, 11),
}


def first_numeric(ws, row):
    for c in range(6, 35):
        v = ws.cell(row, c).value
        if isinstance(v, (int, float)):
            return v
    return None


def row_series(ws, row):
    return {COL2YEAR[c]: ws.cell(row, c).value
            for c in range(6, 35)
            if isinstance(ws.cell(row, c).value, (int, float))}


def extract(xlsx_path):
    wb_val = openpyxl.load_workbook(xlsx_path, data_only=True)   # cached results
    wb_raw = openpyxl.load_workbook(xlsx_path, data_only=False)  # literal constants
    ws_v = wb_val["TAM Solid"]
    ws_r = wb_raw["TAM Solid"]

    seed = {"incidence": {}, "globals": {}, "reference_growth": {},
            "maturity_curve": {}, "cogs": {}}

    # Incidence rates (R520-R538): label "<CODE> Incidence" -> rate
    for r in range(520, 539):
        label = ws_v.cell(r, 4).value
        if label and "Incidence" in str(label):
            code = str(label).replace(" Incidence", "").strip()
            seed["incidence"][code] = first_numeric(ws_v, r)

    # Global scalars
    seed["globals"]["world_population_2010_m"] = (
        ws_v.cell(517, 6).value or first_numeric(ws_v, 517))
    seed["globals"]["population_growth_rate"] = first_numeric(ws_v, 518)

    # Peer-drug growth ramps (cached values are reliable here)
    for r, name in GROWTH_ROWS.items():
        seed["reference_growth"][name] = row_series(ws_v, r)

    # Maturity curves: factor per year-offset (1..29) per tier
    for r, tier in MATURITY_ROWS.items():
        curve = [ws_v.cell(r, c).value for c in range(6, 35)]
        seed["maturity_curve"][tier] = [v for v in curve
                                        if isinstance(v, (int, float))]

    # Xpovio COGS constants (read literal cell, not a formula result)
    for item, (r, c) in COGS_CELLS.items():
        v = ws_r.cell(r, c).value
        seed["cogs"][item] = v if isinstance(v, (int, float)) else None

    return seed


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default="/mnt/c/Users/yzsun/Desktop/DD/CMPX/DCF CMPX.xlsx")
    ap.add_argument("--out", default=os.path.join(os.path.dirname(__file__),
                                                  "seed", "params_seed.json"))
    args = ap.parse_args()

    seed = extract(args.xlsx)
    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    with open(args.out, "w") as f:
        json.dump(seed, f, indent=2)

    print(f"[extract] incidence indications : {len(seed['incidence'])}")
    print(f"[extract] growth reference drugs: {len(seed['reference_growth'])}")
    print(f"[extract] maturity curves       : "
          f"{ {t: len(v) for t, v in seed['maturity_curve'].items()} }")
    print(f"[extract] cogs lines            : "
          f"{sum(v is not None for v in seed['cogs'].values())}/4")
    print(f"[extract] wrote {args.out}")


if __name__ == "__main__":
    main()
