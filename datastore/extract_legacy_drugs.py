#!/usr/bin/env python3
"""
extract_legacy_drugs.py — backfill the legacy solid-tumor drug market (rows
R9-R341 of the live TAM Solid sheet) into the Data Center JSON schema.

These ~80 drugs exist only as cell formulas (core/tam_solid_cells.py + the live
sheet), not as per-drug JSON. We harvest them here so Layer 1 covers the whole
solid market, not just the 14 HL/MM JSON drugs.

Row classification (deterministic, matches how the sheet was built):
  * DRUG HEADER row  -> column-D font has an explicit color (theme1); its cells
                        hold the drug's TOTAL net sales per year.
  * INDICATION row   -> column-D font is default/None; its cells hold the
                        revenue attributed to that indication.

For each drug we store the TOTAL per year (drug_revenue) and a per-indication
`share` = median over years of (indication_value / total). Shares may sum to <1
(the sheet does not always break out 100% of a drug) — that is faithful.

Output: datastore/seed/legacy_drugs.json  (gitignored; regenerate any time).

Usage:
    python datastore/extract_legacy_drugs.py \
        --xlsx "/mnt/c/Users/yzsun/Desktop/DD/CMPX/DCF CMPX.xlsx"
"""
import argparse
import json
import os
import statistics

import openpyxl

COL2YEAR = {c: 2010 + (c - 6) for c in range(6, 35)}   # F=2010 ... AH=2038
ROW_START, ROW_END = 9, 341                              # Oncology Drug Market

# A row in column D is an INDICATION breakdown iff its label is in this closed
# oncology vocabulary; otherwise it is a DRUG HEADER (open vocabulary of brands).
# Built from the frequency analysis of the live sheet (freq>=2 values are
# indications, minus drug brands like "Tafinlar", plus single-use indications).
INDICATIONS = {
    "NSCLC", "BRCA", "CRC", "Melanoma", "TNBC", "RCC", "GC", "BLCA", "HCC", "OV",
    "Melanoma NCAM+", "BTC", "HNSCC", "GBM", "PRAD", "ESCA", "MPM", "ES-SCLC",
    "mCRPC", "EC", "Throid Cancer", "Thyroid Cancer", "SCLC",
    "MM", "HL", "PTCL", "DLBCL", "FL", "MCL", "MZL", "ALCL", "MCC", "MDS",
    "Kaposi", "Other", "Others",
}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default="/mnt/c/Users/yzsun/Desktop/DD/CMPX/DCF CMPX.xlsx")
    ap.add_argument("--out", default=os.path.join(os.path.dirname(__file__),
                                                  "seed", "legacy_drugs.json"))
    ap.add_argument("--report", action="store_true", help="print structure report only")
    args = ap.parse_args()

    ws_v = openpyxl.load_workbook(args.xlsx, data_only=True)["TAM Solid"]  # values

    def series(r):
        return {COL2YEAR[c]: ws_v.cell(r, c).value
                for c in range(6, 35)
                if isinstance(ws_v.cell(r, c).value, (int, float))}

    drugs, current, anomalies = [], None, []
    for r in range(ROW_START, ROW_END + 1):
        d = ws_v.cell(r, 4).value
        if d is None or str(d).strip() == "":
            continue
        name = str(d).strip()
        if name in INDICATIONS:                          # INDICATION breakdown
            if current is None:
                anomalies.append((r, name, "breakdown before any drug header"))
                continue
            current["_breakdowns"].append({"row": r, "name": name, "vals": series(r)})
        else:                                            # DRUG HEADER
            current = {"row": r, "drug_name": name, "revenues": series(r),
                       "_breakdowns": []}
            drugs.append(current)

    # derive per-indication share = median(indication/total) over years total>0
    out = []
    for dr in drugs:
        tot = dr["revenues"]
        inds = []
        for b in dr["_breakdowns"]:
            ratios = [b["vals"][y] / tot[y] for y in b["vals"]
                      if y in tot and tot[y] not in (0, None) and b["vals"][y] is not None]
            if not ratios:
                continue
            inds.append({"name": b["name"], "share": round(statistics.median(ratios), 6)})
        if not tot:
            anomalies.append((dr["row"], dr["drug_name"], "drug header with no revenue"))
            continue
        out.append({"drug_name": dr["drug_name"],
                    "revenues": {str(y): v for y, v in tot.items()},
                    "indications": inds})

    # ---- report ----
    print(f"[legacy] drug headers : {len(drugs)}")
    print(f"[legacy] drugs kept   : {len(out)} (with revenue)")
    print(f"[legacy] total indication splits: {sum(len(d['indications']) for d in out)}")
    no_split = [d["drug_name"] for d in out if not d["indications"]]
    if no_split:
        print(f"[legacy] {len(no_split)} drug(s) with NO indication split: "
              f"{', '.join(no_split[:8])}{'...' if len(no_split) > 8 else ''}")
    if anomalies:
        print(f"[legacy] ANOMALIES ({len(anomalies)}):")
        for a in anomalies[:15]:
            print("   ", a)
    print("[legacy] sample drugs:")
    for d in out[:6]:
        yrs = sorted(int(y) for y in d["revenues"])
        print(f"    {d['drug_name'][:34]:36} {yrs[0]}-{yrs[-1]}  "
              f"inds={[i['name'] for i in d['indications']]}")

    if args.report:
        return
    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    with open(args.out, "w") as f:
        json.dump(out, f, indent=2)
    print(f"[legacy] wrote {args.out}")


if __name__ == "__main__":
    main()
