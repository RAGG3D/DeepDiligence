#!/usr/bin/env python3
"""
extract_peer_views.py — normalize the Peer Views sheet into the Data Center.

Peer Views is a stack of drug-vs-drug clinical-readout tables. Each SECTION is
delimited by a marker "X" in column A; column D of that row is the section title
(indication / setting, e.g. "+ KEYTRUDA 1L Melanoma (NCAM+)"). Within a section:
  * a ticker row   : columns E.. hold "<TICKER> US Equity"
  * a drug row     : columns E.. hold the drug names/codes; column D = the anchor
  * metric rows    : column D = metric label (ORR, Median PFS, Median OS, ...),
                     columns E.. = each drug's value.

The BIC/T1/AVG rating is encoded as each drug column's FILL COLOR
(theme 9 = green = BIC, theme 8 = blue = T1, theme 7 = olive = AVG). We DECODE it
into an explicit text `rating` column so the data no longer depends on cell color.

Output: datastore/seed/peer_views.json  (gitignored). Produces two logical tables
at build time: peer_drug (section, drug, ticker, rating) and peer_metric
(section, drug, metric, value).

Usage:
    python datastore/extract_peer_views.py \
        --xlsx "/mnt/c/Users/yzsun/Desktop/DD/CMPX/DCF CMPX.xlsx"
"""
import argparse
import json
import os

import openpyxl
from openpyxl.utils import get_column_letter

THEME_RATING = {9: "BIC", 8: "T1", 7: "AVG"}   # fill theme -> rating
DRUG_COLS = range(5, 41)                        # E .. AN


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default="/mnt/c/Users/yzsun/Desktop/DD/CMPX/DCF CMPX.xlsx")
    ap.add_argument("--out", default=os.path.join(os.path.dirname(__file__),
                                                  "seed", "peer_views.json"))
    ap.add_argument("--report", action="store_true")
    args = ap.parse_args()

    ws_s = openpyxl.load_workbook(args.xlsx)["Peer Views"]                 # fills
    ws_v = openpyxl.load_workbook(args.xlsx, data_only=True)["Peer Views"]  # values
    maxr = ws_v.max_row

    def val(r, c):
        return ws_v.cell(r, c).value

    def text(r, c):
        v = val(r, c)
        return "" if v is None else str(v).strip()

    def rating(r, c):
        f = ws_s.cell(r, c).fill
        if f is None or f.patternType is None:
            return None
        fg = f.fgColor
        if fg is not None and fg.type == "theme":
            return THEME_RATING.get(fg.theme)
        return None

    # section boundaries: rows where column A == "X"
    heads = [r for r in range(1, maxr + 1) if text(r, 1) == "X"]
    bounds = [(h, (heads[i + 1] - 1 if i + 1 < len(heads) else maxr))
              for i, h in enumerate(heads)]

    sections = []
    for hr, end in bounds:
        title = text(hr, 4) or f"section_R{hr}"

        # ticker row(s): any drug col ends with "US Equity"
        ticker_rows = [r for r in range(hr + 1, end + 1)
                       if any(text(r, c).endswith("US Equity") for c in DRUG_COLS)]
        anchor_after = max(ticker_rows) if ticker_rows else hr
        ticker_row = ticker_rows[0] if ticker_rows else None

        # drug row: first non-blank data row after the ticker block
        drug_row = None
        for r in range(anchor_after + 1, end + 1):
            if any(text(r, c) for c in DRUG_COLS):
                drug_row = r
                break
        if drug_row is None:
            continue

        active = [c for c in DRUG_COLS if text(drug_row, c)]
        if not active:
            continue

        drugs = []
        for c in active:
            drugs.append({
                "col": get_column_letter(c),
                "drug": text(drug_row, c),
                "ticker": (text(ticker_row, c) if ticker_row else "").replace(" US Equity", ""),
                "rating": rating(drug_row, c) or (rating(ticker_row, c) if ticker_row else None),
                "metrics": {},
            })

        for r in range(drug_row + 1, end + 1):
            metric = text(r, 4)
            if not metric:
                continue
            for d, c in zip(drugs, active):
                v = val(r, c)
                if v is not None and str(v).strip() != "":
                    d["metrics"][metric] = str(v).strip()

        sections.append({"section": title, "anchor": text(drug_row, 4),
                         "drugs": drugs})

    # ---- report ----
    n_drugs = sum(len(s["drugs"]) for s in sections)
    n_metrics = sum(len(d["metrics"]) for s in sections for d in s["drugs"])
    rated = [d["rating"] for s in sections for d in s["drugs"]]
    from collections import Counter
    print(f"[peer] sections: {len(sections)} | drugs: {n_drugs} | metric cells: {n_metrics}")
    print(f"[peer] rating distribution: {dict(Counter(rated))}")
    print("[peer] sample sections:")
    for s in sections[:6]:
        ds = ", ".join(f"{d['drug']}({d['rating']})" for d in s["drugs"][:5])
        print(f"    {s['section'][:40]:42} [{len(s['drugs'])} drugs] {ds}")

    if args.report:
        return
    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    with open(args.out, "w") as f:
        json.dump(sections, f, indent=2, default=str)
    print(f"[peer] wrote {args.out}")


if __name__ == "__main__":
    main()
