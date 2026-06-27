#!/usr/bin/env python3
"""
extract_tam_blood.py — backfill the blood-cancer drug roster from the live
TAM Blood sheet into the Data Center JSON schema.

TAM Blood is built differently from TAM Solid: drugs are grouped by MECHANISM /
market segment rather than split into indication shares —
  * CAR-T        (R15-R33)   : Yescarta, Tecartus, Kymriah, Breyanzi
  * BTKi         (R34-R53)   : Imbruvica, Calquence, Brukinsa, Jaypirca
  * Heme-Other   (R54-R179)  : Venclexta, Rituxan, Gazyva, Polivy, Xpovio, ...
  * HL / MM      (R180-R224) : red-font drug -> indication blocks
Year columns: J=2013 ... T=2023 (no projection columns).

We capture each drug's NET SALES (the raw fact) and tag it with its segment as
the indication. The sheet's bespoke list-price market-sizing (the UPPERCASE
"purchases" rows and "Purchase In Total" rows) is Layer-2 math left for later.

Output: datastore/seed/blood_drugs.json  (gitignored). Drugs already in the
datastore (HL/MM JSON set) are deduped at build time, so this mainly adds the
lymphoma/leukemia roster + a few new MM drugs (Velcade, Empliciti, Blenrep).

Usage:
    python datastore/extract_tam_blood.py \
        --xlsx "/mnt/c/Users/yzsun/Desktop/DD/CMPX/DCF CMPX.xlsx"
"""
import argparse
import json
import os
import re

import openpyxl

COL2YEAR = {c: 2013 + (c - 10) for c in range(10, 21)}   # J=2013 ... T=2023

SEGMENTS = [(15, 34, "CAR-T"), (34, 54, "BTKi"), (54, 180, "Heme-Other")]
HLMM_START = 180

DRUG_ROW_RE = re.compile(r"(net sale|sales\s*/|sales\b)", re.IGNORECASE)


def parse_name(full):
    base = re.split(r"\bNet Sale\b|\bSales\b|\(", full)[0].strip().strip(",")
    company = molecule = None
    if "(" in full and ")" in full:
        inside = full[full.index("(") + 1: full.rindex(")")]
        parts = [p.strip() for p in inside.split(",") if p.strip()]
        company = parts[0] if parts else None
        molecule = ", ".join(parts[1:]) if len(parts) > 1 else None
    return base, company, molecule


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default="/mnt/c/Users/yzsun/Desktop/DD/CMPX/DCF CMPX.xlsx")
    ap.add_argument("--out", default=os.path.join(os.path.dirname(__file__),
                                                  "seed", "blood_drugs.json"))
    ap.add_argument("--report", action="store_true")
    args = ap.parse_args()

    ws = openpyxl.load_workbook(args.xlsx, data_only=True)["TAM Blood"]

    def series(r):
        return {COL2YEAR[c]: ws.cell(r, c).value
                for c in range(10, 21)
                if isinstance(ws.cell(r, c).value, (int, float))}

    out = []

    # --- mechanism segments: capture "<Drug> Net Sale/Sales" rows --------------
    for lo, hi, seg in SEGMENTS:
        for r in range(lo, hi):
            d = ws.cell(r, 4).value
            if not d:
                continue
            name = str(d).strip()
            if "Purchase In Total" in name or name.isupper():
                continue
            if DRUG_ROW_RE.search(name):
                base, company, molecule = parse_name(name)
                rev = series(r)
                if base and rev:
                    out.append({"drug_name": base, "company": company,
                                "molecule": molecule, "segment": seg,
                                "revenues": {str(y): v for y, v in rev.items()},
                                "indications": [{"name": seg, "share": 1.0}]})

    # --- HL / MM red-font blocks: drug -> indication (value on the IND row) -----
    current = None
    for r in range(HLMM_START, ws.max_row + 1):
        d = ws.cell(r, 4).value
        if not d:
            continue
        name = str(d).strip()
        if name.startswith("Total") or name.endswith("Incidence"):
            current = None
            continue
        if name in ("HL", "MM"):
            rev = series(r)
            if current and rev:                       # value row for current drug
                base, company, molecule = current
                out.append({"drug_name": base, "company": company,
                            "molecule": molecule, "segment": name,
                            "revenues": {str(y): v for y, v in rev.items()},
                            "indications": [{"name": name, "share": 1.0}]})
                current = None
            # else: section header — ignore
        elif "(" in name:                             # a drug header row
            current = parse_name(name)

    # ---- report ----
    by_seg = {}
    for d in out:
        by_seg.setdefault(d["segment"], []).append(d["drug_name"])
    print(f"[blood] drugs extracted: {len(out)}")
    for seg, names in by_seg.items():
        print(f"  {seg:11}: {len(names):2}  {', '.join(names[:8])}"
              f"{'...' if len(names) > 8 else ''}")

    if args.report:
        return
    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    with open(args.out, "w") as f:
        json.dump(out, f, indent=2)
    print(f"[blood] wrote {args.out}")


if __name__ == "__main__":
    main()
