"""
excel_writer.py – Surgical xlsx patcher for the DCF Excel model.

Writing strategy
─────────────────
We NEVER call openpyxl's .save(), which discards xl/sharedStrings.xml,
xl/calcChain.xml, VBA, drawings, and other parts.  Instead:

1. Open the workbook with openpyxl (read-only) to discover the row/column
   layout of 'FY DATA K USD' and 'FY DATA'.
2. Build a write-plan:  {sheet_name: [(row, col, value), ...]}
3. Use zipfile + ElementTree to surgically patch ONLY the target sheet XML
   entries while copying every other zip entry byte-for-byte.

This ensures sharedStrings.xml, calcChain.xml, styles, VBA, comments,
drawings, and all other parts are preserved exactly as-is.
"""

import logging
import shutil
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

logger = logging.getLogger(__name__)

# ── xlsx XML namespaces ───────────────────────────────────────────────────────
_NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NS_R    = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_NS_REL  = "http://schemas.openxmlformats.org/package/2006/relationships"

# Register so ElementTree uses the correct (non-mangled) prefixes on output
ET.register_namespace("",      _NS_MAIN)
ET.register_namespace("r",     _NS_R)
ET.register_namespace("mc",    "http://schemas.openxmlformats.org/markup-compatibility/2006")
ET.register_namespace("x14ac", "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac")
ET.register_namespace("xr",    "http://schemas.microsoft.com/office/spreadsheetml/2014/revision")
ET.register_namespace("xr2",   "http://schemas.microsoft.com/office/spreadsheetml/2015/revision2")
ET.register_namespace("xr3",   "http://schemas.microsoft.com/office/spreadsheetml/2016/revision3")

# ── sheet / column constants ──────────────────────────────────────────────────
FIRST_DATA_COL = 6   # column F (1-based)
MAX_YEAR_COLS  = 6   # F … K
HEADER_ROW     = 4   # row containing year integers
COL_B, COL_C, COL_D = 2, 3, 4

SHEET_KUSD = "FY DATA K USD"
SHEET_MM   = "FY DATA"


# ── low-level helpers ─────────────────────────────────────────────────────────

def _backup(xlsx_path: Path) -> Path:
    ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
    dst = xlsx_path.with_name(f"{xlsx_path.stem}_backup_{ts}.xlsx")
    shutil.copy2(xlsx_path, dst)
    logger.info(f"Backup created: {dst}")
    return dst


def _is_formula(v) -> bool:
    return isinstance(v, str) and v.startswith("=")


def _cell_addr(col_idx: int, row: int) -> str:
    return f"{get_column_letter(col_idx)}{row}"


def _col_of(addr: str) -> int:
    """Return 1-based column index from a cell address like 'F8'."""
    col_str = "".join(c for c in addr if c.isalpha())
    return column_index_from_string(col_str)


# ── openpyxl read helpers ─────────────────────────────────────────────────────

def _build_year_col_map(ws) -> Dict[int, int]:
    """
    Read the DATA header row to find {year: col_index}.
    Col F contains a raw year integer; G-K have formulas like =F4+1.
    """
    row = ws[HEADER_ROW]
    year_col: Dict[int, int] = {}
    base_year = base_col = None

    for col_idx in range(FIRST_DATA_COL, FIRST_DATA_COL + MAX_YEAR_COLS):
        val = row[col_idx - 1].value
        if isinstance(val, (int, float)) and 2000 <= int(val) <= 2099:
            base_year = int(val)
            base_col  = col_idx
            year_col[base_year] = col_idx
        elif base_year is not None and val is None:
            break
        elif base_year is not None and isinstance(val, str) and val.startswith("="):
            year_col[base_year + (col_idx - base_col)] = col_idx

    return year_col


def _build_row_map(ws_formula, ws_data, year_col: Dict[int, int]) -> Dict[Tuple, int]:
    """
    Build {(col_c, col_b, col_d): row_number} for writable (non-formula) rows.

    ws_formula : data_only=False  – to detect formula rows in data columns
    ws_data    : data_only=True   – to get the actual B/C/D computed values
    """
    first_data_col = min(year_col.values()) if year_col else FIRST_DATA_COL
    row_map: Dict[Tuple, int] = {}

    for row_num in range(1, ws_formula.max_row + 1):
        b_val = ws_data.cell(row=row_num, column=COL_B).value
        c_val = ws_data.cell(row=row_num, column=COL_C).value
        d_val = ws_data.cell(row=row_num, column=COL_D).value
        f_val = ws_formula.cell(row=row_num, column=first_data_col).value

        if c_val is None or d_val is None:
            continue
        if _is_formula(f_val):          # skip subtotal/formula rows
            continue
        if isinstance(c_val, str) and c_val.startswith("="):
            continue                     # unresolved formula in C

        if b_val == "":
            b_val = None                 # normalise empty string → None

        key = (c_val, b_val, d_val)
        if key not in row_map:
            row_map[key] = row_num
        else:
            logger.warning(
                f"Duplicate key {key} at rows {row_map[key]} and {row_num}; keeping first."
            )

    return row_map


# ── write-plan collectors ─────────────────────────────────────────────────────

def _collect_kusd_patches(
    wb:             openpyxl.Workbook,
    wb_data:        openpyxl.Workbook,
    financial_data: Dict[Tuple, Dict[int, Optional[float]]],
    years:          List[int],
    rename_map:     Optional[Dict[Tuple, str]] = None,
    note_details:   Optional[Dict] = None,
    target_sheet:   str = SHEET_KUSD,
) -> Tuple[List[Tuple[int, int, Any]], set]:
    """Return (patches, hidden_rows) for the target data sheet.

    target_sheet: sheet name to target (default: 'FY DATA K USD',
                  use 'FY DATA' for direct MM USD mode)

    patches:     [(row, col, value)] for cells that should be written
    hidden_rows: set of row numbers to hide (blank ISN/BSN rows + SBC rows)

    Handles:
    - Standard numeric data writes (None → 0 for placeholder clearing)
    - Row renames via rename_map (writes text to col D via inlineStr)
    - ISN/BSN notes sections: uses note_details from 10-K HTML parsing
      when available, falling back to aggregate XBRL totals
    - SBC rows: zeroed out and hidden
    - Blank note rows: hidden
    """
    if rename_map is None:
        rename_map = {}
    if note_details is None:
        note_details = {}

    ws      = wb[target_sheet]
    ws_data = wb_data[target_sheet]
    year_col = _build_year_col_map(ws)
    row_map  = _build_row_map(ws, ws_data, year_col)

    # ── Determine ALL year columns (including ones beyond `years` list) ────
    # E.g. template has 2020-2025 but `years` = [2020..2024] → also write 0
    # for 2025 to clear template placeholders.
    all_template_years = sorted(year_col.keys())

    logger.info(
        f"K USD sheet: {len(year_col)} year columns {all_template_years}, "
        f"{len(row_map)} writable rows"
    )

    # ── Build reverse rename: financial_data key → Excel row_map key ────────
    new_to_original: Dict[str, Tuple] = {}
    for orig_key, new_name in rename_map.items():
        if len(orig_key) == 3:
            col_c, col_b, original_d = orig_key
            new_to_original[f"{col_c}|{col_b}|{new_name}"] = orig_key

    # ── Discover ISN/BSN note row positions ─────────────────────────────────
    # Scan the sheet directly instead of using row_map (which deduplicates
    # rows that were previously renamed to "").
    first_data_col = min(year_col.values()) if year_col else FIRST_DATA_COL
    isn_rows = []
    bsn_rows = []
    for row_num in range(1, ws.max_row + 1):
        c_val = ws_data.cell(row=row_num, column=COL_C).value
        if c_val not in ("ISN", "BSN"):
            continue
        b_val = ws_data.cell(row=row_num, column=COL_B).value
        if b_val == "":
            b_val = None
        if b_val is not None:
            continue  # ISN/BSN data rows have col_b=None; skip SUM rows
        f_val = ws.cell(row=row_num, column=first_data_col).value
        if _is_formula(f_val):
            continue  # skip SUM formula rows
        d_val = ws_data.cell(row=row_num, column=COL_D).value
        if c_val == "ISN":
            isn_rows.append((b_val, row_num, d_val))
        elif c_val == "BSN":
            bsn_rows.append((b_val, row_num, d_val))

    isn_rd_rows = [r for r in isn_rows if r[1] < 40]
    isn_ga_rows = [r for r in isn_rows if r[1] >= 40]
    bsn_ppe_rows = [r for r in bsn_rows if r[1] < 100]
    bsn_acc_rows = [r for r in bsn_rows if r[1] >= 100]

    logger.info(
        f"  Notes rows: ISN R&D={len(isn_rd_rows)}, ISN G&A={len(isn_ga_rows)}, "
        f"BSN PP&E={len(bsn_ppe_rows)}, BSN Accrued={len(bsn_acc_rows)}"
    )

    patches: List[Tuple[int, int, Any]] = []
    hidden_rows: set = set()
    show_rows: set = set()  # rows to explicitly unhide (may have been hidden before)

    # ── Discover and zero-out Share-Based Compensation rows ───────────────
    for row_num in range(1, ws.max_row + 1):
        c_val = ws_data.cell(row=row_num, column=COL_C).value
        if c_val != "ISN":
            continue
        d_val = ws_data.cell(row=row_num, column=COL_D).value
        if not isinstance(d_val, str):
            continue
        if "share-based compensation" not in d_val.lower():
            continue
        # Zero all year columns, blank col D, and hide
        patches.append((row_num, COL_D, ("text", "")))
        for year in all_template_years:
            if year in year_col:
                patches.append((row_num, year_col[year], 0))
        hidden_rows.add(row_num)
        logger.info(f"  SBC row R{row_num} zeroed and will be hidden")

    # ── Helper: write data + rename for a notes section ─────────────────────
    def _write_notes_section(
        section_rows: list,
        data_map: dict,
    ):
        """Write data and renames for a notes section.

        data_map: {0-based position: (new_col_d_name, year_vals_dict_or_None)}
        If year_vals is None, write 0 for all years.
        year_vals can be a dict {year: value} or a financial_data key tuple.
        Unused rows (not in data_map) are zeroed and hidden.
        """
        for idx, (sub, row_num, col_d) in enumerate(section_rows):
            if idx in data_map:
                new_name, year_vals = data_map[idx]
                show_rows.add(row_num)  # ensure visible (may have been hidden)
            else:
                # Keep template placeholder name if it exists;
                # if D is blank (old template), set a filler name
                if col_d and str(col_d).strip():
                    new_name = None   # keep existing placeholder name
                else:
                    new_name = f"Reserved {idx+1}"  # fill blank D
                year_vals = None
                hidden_rows.add(row_num)  # unused row → hide

            # Rename col D
            if new_name is not None:
                patches.append((row_num, COL_D, ("text", new_name)))
                logger.debug(f"  Rename R{row_num} col D → '{new_name}'")

            # Write numeric data for ALL template year columns
            for year in all_template_years:
                if year not in year_col:
                    continue
                col_idx = year_col[year]
                cell = ws.cell(row=row_num, column=col_idx)
                if _is_formula(cell.value):
                    continue

                val = 0
                if year_vals is not None:
                    if isinstance(year_vals, dict):
                        val = year_vals.get(year, 0) or 0
                    elif isinstance(year_vals, tuple) and year_vals in financial_data:
                        v = financial_data[year_vals].get(year)
                        val = v if v is not None else 0
                patches.append((row_num, col_idx, val))

    # ── Process ISN R&D Notes ───────────────────────────────────────────────
    if isn_rd_rows:
        rd_detail = note_details.get("rd")
        rd_total_key = ("IS", None, "Research And Development")
        rd_total_vals = financial_data.get(rd_total_key, {})
        if rd_detail:
            # Use parsed detail items from 10-K; reserve LAST row for remainder
            n_rows = len(isn_rd_rows)
            max_detail = n_rows - 1  # reserve last slot for "Other"
            data_map = {}
            for i, (name, yr_vals) in enumerate(rd_detail):
                if i >= max_detail:
                    break
                data_map[i] = (name, yr_vals)
            # Compute remainder = IS_total - SUM(detail items written)
            remainder: Dict[int, float] = {}
            for year in all_template_years:
                total = rd_total_vals.get(year) or 0
                detail_sum = sum(
                    (data_map[j][1].get(year, 0) if isinstance(data_map[j][1], dict) else 0)
                    for j in data_map
                )
                remainder[year] = total - detail_sum
            data_map[n_rows - 1] = ("Other Research And Development", remainder)
            logger.info(
                f"  R&D notes: {len(data_map) - 1} detail items from 10-K + "
                f"1 remainder row (of {n_rows} available rows)"
            )
            _write_notes_section(isn_rd_rows, data_map)
        else:
            # Fallback: aggregate total from XBRL
            rd_data_key = ("ISN", 1, "Research And Development Expenses")
            _write_notes_section(isn_rd_rows, {
                0: ("Research And Development Expenses", rd_data_key),
            })

    # ── Process ISN G&A Notes ───────────────────────────────────────────────
    if isn_ga_rows:
        ga_detail = note_details.get("ga")
        ga_total_key = ("IS", None, "General And Administrative")
        ga_total_vals = financial_data.get(ga_total_key, {})
        if ga_detail:
            n_rows = len(isn_ga_rows)
            max_detail = n_rows - 1  # reserve last slot for "Other"
            data_map = {}
            for i, (name, yr_vals) in enumerate(ga_detail):
                if i >= max_detail:
                    break
                data_map[i] = (name, yr_vals)
            # Compute remainder = IS_total - SUM(detail items written)
            ga_remainder: Dict[int, float] = {}
            for year in all_template_years:
                total = ga_total_vals.get(year) or 0
                detail_sum = sum(
                    (data_map[j][1].get(year, 0) if isinstance(data_map[j][1], dict) else 0)
                    for j in data_map
                )
                ga_remainder[year] = total - detail_sum
            data_map[n_rows - 1] = ("Other General And Administrative", ga_remainder)
            logger.info(
                f"  G&A notes: {len(data_map) - 1} detail items from 10-K + "
                f"1 remainder row"
            )
            _write_notes_section(isn_ga_rows, data_map)
        else:
            # Fallback: aggregate total from XBRL
            ga_data_key = ("ISN", 1, "General And Administrative Expenses")
            _write_notes_section(isn_ga_rows, {
                0: ("General And Administrative Expenses", ga_data_key),
            })

    # ── Process BSN PP&E Notes ──────────────────────────────────────────────
    if bsn_ppe_rows:
        ppe_detail = note_details.get("ppe")
        if ppe_detail:
            n_rows = len(bsn_ppe_rows)
            data_map = {}
            # Separate depreciation (negative) from gross items
            dep_items = []
            gross_items = []
            for name, yr_vals in ppe_detail:
                name_lower = name.lower()
                if "depreciat" in name_lower or "amortiz" in name_lower:
                    dep_items.append((name, yr_vals))
                else:
                    gross_items.append((name, yr_vals))

            # Gross items fill from position 0
            for i, (name, yr_vals) in enumerate(gross_items):
                if i >= n_rows - len(dep_items):
                    break
                data_map[i] = (name, yr_vals)

            # Depreciation items fill from the end
            for j, (name, yr_vals) in enumerate(dep_items):
                pos = n_rows - len(dep_items) + j
                if pos >= 0:
                    data_map[pos] = (name, yr_vals)

            logger.info(
                f"  PP&E notes: {len(gross_items)} gross + "
                f"{len(dep_items)} depreciation items from 10-K"
            )
            _write_notes_section(bsn_ppe_rows, data_map)
        else:
            # Fallback: XBRL gross + depreciation
            # If XBRL gross is all None, compute: Gross = Net + |Depreciation|
            ppe_gross_key = ("BSN", 1, "Property, Plant And Equipment, Gross")
            ppe_dep_key = ("BSN", 2, "Accumulated Depreciation")
            ppe_net_key = ("BS", None, "Property And Equipment, Net")

            ppe_gross_vals = financial_data.get(ppe_gross_key, {})
            ppe_dep_vals = financial_data.get(ppe_dep_key, {})
            ppe_net_vals = financial_data.get(ppe_net_key, {})

            gross_all_none = all(
                ppe_gross_vals.get(y) is None for y in all_template_years
            )
            if gross_all_none and ppe_net_vals:
                # Compute gross = net - dep (dep stored as negative)
                computed_gross: Dict[int, float] = {}
                for year in all_template_years:
                    net = ppe_net_vals.get(year) or 0
                    dep = ppe_dep_vals.get(year) or 0  # negative
                    computed_gross[year] = net - dep  # net + |dep|
                logger.info("  PP&E: computing Gross = Net + |Depreciation| (XBRL gross missing)")
                n = len(bsn_ppe_rows)
                _write_notes_section(bsn_ppe_rows, {
                    0: ("Property, Plant And Equipment, Gross", computed_gross),
                    n - 1: ("Accumulated Depreciation", ppe_dep_vals),
                })
            else:
                n = len(bsn_ppe_rows)
                _write_notes_section(bsn_ppe_rows, {
                    0: ("Property, Plant And Equipment, Gross", ppe_gross_key),
                    n - 1: ("Accumulated Depreciation", ppe_dep_key),
                })

    # ── Process BSN Accrued Notes ───────────────────────────────────────────
    if bsn_acc_rows:
        acc_detail = note_details.get("accrued")
        acc_total_key = ("BS", None, "Accrued Expenses And Other Current Liabilities")
        acc_total_vals = financial_data.get(acc_total_key, {})
        if acc_detail:
            n_rows = len(bsn_acc_rows)
            max_detail = n_rows - 1  # reserve last slot for remainder

            # If more items than available slots, combine smallest into "Other"
            if len(acc_detail) > max_detail:
                sorted_items = sorted(
                    acc_detail,
                    key=lambda x: sum(abs(v) for v in x[1].values()),
                    reverse=True,
                )
                keep = sorted_items[: max_detail - 1]
                merge = sorted_items[max_detail - 1 :]
                other_vals: Dict[int, float] = {}
                for _, yr_vals in merge:
                    for y, v in yr_vals.items():
                        other_vals[y] = other_vals.get(y, 0) + v
                existing_other_idx = None
                for ki, (kn, _) in enumerate(keep):
                    if kn.lower() == "other":
                        existing_other_idx = ki
                        break
                if existing_other_idx is not None:
                    old_name, old_vals = keep[existing_other_idx]
                    merged = {
                        y: old_vals.get(y, 0) + other_vals.get(y, 0)
                        for y in set(old_vals) | set(other_vals)
                    }
                    keep[existing_other_idx] = (old_name, merged)
                else:
                    keep.append(("Other", other_vals))
                acc_detail = keep

            data_map = {}
            for i, (name, yr_vals) in enumerate(acc_detail):
                if i >= max_detail:
                    break
                data_map[i] = (name, yr_vals)

            # Compute remainder = BS total - SUM(detail items)
            acc_remainder: Dict[int, float] = {}
            for year in all_template_years:
                total = acc_total_vals.get(year) or 0
                detail_sum = sum(
                    (data_map[j][1].get(year, 0) if isinstance(data_map[j][1], dict) else 0)
                    for j in data_map
                )
                acc_remainder[year] = total - detail_sum
            data_map[n_rows - 1] = ("Other Accrued Expenses", acc_remainder)

            logger.info(
                f"  Accrued notes: {len(data_map) - 1} detail items from 10-K + "
                f"1 remainder row (of {n_rows} available rows)"
            )
            _write_notes_section(bsn_acc_rows, data_map)
        else:
            # Fallback: XBRL employee + balancing + other
            emp_key = ("BSN", 1, "Accrued Employee Benefits")
            bal_key = ("BSN", 3, "Other Accrued Liabilities (Balancing)")
            other_key = ("BSN", 2, "Other Accrued Liabilities")
            n = len(bsn_acc_rows)
            _write_notes_section(bsn_acc_rows, {
                0: ("Accrued Employee Benefits", emp_key),
                1: ("Other Accrued Liabilities (Balancing)", bal_key),
                n - 1: ("Other Accrued Liabilities", other_key),
            })

    # ── Mark ISN/BSN data keys as handled (skip in main loop below) ─────────
    handled_keys = set()
    for key in financial_data:
        col_c = key[0]
        if col_c in ("ISN", "BSN"):
            handled_keys.add(key)
    for key in row_map:
        col_c = key[0]
        if col_c in ("ISN", "BSN"):
            handled_keys.add(key)

    # ── Build set of original keys whose rows are being repurposed ──────────
    repurposed_originals = set()
    for orig_key, new_name in rename_map.items():
        if len(orig_key) == 3 and new_name:
            col_c, col_b, orig_d = orig_key
            new_data_key = (col_c, col_b, new_name)
            if new_data_key in financial_data and orig_key in financial_data:
                repurposed_originals.add(orig_key)
                logger.info(
                    f"  Row repurposed: '{orig_d}' → '{new_name}' "
                    f"(data from {new_data_key})"
                )

    # ── Process standard rows (IS, BS, CFS) ─────────────────────────────────
    for (col_c, col_b, col_d), year_vals in financial_data.items():
        key = (col_c, col_b, col_d)
        if key in handled_keys:
            continue

        if key in repurposed_originals:
            continue

        excel_key = key
        if key not in row_map:
            lookup = f"{col_c}|{col_b}|{col_d}"
            if lookup in new_to_original:
                excel_key = new_to_original[lookup]

        if excel_key not in row_map:
            logger.debug(f"  Row not found in K USD sheet: {key}")
            continue

        row_num = row_map[excel_key]

        # If this row is being renamed, write the rename text patch
        if excel_key in rename_map and len(excel_key) == 3:
            new_name = rename_map[excel_key]
            patches.append((row_num, COL_D, ("text", new_name)))
            logger.info(f"  Renaming R{row_num} col D: '{excel_key[2]}' → '{new_name}'")

        # Write values for ALL template year columns
        # None → 0 to replace template placeholders
        for year in all_template_years:
            if year not in year_col:
                continue
            col_idx = year_col[year]
            cell    = ws.cell(row=row_num, column=col_idx)
            if _is_formula(cell.value):
                continue

            val = year_vals.get(year)
            if val is None:
                val = 0  # Replace placeholder with 0
            patches.append((row_num, col_idx, val))
            logger.debug(
                f"  K USD [{col_c}|{col_b}|{col_d[:30]}] year={year} → {val:,}"
            )

    # ── Fix template D column label mismatches ──────────────────────────────
    # SUM rows (ISN/BSN with B≠None and formula in F) and Check rows
    # (C=None, D starts with "Check") may have wrong labels from the template.
    # Correct labels based on which section each row belongs to.
    SECTION_LABELS = {
        # (col_c, section_key): (sum_label, check_label)
        ("ISN", "rd"):      ("Total Research And Development",
                             "Check - Research And Development"),
        ("ISN", "ga"):      ("Total General And Administrative",
                             "Check - General And Administrative"),
        ("BSN", "ppe"):     ("Property And Equipment, Net",
                             "Check - Property And Equipment"),
        ("BSN", "accrued"): ("Accrued Expenses And Other Current Liabilities",
                             "Check - Accrued Expenses And Other Current Liabilities"),
    }

    # Find SUM rows: ISN/BSN with B≠None and formula in first data column.
    # Also fix the corresponding Check row (always 2 rows below the SUM row).
    for row_num in range(1, ws.max_row + 1):
        c_val = ws_data.cell(row=row_num, column=COL_C).value
        if c_val not in ("ISN", "BSN"):
            continue
        b_val = ws_data.cell(row=row_num, column=COL_B).value
        if b_val is None or b_val == "":
            continue
        f_val = ws.cell(row=row_num, column=first_data_col).value
        if not _is_formula(f_val):
            continue
        d_val = ws_data.cell(row=row_num, column=COL_D).value

        # Determine section by row position
        if c_val == "ISN":
            section_key = "rd" if row_num < 40 else "ga"
        else:
            section_key = "ppe" if row_num < 100 else "accrued"

        sum_label, check_label = SECTION_LABELS[(c_val, section_key)]

        # Fix SUM row label
        if d_val != sum_label:
            patches.append((row_num, COL_D, ("text", sum_label)))
            logger.info(f"  Label fix R{row_num}: '{d_val}' → '{sum_label}'")

        # Fix Check row label (always 2 rows below SUM)
        check_row = row_num + 2
        check_d = ws_data.cell(row=check_row, column=COL_D).value
        if isinstance(check_d, str) and check_d.startswith("Check") and check_d != check_label:
            patches.append((check_row, COL_D, ("text", check_label)))
            logger.info(f"  Label fix R{check_row}: '{check_d}' → '{check_label}'")

    # ── Fill blank D cells for any ISN/BSN rows (legacy template cleanup) ────
    for row_num in range(1, ws.max_row + 1):
        c_val = ws_data.cell(row=row_num, column=COL_C).value
        if c_val not in ("ISN", "BSN"):
            continue
        d_val = ws_data.cell(row=row_num, column=COL_D).value
        if d_val and str(d_val).strip():
            continue  # already has a name
        # Check if this row is already being patched with a new name
        already_patched = any(
            p[0] == row_num and p[1] == COL_D for p in patches
        )
        if already_patched:
            continue
        patches.append((row_num, COL_D, ("text", f"Reserved")))
        hidden_rows.add(row_num)
        logger.debug(f"  Blank-D cleanup: R{row_num} → 'Reserved', hidden")

    # Remove show_rows from hidden_rows (in case re-run with different data)
    hidden_rows -= show_rows
    logger.info(f"K USD: {len(patches)} cells to write, {len(hidden_rows)} rows to hide, {len(show_rows)} rows to show")
    return patches, hidden_rows, show_rows


def _collect_mm_patches(
    wb:      openpyxl.Workbook,
    wb_data: openpyxl.Workbook,
    years:   List[int],
) -> List[Tuple[int, int, Any]]:
    """Return [(row, col, formula)] for blank FY DATA cells needing SUMIFS."""
    ws_mm      = wb[SHEET_MM]
    ws_mm_data = wb_data[SHEET_MM]
    year_col_mm = _build_year_col_map(ws_mm)
    row_map_mm  = _build_row_map(ws_mm, ws_mm_data, year_col_mm)

    patches: List[Tuple[int, int, Any]] = []
    for row_num in set(row_map_mm.values()):
        for year in years:
            if year not in year_col_mm:
                continue
            col_idx    = year_col_mm[year]
            col_letter = get_column_letter(col_idx)
            if ws_mm.cell(row=row_num, column=col_idx).value is not None:
                continue
            formula = (
                f"=SUMIFS('FY DATA K USD'!{col_letter}:{col_letter},"
                f"'FY DATA K USD'!$D:$D,'FY DATA'!$D{row_num},"
                f"'FY DATA K USD'!$C:$C,'FY DATA'!$C{row_num},"
                f"'FY DATA K USD'!$B:$B,'FY DATA'!$B{row_num})/1000"
            )
            patches.append((row_num, col_idx, formula))

    logger.info(f"FY DATA: {len(patches)} SUMIFS formulas to inject")
    return patches


# ── surgical zip patcher ──────────────────────────────────────────────────────

def _get_sheet_zip_paths(xlsx_path: Path) -> Dict[str, str]:
    """Return {sheet_name: zip_entry_path} for all worksheets."""
    with zipfile.ZipFile(xlsx_path) as zf:
        wb_xml   = ET.fromstring(zf.read("xl/workbook.xml"))
        rels_xml = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))

    rid_to_path: Dict[str, str] = {}
    for rel in rels_xml:
        if "worksheet" in rel.get("Type", ""):
            rid    = rel.get("Id", "")
            target = rel.get("Target", "")
            # Target is relative to xl/: e.g. "worksheets/sheet18.xml"
            rid_to_path[rid] = (
                f"xl/{target}" if not target.startswith("/") else target.lstrip("/")
            )

    sheet_map: Dict[str, str] = {}
    for sheet_elem in wb_xml.findall(f".//{{{_NS_MAIN}}}sheet"):
        name = sheet_elem.get("name", "")
        rid  = sheet_elem.get(f"{{{_NS_R}}}id", "")
        if rid in rid_to_path:
            sheet_map[name] = rid_to_path[rid]

    return sheet_map


def _num_str(value: Any) -> str:
    """Format a numeric value for the XML <v> element."""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value)


def _patch_numeric_cell(xml: str, addr: str, val_str: str) -> str:
    """
    Find cell r="ADDR" in the sheet XML string and update its <v> content.
    The original XML bytes are preserved character-for-character except for
    the value text — no re-serialization, no namespace changes.
    """
    search = f'r="{addr}"'
    start  = 0
    while True:
        pos = xml.find(search, start)
        if pos == -1:
            logger.warning(f"  Cell {addr} not found in sheet XML; skipping")
            return xml

        # Walk backwards to the '<' that opens this element
        lt = xml.rfind("<", 0, pos)
        if lt == -1:
            start = pos + 1
            continue
        # Only act on <c elements  (not <row r=…> or anything else)
        if xml[lt + 1] != "c" or xml[lt + 2] not in (" ", "\t", "\n", "/", ">"):
            start = pos + 1
            continue

        # Find the end of the opening tag
        tag_end = xml.index(">", lt) + 1

        # Self-closing?  <c r="F9" s="21" />  → expand it
        if xml[tag_end - 2 : tag_end] == "/>":
            return (
                xml[: tag_end - 2]
                + f"><v>{val_str}</v></c>"
                + xml[tag_end:]
            )

        # Find </c>
        c_end = xml.index("</c>", tag_end)
        cell_body = xml[tag_end:c_end]

        if "<v>" in cell_body:
            # Replace existing <v>…</v>
            v_start = tag_end + cell_body.index("<v>")
            v_end   = tag_end + cell_body.index("</v>") + 4
            return xml[:v_start] + f"<v>{val_str}</v>" + xml[v_end:]
        else:
            # No <v> yet – insert one before </c>
            return xml[:c_end] + f"<v>{val_str}</v>" + xml[c_end:]


def _patch_formula_cell(xml: str, row_num: int, addr: str, formula_body: str) -> str:
    """
    Insert a formula cell into an existing row, or update it if it already
    has a <v> or <f> element.  'formula_body' is the formula without '='.

    For blank cells (not present in the XML), a new <c> element is inserted
    in the correct column order within the row.
    """
    # ── Does the cell already exist? ──
    search = f'r="{addr}"'
    pos = xml.find(search)
    if pos != -1:
        lt = xml.rfind("<", 0, pos)
        if lt != -1 and xml[lt + 1] == "c" and xml[lt + 2] in (" ", "\t", "\n", "/", ">"):
            tag_end = xml.index(">", lt) + 1
            if xml[tag_end - 2 : tag_end] == "/>":
                # Self-closing → expand
                return (
                    xml[: tag_end - 2]
                    + f"><f>{formula_body}</f></c>"
                    + xml[tag_end:]
                )
            c_end = xml.index("</c>", tag_end)
            # Replace everything between open tag and </c>
            return xml[:tag_end] + f"<f>{formula_body}</f>" + xml[c_end:]

    # ── Cell doesn't exist: insert into the row ──
    row_search = f'r="{row_num}"'
    row_pos = 0
    while True:
        rp = xml.find(row_search, row_pos)
        if rp == -1:
            logger.warning(f"  Row {row_num} not found; cannot inject formula at {addr}")
            return xml
        lt = xml.rfind("<", 0, rp)
        if lt != -1 and xml[lt + 1 : lt + 4] in ("row", "ROW"):
            break
        row_pos = rp + 1

    row_tag_end  = xml.index(">", lt) + 1
    row_end      = xml.index("</row>", row_tag_end)
    row_body     = xml[row_tag_end:row_end]

    new_cell = f'<c r="{addr}"><f>{formula_body}</f></c>'
    col_idx  = _col_of(addr)

    # Insert in column order among existing cells
    import re as _re
    cells = list(_re.finditer(r'<c\b[^>]*\br="([A-Z]+\d+)"', row_body))
    insert_at = len(row_body)  # default: append
    for m in cells:
        if _col_of(m.group(1)) > col_idx:
            insert_at = m.start()
            break

    new_body = row_body[:insert_at] + new_cell + row_body[insert_at:]
    return xml[:row_tag_end] + new_body + xml[row_end:]


_EMPTY_CALC_CHAIN = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
    '<calcChain xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"></calcChain>'
)


def _strip_formula_cache(xml: str) -> str:
    """Remove cached ERROR values from formula cells to force Excel recalculation.

    Cells with t="e" and <f>: strip t="e" and <v> (Excel recalculates).
    Cells with t="e" but NO <f>: convert to empty cell (self-closing <c/>).
    Leaves valid cached values (numbers, strings) intact.
    """
    import re as _re

    def _fix_error_cell(m):
        tag = m.group(0)
        if '<f' in tag:
            tag = _re.sub(r' t="e"', '', tag)
            tag = _re.sub(r'<v>[^<]*</v>', '', tag)
        else:
            open_tag = _re.match(r'<c\b[^>]*', tag).group(0)
            open_tag = _re.sub(r' t="e"', '', open_tag)
            tag = open_tag + '/>'
        return tag

    return _re.sub(r'<c\b[^>]* t="e"[^>]*>.*?</c>', _fix_error_cell, xml)


def _normalize_formula_xml(xml: str) -> str:
    """Fix formula XML issues that cause Excel's 'Removed Records: Formula' repair.

    1. Double-escaped entities: &amp;gt; → &gt;, &amp;lt; → &lt;, &amp;quot; → "
    2. &quot; in <f> element content: replace with literal ".
    3. <v/> self-closing: replace with <v></v> for Excel compatibility.
    """
    import re as _re

    def _fix_formula(m):
        f = m.group(0)
        f = f.replace('&amp;gt;', '&gt;')
        f = f.replace('&amp;lt;', '&lt;')
        f = f.replace('&amp;amp;', '&amp;')
        f = f.replace('&amp;quot;', '"')
        f = f.replace('&quot;', '"')
        return f

    xml = _re.sub(r'<f\b[^>]*>.*?</f>', _fix_formula, xml, flags=_re.DOTALL)
    xml = xml.replace('<v/>', '<v></v>')
    return xml


def _xml_escape(text: str) -> str:
    """Escape XML special characters in text content."""
    return (
        text.replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;")
            .replace("'", "&apos;")
    )


def _patch_text_cell(xml: str, addr: str, new_text: str) -> str:
    """
    Replace a cell's content with an inlineStr text value.
    Converts <c r="D28" t="s"><v>17</v></c> (shared string ref)
    to <c r="D28" t="inlineStr"><is><t>new_text</t></is></c>.
    """
    import re as _re
    new_text = _xml_escape(new_text)
    search = f'r="{addr}"'
    start = 0
    while True:
        pos = xml.find(search, start)
        if pos == -1:
            logger.warning(f"  Cell {addr} not found in sheet XML for text patch; skipping")
            return xml

        lt = xml.rfind("<", 0, pos)
        if lt == -1:
            start = pos + 1
            continue
        if xml[lt + 1] != "c" or xml[lt + 2] not in (" ", "\t", "\n", "/", ">"):
            start = pos + 1
            continue

        # Remove any existing t="..." attribute from the opening tag
        tag_end = xml.index(">", lt) + 1

        if xml[tag_end - 2: tag_end] == "/>":
            # Self-closing: <c r="D28" s="5" t="s" />
            open_tag = xml[lt:tag_end - 2]
            open_tag = _re.sub(r'\s+t="[^"]*"', '', open_tag)
            return (
                xml[:lt]
                + open_tag
                + f' t="inlineStr"><is><t>{new_text}</t></is></c>'
                + xml[tag_end:]
            )

        # Find </c>
        c_end = xml.index("</c>", tag_end)
        open_tag = xml[lt:tag_end - 1]  # everything up to the '>'
        open_tag = _re.sub(r'\s+t="[^"]*"', '', open_tag)

        return (
            xml[:lt]
            + open_tag
            + f' t="inlineStr"><is><t>{new_text}</t></is></c>'
            + xml[c_end + 4:]
        )


def _hide_rows(xml: str, row_numbers: set, show_numbers: Optional[set] = None) -> str:
    """Add hidden="1" to rows in row_numbers; remove hidden="1" from show_numbers."""
    import re as _re
    for rn in row_numbers:
        pattern = f'(<row\\b[^>]*?\\br="{rn}"[^>]*>)'
        match = _re.search(pattern, xml)
        if match:
            full_tag = match.group(1)
            if 'hidden="1"' not in full_tag:
                new_tag = full_tag.replace('<row ', '<row hidden="1" ', 1)
                xml = xml[:match.start()] + new_tag + xml[match.end():]
    if show_numbers:
        for rn in show_numbers:
            pattern = f'(<row\\b[^>]*?\\br="{rn}"[^>]*>)'
            match = _re.search(pattern, xml)
            if match:
                full_tag = match.group(1)
                if 'hidden="1"' in full_tag:
                    new_tag = full_tag.replace(' hidden="1"', '', 1)
                    xml = xml[:match.start()] + new_tag + xml[match.end():]
    return xml


def _patch_sheet_xml(
    xml_bytes:   bytes,
    cell_writes: List[Tuple[int, int, Any]],
    hidden_rows: Optional[set] = None,
    show_rows:   Optional[set] = None,
) -> bytes:
    """
    Apply (row, col, value) writes to a worksheet XML using direct string
    manipulation — the original XML is preserved byte-for-byte except for
    the exact cell values changed.  No re-serialization; no namespace issues.
    """
    xml = xml_bytes.decode("utf-8")

    for (row_num, col_idx, value) in cell_writes:
        addr = _cell_addr(col_idx, row_num)

        if isinstance(value, tuple) and len(value) == 2 and value[0] == "text":
            # Text cell write (inlineStr)
            xml = _patch_text_cell(xml, addr, value[1])
        elif isinstance(value, str) and value.startswith("="):
            # Formula write
            xml = _patch_formula_cell(xml, row_num, addr, value[1:])
        else:
            # Numeric write
            xml = _patch_numeric_cell(xml, addr, _num_str(value))

    if hidden_rows or show_rows:
        xml = _hide_rows(xml, hidden_rows or set(), show_rows)

    return xml.encode("utf-8")


def _apply_xlsx_patches(
    xlsx_path:   Path,
    patches:     Dict[str, List[Tuple[int, int, Any]]],
    hidden_rows: Optional[Dict[str, set]] = None,
    show_rows:   Optional[Dict[str, set]] = None,
) -> None:
    """
    Surgically write {sheet_name: [(row, col, value)]} patches into xlsx.

    hidden_rows: {sheet_name: set_of_row_numbers} to hide in the output.
    show_rows:   {sheet_name: set_of_row_numbers} to unhide in the output.

    Every zip entry that is NOT a patched sheet is copied byte-for-byte
    (sharedStrings.xml, calcChain.xml, styles, VBA, drawings, etc.).
    """
    if hidden_rows is None:
        hidden_rows = {}
    if show_rows is None:
        show_rows = {}
    sheet_zip_paths = _get_sheet_zip_paths(xlsx_path)

    # Build modified sheet XMLs
    # Process all sheets that have patches OR hidden_rows OR show_rows
    all_sheet_names = set(patches.keys()) | set(hidden_rows.keys()) | set(show_rows.keys())
    modified: Dict[str, bytes] = {}
    with zipfile.ZipFile(xlsx_path) as zf:
        for sheet_name in all_sheet_names:
            cell_writes = patches.get(sheet_name, [])
            sheet_hidden = hidden_rows.get(sheet_name)
            sheet_show = show_rows.get(sheet_name)
            if not cell_writes and not sheet_hidden and not sheet_show:
                continue
            zip_path = sheet_zip_paths.get(sheet_name)
            if not zip_path:
                logger.warning(f"Sheet '{sheet_name}' not found in zip; skipping")
                continue
            xml_bytes = zf.read(zip_path)
            modified[zip_path] = _patch_sheet_xml(
                xml_bytes, cell_writes, sheet_hidden, sheet_show
            )
            logger.info(
                f"Patched '{sheet_name}' ({zip_path}): {len(cell_writes)} cells"
                + (f", {len(sheet_hidden)} rows hidden" if sheet_hidden else "")
                + (f", {len(sheet_show)} rows shown" if sheet_show else "")
            )

    if not modified:
        logger.info("No patches to apply.")
        return

    # ── Patch workbook.xml: add fullCalcOnLoad="1" ──────────────────────────
    # This forces Excel to recalculate ALL formulas on open
    with zipfile.ZipFile(xlsx_path) as zf:
        wb_xml_bytes = zf.read("xl/workbook.xml")
    wb_xml_str = wb_xml_bytes.decode("utf-8")
    if 'fullCalcOnLoad' not in wb_xml_str:
        # Add fullCalcOnLoad="1" to existing <calcPr element
        wb_xml_str = wb_xml_str.replace("<calcPr", '<calcPr fullCalcOnLoad="1"', 1)
        logger.info("Added fullCalcOnLoad='1' to workbook.xml")
    modified["xl/workbook.xml"] = wb_xml_str.encode("utf-8")

    # Remove calcChain.xml: strip its references from Content_Types + workbook.rels.
    # An empty/absent calcChain causes Excel "Catastrophic failure" XML parse error.
    # Correct fix: delete the file AND its <Override>/<Relationship> entries so
    # Excel never looks for it and rebuilds the chain on open (fullCalcOnLoad=1).
    import re as _re
    with zipfile.ZipFile(xlsx_path) as zf:
        ct = zf.read("[Content_Types].xml").decode("utf-8")
        wr = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    ct = _re.sub(r'<Override[^>]*/xl/calcChain\.xml[^>]*/>', '', ct)
    wr = _re.sub(r'<Relationship[^>]*calcChain[^>]*/>', '', wr)
    modified["[Content_Types].xml"] = ct.encode("utf-8")
    modified["xl/_rels/workbook.xml.rels"] = wr.encode("utf-8")

    # Write a new zip, replacing only the patched sheet XMLs
    tmp_path = xlsx_path.with_suffix(".~patch.xlsx")
    with zipfile.ZipFile(xlsx_path, "r") as zin:
        with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                if item.filename == "xl/calcChain.xml":
                    continue  # omitted — references stripped from CT + rels above
                if item.filename in modified:
                    raw = modified[item.filename]
                    if (item.filename.startswith("xl/worksheets/sheet")
                            and item.filename.endswith(".xml")):
                        s = raw.decode("utf-8")
                        s = _strip_formula_cache(s)
                        s = _normalize_formula_xml(s)
                        raw = s.encode("utf-8")
                    zout.writestr(item, raw)
                elif (item.filename.startswith("xl/worksheets/sheet")
                      and item.filename.endswith(".xml")):
                    s = zin.read(item.filename).decode("utf-8")
                    s = _strip_formula_cache(s)
                    s = _normalize_formula_xml(s)
                    zout.writestr(item, s.encode("utf-8"))
                else:
                    zout.writestr(item, zin.read(item.filename))

    tmp_path.replace(xlsx_path)
    logger.info(f"Surgical patch applied → {xlsx_path}")


# ── public entry point ────────────────────────────────────────────────────────

def update_workbook(
    ticker:         str,
    financial_data: Dict[Tuple, Dict[int, Optional[float]]],
    years:          List[int],
    excel_path:     Optional[str] = None,
    rename_map:     Optional[Dict[Tuple, str]] = None,
    note_details:   Optional[Dict] = None,
    reporting_unit: str = "K",
) -> Path:
    """
    Main entry point: backup → read layout → build patch plan → surgical write.

    Never calls openpyxl .save(); all non-sheet parts are preserved intact.

    Parameters
    ----------
    ticker         : stock ticker (used to locate the file if excel_path is None)
    financial_data : output of SECFetcher.build_financial_data()
    years          : list of fiscal years to populate
    excel_path     : explicit path override
    rename_map     : {(col_c, col_b, original_col_d): new_col_d_text}
    note_details   : {section: [(item_name, {year: value})] or None}
    reporting_unit : 'K' (thousands) or 'MM' (millions)
    """
    path = (
        Path(excel_path)
        if excel_path
        else Path(f"/mnt/c/Users/yzsun/Desktop/DD/{ticker}/DCF {ticker}.xlsx")
    )

    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    # ── Backup ────────────────────────────────────────────────────────────────
    _backup(path)

    # ── Open two workbooks (read-only) ────────────────────────────────────────
    # wb      : data_only=False  → detects formula rows
    # wb_data : data_only=True   → reads actual computed B/C/D values
    logger.info(f"Opening workbook (read-only): {path}")
    wb      = openpyxl.load_workbook(str(path), data_only=False)
    wb_data = openpyxl.load_workbook(str(path), data_only=True)

    # ── Build write plans ─────────────────────────────────────────────────────
    all_patches: Dict[str, List] = {}
    all_hidden:  Dict[str, set] = {}
    all_show:    Dict[str, set] = {}

    if reporting_unit == "MM":
        # ── MM direct mode: write data to FY DATA sheet ──────────────────
        if SHEET_MM not in wb.sheetnames:
            raise ValueError(f"Sheet '{SHEET_MM}' not found in workbook.")
        data_patches, data_hidden, data_show = _collect_kusd_patches(
            wb, wb_data, financial_data, years, rename_map, note_details,
            target_sheet=SHEET_MM,
        )
        if data_patches:
            all_patches[SHEET_MM] = data_patches
        if data_hidden:
            all_hidden[SHEET_MM] = data_hidden
        if data_show:
            all_show[SHEET_MM] = data_show
        logger.info(f"MM direct mode: {len(data_patches)} cells to write")
    else:
        # ── K USD mode: write to K USD sheet, optionally inject SUMIFS ───
        if SHEET_KUSD not in wb.sheetnames:
            raise ValueError(f"Sheet '{SHEET_KUSD}' not found in workbook.")
        kusd_patches, kusd_hidden, kusd_show = _collect_kusd_patches(
            wb, wb_data, financial_data, years, rename_map, note_details,
        )
        if kusd_patches:
            all_patches[SHEET_KUSD] = kusd_patches
        if kusd_hidden:
            all_hidden[SHEET_KUSD] = kusd_hidden
        if kusd_show:
            all_show[SHEET_KUSD] = kusd_show

        if SHEET_MM in wb.sheetnames:
            mm_patches = _collect_mm_patches(wb, wb_data, years)
            if mm_patches:
                all_patches[SHEET_MM] = mm_patches
            # Apply same hidden/show rows to FY DATA sheet
            if kusd_hidden:
                all_hidden[SHEET_MM] = kusd_hidden
            if kusd_show:
                all_show[SHEET_MM] = kusd_show
        else:
            mm_patches = []
            logger.info(f"Sheet '{SHEET_MM}' not found; skipping SUMIFS injection.")

        logger.info(
            f"K USD mode: {len(kusd_patches)} K USD cells, "
            f"{len(mm_patches)} FY DATA formulas"
        )

    wb.close()
    wb_data.close()

    # ── Apply surgically ──────────────────────────────────────────────────────
    _apply_xlsx_patches(path, all_patches, all_hidden, all_show)

    logger.info(f"Done. Workbook updated: {path}")
    return path
