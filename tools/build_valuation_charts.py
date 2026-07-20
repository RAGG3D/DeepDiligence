#!/usr/bin/env python3
"""Update VALUATION waterfall table and build valuation charts with Excel COM."""

from __future__ import annotations

import argparse
import html
import re
import shutil
import subprocess
import sys
import time
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl


NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


def _default_path(ticker: str) -> Path:
    return Path(f"/mnt/c/Users/yzsun/Desktop/DD/{ticker}/DCF {ticker}.xlsx")


def _sheet_zip_path(parts: dict[str, bytes], sheet_name: str) -> str:
    wb = ET.fromstring(parts["xl/workbook.xml"])
    rels = ET.fromstring(parts["xl/_rels/workbook.xml.rels"])
    relmap = {rel.get("Id"): rel.get("Target") for rel in rels}
    for sheet in wb.findall(f".//{{{NS_MAIN}}}sheet"):
        if sheet.get("name") == sheet_name:
            target = relmap[sheet.get(f"{{{NS_R}}}id")]
            target = target.lstrip("/")
            return target if target.startswith("xl/") else "xl/" + target
    raise RuntimeError(f"Sheet not found: {sheet_name}")


def _read_parts(path: Path) -> tuple[dict[str, bytes], list[str]]:
    with zipfile.ZipFile(path, "r") as zf:
        return {name: zf.read(name) for name in zf.namelist()}, zf.namelist()


def _write_parts(path: Path, parts: dict[str, bytes], order: list[str]) -> None:
    tmp = path.with_suffix(".~valuation.xlsx")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        written = set()
        for name in order:
            if name in parts:
                zout.writestr(name, parts[name])
                written.add(name)
        for name, data in parts.items():
            if name not in written:
                zout.writestr(name, data)
    tmp.replace(path)


def _cell_col(addr: str) -> int:
    col = re.match(r"([A-Z]+)", addr).group(1)
    out = 0
    for ch in col:
        out = out * 26 + ord(ch) - 64
    return out


def _cell_style(xml: str, addr: str, default: str) -> str:
    m = re.search(
        rf'<c\b(?=[^>]*\br="{addr}")[^>]*/>|'
        rf'<c\b(?=[^>]*\br="{addr}")[^>]*>.*?</c>',
        xml,
        re.S,
    )
    if not m:
        return default
    sm = re.search(r'\bs="(\d+)"', m.group(0))
    return sm.group(1) if sm else default


def _replace_cell(xml: str, addr: str, cell_xml: str) -> str:
    m = re.search(
        rf'<c\b(?=[^>]*\br="{addr}")[^>]*/>|'
        rf'<c\b(?=[^>]*\br="{addr}")[^>]*>.*?</c>',
        xml,
        re.S,
    )
    if m:
        return xml[:m.start()] + cell_xml + xml[m.end():]

    row_num = int(re.search(r"\d+", addr).group(0))
    row_m = re.search(rf'(<row\b(?=[^>]*\br="{row_num}")[^>]*>)(.*?)(</row>)', xml, re.S)
    if not row_m:
        return xml
    body = row_m.group(2)
    insert_at = len(body)
    target_col = _cell_col(addr)
    for cm in re.finditer(r'<c\b[^>]*\br="([A-Z]+\d+)"', body):
        if _cell_col(cm.group(1)) > target_col:
            insert_at = cm.start()
            break
    new_body = body[:insert_at] + cell_xml + body[insert_at:]
    return xml[:row_m.start(2)] + new_body + xml[row_m.end(2):]


def _text_cell(addr: str, text: str, style: str) -> str:
    return f'<c r="{addr}" s="{style}" t="inlineStr"><is><t>{html.escape(text, quote=False)}</t></is></c>'


def _formula_cell(addr: str, formula: str, style: str) -> str:
    return f'<c r="{addr}" s="{style}"><f>{formula}</f></c>'


def _number_cell(addr: str, value: float | int, style: str) -> str:
    return f'<c r="{addr}" s="{style}"><v>{value}</v></c>'


def _datatable_cell(addr: str, ref: str, style: str) -> str:
    return (
        f'<c r="{addr}" s="{style}"><f t="dataTable" ref="{ref}" '
        f'dt2D="1" dtr="1" r1="C5" r2="C3" ca="1"/><v>0</v></c>'
    )


def _empty_cell(addr: str, style: str) -> str:
    return f'<c r="{addr}" s="{style}"/>'


def _breakdown_scenarios(path: Path) -> list[tuple[int, str]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        ws = wb["Scenarios"]
        out: list[tuple[int, str]] = []
        in_breakdown = False
        for row in range(1, ws.max_row + 1):
            label = ws.cell(row, 3).value
            if label == "Break Down":
                in_breakdown = True
                continue
            if label == "Catalyst Scenarios":
                break
            if not in_breakdown:
                continue
            scen = ws.cell(row, 2).value
            if isinstance(scen, (int, float)) and isinstance(label, str) and label.strip():
                out.append((int(scen), label.strip()))
        return out
    finally:
        wb.close()


def update_waterfall_table(path: Path) -> int:
    scenarios = _breakdown_scenarios(path)
    if not scenarios:
        print("No Scenarios breakdown rows found; valuation waterfall unchanged")
        return 0

    entries = scenarios + [(1, "Base")]
    last_row = 25 + len(entries)

    parts, order = _read_parts(path)
    sheet_path = _sheet_zip_path(parts, "VALUATION")
    xml = parts[sheet_path].decode("utf-8", "ignore")

    styles = {
        "title": _cell_style(xml, "F23", "137"),
        "subtitle": _cell_style(xml, "F24", "170"),
        "head": _cell_style(xml, "F25", "29"),
        "head_value": _cell_style(xml, "G25", "171"),
        "label": _cell_style(xml, "F26", "154"),
        "value": _cell_style(xml, "G26", "148"),
        "input": _cell_style(xml, "I26", "138"),
        "tg": _cell_style(xml, "J25", "145"),
    }

    xml = _replace_cell(xml, "F23", _text_cell("F23", "Valuation Waterfall Data Table", styles["title"]))
    xml = _replace_cell(
        xml,
        "F24",
        _formula_cell(
            "F24",
            'D3&amp;" Scenario with Terminal Growth @"&amp;TEXT($J$25,"0.00%")&amp;" as at "&amp;TEXT($C$4,"DD MMM YYYY")',
            styles["subtitle"],
        ),
    )
    xml = _replace_cell(xml, "F25", _text_cell("F25", "Valuation Component", styles["head"]))
    xml = _replace_cell(xml, "G25", _text_cell("G25", "[USD/share]", styles["head_value"]))
    xml = _replace_cell(xml, "I25", _formula_cell("I25", "IFERROR(C48,0)", styles["input"]))
    xml = _replace_cell(xml, "J25", _number_cell("J25", 0.03, styles["tg"]))

    for idx, (scenario_id, _name) in enumerate(entries):
        row = 26 + idx
        xml = _replace_cell(
            xml,
            f"F{row}",
            _formula_cell(
                f"F{row}",
                f'IFERROR(INDEX(Scenarios!$B:$B,MATCH($I{row},Scenarios!$A:$A,0)),"NO SCENARIO SELECTED")',
                styles["label"],
            ),
        )
        if idx == 0 or scenario_id == 1:
            g_formula = f"IFERROR(J{row},0)"
        else:
            g_formula = f"IFERROR(J{row}-J{row-1},0)"
        xml = _replace_cell(xml, f"G{row}", _formula_cell(f"G{row}", g_formula, styles["value"]))
        xml = _replace_cell(xml, f"I{row}", _number_cell(f"I{row}", scenario_id, styles["input"]))
        if idx == 0:
            xml = _replace_cell(xml, f"J{row}", _datatable_cell(f"J{row}", f"J26:J{last_row}", styles["value"]))
        else:
            xml = _replace_cell(xml, f"J{row}", _number_cell(f"J{row}", 0, styles["value"]))

    for row in range(last_row + 1, 61):
        for col in ("F", "G", "I", "J"):
            # Trailing empties below the last live row: never the yellow input
            # style, so no stray yellow cells are left in the waterfall frame.
            style = styles["label"] if col == "F" else styles["value"]
            xml = _replace_cell(xml, f"{col}{row}", _empty_cell(f"{col}{row}", style))

    wb_xml = parts["xl/workbook.xml"].decode("utf-8", "ignore")
    if "fullCalcOnLoad" not in wb_xml:
        wb_xml = wb_xml.replace("<calcPr", '<calcPr fullCalcOnLoad="1"', 1)
        parts["xl/workbook.xml"] = wb_xml.encode("utf-8")
    parts[sheet_path] = xml.encode("utf-8")
    _write_parts(path, parts, order)
    print(f"Valuation waterfall table updated: {len(scenarios)} breakdown rows + Base, rows 26-{last_row}")
    return len(entries)


def _to_windows_path(path: Path) -> str:
    text = str(path)
    match = re.match(r"^/mnt/([a-zA-Z])/(.*)$", text)
    if match:
        drive = match.group(1).upper()
        rest = match.group(2).replace("/", "\\")
        return f"{drive}:\\{rest}"
    return text


def build_charts_with_excel(path: Path) -> None:
    script = Path(__file__).with_name("build_valuation_charts.ps1")
    if not script.exists():
        raise FileNotFoundError(script)
    subprocess.run(
        [
            "powershell.exe",
            "-NoProfile",
            "-ExecutionPolicy",
            "Bypass",
            "-File",
            str(script),
            "-Path",
            _to_windows_path(path),
        ],
        check=True,
        timeout=600,
    )
    wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        charts = wb["VALUATION"]._charts

        def _title(chart) -> str:
            try:
                return " ".join(
                    run.t or ""
                    for paragraph in chart.title.tx.rich.p
                    for run in (paragraph.r or [])
                ).strip()
            except Exception:
                return ""

        titles = [_title(chart) for chart in charts]
        if not any("Waterfall" in title for title in titles):
            raise RuntimeError(
                f"Excel chart build completed without Valuation Waterfall: {titles}"
            )
    finally:
        wb.close()


def repair_save_with_excel(path: Path) -> None:
    script = Path(__file__).with_name("excel_repair_saveas.ps1")
    if not script.exists():
        raise FileNotFoundError(script)
    subprocess.run(
        [
            "powershell.exe",
            "-NoProfile",
            "-ExecutionPolicy",
            "Bypass",
            "-File",
            str(script),
            "-Path",
            _to_windows_path(path),
        ],
        check=True,
        timeout=600,
    )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--ticker")
    parser.add_argument("--path")
    parser.add_argument("--no-backup", action="store_true")
    parser.add_argument("--no-charts", action="store_true")
    args = parser.parse_args()

    if args.path:
        path = Path(args.path)
    elif args.ticker:
        path = _default_path(args.ticker)
    else:
        raise SystemExit("--ticker or --path required")
    if not path.exists():
        raise FileNotFoundError(path)

    if not args.no_backup:
        backup = path.with_name(f"{path.stem}_pre_valuation_charts_{time.strftime('%Y%m%d_%H%M%S')}.xlsx")
        shutil.copy2(path, backup)
        print(f"Backup: {backup}")

    update_waterfall_table(path)
    repair_save_with_excel(path)
    if not args.no_charts:
        build_charts_with_excel(path)


if __name__ == "__main__":
    main()
