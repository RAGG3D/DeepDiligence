#!/usr/bin/env python3
"""Score a frozen Test-EVENT prediction after the public clinical release.

This is deliberately separate from ``test_catalyst_event.py``.  The blind
clinical artifact and workbook prediction are hashed before any post-release
market data are accepted.  A second agent may then provide exactly three
eligible raw closes and one integer 1-10 score per active target.  This tool
validates the closes, writes only the post-release score overlay, and proves
that the frozen Test prediction cells did not change.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import math
import shutil
import subprocess
import sys
import zipfile
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import yfinance as yf
from openpyxl import load_workbook
from openpyxl.worksheet.formula import DataTableFormula

REPO = Path(__file__).resolve().parents[1]
if str(REPO) not in sys.path:
    sys.path.insert(0, str(REPO))

from datastore.research_fact_store import upsert_research_facts  # noqa: E402


def default_workbook(ticker: str) -> Path:
    return Path(f"/mnt/c/Users/yzsun/Desktop/DD/{ticker}/DCF {ticker}.xlsx")


def event_slug(event: str) -> str:
    import re

    slug = re.sub(r"[^A-Za-z0-9_-]+", "", event.strip())
    if not slug:
        raise ValueError("event must contain at least one letter or digit")
    return slug[:26]


def sha256_file(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def canonical_sha256(value: Any) -> str:
    payload = json.dumps(
        value, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str
    ).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def windows_path(path: Path) -> str:
    return subprocess.run(
        ["wslpath", "-w", str(path)], check=True, capture_output=True, text=True
    ).stdout.strip()


def blind_prediction_payload(clinical: dict[str, Any]) -> dict[str, Any]:
    return {
        "ticker": clinical["ticker"],
        "event_name": clinical["event_name"],
        "event_date": clinical["event_date"],
        "public_disclosure_date": clinical["public_disclosure_date"],
        "relevant_targets": clinical["relevant_targets"],
        "target_assessments": clinical["target_assessments"],
        "pre_event_price_calibration": clinical["pre_event_price_calibration"],
    }


def _cell_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    attrs = getattr(value, "__dict__", None)
    if attrs is not None:
        return {"type": type(value).__name__, "attrs": attrs}
    return {"type": type(value).__name__, "value": str(value)}


def frozen_sheet_digest(path: Path, sheet_name: str, max_row: int | None = None) -> tuple[str, int]:
    """Hash every original Test value/formula from row 3 down.

    Row 2 is the only main-table row the scoring overlay may edit.  New audit
    rows are appended below the original max row and are intentionally outside
    this digest.
    """
    wb = load_workbook(path, data_only=False, read_only=False, keep_links=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise RuntimeError(f"missing Test sheet: {sheet_name}")
        ws = wb[sheet_name]
        if not isinstance(ws["C9"].value, DataTableFormula):
            raise RuntimeError("native Test What-If Data Table was lost during scoring")
        boundary = max_row or ws.max_row
        cells = []
        for row in range(3, boundary + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row, col)
                # Excel may refresh the cached results of the native What-If
                # Data Table when COM opens/saves the workbook.  Those C-column
                # caches are derived outputs, not blind inputs or formulas; C9
                # itself is audited separately as a native DataTableFormula.
                if (
                    col == 3
                    and row >= 9
                    and isinstance(ws.cell(row, 2).value, (int, float))
                ):
                    continue
                if cell.value is not None:
                    cells.append([cell.coordinate, cell.data_type, _cell_value(cell.value)])
        merged = sorted(
            str(item) for item in ws.merged_cells.ranges if item.max_row <= boundary
        )
        return canonical_sha256({"cells": cells, "merged": merged}), boundary
    finally:
        wb.close()


def fetch_post_release_closes(
    market_ticker: str,
    disclosure_date: date,
    include_disclosure_date: bool,
) -> list[dict[str, Any]]:
    start = disclosure_date if include_disclosure_date else disclosure_date + timedelta(days=1)
    frame = yf.download(
        market_ticker,
        start=start.isoformat(),
        end=(disclosure_date + timedelta(days=14)).isoformat(),
        auto_adjust=False,
        actions=False,
        progress=False,
        threads=False,
    )
    if frame.empty or "Close" not in frame:
        raise RuntimeError(f"no post-release raw closes returned for {market_ticker}")
    close = frame["Close"]
    if getattr(close, "ndim", 1) == 2:
        close = close.iloc[:, 0]
    rows = [
        {"date": index.date().isoformat(), "close": float(value)}
        for index, value in close.dropna().items()
        if index.date() >= start
    ][:3]
    if len(rows) != 3:
        raise RuntimeError(f"expected three completed post-release sessions, got {len(rows)}")
    return rows


def fetch_post_release_intraday_highs(
    market_ticker: str,
    session_dates: list[str],
) -> list[dict[str, Any]]:
    start = date.fromisoformat(session_dates[0])
    end = date.fromisoformat(session_dates[-1]) + timedelta(days=1)
    frame = yf.download(
        market_ticker,
        start=start.isoformat(),
        end=end.isoformat(),
        interval="60m",
        auto_adjust=False,
        actions=False,
        prepost=False,
        progress=False,
        threads=False,
    )
    if frame.empty or "High" not in frame:
        raise RuntimeError(f"no 60m post-release highs returned for {market_ticker}")
    high = frame["High"]
    if getattr(high, "ndim", 1) == 2:
        high = high.iloc[:, 0]
    expected = set(session_dates)
    grouped: dict[str, list[tuple[Any, float]]] = {item: [] for item in session_dates}
    for index, value in high.dropna().items():
        local_index = index
        if getattr(index, "tzinfo", None) is not None:
            local_index = index.tz_convert("America/New_York")
        key = local_index.date().isoformat()
        if key in expected:
            grouped[key].append((local_index, float(value)))
    result = []
    for session_date in session_dates:
        bars = grouped[session_date]
        if not bars:
            raise RuntimeError(f"no 60m bar returned for eligible session {session_date}")
        bar_start, raw_high = max(bars, key=lambda item: item[1])
        result.append({
            "date": session_date,
            "raw_high": raw_high,
            "bar_start": bar_start.isoformat(),
            "bar_count": len(bars),
        })
    return result


def validate_intraday_payload(
    score: dict[str, Any],
    clinical: dict[str, Any],
    sessions: list[dict[str, Any]],
) -> dict[str, Any]:
    supplied = score.get("intraday") or {}
    if supplied.get("granularity") != "60m":
        raise ValueError("independent scorer must supply 60m intraday validation")
    source_url = str(supplied.get("source_url") or "")
    if not source_url.startswith("https://query2.finance.yahoo.com/v8/finance/chart/"):
        raise ValueError("intraday source must be an auditable Yahoo chart endpoint")
    market_ticker = str(
        supplied.get("market_ticker") or score["post_release"].get("market_ticker")
        or clinical["ticker"]
    ).upper()
    expected = fetch_post_release_intraday_highs(
        market_ticker, [item["date"] for item in sessions]
    )
    provided = supplied.get("daily_highs") or []
    if len(provided) != 3:
        raise ValueError("exactly three daily intraday highs are required")
    for actual, claimed in zip(expected, provided):
        if claimed.get("date") != actual["date"]:
            raise ValueError(f"intraday date mismatch: {claimed} vs {actual}")
        if abs(float(claimed.get("raw_high")) - actual["raw_high"]) > 1e-6:
            raise ValueError(f"intraday high mismatch: {claimed} vs {actual}")
    peak = max(expected, key=lambda item: item["raw_high"])
    return {
        "market_ticker": market_ticker,
        "granularity": "60m",
        "regular_session_only": True,
        "price_field": "unadjusted High",
        "source_url": source_url,
        "daily_highs": expected,
        "three_day_peak": peak,
        "independent_agent_note": str(supplied.get("independent_agent_note") or "").strip(),
    }


def evaluate_highest_rjconv_scenarios(
    workbook: Path,
    clinical: dict[str, Any],
    intraday: dict[str, Any],
) -> dict[str, Any]:
    active_targets = clinical["relevant_targets"]
    conviction = {
        item["target"]: {row["outcome"]: float(row["conviction"]) for row in item["outcomes"]}
        for item in clinical["target_assessments"]
    }
    allowed = {
        target: [outcome for outcome, value in conviction[target].items() if value >= 0.10]
        for target in active_targets
    }
    scenario_count = math.prod(len(allowed[target]) for target in active_targets)
    formulas_wb = load_workbook(workbook, data_only=False, read_only=False, keep_links=True)
    values_wb = load_workbook(workbook, data_only=True, read_only=False, keep_links=True)
    try:
        ws = formulas_wb[clinical["sheet_name"]]
        values_ws = values_wb[clinical["sheet_name"]]
        if [ws.cell(7, col).value for col in range(2, 7)] != [
            "Scenario", "Base Case (USD/Share)", "Final Market Price", "Upside", "RJConv."
        ]:
            raise RuntimeError("Test scenario table lacks the required RJConv. column")
        rows = []
        for row in range(10, 10 + scenario_count):
            outcomes = [ws.cell(row, 7 + index).value for index in range(len(active_targets))]
            probability = math.prod(
                conviction[target][outcome]
                for target, outcome in zip(active_targets, outcomes)
            )
            blind_price = values_ws.cell(row, 4).value
            cached_probability = values_ws.cell(row, 6).value
            if not isinstance(blind_price, (int, float)):
                raise RuntimeError(f"blind Final Market Price cache missing at D{row}")
            if not isinstance(cached_probability, (int, float)) or abs(
                float(cached_probability) - probability
            ) > 1e-10:
                raise RuntimeError(f"RJConv. cache mismatch at F{row}")
            rows.append({
                "row": row,
                "scenario_id": int(ws.cell(row, 2).value),
                "outcomes": dict(zip(active_targets, outcomes)),
                "rjconv": probability,
                "blind_final_market_price": float(blind_price),
            })
    finally:
        formulas_wb.close()
        values_wb.close()
    highest = max(item["rjconv"] for item in rows)
    peak = float(intraday["three_day_peak"]["raw_high"])
    winners = []
    for item in rows:
        if abs(item["rjconv"] - highest) > 1e-12:
            continue
        blind = item["blind_final_market_price"]
        reached = peak >= blind
        difference = (peak - blind) if reached else (blind - peak)
        item["result"] = "REACHED" if reached else "MISS"
        item["difference_usd_per_share"] = difference
        item["difference_as_pct_of_real_peak"] = difference / peak
        winners.append(item)
    return {
        "rule": "max RJConv. (raw Table-3 conviction product); all exact ties retained",
        "three_day_real_peak": peak,
        "three_day_peak_bar_start": intraday["three_day_peak"]["bar_start"],
        "highest_rjconv": highest,
        "scenarios": winners,
    }


def validate_score_payload(
    score: dict[str, Any],
    clinical: dict[str, Any],
    blind_snapshot: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    if str(score.get("ticker") or "").upper() != clinical["ticker"]:
        raise ValueError("score ticker does not match frozen clinical artifact")
    if score.get("event") != clinical["event_name"]:
        raise ValueError("score event does not match frozen clinical artifact")
    if score.get("scoring_agent_role") != "independent_post_release_agent":
        raise ValueError("score payload is not marked as an independent post-release agent")
    for key in ("clinical_artifact_sha256", "blind_prediction_sha256"):
        if score.get(key) != blind_snapshot[key]:
            raise ValueError(f"score payload does not reference frozen {key}")

    post = score.get("post_release") or {}
    if post.get("release_date") != clinical["public_disclosure_date"]:
        raise ValueError("post-release date must equal the frozen earliest public disclosure")
    include_release = post.get("include_release_date_close")
    if not isinstance(include_release, bool):
        raise ValueError("include_release_date_close must be explicit")
    if not str(post.get("release_timing_basis") or "").strip():
        raise ValueError("release_timing_basis is required")
    source_url = str(post.get("source_url") or "")
    if not source_url.startswith("https://finance.yahoo.com/quote/"):
        raise ValueError("post-release price source must be the auditable Yahoo history page")

    expected_sessions = fetch_post_release_closes(
        str(post.get("market_ticker") or clinical["ticker"]).upper(),
        date.fromisoformat(clinical["public_disclosure_date"]),
        include_release,
    )
    supplied_sessions = post.get("sessions") or []
    if len(supplied_sessions) != 3:
        raise ValueError("exactly three post-release closes are required")
    for expected, supplied in zip(expected_sessions, supplied_sessions):
        if supplied.get("date") != expected["date"]:
            raise ValueError(f"post-release session date mismatch: {supplied} vs {expected}")
        if abs(float(supplied.get("close")) - expected["close"]) > 1e-6:
            raise ValueError(f"post-release raw close mismatch: {supplied} vs {expected}")

    target_scores = score.get("target_scores") or []
    expected_targets = clinical["relevant_targets"]
    if [item.get("target") for item in target_scores] != expected_targets:
        raise ValueError("target scores must follow the frozen active-target order exactly")
    for item in target_scores:
        value = item.get("score")
        if isinstance(value, bool) or not isinstance(value, int) or not 1 <= value <= 10:
            raise ValueError(f"score must be an integer from 1 to 10: {item}")
        rationale = str(item.get("rationale") or "").strip()
        if not rationale or len(rationale) > 280:
            raise ValueError("each score needs a brief rationale of at most 280 characters")
    return expected_sessions, target_scores


def facts_from_post_score(
    ticker: str,
    clinical: dict[str, Any],
    score: dict[str, Any],
) -> list[dict[str, Any]]:
    """Route post-release prices and evaluations outside clinical-only context."""
    post = score["post_release"]
    price_source = post["source_url"]
    last_session = post["sessions"][-1]["date"]
    facts: list[dict[str, Any]] = []
    for session in post["sessions"]:
        facts.append({
            "subject": ticker,
            "metric_group": "price",
            "metric": "historical_close",
            "value": session["close"],
            "unit": "USD",
            "population": f"post-release scoring window for {clinical['event_name']}",
            "as_of_date": session["date"],
            "source_url": price_source,
            "source_kind": "market_data_vendor",
            "classification": "Market Data",
            "status": "reported",
        })
    for item in score["target_scores"]:
        facts.append({
            "subject": item["target"],
            "indication": item["target"].rsplit(" - ", 1)[-1],
            "metric_group": "backtest",
            "metric": "post_release_prediction_score",
            "value": item["score"],
            "unit": "/10",
            "population": f"{clinical['event_name']} | {item['rationale']}",
            "as_of_date": last_session,
            "source_url": post["official_release_url"],
            "source_kind": "analyst_backtest",
            "classification": "Analyst Evaluation",
            "status": "estimated",
        })
    intraday = score.get("intraday") or {}
    for item in intraday.get("daily_highs") or []:
        facts.append({
            "subject": ticker,
            "metric_group": "price",
            "metric": "post_release_intraday_high_60m",
            "value": item["raw_high"],
            "unit": "USD",
            "population": f"regular session | {item['bar_count']} bars | {clinical['event_name']}",
            "as_of_date": item["date"],
            "source_url": intraday["source_url"],
            "source_kind": "market_data_vendor",
            "classification": "Market Data",
            "status": "reported",
        })
    for item in (score.get("highest_rjconv_scenario_test") or {}).get("scenarios") or []:
        facts.append({
            "subject": ticker,
            "metric_group": "backtest",
            "metric": "highest_rjconv_scenario_peak_test",
            "value": item["result"],
            "unit": "result",
            "population": (
                f"{clinical['event_name']} scenario {item['scenario_id']} | "
                f"blind={item['blind_final_market_price']:.6f} | "
                f"diff={item['difference_usd_per_share']:.6f} | "
                f"pct_peak={item['difference_as_pct_of_real_peak']:.6%}"
            ),
            "as_of_date": last_session,
            "source_url": intraday["source_url"],
            "source_kind": "analyst_backtest",
            "classification": "Analyst Evaluation",
            "status": "estimated",
        })
    return facts


def audit_overlay(
    workbook: Path,
    clinical: dict[str, Any],
    score: dict[str, Any],
    original_digest: str,
    original_max_row: int,
) -> dict[str, Any]:
    sheet_name = clinical["sheet_name"]
    post = score["post_release"]
    sessions = post["sessions"]
    targets = clinical["relevant_targets"]
    scores = {item["target"]: item for item in score["target_scores"]}
    digest_after, _ = frozen_sheet_digest(workbook, sheet_name, original_max_row)
    if digest_after != original_digest:
        raise RuntimeError("blind Test prediction changed while writing post-release overlay")
    wb = load_workbook(workbook, data_only=False, read_only=False, keep_links=True)
    try:
        ws = wb[sheet_name]
        expected_price_text = [
            f"{row['date'][5:].replace('-', '/')} ${float(row['close']):.4f}" for row in sessions
        ]
        actual_price_text = [ws.cell(2, col).value for col in range(4, 7)]
        if actual_price_text != expected_price_text:
            raise RuntimeError(f"row-2 post-release closes mismatch: {actual_price_text}")
        score_cells = {}
        for target in targets:
            group_col = next((
                col for col in range(6, ws.max_column - 2)
                if ws.cell(7, col).value == target
                and ws.cell(8, col).value == "USD/Share"
                and ws.cell(8, col + 3).value == "LOA"
            ), None)
            if group_col is None or ws.cell(8, group_col + 3).value != "LOA":
                raise RuntimeError(f"cannot resolve frozen LOA result column for {target}")
            coordinate = ws.cell(2, group_col + 3).coordinate
            expected = f"{scores[target]['score']}/10"
            if ws[coordinate].value != expected:
                raise RuntimeError(f"score placement mismatch at {coordinate}")
            score_cells[target] = coordinate
        column_b = {ws.cell(row, 2).value for row in range(1, ws.max_row + 1)}
        for required_label in (
            "Independent Post-Release Backtest",
            "Three-Day Intraday High Test",
            "Highest-RJConv. Scenario",
        ):
            if required_label not in column_b:
                raise RuntimeError(f"post-release audit section missing: {required_label}")
        strikes = [
            cell.coordinate for row in ws.iter_rows() for cell in row
            if cell.value is not None and cell.font and cell.font.strike
        ]
        if strikes:
            raise RuntimeError(f"post-release overlay introduced strikethrough: {strikes[:20]}")
        return {
            "blind_prediction_unchanged": True,
            "blind_sheet_digest": original_digest,
            "post_release_close_cells": ["D2", "E2", "F2"],
            "score_cells": score_cells,
            "post_release_sessions": sessions,
            "score_count": len(score_cells),
            "intraday_high_count": len(score["intraday"]["daily_highs"]),
            "highest_rjconv_scenario_count": len(
                score["highest_rjconv_scenario_test"]["scenarios"]
            ),
            "strikethrough_cells": 0,
        }
    finally:
        wb.close()


def main() -> None:
    parser = argparse.ArgumentParser(description="Score a frozen Test-EVENT prediction")
    parser.add_argument("--ticker", required=True)
    parser.add_argument("--event", required=True)
    parser.add_argument("--scores-file", required=True)
    parser.add_argument("--path")
    parser.add_argument("--no-backup", action="store_true")
    args = parser.parse_args()

    ticker = args.ticker.upper()
    slug = event_slug(args.event)
    artifact_dir = REPO / "artifacts" / ticker
    manifest_path = artifact_dir / f"{ticker}_test_{slug}_manifest.json"
    clinical_path = artifact_dir / f"{ticker}_test_{slug}_clinical.json"
    workbook = Path(args.path) if args.path else default_workbook(ticker)
    score_path = Path(args.scores_file)
    for path in (manifest_path, clinical_path, workbook, score_path):
        if not path.exists():
            raise SystemExit(f"required scoring input not found: {path}")

    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    clinical = json.loads(clinical_path.read_text(encoding="utf-8"))
    score = json.loads(score_path.read_text(encoding="utf-8"))
    if (manifest.get("post_release_scoring") or {}).get("status") == "complete":
        raise RuntimeError(
            "post-release scoring is already complete; rebuild the blind Test stage before rescoring"
        )
    if manifest.get("clinical_interpretation_price_blind") is not True:
        raise RuntimeError("blind clinical interpretation is not locked")
    if manifest.get("same_day_or_post_event_price_data_used") is not False:
        raise RuntimeError("blind stage already contains post-event market data")
    if clinical.get("same_day_or_post_event_price_data_used") is not False:
        raise RuntimeError("clinical artifact already contains post-event market data")

    clinical_sha = sha256_file(clinical_path)
    prediction_sha = canonical_sha256(blind_prediction_payload(clinical))
    existing_snapshot = manifest.get("blind_snapshot") or {}
    blind_snapshot = {
        "frozen_before_post_release_fetch": True,
        "frozen_at": existing_snapshot.get("frozen_at") or datetime.now().astimezone().isoformat(),
        "original_manifest_sha256": existing_snapshot.get("original_manifest_sha256") or sha256_file(manifest_path),
        "clinical_artifact_sha256": clinical_sha,
        "blind_prediction_sha256": prediction_sha,
        "clinical_interpretation_price_blind": True,
        "blind_stage_post_release_market_data_used": False,
    }
    sessions, target_scores = validate_score_payload(score, clinical, blind_snapshot)
    score["post_release"]["sessions"] = sessions
    score["target_scores"] = target_scores
    score["intraday"] = validate_intraday_payload(score, clinical, sessions)
    score["highest_rjconv_scenario_test"] = evaluate_highest_rjconv_scenarios(
        workbook, clinical, score["intraday"]
    )
    stored_facts = upsert_research_facts(
        facts_from_post_score(ticker, clinical, score), ticker
    )
    print(f"database scan: stored/refreshed {len(stored_facts)} post-release facts")

    original_digest, original_max_row = frozen_sheet_digest(workbook, clinical["sheet_name"])
    blind_snapshot["blind_sheet_digest"] = original_digest
    blind_snapshot["blind_sheet_original_max_row"] = original_max_row

    if not args.no_backup:
        backup = workbook.with_name(
            f"{workbook.stem}_pre_post_score_{slug}_{datetime.now():%Y%m%d_%H%M%S}{workbook.suffix}"
        )
        shutil.copy2(workbook, backup)
        print(f"backup → {backup}")

    normalized_score_path = artifact_dir / f"{ticker}_test_{slug}_post_release_score.json"
    normalized_score_path.write_text(
        json.dumps(score, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    powershell = Path("/mnt/c/Windows/System32/WindowsPowerShell/v1.0/powershell.exe")
    script = REPO / "tools" / "score_test_catalyst_event.ps1"
    subprocess.run([
        str(powershell), "-NoProfile", "-ExecutionPolicy", "Bypass", "-File",
        windows_path(script), "-Path", windows_path(workbook),
        "-ScorePath", windows_path(normalized_score_path),
    ], check=True)
    subprocess.run([
        sys.executable, str(REPO / "tools" / "normalize_calc_state.py"),
        "--path", str(workbook), "--no-backup",
    ], check=True)

    audit = audit_overlay(
        workbook, clinical, score, original_digest, original_max_row
    )
    with zipfile.ZipFile(workbook) as archive:
        if archive.testzip() is not None:
            raise RuntimeError("scored workbook ZIP integrity failure")

    manifest["blind_snapshot"] = blind_snapshot
    manifest["blind_stage_same_day_or_post_event_price_data_used"] = False
    manifest["post_release_scoring_market_data_used"] = True
    manifest["post_release_scoring"] = {
        "status": "complete",
        "agent_role": score["scoring_agent_role"],
        "release_date": clinical["public_disclosure_date"],
        "release_timing_basis": score["post_release"]["release_timing_basis"],
        "include_release_date_close": score["post_release"]["include_release_date_close"],
        "price_field": "unadjusted Close",
        "source_url": score["post_release"]["source_url"],
        "sessions": sessions,
        "intraday": score["intraday"],
        "highest_rjconv_scenario_test": score["highest_rjconv_scenario_test"],
        "target_scores": target_scores,
        "score_artifact": str(normalized_score_path),
        "audit": audit,
    }
    manifest_path.write_text(
        json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    print(
        f"independent post-release scoring complete → {clinical['sheet_name']}; "
        f"three closes, {len(target_scores)} target scores"
    )


if __name__ == "__main__":
    main()
