#!/usr/bin/env python3
"""Finish the single-event Catalyst workflow and repair related formats.

This is intentionally an OOXML patcher.  It does not save the workbook with
openpyxl, so Excel data tables, drawings, charts, and user-entered Catalyst
inputs survive unchanged.

The tool:
  * assigns the Catalyst outcome rows scenario IDs immediately after Breakdown;
  * rebuilds Scenarios / Catalyst Scenarios with one block per outcome;
  * sends the selected Catalyst market share to each new scenario block;
  * delegates construction of the embedded Catalyst B:C Excel data table;
  * links Catalyst's DCF comparison column to that table;
  * replaces the valuation-date cash embedded in the first waterfall component
    with positive current-year RCFS Ending Cash / shares outstanding; and
  * repairs the generated Scenarios and RBS number formats without changing
    their formulas or values.

Usage:
    python tools/sync_catalyst_model.py --ticker TARA
    python tools/sync_catalyst_model.py --path "/path/DCF TARA.xlsx"
"""

from __future__ import annotations

import argparse
import html
import json
import re
import shutil
import subprocess
import sys
import time
import zipfile
from pathlib import Path
from types import SimpleNamespace
from xml.etree import ElementTree as ET

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from core import _openpyxl_compat  # noqa: F401,E402
import openpyxl  # noqa: E402

from generate.generate_scenarios import (  # noqa: E402
    _add_scenario_block,
    _spacer_row,
)


NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


def _default_path(ticker: str) -> Path:
    return Path(f"/mnt/c/Users/yzsun/Desktop/DD/{ticker}/DCF {ticker}.xlsx")


def _read_parts(path: Path) -> tuple[dict[str, bytes], list[str]]:
    with zipfile.ZipFile(path, "r") as zf:
        return {name: zf.read(name) for name in zf.namelist()}, zf.namelist()


def _write_parts(path: Path, parts: dict[str, bytes], order: list[str]) -> None:
    tmp = path.with_suffix(".~catalyst_sync.xlsx")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        written: set[str] = set()
        for name in order:
            if name in parts:
                zout.writestr(name, parts[name])
                written.add(name)
        for name, payload in parts.items():
            if name not in written:
                zout.writestr(name, payload)
    tmp.replace(path)


def _sheet_zip_path(parts: dict[str, bytes], sheet_name: str) -> str:
    wb = ET.fromstring(parts["xl/workbook.xml"])
    rels = ET.fromstring(parts["xl/_rels/workbook.xml.rels"])
    relmap = {rel.get("Id"): rel.get("Target") for rel in rels}
    for sheet in wb.findall(f".//{{{NS_MAIN}}}sheet"):
        if html.unescape(sheet.get("name", "")) != sheet_name:
            continue
        target = relmap[sheet.get(f"{{{NS_R}}}id")].lstrip("/")
        return target if target.startswith("xl/") else "xl/" + target
    raise RuntimeError(f"Sheet not found: {sheet_name}")


def _cell_col(addr: str) -> int:
    letters = re.match(r"([A-Z]+)", addr).group(1)
    value = 0
    for char in letters:
        value = value * 26 + ord(char) - 64
    return value


def _cell_xml(xml: str, addr: str) -> str | None:
    match = re.search(
        rf'<c\b(?=[^>]*\br="{addr}")[^>]*/>|'
        rf'<c\b(?=[^>]*\br="{addr}")[^>]*>.*?</c>',
        xml,
        re.S,
    )
    return match.group(0) if match else None


def _replace_cell(xml: str, addr: str, cell_xml: str) -> str:
    match = re.search(
        rf'<c\b(?=[^>]*\br="{addr}")[^>]*/>|'
        rf'<c\b(?=[^>]*\br="{addr}")[^>]*>.*?</c>',
        xml,
        re.S,
    )
    if match:
        return xml[:match.start()] + cell_xml + xml[match.end():]

    row_num = int(re.search(r"\d+", addr).group(0))
    row_match = re.search(
        rf'(<row\b(?=[^>]*\br="{row_num}")[^>]*>)(.*?)(</row>)',
        xml,
        re.S,
    )
    if not row_match:
        raise RuntimeError(f"Row {row_num} not found while inserting {addr}")
    body = row_match.group(2)
    insert_at = len(body)
    target_col = _cell_col(addr)
    for candidate in re.finditer(r'<c\b[^>]*\br="([A-Z]+\d+)"', body):
        if _cell_col(candidate.group(1)) > target_col:
            insert_at = candidate.start()
            break
    body = body[:insert_at] + cell_xml + body[insert_at:]
    return xml[:row_match.start(2)] + body + xml[row_match.end(2):]


def _cell_style(xml: str, addr: str, fallback: str = "0") -> str:
    cell = _cell_xml(xml, addr)
    if not cell:
        return fallback
    match = re.search(r'\bs="(\d+)"', cell)
    return match.group(1) if match else fallback


def _formula_cell(addr: str, formula: str, style: str) -> str:
    return f'<c r="{addr}" s="{style}"><f>{html.escape(formula, quote=False)}</f></c>'


def _number_cell(addr: str, value: float | int, style: str) -> str:
    return f'<c r="{addr}" s="{style}"><v>{value}</v></c>'


def _empty_cell(addr: str, style: str) -> str:
    return f'<c r="{addr}" s="{style}"/>'


def _datatable_cell(addr: str, ref: str, style: str) -> str:
    return (
        f'<c r="{addr}" s="{style}"><f t="dataTable" ref="{ref}" '
        f'dt2D="1" dtr="1" r1="C5" r2="C3" ca="1"/><v>0</v></c>'
    )


class _StyleFormats:
    """Append number-format variants while preserving every visual XF field."""

    def __init__(self, styles_xml: str):
        self.xml = styles_xml
        self.cache: dict[tuple[str, str], str] = {}

    def _numfmt_id(self, code: str) -> str:
        for match in re.finditer(
            r'<numFmt\b[^>]*numFmtId="(\d+)"[^>]*formatCode="([^"]*)"[^>]*/>',
            self.xml,
        ):
            if html.unescape(match.group(2)) == code:
                return match.group(1)

        used = [int(x) for x in re.findall(r'<numFmt\b[^>]*numFmtId="(\d+)"', self.xml)]
        new_id = str(max([163] + used) + 1)
        tag = f'<numFmt numFmtId="{new_id}" formatCode="{html.escape(code, quote=True)}"/>'
        collection = re.search(r'<numFmts\b([^>]*)>(.*?)</numFmts>', self.xml, re.S)
        if collection:
            attrs, body = collection.group(1), collection.group(2)
            count = len(re.findall(r'<numFmt\b', body)) + 1
            attrs = re.sub(r'\bcount="\d+"', f'count="{count}"', attrs)
            if 'count="' not in attrs:
                attrs += f' count="{count}"'
            replacement = f'<numFmts{attrs}>{body}{tag}</numFmts>'
            self.xml = self.xml[:collection.start()] + replacement + self.xml[collection.end():]
        else:
            self.xml = self.xml.replace('<fonts', f'<numFmts count="1">{tag}</numFmts><fonts', 1)
        return new_id

    def style(self, source_style: str, code: str) -> str:
        key = (source_style, code)
        if key in self.cache:
            return self.cache[key]
        numfmt_id = self._numfmt_id(code)
        collection = re.search(r'<cellXfs\b([^>]*)>(.*?)</cellXfs>', self.xml, re.S)
        if not collection:
            raise RuntimeError("styles.xml has no cellXfs")
        xfs = re.findall(r'<xf\b[^>]*/>|<xf\b[^>]*>.*?</xf>', collection.group(2), re.S)
        source_idx = int(source_style)
        if source_idx >= len(xfs):
            raise RuntimeError(f"Cell style {source_idx} not found (cellXfs={len(xfs)})")
        xf = xfs[source_idx]
        current_numfmt = re.search(r'\bnumFmtId="(\d+)"', xf)
        if current_numfmt and current_numfmt.group(1) == numfmt_id:
            self.cache[key] = source_style
            return source_style
        if re.search(r'\bnumFmtId="\d+"', xf):
            xf = re.sub(r'\bnumFmtId="\d+"', f'numFmtId="{numfmt_id}"', xf, count=1)
        else:
            xf = xf.replace('<xf ', f'<xf numFmtId="{numfmt_id}" ', 1)
        if re.search(r'\bapplyNumberFormat="[^"]*"', xf):
            xf = re.sub(r'\bapplyNumberFormat="[^"]*"', 'applyNumberFormat="1"', xf, count=1)
        else:
            xf = xf.replace('<xf ', '<xf applyNumberFormat="1" ', 1)

        attrs, body = collection.group(1), collection.group(2)
        new_style = str(len(xfs))
        count = len(xfs) + 1
        attrs = re.sub(r'\bcount="\d+"', f'count="{count}"', attrs)
        if 'count="' not in attrs:
            attrs += f' count="{count}"'
        replacement = f'<cellXfs{attrs}>{body}{xf}</cellXfs>'
        self.xml = self.xml[:collection.start()] + replacement + self.xml[collection.end():]
        self.cache[key] = new_style
        return new_style


def _set_number_format(xml: str, addr: str, code: str, formats: _StyleFormats) -> str:
    cell = _cell_xml(xml, addr)
    if not cell:
        return xml
    old_style = _cell_style(xml, addr)
    new_style = formats.style(old_style, code)
    if re.search(r'\bs="\d+"', cell):
        replacement = re.sub(r'\bs="\d+"', f's="{new_style}"', cell, count=1)
    else:
        replacement = cell.replace('<c ', f'<c s="{new_style}" ', 1)
    return xml.replace(cell, replacement, 1)


def _indication_from_formula(value: object) -> str:
    text = str(value or "")
    match = re.search(r'&"\s*(.*?)\s+Market Share"', text)
    if match:
        return match.group(1).strip()
    match = re.search(r'\)\s+(.*?)\s+Market Share$', text)
    return match.group(1).strip() if match else "All"


def _read_model_layout(path: Path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        scenarios = wb["Scenarios"]
        catalyst = wb["Catalyst"]

        abs_first = None
        abs_last = None
        abs_asset_rows: dict[str, int] = {}
        abs_ms_rows: dict[tuple[str, str], int] = {}
        assets: list[SimpleNamespace] = []
        current_asset = None
        target_ms_row_match = re.search(r'Y\$(\d+)', str(catalyst["AD9"].value or ""))
        target_ms_row = int(target_ms_row_match.group(1)) if target_ms_row_match else None
        target_pair = None

        for row in range(1, scenarios.max_row + 1):
            a = scenarios.cell(row, 1).value
            b = scenarios.cell(row, 2).value
            c = scenarios.cell(row, 3).value
            d = scenarios.cell(row, 4).value
            if a != 4 or not isinstance(b, str) or b.strip() != "Absolute":
                continue
            abs_first = row if abs_first is None else abs_first
            abs_last = row
            if d == "[%]":
                if current_asset is None:
                    raise RuntimeError(f"Scenarios row {row}: market share has no parent asset")
                indication = _indication_from_formula(c)
                current_asset.market_shares[indication] = {}
                abs_ms_rows[(current_asset.name, indication)] = row
                if row == target_ms_row:
                    target_pair = (current_asset.name, indication)
            else:
                short_name = str(c).split(" (")[0].strip()
                current_asset = SimpleNamespace(name=short_name, market_shares={})
                assets.append(current_asset)
                abs_asset_rows[short_name] = row

        if not assets or abs_first is None or abs_last is None:
            raise RuntimeError("No Absolute scenario assets found")
        if target_pair is None:
            first = assets[0]
            first_indication = next(iter(first.market_shares), "All")
            target_pair = (first.name, first_indication)

        divider_row = None
        max_breakdown = 4
        in_breakdown = False
        for row in range(1, scenarios.max_row + 1):
            label = scenarios.cell(row, 3).value
            if label == "Break Down":
                in_breakdown = True
                continue
            if label == "Catalyst Scenarios":
                divider_row = row
                break
            if in_breakdown:
                scenario_id = scenarios.cell(row, 2).value
                if isinstance(scenario_id, (int, float)):
                    max_breakdown = max(max_breakdown, int(scenario_id))
        if divider_row is None:
            raise RuntimeError("Catalyst Scenarios divider not found")

        outcome_rows: list[tuple[int, str]] = []
        for row in range(10, min(catalyst.max_row, 200) + 1):
            outcome = catalyst.cell(row, 2).value or catalyst.cell(row, 28).value
            if isinstance(outcome, str) and outcome.startswith("="):
                outcome = catalyst.cell(row, 28).value
            if outcome in (None, ""):
                continue
            outcome_rows.append((row, str(outcome)))
        if not outcome_rows:
            raise RuntimeError("No Catalyst outcome rows found")

        event_name = str(catalyst["C7"].value or assets[0].name)
        existing_headers: set[int] = set()
        existing_asset_rows: set[int] = set()
        existing_market_share_rows: set[int] = set()
        for row in range(10, divider_row):
            a = scenarios.cell(row, 1).value
            b = scenarios.cell(row, 2).value
            c = scenarios.cell(row, 3).value
            d = scenarios.cell(row, 4).value
            if d == "[%]":
                existing_market_share_rows.add(row)
            elif c not in (None, "") and (
                a == 4 or isinstance(a, str) and a.startswith("=")
            ):
                existing_asset_rows.add(row)
            elif isinstance(b, (int, float)) and c not in (None, ""):
                existing_headers.add(row)
        return {
            "assets": assets,
            "abs_first": abs_first,
            "abs_last": abs_last,
            "abs_asset_rows": abs_asset_rows,
            "abs_ms_rows": abs_ms_rows,
            "target_pair": target_pair,
            "divider_row": divider_row,
            "scenario_start": max_breakdown + 1,
            "outcome_rows": outcome_rows,
            "event_name": event_name,
            "existing_headers": existing_headers,
            "existing_asset_rows": existing_asset_rows,
            "existing_market_share_rows": existing_market_share_rows,
        }
    finally:
        wb.close()


def _rebuild_catalyst_scenarios(xml: str, layout: dict) -> tuple[str, list[int], set[int], set[int], set[int]]:
    assets = layout["assets"]
    abs_ms_rows = layout["abs_ms_rows"]
    target_pair = layout["target_pair"]
    new_rows: list[str] = []
    scenario_ids: list[int] = []
    header_rows: set[int] = set()
    asset_rows: set[int] = set()
    ms_rows: set[int] = set()
    current = layout["divider_row"] + 1

    for idx, (catalyst_row, outcome) in enumerate(layout["outcome_rows"]):
        scenario_id = layout["scenario_start"] + idx
        scenario_ids.append(scenario_id)
        new_rows.append(_spacer_row(current))
        current += 1
        start = current
        header_rows.add(current)
        cursor = current + 1
        for asset in assets:
            asset_rows.add(cursor)
            cursor += 1
            for _indication in asset.market_shares:
                ms_rows.add(cursor)
                cursor += 1

        def peak_provider(asset_name, indication, source_row=catalyst_row):
            pair = (asset_name, indication)
            if pair == target_pair:
                return f"=Catalyst!$AD${source_row}"
            absolute_row = abs_ms_rows.get(pair)
            return f"=$Y${absolute_row}" if absolute_row else 0

        label = f"{layout['event_name']} - {outcome}"
        current = _add_scenario_block(
            new_rows,
            start,
            scenario_id,
            label,
            assets,
            layout["abs_first"],
            layout["abs_last"],
            layout["abs_asset_rows"],
            abs_ms_rows,
            peak_provider,
        )

    divider_match = re.search(
        rf'<row\b(?=[^>]*\br="{layout["divider_row"]}")[^>]*>.*?</row>',
        xml,
        re.S,
    )
    if not divider_match:
        raise RuntimeError("Could not locate Catalyst Scenarios divider XML")
    sheetdata_end = xml.find("</sheetData>")
    if sheetdata_end < 0:
        raise RuntimeError("Scenarios sheetData end not found")
    xml = xml[:divider_match.end()] + "\n" + "\n".join(new_rows) + "\n" + xml[sheetdata_end:]
    new_max = current - 1
    xml = re.sub(
        r'<dimension\b[^>]*ref="[^"]+"[^>]*/>',
        f'<dimension ref="A1:AE{new_max}"/>',
        xml,
        count=1,
    )
    return xml, scenario_ids, header_rows, asset_rows, ms_rows


def _repair_scenarios_formats(
    xml: str,
    formats: _StyleFormats,
    header_rows: set[int],
    asset_rows: set[int],
    market_share_rows: set[int],
) -> str:
    for row in sorted(header_rows):
        xml = _set_number_format(xml, f"B{row}", '"Scenario "0', formats)
    for row in sorted(asset_rows):
        for col in range(5, 25):
            addr = f"{openpyxl.utils.get_column_letter(col)}{row}"
            xml = _set_number_format(
                xml, addr, '"Stage "#,##0_);"(Stage "#,##0\\);""', formats
            )
        xml = _set_number_format(xml, f"Y{row}", "0.0%", formats)
    for row in sorted(market_share_rows):
        for col in range(5, 25):
            addr = f"{openpyxl.utils.get_column_letter(col)}{row}"
            xml = _set_number_format(xml, addr, "0%", formats)
        xml = _set_number_format(xml, f"Y{row}", "0.0%", formats)
    return xml


def _repair_rbs_formats(xml: str, formats: _StyleFormats) -> str:
    one_decimal = set(range(11, 17)) | set(range(19, 28))
    whole = {
        6, 32, 33, 34, 36, 41, 42, 43, 44, 45, 47,
        50, 51, 55, 56, 58, 61, 62, 63, 64, 65, 66, 67,
    }
    two_decimal = {52, 54}
    percent = {53}
    for row in range(1, 68):
        if row in one_decimal:
            code = '#,##0.0_);(#,##0.0);0.0'
        elif row in whole:
            code = '#,##0_);(#,##0);0'
        elif row in two_decimal:
            code = '#,##0.00_);(#,##0.00);0.00'
        elif row in percent:
            code = '0%'
        else:
            continue
        for col in range(6, 24):
            addr = f"{openpyxl.utils.get_column_letter(col)}{row}"
            xml = _set_number_format(xml, addr, code, formats)
    for col in range(6, 11):
        xml = _set_number_format(xml, f"{openpyxl.utils.get_column_letter(col)}4", "0", formats)
    for col in range(11, 24):
        xml = _set_number_format(xml, f"{openpyxl.utils.get_column_letter(col)}4", '0"E"', formats)
    return xml


def _patch_catalyst(xml: str, scenario_ids: list[int], outcome_rows: list[tuple[int, str]]) -> str:
    cash_per_share = (
        'MAX(0,INDEX(RCFS!$G$38:$W$38,1,'
        'MATCH(YEAR($B$9),RCFS!$G$4:$W$4,0)))/VALUATION!$C$47'
    )
    old_cash_per_share = 'MAX(0,VALUATION!$C$45/VALUATION!$C$47)'

    xml = _replace_cell(
        xml,
        "F7",
        f'<c r="F7" s="{_cell_style(xml, "F7")}" t="inlineStr">'
        '<is><t>Post-Catalyst DCF (USD/Share)</t></is></c>',
    )
    xml = _replace_cell(
        xml,
        "F9",
        _formula_cell(
            "F9",
            f'IFERROR(VALUATION!G$30-{old_cash_per_share}+{cash_per_share},0)',
            _cell_style(xml, "F9"),
        ),
    )
    xml = _replace_cell(
        xml,
        "G9",
        _formula_cell(
            "G9",
            f'IFERROR(MAX(0,VALUATION!G26-{old_cash_per_share}),0)',
            _cell_style(xml, "G9"),
        ),
    )
    xml = _replace_cell(
        xml,
        "X9",
        _formula_cell("X9", "IFERROR(VALUATION!$C$52,0)", _cell_style(xml, "X9")),
    )

    for idx, ((row, _outcome), scenario_id) in enumerate(zip(outcome_rows, scenario_ids)):
        xml = _replace_cell(
            xml, f"AA{row}", _number_cell(f"AA{row}", scenario_id, _cell_style(xml, f"AA{row}"))
        )
        valuation_row = 5 + idx
        xml = _replace_cell(
            xml,
            f"F{row}",
            _formula_cell(
                f"F{row}",
                f'IFERROR(VALUATION!P{valuation_row}-{old_cash_per_share}+{cash_per_share},0)',
                _cell_style(xml, f"F{row}"),
            ),
        )

    for row in [9] + [row for row, _ in outcome_rows]:
        formula = f'IFERROR(I{row}+M{row}+Q{row}+U{row}+{cash_per_share},0)'
        xml = _replace_cell(
            xml, f"W{row}", _formula_cell(f"W{row}", formula, _cell_style(xml, f"W{row}"))
        )
    return xml


def _patch_post_catalyst_table(xml: str, scenario_ids: list[int]) -> str:
    if not scenario_ids:
        return xml
    last_row = 4 + len(scenario_ids)
    o_style = _cell_style(xml, "O5")
    p_style = _cell_style(xml, "P5")
    for idx, scenario_id in enumerate(scenario_ids):
        row = 5 + idx
        xml = _replace_cell(xml, f"O{row}", _number_cell(f"O{row}", scenario_id, o_style))
        if idx == 0:
            xml = _replace_cell(
                xml, f"P{row}", _datatable_cell(f"P{row}", f"P5:P{last_row}", p_style)
            )
        else:
            xml = _replace_cell(xml, f"P{row}", _number_cell(f"P{row}", 0, p_style))
    for row in range(last_row + 1, 26):
        xml = _replace_cell(xml, f"O{row}", _empty_cell(f"O{row}", o_style))
        xml = _replace_cell(xml, f"P{row}", _empty_cell(f"P{row}", p_style))
    return xml


def sync(path: Path, backup: bool = True) -> dict:
    """Sync through the Excel-native embedded Catalyst builder.

    The builder preserves the analyst's Table-3 inputs and is the single source
    of truth for Catalyst and Scenarios catalyst blocks. The What-If table lives
    in Catalyst B:C; delegating prevents this legacy OOXML path from recreating
    Table 2 or the retired VALUATION O:P table.
    """
    if not path.exists():
        raise FileNotFoundError(path)
    if backup:
        backup_path = path.with_name(f"{path.stem}_pre_catalyst_sync_{time.strftime('%Y%m%d_%H%M%S')}.xlsx")
        shutil.copy2(path, backup_path)
    else:
        backup_path = None

    ticker = path.stem[4:] if path.stem.upper().startswith("DCF ") else path.parent.name
    subprocess.run(
        [
            sys.executable,
            str(REPO_ROOT / "generate" / "build_catalyst_framework.py"),
            "--ticker", ticker.upper(), "--path", str(path),
        ],
        check=True,
    )
    manifest_path = (
        REPO_ROOT / "artifacts" / ticker.upper() /
        f"{ticker.upper()}_catalyst_manifest.json"
    )
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    layout = manifest["layout"]
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        ws = wb["Catalyst"]
        scenario_ids = [
            int(ws.cell(row, 2).value)
            for row in range(layout["main_scenario_first"], layout["main_scenario_last"] + 1)
        ]
        outcome_col = int(layout.get("outcome_first_col") or 4)
        outcomes = [
            str(ws.cell(row, outcome_col).value)
            for row in range(layout["main_scenario_first"], layout["main_scenario_last"] + 1)
        ]
    finally:
        wb.close()

    return {
        "path": path,
        "backup": backup_path,
        "scenario_ids": scenario_ids,
        "target_pair": None,
        "outcomes": outcomes,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Sync Catalyst outcomes and embedded What-If results")
    parser.add_argument("--ticker")
    parser.add_argument("--path")
    parser.add_argument("--no-backup", action="store_true")
    args = parser.parse_args()
    if not args.path and not args.ticker:
        parser.error("provide --path or --ticker")
    path = Path(args.path) if args.path else _default_path(args.ticker)
    result = sync(path, backup=not args.no_backup)
    print(f"Catalyst model synced: {result['path']}")
    print(f"  Scenario IDs: {result['scenario_ids']}")
    print(f"  Outcomes: {result['outcomes']}")
    print(f"  Target market share: {result['target_pair']}")
    if result["backup"]:
        print(f"  Backup: {result['backup']}")


if __name__ == "__main__":
    main()
