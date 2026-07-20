#!/usr/bin/env python3
"""Strict visual/structural audit against the approved MOLN model styles."""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.formula import DataTableFormula

REPO = Path(__file__).resolve().parents[1]
RULES_PATH = REPO / "information" / "MODEL_FORMAT_RULES.json"


def color_sig(color) -> tuple:
    if color is None:
        return ()
    # Excel freely rewrites equivalent colours while copying formats between
    # workbooks (RGB -> indexed palette; automatic -> indexed 64).  Audit the
    # rendered colour, not the package-level encoding chosen on save.
    if color.type == "rgb" and isinstance(color.rgb, str):
        return ("rgb", color.rgb[-6:].upper())
    if color.type == "indexed" and isinstance(color.indexed, int):
        if color.indexed == 64:
            return ("auto",)
        if 0 <= color.indexed < len(COLOR_INDEX):
            return ("rgb", COLOR_INDEX[color.indexed][-6:].upper())
    if color.type == "auto" or color.auto is True:
        return ("auto",)
    if color.type == "theme":
        # In the standard Office theme used by both workbooks, theme 1 is the
        # dark/text colour.  Excel may materialize it as explicit RGB black.
        if color.theme == 1 and not color.tint:
            return ("rgb", "000000")
        return ("theme", color.theme, round(float(color.tint or 0), 4))
    return (color.type,)


def side_sig(side) -> tuple:
    if not side.style:
        return (None, ())
    colour = color_sig(side.color)
    # Excel alternates between an omitted border colour and explicit Automatic;
    # both render with the same default line colour.
    if colour in {(), ("auto",)}:
        colour = ("auto",)
    return (side.style, colour)


def style_sig(
    cell, *, ignore_fill: bool = False, ignore_number_format: bool = False
) -> tuple:
    font = cell.font
    alignment = cell.alignment
    border = cell.border
    if ignore_fill:
        fill = ()
    elif cell.fill.fill_type == "solid":
        # The background colour is not rendered for a solid fill and Excel
        # often normalizes it to indexed 64 during PasteSpecial.
        fill = ("solid", color_sig(cell.fill.fgColor))
    elif cell.fill.fill_type:
        fill = (cell.fill.fill_type, color_sig(cell.fill.fgColor), color_sig(cell.fill.bgColor))
    else:
        fill = ()
    return (
        font.name, font.sz, font.bold, font.italic, font.underline,
        font.strike, color_sig(font.color), fill,
        side_sig(border.left), side_sig(border.right), side_sig(border.top),
        side_sig(border.bottom), side_sig(border.diagonal),
        None if alignment.horizontal in (None, "general") else alignment.horizontal,
        None if alignment.vertical in (None, "bottom") else alignment.vertical,
        alignment.text_rotation,
        bool(alignment.wrap_text), bool(alignment.shrink_to_fit), alignment.indent,
        None if ignore_number_format else cell.number_format,
    )


def rating_key(value: Any) -> str:
    text = str(value or "")
    if re.search(r"BIC|Best", text, re.I):
        return "BIC"
    if re.search(r"T1|Tier One", text, re.I):
        return "T1"
    return "AVG"


def scenario_source_row(ws, row: int) -> int:
    a, b, c, d = [ws.cell(row, col).value for col in range(1, 5)]
    if not any(ws.cell(row, col).value not in (None, "") for col in range(1, 32)):
        return 19
    if c == "Break Down":
        return 53
    if c == "Catalyst Scenarios" or str(c or "").startswith("Test Scenarios - "):
        return 121
    if isinstance(b, (int, float)) and c not in (None, ""):
        return 20
    if d == "[%]" and a == 4:
        return 11
    if d == "[%]":
        return 22
    if a == 4 and str(b) == "Absolute":
        return 10
    return 21


def pipeline_source_row(ws, row: int) -> int:
    a = str(ws.cell(row, 1).value or "")
    d = str(ws.cell(row, 4).value or "")
    text = d
    if not text:
        return 15
    if "Market Share" in text:
        return 11
    if " TAM" in text:
        return 10
    if "List Price" in text:
        return 12
    if "Revenue" in text:
        return 13
    if "COGS" in text:
        return 14
    return 9 if a == "X" else 15


def add_mismatch(items: list, sheet: str, address: str, kind: str) -> None:
    if len(items) < 200:
        items.append({"sheet": sheet, "address": address, "kind": kind})


def visible_blank_format(cell) -> bool:
    """Return whether a blank cell has a rendered decoration worth auditing.

    Font, alignment and number-format differences on an empty cell cannot be
    seen.  Excel also commonly materializes a white fill while normalizing a
    workbook, which is visually identical to no fill on these model sheets.
    Black/coloured fills and borders remain strict because they implement the
    model's separator and decoration rows.
    """
    if cell.fill.fill_type:
        colour = color_sig(cell.fill.fgColor)
        if colour not in {(), ("rgb", "FFFFFF"), ("auto",)}:
            return True
    border = cell.border
    return any(getattr(border, side).style for side in (
        "left", "right", "top", "bottom", "diagonal"
    ))


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--ticker", required=True)
    parser.add_argument("--path")
    parser.add_argument("--reference")
    parser.add_argument("--rules", default=str(RULES_PATH))
    args = parser.parse_args()
    ticker = args.ticker.upper()
    rules = json.loads(Path(args.rules).read_text(encoding="utf-8"))
    path = Path(args.path) if args.path else Path(
        f"/mnt/c/Users/yzsun/Desktop/DD/{ticker}/DCF {ticker}.xlsx"
    )
    reference = Path(args.reference or rules["reference_workbook"])
    pipeline_reference = Path(rules["pipeline_reference_workbook"])
    target = load_workbook(path, data_only=False, read_only=False)
    ref = load_workbook(reference, data_only=False, read_only=False)
    pipeline_ref = load_workbook(
        pipeline_reference, data_only=False, read_only=False
    )
    mismatches: list[dict[str, str]] = []
    mismatch_count = 0
    manifest_path = REPO / "artifacts" / ticker / f"{ticker}_catalyst_manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    layout = manifest["layout"]
    framework_version = int(manifest.get("framework_version") or 0)

    # No delivered cell or conditional-format differential style may carry a
    # literal strike flag. Excel's separate stale-value strikethrough is gated
    # by workbook_audit's calc-state check.
    for ws in target.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.font and cell.font.strike:
                    mismatch_count += 1
                    add_mismatch(mismatches, ws.title, cell.coordinate,
                                 "stored font strikethrough")
    for index, dxf in enumerate(target._differential_styles.styles):
        if dxf.font and dxf.font.strike:
            mismatch_count += 1
            add_mismatch(mismatches, "Workbook", f"dxf:{index}",
                         "conditional-format strikethrough")

    def compare_cell(
        sheet: str,
        dst_cell,
        src_cell,
        ignore_fill: bool = False,
        ignore_number_format: bool = False,
    ):
        nonlocal mismatch_count
        if dst_cell.value in (None, "") and src_cell.value in (None, "") and \
                not visible_blank_format(dst_cell) and not visible_blank_format(src_cell):
            return
        if style_sig(
            dst_cell,
            ignore_fill=ignore_fill,
            ignore_number_format=ignore_number_format,
        ) != style_sig(
            src_cell,
            ignore_fill=ignore_fill,
            ignore_number_format=ignore_number_format,
        ):
            mismatch_count += 1
            add_mismatch(mismatches, sheet, dst_cell.coordinate, "cell style")

    def effective_row_height(ws, row: int) -> float:
        value = ws.row_dimensions[row].height
        if value is None:
            value = ws.sheet_format.defaultRowHeight or 15
        return round(float(value), 4)

    def effective_col_width(ws, col: str) -> float:
        if col in ws.column_dimensions and ws.column_dimensions[col].width is not None:
            value = ws.column_dimensions[col].width
        else:
            value = ws.sheet_format.defaultColWidth or 8.43
        return round(float(value), 4)

    for sheet, spec in rules["static_sheets"].items():
        dst, src = target[sheet], ref[sheet]
        for row in range(1, int(spec["rows"]) + 1):
            for col in range(1, int(spec["cols"]) + 1):
                if sheet == "VALUATION" and 23 <= row <= 60 and 6 <= col <= 10:
                    # Dynamic waterfall rows scale with the ticker's number of
                    # Breakdown scenarios and are checked semantically below.
                    continue
                if framework_version >= 5 and sheet == "VALUATION" and \
                        4 <= row <= 200 and col in (15, 16):
                    # The retired Catalyst O:P table is intentionally blank;
                    # its old reference formatting is not part of v5.
                    continue
                special_valuation_numfmt = sheet == "VALUATION" and (
                    (row == 4 and col == 12)
                    or (5 <= row <= 60 and col in (13, 16))
                )
                compare_cell(
                    sheet,
                    dst.cell(row, col),
                    src.cell(row, col),
                    ignore_number_format=special_valuation_numfmt,
                )
            if effective_row_height(dst, row) != effective_row_height(src, row):
                mismatch_count += 1; add_mismatch(mismatches, sheet, str(row), "row height")
        for col in range(1, int(spec["cols"]) + 1):
            dl = dst.cell(1, col).column_letter
            # Excel's 8.43 default width may serialize as 8.7109375 depending
            # on whether it is inherited or explicitly materialized.
            if abs(effective_col_width(dst, dl) - effective_col_width(src, dl)) > 0.35 or \
                    dst.column_dimensions[dl].hidden != src.column_dimensions[dl].hidden:
                mismatch_count += 1; add_mismatch(mismatches, sheet, dl, "column dimension")

    # Scenarios semantic rows.
    dst, src = target["Scenarios"], ref["Scenarios"]
    scenario_ws = dst
    for row in range(1, min(9, dst.max_row) + 1):
        for col in range(1, 32):
            compare_cell("Scenarios", dst.cell(row, col), src.cell(row, col))
    for row in range(10, dst.max_row + 1):
        source_row = scenario_source_row(dst, row)
        for col in range(1, 32):
            compare_cell("Scenarios", dst.cell(row, col), src.cell(source_row, col))

    # No numeric-only Catalyst header rows may survive.
    divider = next((r for r in range(1, dst.max_row + 1)
                    if dst.cell(r, 3).value == "Catalyst Scenarios"), None)
    if divider:
        for row in range(divider + 1, dst.max_row + 1):
            if isinstance(dst.cell(row, 2).value, (int, float)) and dst.cell(row, 3).value in (None, ""):
                mismatch_count += 1
                add_mismatch(mismatches, "Scenarios", str(row), "numeric-only Catalyst header")

    # Pipeline semantic rows.
    dst, src = target["Pipeline"], pipeline_ref["Pipeline"]
    for row in range(1, min(8, dst.max_row) + 1):
        for col in range(1, 35):
            compare_cell("Pipeline", dst.cell(row, col), src.cell(row, col))
    for row in range(9, dst.max_row + 1):
        source_row = pipeline_source_row(dst, row)
        for col in range(1, 35):
            compare_cell("Pipeline", dst.cell(row, col), src.cell(source_row, col))
    actual_cf = {
        str(cell_range)
        for cf in dst.conditional_formatting._cf_rules
        for cell_range in cf.sqref.ranges
    }
    expected_cf = {"B1:B6"} | {
        f"F{row}:AH{row}"
        for row in range(9, dst.max_row + 1)
        if pipeline_source_row(dst, row) == 11
    }
    if actual_cf != expected_cf:
        mismatch_count += 1
        add_mismatch(
            mismatches,
            "Pipeline",
            "conditional formatting",
            f"expected {sorted(expected_cf)}; got {sorted(actual_cf)}",
        )

    # Peer View company/peer + rating semantics.  When auditing the active MOLN
    # authority itself, its row-level Peer View variants are authoritative; do
    # not remap one MOLN row onto a different MOLN semantic exemplar.  Pipeline
    # remains independently checked against its locked historical authority.
    dst, src = target["Peer View"], ref["Peer View"]
    auditing_reference_itself = path.resolve() == reference.resolve()
    company_rows, peer_rows = {}, {}
    for row in range(8, src.max_row + 1):
        key = rating_key(src.cell(row, 7).value)
        bucket = company_rows if "MOLN" in str(src.cell(row, 6).value or "") else peer_rows
        bucket.setdefault(key, row)
    for row in range(1, min(7, dst.max_row) + 1):
        for col in range(1, 18):
            compare_cell("Peer View", dst.cell(row, col), src.cell(row, col))
    for row in range(8, dst.max_row + 1):
        if auditing_reference_itself:
            continue
        key = rating_key(dst.cell(row, 7).value)
        is_company = str(dst.cell(row, 6).value or "").endswith(" US Equity")
        source_row = (company_rows if is_company else peer_rows).get(key, 8 if is_company else 9)
        for col in range(1, 18):
            source_cell = src.cell(source_row, col)
            # Some MOLN peer rows intentionally leave narrative/result columns
            # blank.  If CMPX has content, compare it with the same semantic
            # company/peer + rating cell that actually contains that data type.
            if dst.cell(row, col).value not in (None, "") and source_cell.value in (None, ""):
                candidates = []
                for rr in range(8, src.max_row + 1):
                    ref_company = "MOLN" in str(src.cell(rr, 6).value or "")
                    if ref_company == is_company and rating_key(src.cell(rr, 7).value) == key and \
                            src.cell(rr, col).value not in (None, ""):
                        candidates.append(rr)
                if not candidates:
                    candidates = [rr for rr in range(8, src.max_row + 1)
                                  if src.cell(rr, col).value not in (None, "")]
                if candidates:
                    source_cell = src.cell(candidates[0], col)
            compare_cell("Peer View", dst.cell(row, col), source_cell)

    # Catalyst v3-v7 formula/layout gates. An active run keeps every target
    # visible and masks unrelated targets with grey background + grey data.
    cws = target["Catalyst"]
    active_state_path = REPO / "artifacts" / ticker / f"{ticker}_catalyst_active_state.json"
    active_mask_expected = active_state_path.exists()
    expected = len(manifest["targets"])
    expected_scenarios = (
        int(layout.get("scenario_count") or 0)
        if framework_version >= 4 else expected * 4
    )
    if divider:
        actual_end = next(
            (row for row in range(divider + 1, scenario_ws.max_row + 1)
             if str(scenario_ws.cell(row, 3).value or "").startswith("Test Scenarios - ")),
            scenario_ws.max_row + 1,
        )
        catalyst_title_ids = [
            scenario_ws.cell(row, 2).value
            for row in range(divider + 1, actual_end)
            if isinstance(scenario_ws.cell(row, 2).value, (int, float)) and
            scenario_ws.cell(row, 3).value not in (None, "")
        ]
        if len(catalyst_title_ids) != expected_scenarios or \
                len(catalyst_title_ids) != len(set(catalyst_title_ids)):
            mismatch_count += 1
            add_mismatch(mismatches, "Scenarios", str(divider + 1),
                         "Catalyst title count/ID uniqueness")
    for i, item in enumerate(manifest["targets"]):
        g = int(layout.get("target_group_first_col") or 7) + 4 * i
        ms = column_index_from_string(item["market_share_change_col"])
        if cws.cell(7, g).value != item["name"] or cws.cell(7, g + 2).value != "Market Price":
            mismatch_count += 1; add_mismatch(mismatches, "Catalyst", cws.cell(7, g).coordinate, "main target header")
        expected_ms = g if framework_version >= 4 else g + 2
        if ms != expected_ms or cws.cell(layout["table3_header_row"], ms).value != "Market Share Change" or \
                cws.cell(layout["table3_header_row"], ms + 1).value != "LOA Change" or \
                cws.cell(layout["table3_header_row"], ms + 2).value != "Conv.":
            mismatch_count += 1; add_mismatch(mismatches, "Catalyst", cws.cell(layout["table3_header_row"], ms).coordinate, "Table 3 alignment")
        ratio = str(cws.cell(6, g).value or "")
        base_col = int(layout.get("base_col") or 6)
        normalized_ratio = ratio.replace("$", "")
        if f"{cws.cell(9, g).coordinate}/{cws.cell(9, base_col).coordinate}" not in normalized_ratio:
            mismatch_count += 1; add_mismatch(mismatches, "Catalyst", cws.cell(6, g).coordinate, "Scenario-1 ratio lock")
        if framework_version >= 4:
            for col in range(g, g + 4):
                if bool(cws.column_dimensions[get_column_letter(col)].hidden):
                    mismatch_count += 1
                    add_mismatch(mismatches, "Catalyst", cws.cell(7, col).coordinate,
                                 "target breakdown column must remain visible")
                    break
            first = int(layout["main_scenario_first"])
            last = int(layout["main_scenario_last"])
            base_letter = get_column_letter(base_col)
            group_letter = get_column_letter(g)
            market_letter = get_column_letter(g + 2)
            loa_letter = get_column_letter(g + 3)
            for row in range(first, last + 1):
                breakdown = str(cws.cell(row, g).value or "").replace(" ", "")
                if f"${base_letter}{row}*${group_letter}$6" not in breakdown:
                    mismatch_count += 1
                    add_mismatch(mismatches, "Catalyst", cws.cell(row, g).coordinate,
                                 "breakdown must equal scenario Base Case x locked row-6 ratio")
                    break
                market_price = str(cws.cell(row, g + 2).value or "").replace(" ", "")
                if f"${group_letter}{row}*${loa_letter}{row}" not in market_price:
                    mismatch_count += 1
                    add_mismatch(mismatches, "Catalyst", cws.cell(row, g + 2).coordinate,
                                 "target Market Price must apply scenario LOA")
                    break
            if not item.get("active"):
                for row in range(first, last + 1):
                    expected_loa = f"=${loa_letter}$9"
                    actual_loa = str(cws.cell(row, g + 3).value or "").replace(" ", "")
                    if actual_loa != expected_loa:
                        mismatch_count += 1
                        add_mismatch(mismatches, "Catalyst", cws.cell(row, g + 3).coordinate,
                                     "non-catalyst LOA must remain at base")
                        break
            if item.get("masked") and active_mask_expected:
                bad_mask = None
                for cell_range in item.get("display_ranges") or []:
                    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
                    for row in range(min_row, max_row + 1):
                        for col in range(min_col, max_col + 1):
                            cell = cws.cell(row, col)
                            if isinstance(cell, MergedCell):
                                continue
                            if cell.fill.fill_type != "solid" or \
                                    color_sig(cell.fill.fgColor) != ("rgb", "E7E6E6") or \
                                    color_sig(cell.font.color) != ("rgb", "7F7F7F"):
                                bad_mask = cell.coordinate
                                break
                        if bad_mask:
                            break
                    if bad_mask:
                        break
                if bad_mask:
                    mismatch_count += 1
                    add_mismatch(mismatches, "Catalyst", bad_mask,
                                 "non-catalyst group requires grey background and grey data")
    if framework_version >= 4:
        active = manifest.get("active_targets") or []
        outcome_first = int(layout.get("outcome_first_col") or 3)
        for index, name in enumerate(active):
            if cws.cell(7, outcome_first + index).value != name:
                mismatch_count += 1
                add_mismatch(mismatches, "Catalyst", cws.cell(7, outcome_first + index).coordinate,
                             "active outcome header")
        base_col = int(layout["base_col"])
        final_col = int(layout["final_market_col"])
        if final_col != base_col + 1 or cws.cell(7, final_col).value != "Final Market Price":
            mismatch_count += 1
            add_mismatch(mismatches, "Catalyst", cws.cell(7, final_col).coordinate,
                         "Final Market must immediately follow Base Case")
        rjconv_col = int(layout.get("rjconv_col") or 6)
        if cws.cell(7, rjconv_col).value != "RJConv.":
            mismatch_count += 1
            add_mismatch(mismatches, "Catalyst", cws.cell(7, rjconv_col).coordinate,
                         "scenario RJConv. header")
        market_price_cols = [
            int(layout["target_group_first_col"]) + 4 * i + 2
            for i in range(len(manifest["targets"]))
        ]
        for row in range(int(layout["main_scenario_first"]),
                         int(layout["main_scenario_last"]) + 1):
            final_formula = str(cws.cell(row, final_col).value or "").replace(" ", "")
            missing = [
                cws.cell(row, col).coordinate
                for col in market_price_cols
                if f"${get_column_letter(col)}{row}" not in final_formula
            ]
            if missing:
                mismatch_count += 1
                add_mismatch(mismatches, "Catalyst", cws.cell(row, final_col).coordinate,
                             "Final Market must sum every target Market Price")
                break
    if any(str(cws.cell(8, col).value or "") == "Scenario" for col in range(27, cws.max_column + 1)):
        mismatch_count += 1; add_mismatch(mismatches, "Catalyst", "row 8", "legacy Table 2")
    ids = [cws.cell(r, 2).value for r in range(layout["main_scenario_first"], layout["main_scenario_last"] + 1)]
    if len(ids) != expected_scenarios or not all(isinstance(x, (int, float)) for x in ids):
        mismatch_count += 1; add_mismatch(mismatches, "Catalyst", "B10", "scenario ID count")
    if divider and catalyst_title_ids != ids:
        mismatch_count += 1
        add_mismatch(mismatches, "Scenarios", str(divider + 1),
                     "Catalyst title IDs differ from main table")

    # The Absolute Value table stays on VALUATION. Framework v5+ embeds the
    # Catalyst What-If table directly in Catalyst B:C and retires O:P.
    vws = target["VALUATION"]
    l4 = str(vws["L4"].value or "").upper().replace("$", "")
    if "C48" not in l4 or "RIS!" in l4:
        mismatch_count += 1
        add_mismatch(mismatches, "VALUATION", "L4", "Absolute Value output is not C48")
    if not isinstance(vws["M5"].value, DataTableFormula):
        mismatch_count += 1
        add_mismatch(mismatches, "VALUATION", "M5", "Absolute Value data table missing")
    if vws["L4"].number_format != '"Abs. Value"':
        mismatch_count += 1
        add_mismatch(mismatches, "VALUATION", "L4", "Absolute Value label format")
    for addr in ("M5",):
        if vws[addr].number_format != "0.00":
            mismatch_count += 1
            add_mismatch(mismatches, "VALUATION", addr, "price/share precision format")
    if framework_version >= 5:
        embedded = cws["C9"].value
        if not isinstance(embedded, DataTableFormula):
            mismatch_count += 1
            add_mismatch(mismatches, "Catalyst", "C9", "embedded Catalyst data table missing")
        elif getattr(embedded, "r1", None) != "C6" or getattr(embedded, "r2", None) != "B6":
            mismatch_count += 1
            add_mismatch(mismatches, "Catalyst", "C9", "embedded input bridge mismatch")
        if cws["C8"].value != 0.03 or cws["C8"].number_format != "0.0%":
            mismatch_count += 1
            add_mismatch(mismatches, "Catalyst", "C8", "embedded terminal growth must be 3.0%")
        if cws["B8"].number_format != '"ID"' or cws["B9"].number_format != '"Base"':
            mismatch_count += 1
            add_mismatch(mismatches, "Catalyst", "B8:B9", "embedded table labels")
        legacy = [
            vws.cell(row, col).value
            for row in range(4, 201) for col in (15, 16)
            if vws.cell(row, col).value not in (None, "")
        ]
        if legacy:
            mismatch_count += 1
            add_mismatch(mismatches, "VALUATION", "O4:P200", "legacy Catalyst table remains")
        c3 = str(vws["C3"].value or "").replace("$", "").upper()
        c5 = str(vws["C5"].value or "").replace("$", "").upper()
        if "CATALYST!B6" not in c3 or "CATALYST!C6" not in c5:
            mismatch_count += 1
            add_mismatch(mismatches, "VALUATION", "C3/C5", "embedded input bridges missing")
    elif ids and not isinstance(vws["P5"].value, DataTableFormula):
        mismatch_count += 1
        add_mismatch(mismatches, "VALUATION", "P5", "Catalyst data table missing")

    waterfall_last = max(
        (row for row in range(26, 61) if vws.cell(row, 9).value not in (None, "")),
        default=25,
    )
    waterfall_ref = ref["VALUATION"]
    for row in range(23, waterfall_last + 1):
        source_row = row if row <= 25 else 26
        for col in range(6, 11):
            compare_cell(
                "VALUATION",
                vws.cell(row, col),
                waterfall_ref.cell(source_row, col),
            )

    def chart_title(chart) -> str:
        try:
            rich = chart.title.tx.rich
            return " ".join(
                run.t or ""
                for paragraph in rich.p
                for run in (paragraph.r or [])
            ).strip()
        except Exception:
            return ""

    chart_titles = [chart_title(chart) for chart in vws._charts]
    if not any("Waterfall" in title for title in chart_titles):
        mismatch_count += 1
        add_mismatch(mismatches, "VALUATION", "chart", "Valuation Waterfall missing")

    report = {
        "ticker": ticker,
        "workbook": str(path),
        "reference": str(reference),
        "pipeline_reference": str(pipeline_reference),
        "mismatch_count": mismatch_count,
        "first_mismatches": mismatches,
    }
    out = REPO / "artifacts" / ticker / f"{ticker}_style_audit.json"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    target.close(); ref.close(); pipeline_ref.close()
    if mismatch_count:
        print(json.dumps(report, indent=2, ensure_ascii=False))
        raise SystemExit(1)
    print(f"Workbook style audit OK: MOLN rules matched across all 15 delivered tabs; report → {out}")


if __name__ == "__main__":
    main()
