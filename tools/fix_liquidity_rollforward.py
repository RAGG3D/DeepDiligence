#!/usr/bin/env python3
"""Use one auditable unrestricted-liquidity definition across RBS/RCFS/Valuation.

For pre-revenue biotech issuers, current and non-current marketable securities
fund operations just like cash. This tool prevents three recurring errors:
non-current securities buried in Other Assets, RCFS forecast cash rolled from a
stale modeled historical balance, and Valuation adding only bank cash while
omitting marketable securities.
"""
from __future__ import annotations

import argparse
import json
import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parents[1]
if str(REPO) not in sys.path:
    sys.path.insert(0, str(REPO))

from core.excel_writer import _apply_xlsx_patches


def _row_by_label(ws, *labels: str) -> int:
    for row in range(1, ws.max_row + 1):
        if str(ws.cell(row, 4).value or "").strip() in labels:
            return row
    raise RuntimeError(f"{ws.title}: label not found: {labels}")


def _years(ws, row: int = 4) -> List[Tuple[int, int]]:
    out = []
    for col in range(6, 24):
        value = ws.cell(row, col).value
        if isinstance(value, (int, float)):
            out.append((col, int(value)))
        elif (
            isinstance(value, str)
            and value.startswith("=")
            and out
            and value.replace("$", "").upper()
            == f"={get_column_letter(col - 1)}{row}+1"
        ):
            out.append((col, out[-1][1] + 1))
    return out


def _usd_m(value: Any, unit: str = ""):
    if isinstance(value, dict):
        unit = str(value.get("unit") or unit)
        value = value.get("value")
    if value is None:
        return None
    number = float(value)
    low = unit.lower()
    if "thousand" in low or low in {"k", "usd k"}:
        return number / 1000
    if "billion" in low or low in {"bn", "usd bn"}:
        return number * 1000
    return number


def build_patches(path: Path, ticker: str) -> Dict[str, List[Tuple[int, int, Any]]]:
    facts_path = REPO / "artifacts" / ticker / f"{ticker}_company_facts.json"
    facts = json.loads(facts_path.read_text(encoding="utf-8")) if facts_path.exists() else {}
    prior = facts.get("prior_year_reported_liquidity") or {}
    latest = facts.get("latest_reported_liquidity") or {}
    noncurrent = _usd_m(
        prior.get("noncurrent_marketable_debt_securities_usd_m",
                  prior.get("noncurrent_marketable_securities")),
        str(prior.get("unit") or "USD m"),
    )

    wb = load_workbook(path, data_only=False, read_only=True)
    wb_values = load_workbook(path, data_only=True, read_only=True)
    try:
        rbs = wb["RBS"]; rcfs = wb["RCFS"]
        # Newly generated RBS year headers are formula chains and may not yet
        # have cached values.  _years resolves the simple prior-year + 1 chain.
        rbs_years = _years(rbs)
        if not rbs_years:
            raise RuntimeError("RBS contains no year headers")
        fy_years = _years(wb_values["FY DATA"])
        historical_years = {year for _, year in fy_years}
        latest_hist_year = max(historical_years)
        latest_hist_col = next(col for col, year in rbs_years if year == latest_hist_year)
        first_forecast_idx = next(i for i, (_, year) in enumerate(rbs_years) if year > latest_hist_year)

        marketable_row = _row_by_label(rbs, "Marketable Securities")
        other_assets_row = _row_by_label(rbs, "Other Assets")
        total_financial_row = _row_by_label(rbs, "Total Financial Assets")
        rcfs_net_row = _row_by_label(rcfs, "Net Cash Flow [As Modelled]", "Net Unrestricted Liquidity Change [RBS-Derived]")
        rcfs_begin_row = _row_by_label(rcfs, "Beginning Cash", "Beginning Unrestricted Liquidity")
        rcfs_end_row = _row_by_label(rcfs, "Ending Cash", "Ending Unrestricted Liquidity")

        patches: Dict[str, List[Tuple[int, int, Any]]] = {"RBS": [], "RCFS": [], "VALUATION": []}

        # Correct the latest reported BS only when company facts prove the
        # non-current tranche was buried in the FY DATA Other Assets residual.
        if noncurrent is not None and float(noncurrent) > 0:
            m_addr = f"{get_column_letter(latest_hist_col)}{marketable_row}"
            o_addr = f"{get_column_letter(latest_hist_col)}{other_assets_row}"
            market_formula = str(rbs[m_addr].value or "0").lstrip("=")
            other_formula = str(rbs[o_addr].value or "0").lstrip("=")
            token = str(float(noncurrent))
            if f"+{token}" not in market_formula.replace(" ", ""):
                patches["RBS"].append((marketable_row, latest_hist_col, f"={market_formula}+{token}"))
            if f"-{token}" not in other_formula.replace(" ", ""):
                patches["RBS"].append((other_assets_row, latest_hist_col, f"=MAX(0,{other_formula}-{token})"))

        # RCFS now explicitly rolls unrestricted liquidity (cash + current and
        # non-current marketable securities), not an ambiguous cash subtotal.
        patches["RCFS"].extend([
            (rcfs_net_row, 4, ("text", "Net Unrestricted Liquidity Change [RBS-Derived]")),
            (rcfs_begin_row, 4, ("text", "Beginning Unrestricted Liquidity")),
            (rcfs_end_row, 4, ("text", "Ending Unrestricted Liquidity")),
        ])
        for i, (col, _year) in enumerate(rbs_years):
            letter = get_column_letter(col)
            patches["RCFS"].append((rcfs_end_row, col, f"=RBS!{letter}{total_financial_row}"))
            if i > 0:
                prev = get_column_letter(rbs_years[i - 1][0])
                patches["RCFS"].append((rcfs_begin_row, col, f"={prev}{rcfs_end_row}"))
            if i >= first_forecast_idx:
                patches["RCFS"].append(
                    (rcfs_net_row, col, f"={letter}{rcfs_end_row}-{letter}{rcfs_begin_row}")
                )

        hist_letter = get_column_letter(latest_hist_col)
        latest_unit = str(latest.get("unit") or "USD m")
        latest_cash = _usd_m(latest.get("cash_and_cash_equivalents_usd_m",
                                        latest.get("cash_and_cash_equivalents")), latest_unit)
        latest_current = _usd_m(latest.get("current_marketable_debt_securities_usd_m",
                                           latest.get("current_marketable_securities")), latest_unit)
        latest_noncurrent = _usd_m(latest.get("noncurrent_marketable_debt_securities_usd_m",
                                              latest.get("noncurrent_marketable_securities")), latest_unit)
        latest_total = _usd_m(
            latest.get("unrestricted_liquidity_usd_m",
                       latest.get("unrestricted_liquidity")),
            latest_unit,
        )
        components = [x for x in (latest_cash, latest_current, latest_noncurrent) if x is not None]
        model_year = datetime.now().year
        model_year_cols = [col for col, year in rbs_years if year == model_year]
        if model_year_cols:
            model_letter = get_column_letter(model_year_cols[0])
            # The valuation includes only positive residual liquidity.  A cash
            # deficit is already reflected through operating/financing needs
            # and must never be subtracted a second time from enterprise value.
            valuation_liquidity = f"=MAX(0,RCFS!{model_letter}{rcfs_end_row})"
            valuation_label = f"(+) Forecast Ending Unrestricted Liquidity ({model_year}E; floor at zero)"
        elif latest_total is not None:
            valuation_liquidity = f"={float(latest_total)}"
            as_of = str(latest.get("as_of") or "latest reported")
            valuation_label = f"(+) Latest Reported Unrestricted Liquidity ({as_of})"
        elif len(components) == 3:
            valuation_liquidity = "=" + "+".join(str(float(x)) for x in components)
            as_of = str(latest.get("as_of") or "latest reported")
            valuation_label = f"(+) Latest Reported Unrestricted Liquidity ({as_of})"
        else:
            valuation_liquidity = f"=RBS!{hist_letter}{total_financial_row}"
            valuation_label = "(+) Unrestricted Cash & Marketable Securities"
        patches["VALUATION"].extend([
            (45, 2, ("text", valuation_label)),
            (45, 3, valuation_liquidity),
        ])
        return patches
    finally:
        wb.close()
        wb_values.close()


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--ticker", required=True)
    ap.add_argument("--path")
    args = ap.parse_args()
    ticker = args.ticker.upper()
    path = Path(args.path) if args.path else Path(
        f"/mnt/c/Users/yzsun/Desktop/DD/{ticker}/DCF {ticker}.xlsx"
    )
    backup = path.with_name(f"{path.stem}_pre_liquidityfix_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
    shutil.copy2(path, backup)
    _apply_xlsx_patches(path, build_patches(path, ticker))
    print(f"Liquidity roll-forward corrected; backup: {backup}")


if __name__ == "__main__":
    main()
