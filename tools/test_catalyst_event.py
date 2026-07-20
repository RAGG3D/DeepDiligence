#!/usr/bin/env python3
"""Clinical-interpretation historical catalyst backtest.

Natural-language front ends map ``TICKER test EVENT`` to this command.  The
clinical judgement stage reads only medical evidence.  After that judgement is
locked, a separate calibration stage uses raw closes strictly before the first
public disclosure to estimate baseline breakdown LOAs.  It never retrieves or
uses same-day/post-event prices or a price reaction.  A different agent and the
separate ``score_test_catalyst_event.py`` tool may score the frozen prediction
only after this blind stage and its hashes are complete.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import math
import os
import re
import shutil
import subprocess
import sys
import time
import zipfile
from datetime import date, datetime, timedelta
from itertools import product
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urlparse

import duckdb
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.formula import DataTableFormula
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parents[1]
if str(REPO) not in sys.path:
    sys.path.insert(0, str(REPO))

from generate.adapt_catalyst import read_assets  # noqa: E402
from datastore.research_fact_store import upsert_research_facts  # noqa: E402

OUTCOMES = ("Increase", "Remain", "Decrease", "Suspension")
SOURCE_KINDS = {
    "company_clinical_release",
    "conference_abstract",
    "conference_poster",
    "clinical_trial_registry",
    "peer_reviewed_publication",
    "regulator_clinical_document",
    "competitor_clinical_release",
}
FORBIDDEN_KEY = re.compile(
    r"(?:^|_)(?:stock|share_price|price_reaction|price_movement|market_cap|"
    r"trading_volume|trading_return|abnormal_return|price_target|beta)(?:_|$)",
    re.I,
)
FORBIDDEN_TEXT = re.compile(
    r"\bstock\b|share\s+price|price\s+reaction|price\s+movement|price\s+target|"
    r"market\s+cap(?:italization)?|trading\s+volume|trading\s+return|"
    r"abnormal\s+return|\bbeta\b|\byfinance\b|finance\.yahoo\.com|"
    r"bloomberg\.com/quote|nasdaq\.com/market-activity|"
    r"marketwatch\.com/investing/stock",
    re.I,
)
FORBIDDEN_FORMULA = re.compile(
    r"BBG DAPI|Historical Events|price_reaction|share_price|\[[^\]]+\]",
    re.I,
)
FORBIDDEN_METRIC = re.compile(
    r"stock|price|sale|revenue|market cap|volume|return|beta|commercial",
    re.I,
)
CLINICAL_METRIC = re.compile(
    r"ORR|DCR|\bCR\b|\bPR\b|PFS|\bOS\b|DFS|DoR|duration|follow[- ]?up|"
    r"patient|evaluable|safety|adverse|\bAE\b|SAE|DLT|grade|G3|dose|dosing|"
    r"route|phase|result|response|endpoint|target|innovation|treatment line|"
    r"NCT|conference|date|biomarker|hazard ratio|confidence interval",
    re.I,
)
STANDARD_CLINICAL_DOMAINS = [
    "clinicaltrials.gov", "pubmed.ncbi.nlm.nih.gov", "ncbi.nlm.nih.gov",
    "fda.gov", "accessdata.fda.gov", "asco.org", "meetings.asco.org",
    "ascopubs.org", "hematology.org", "ashpublications.org", "esmo.org",
    "annalsofoncology.org", "aacr.org", "aacrjournals.org", "nejm.org",
    "thelancet.com", "nature.com", "jamanetwork.com", "who.int", "ctis.eu",
]


def default_workbook(ticker: str) -> Path:
    return Path(f"/mnt/c/Users/yzsun/Desktop/DD/{ticker}/DCF {ticker}.xlsx")


def company_facts(ticker: str) -> dict[str, Any]:
    path = REPO / "artifacts" / ticker.upper() / f"{ticker.upper()}_company_facts.json"
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def resolve_company_name(ticker: str, explicit: str | None) -> str:
    if explicit:
        return explicit
    facts = company_facts(ticker)
    for key in ("legal_name", "company_name", "issuer_name", "name"):
        if str(facts.get(key) or "").strip():
            return str(facts[key]).strip()
    raise SystemExit(
        "company name is missing; create the company-facts artifact or pass --company-name"
    )


def clinical_allowed_domains(ticker: str) -> list[str]:
    """Domain allowlist enforced by the research tool before any page retrieval."""
    domains = list(STANDARD_CLINICAL_DOMAINS)
    facts = company_facts(ticker)
    for key in ("ir_base_url", "ir_news_url"):
        host = (urlparse(str(facts.get(key) or "")).hostname or "").lower()
        if host:
            domains.append(host[4:] if host.startswith("www.") else host)
    return list(dict.fromkeys(domains))


def event_slug(event: str) -> str:
    slug = re.sub(r"[^A-Za-z0-9_-]+", "", event.strip())
    if not slug:
        raise ValueError("event must contain at least one letter or digit")
    return slug[:26]


def canonical_sha256(value: Any) -> str:
    raw = json.dumps(
        value, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str
    ).encode("utf-8")
    return hashlib.sha256(raw).hexdigest()


def blind_prediction_payload(payload: dict[str, Any]) -> dict[str, Any]:
    return {
        "ticker": payload["ticker"],
        "event_name": payload["event_name"],
        "event_date": payload["event_date"],
        "public_disclosure_date": payload["public_disclosure_date"],
        "relevant_targets": payload["relevant_targets"],
        "target_assessments": payload["target_assessments"],
        "pre_event_price_calibration": payload["pre_event_price_calibration"],
    }


def test_sheet_name(event: str) -> str:
    return f"Test-{event_slug(event)}"[:31]


def load_api_key() -> str:
    key = os.environ.get("OPENAI_API_KEY")
    if not key:
        env_path = REPO / ".env"
        if env_path.exists():
            for line in env_path.read_text(encoding="utf-8").splitlines():
                if line.strip().startswith("OPENAI_API_KEY") and "=" in line:
                    key = line.split("=", 1)[1].strip().strip('"').strip("'")
                    break
    if not key:
        raise SystemExit("clinical event research requires OPENAI_API_KEY")
    return key


def _target_indications(targets: Iterable[str]) -> set[str]:
    result: set[str] = set()
    for target in targets:
        if " - " in target:
            result.add(target.rsplit(" - ", 1)[1].upper())
    return result


def _section_matches(section: str, indications: set[str]) -> bool:
    upper = section.upper()
    tokens = set(re.findall(r"[A-Z0-9+/-]+", upper))
    for indication in indications:
        aliases = {indication}
        if indication == "HL":
            aliases |= {"CHL", "HODGKIN"}
        if indication == "MEL":
            aliases |= {"MELANOMA"}
        if indication == "MCRC":
            aliases |= {"CRC", "COLORECTAL"}
        if aliases & tokens or any(alias in upper for alias in aliases if len(alias) > 3):
            return True
    return False


def clinical_db_context(ticker: str, targets: list[str], db_path: Path) -> list[dict[str, Any]]:
    """Return only clinical peer metrics; financial/price metrics never leave SQL context."""
    if not db_path.exists():
        return []
    indications = _target_indications(targets)
    con = duckdb.connect(str(db_path), read_only=True)
    fact_rows: list[tuple[Any, ...]] = []
    try:
        rows = con.execute(
            """
            SELECT d.section_id, d.section, d.col, d.drug, d.ticker, d.rating,
                   m.metric, m.value
            FROM peer_drug d
            JOIN peer_metric m USING (section_id, col)
            ORDER BY d.section_id, d.col, m.metric
            """
        ).fetchall()
        tables = {row[0] for row in con.execute("SHOW TABLES").fetchall()}
        if "research_fact" in tables:
            fact_rows = con.execute(
                """
                SELECT indication, subject, comparator, metric, value, population,
                       dose, as_of_date, source_url
                FROM research_fact
                WHERE context_ticker=?
                  AND metric_group IN ('efficacy','safety','dose','trial')
                ORDER BY as_of_date DESC, subject, metric
                LIMIT 600
                """,
                [ticker.upper()],
            ).fetchall()
    finally:
        con.close()
    groups: dict[tuple[int, str], list[dict[str, Any]]] = {}
    for section_id, section, col, drug, peer_ticker, rating, metric, value in rows:
        metric = str(metric or "")
        value = str(value or "")
        if not _section_matches(str(section or ""), indications):
            continue
        if FORBIDDEN_METRIC.search(metric) or not CLINICAL_METRIC.search(metric):
            continue
        if FORBIDDEN_TEXT.search(value):
            continue
        groups.setdefault((int(section_id), str(col)), []).append({
            "section_id": int(section_id),
            "section": section,
            "drug": drug,
            "ticker": peer_ticker,
            "rating": rating,
            "metric": metric,
            "value": value,
            "origin": "dd.duckdb.peer_metric",
        })
    # Round-robin across section+competitor groups. A simple leading slice can
    # exhaust the context on one NSCLC table and omit other event indications.
    result: list[dict[str, Any]] = []
    ordered_groups = [groups[key] for key in sorted(groups)]
    for metric_index in range(25):
        for group in ordered_groups:
            if metric_index < len(group):
                result.append(group[metric_index])
            if len(result) >= 600:
                return result
    for indication, subject, comparator, metric, value, population, dose, as_of, source in fact_rows:
        if len(result) >= 600:
            break
        result.append({
            "section": indication,
            "drug": subject,
            "competitor": comparator,
            "metric": metric,
            "value": value,
            "population": population,
            "dose": dose,
            "as_of_date": as_of,
            "source_url": source,
            "origin": "dd.duckdb.research_fact",
        })
    return result


def _facts_from_test_payload(payload: dict[str, Any]) -> list[dict[str, Any]]:
    """Materialize newly validated Test evidence into the shared fact ledger."""
    source_by_id = {str(item["id"]): item for item in payload.get("sources") or []}
    facts: list[dict[str, Any]] = []
    raw_cutoff = str(payload.get("data_cutoff") or "")
    fact_as_of = (
        raw_cutoff if re.fullmatch(r"\d{4}-\d{2}-\d{2}", raw_cutoff)
        else payload["public_disclosure_date"]
    )

    def source_for(row: dict[str, Any]) -> dict[str, Any]:
        for source_id in row.get("source_ids") or []:
            if str(source_id) in source_by_id:
                return source_by_id[str(source_id)]
        raise ValueError(f"validated row has no resolved source: {row}")

    for row in payload.get("reported_data") or []:
        source = source_for(row)
        common = {
            "subject": row["target"],
            "indication": row["target"].rsplit(" - ", 1)[-1],
            "population": row["population"],
            "as_of_date": fact_as_of,
            "source_url": source["url"],
            "source_kind": source["source_kind"],
            "classification": "Reported Fact",
            "status": "reported",
        }
        for group, metric, value, unit in (
            ("trial", "n_enrolled", row["n_enrolled"], "patients"),
            ("trial", "n_evaluable", row["n_evaluable"], "patients"),
            ("efficacy", "efficacy_summary", row["efficacy"], None),
            ("safety", "safety_overview", row["safety"], None),
            ("trial", "follow_up", row["follow_up"], None),
        ):
            facts.append({**common, "metric_group": group, "metric": metric, "value": value, "unit": unit})
    for row in payload.get("competitor_comparisons") or []:
        source = source_for(row)
        facts.extend([
            {
                "subject": row["target"],
                "indication": row["target"].rsplit(" - ", 1)[-1],
                "comparator": row["competitor"],
                "metric_group": "efficacy",
                "metric": "clinical_comparison",
                "value": row["clinical_comparison"],
                "population": row["matched_setting"],
                "as_of_date": payload["public_disclosure_date"],
                "source_url": source["url"],
                "source_kind": source["source_kind"],
                "classification": "Reported Fact",
                "status": "reported",
            },
            {
                "subject": row["target"],
                "indication": row["target"].rsplit(" - ", 1)[-1],
                "comparator": row["competitor"],
                "metric_group": "trial",
                "metric": "cross_trial_limitations",
                "value": row["limitations"],
                "population": row["matched_setting"],
                "as_of_date": payload["public_disclosure_date"],
                "source_url": source["url"],
                "source_kind": source["source_kind"],
                "classification": "Analyst Assumption",
                "status": "reported",
            },
        ])
    calibration = payload.get("pre_event_price_calibration") or {}
    for session in calibration.get("sessions") or []:
        facts.append({
            "subject": calibration["market_ticker"],
            "metric_group": "price",
            "metric": "historical_close",
            "value": session["close"],
            "unit": calibration["currency"],
            "as_of_date": session["date"],
            "source_url": calibration["source_url"],
            "source_kind": "market_data_vendor",
            "classification": "Market Data",
            "status": "reported",
        })
    facts.append({
        "subject": calibration["market_ticker"],
        "metric_group": "price",
        "metric": "pre_disclosure_average_close",
        "value": calibration["average_close"],
        "unit": calibration["currency"],
        "population": f"{calibration['window_start']} to {calibration['window_end_exclusive']} exclusive",
        "as_of_date": payload["public_disclosure_date"],
        "source_url": calibration["source_url"],
        "source_kind": "market_data_vendor",
        "classification": "Market Data",
        "status": "reported",
    })
    return facts


def assert_clinical_only(payload: dict[str, Any]) -> None:
    """Reject security-market content in keys, narrative, URLs and source titles."""
    violations: list[str] = []

    def walk(value: Any, path: str = "$") -> None:
        if isinstance(value, dict):
            for key, item in value.items():
                key_text = str(key)
                # An explicit false compliance flag is allowed in the output
                # manifest; it is not a data field or a retrieved observation.
                if key_text != "price_data_used" and FORBIDDEN_KEY.search(key_text):
                    violations.append(f"{path}.{key_text} forbidden key")
                walk(item, f"{path}.{key_text}")
        elif isinstance(value, list):
            for index, item in enumerate(value):
                walk(item, f"{path}[{index}]")
        elif isinstance(value, str) and FORBIDDEN_TEXT.search(value):
            violations.append(f"{path} forbidden text={value[:100]!r}")

    walk(payload)
    if violations:
        raise ValueError("stock/price data is forbidden in test-event workflow: " + "; ".join(violations[:20]))


def _as_number(value: Any, label: str, low: float, high: float) -> float:
    if not isinstance(value, (int, float)) or isinstance(value, bool):
        raise ValueError(f"{label} must be numeric")
    number = float(value)
    if not low <= number <= high:
        raise ValueError(f"{label} must be in [{low}, {high}], got {number}")
    return number


def validate_payload(
    payload: dict[str, Any], ticker: str, event: str, target_universe: list[str]
) -> dict[str, Any]:
    assert_clinical_only(payload)
    if str(payload.get("ticker", "")).upper() != ticker.upper():
        raise ValueError("research ticker mismatch")
    if event_slug(str(payload.get("event", ""))).lower() != event_slug(event).lower():
        raise ValueError("research event mismatch")
    try:
        event_date = date.fromisoformat(str(payload.get("event_date") or ""))
    except ValueError as exc:
        raise ValueError("event_date must be a completed event in YYYY-MM-DD format") from exc
    if event_date > date.today():
        raise ValueError(f"test-event requires a completed disclosure; got future date {event_date}")
    try:
        public_disclosure_date = date.fromisoformat(
            str(payload.get("public_disclosure_date") or event_date.isoformat())
        )
    except ValueError as exc:
        raise ValueError("public_disclosure_date must be YYYY-MM-DD") from exc
    if public_disclosure_date > event_date:
        raise ValueError("public_disclosure_date cannot be after event_date")
    if public_disclosure_date > date.today():
        raise ValueError("public_disclosure_date must already have occurred")
    relevant = list(dict.fromkeys(payload.get("relevant_targets") or []))
    if not relevant or any(target not in target_universe for target in relevant):
        raise ValueError("relevant_targets must be a non-empty subset of the workbook target universe")
    assessments = payload.get("target_assessments") or []
    by_target = {item.get("target"): item for item in assessments if isinstance(item, dict)}
    if len(assessments) != len(relevant) or set(by_target) != set(relevant):
        raise ValueError("target_assessments must contain every relevant target exactly once")
    for target in relevant:
        item = by_target[target]
        if item.get("observed_outcome") not in OUTCOMES:
            raise ValueError(f"{target}: invalid observed_outcome")
        if not str(item.get("data_quality") or "").strip() or \
                not str(item.get("clinical_interpretation") or "").strip():
            raise ValueError(f"{target}: data_quality and clinical_interpretation are required")
        rows = item.get("outcomes") or []
        outcome_map = {row.get("outcome"): row for row in rows if isinstance(row, dict)}
        if len(rows) != len(OUTCOMES) or set(outcome_map) != set(OUTCOMES):
            raise ValueError(f"{target}: outcomes must be exactly {OUTCOMES}")
        total = 0.0
        for outcome in OUTCOMES:
            row = outcome_map[outcome]
            row["market_share_change"] = _as_number(
                row.get("market_share_change"), f"{target}/{outcome}/market_share_change", -1, 1
            )
            row["loa_change"] = _as_number(
                row.get("loa_change"), f"{target}/{outcome}/loa_change", -1, 1
            )
            row["conviction"] = _as_number(
                row.get("conviction"), f"{target}/{outcome}/conviction", 0, 1
            )
            total += row["conviction"]
            if not str(row.get("rationale") or "").strip():
                raise ValueError(f"{target}/{outcome}: rationale is required")
        if abs(total - 1.0) > 0.005:
            raise ValueError(f"{target}: convictions must sum to 1.0, got {total}")
        if not any(row["conviction"] >= 0.10 for row in outcome_map.values()):
            raise ValueError(f"{target}: no outcome survives the 10% filter")
        if outcome_map[item["observed_outcome"]]["conviction"] < 0.10:
            raise ValueError(f"{target}: observed outcome must survive the 10% filter")
    reported = payload.get("reported_data") or []
    reported_targets = {row.get("target") for row in reported if isinstance(row, dict)}
    if not set(relevant).issubset(reported_targets):
        raise ValueError("reported_data must include every relevant target")
    for row in reported:
        enrolled = _as_number(
            row.get("n_enrolled"), "reported_data/n_enrolled", 0, 1_000_000
        )
        evaluable = _as_number(
            row.get("n_evaluable"), "reported_data/n_evaluable", 0, 1_000_000
        )
        if evaluable > enrolled:
            raise ValueError("reported_data n_evaluable cannot exceed n_enrolled")
        for key in ("population", "efficacy", "safety", "follow_up"):
            if not str(row.get(key) or "").strip():
                raise ValueError(f"reported_data/{key} is required")
    comparisons = payload.get("competitor_comparisons") or []
    compared_targets = {row.get("target") for row in comparisons if isinstance(row, dict)}
    if not set(relevant).issubset(compared_targets):
        raise ValueError("competitor_comparisons must include every relevant target")
    for row in comparisons:
        for key in ("competitor", "matched_setting", "clinical_comparison", "limitations"):
            if not str(row.get(key) or "").strip():
                raise ValueError(f"competitor_comparisons/{key} is required")
    sources = payload.get("sources") or []
    if not sources:
        raise ValueError("at least one clinical source is required")
    for source in sources:
        if not isinstance(source, dict) or source.get("source_kind") not in SOURCE_KINDS:
            raise ValueError(f"invalid clinical source kind: {source}")
        if not str(source.get("url") or "").startswith(("http://", "https://")):
            raise ValueError("every source requires an http(s) URL")
    if not any(source["source_kind"] == "company_clinical_release" for source in sources):
        raise ValueError("full company clinical release is required")
    if not any(source["source_kind"] in {
        "conference_abstract", "conference_poster", "clinical_trial_registry",
        "peer_reviewed_publication", "regulator_clinical_document",
    } for source in sources):
        raise ValueError("independent/data-bearing primary clinical source missing")
    source_ids = {str(source.get("id") or "") for source in sources}
    if "" in source_ids or len(source_ids) != len(sources):
        raise ValueError("clinical source ids must be non-empty and unique")
    for collection in (reported, comparisons):
        for row in collection:
            cited = {str(item) for item in row.get("source_ids") or []}
            if not cited or not cited.issubset(source_ids):
                raise ValueError(f"unresolved clinical source_ids in {row}")
    payload["relevant_targets"] = relevant
    payload["target_universe"] = target_universe
    payload["sheet_name"] = test_sheet_name(event)
    payload["public_disclosure_date"] = public_disclosure_date.isoformat()
    payload["clinical_only"] = True
    payload["price_data_used"] = False
    assert_clinical_only(payload)
    return payload


def _pre_disclosure_average_close(
    market_ticker: str,
    public_disclosure_date: date,
    window_days: int,
    currency: str,
) -> dict[str, Any]:
    """Fetch only raw closes strictly before the first public disclosure.

    ``end`` is exclusive in yfinance.  The bounded request therefore cannot
    retrieve the disclosure session or any later reaction data.
    """
    if window_days < 2 or window_days > 30:
        raise ValueError("price-window-days must be between 2 and 30")
    window_start = public_disclosure_date - timedelta(days=window_days)
    data = yf.download(
        market_ticker,
        start=window_start.isoformat(),
        end=public_disclosure_date.isoformat(),
        auto_adjust=False,
        progress=False,
        threads=False,
    )
    if data is None or data.empty or "Close" not in data:
        raise RuntimeError(f"no pre-disclosure close history for {market_ticker}")
    close = data["Close"]
    if hasattr(close, "columns"):
        close = close.iloc[:, 0]
    sessions = [
        {"date": index.date().isoformat(), "close": float(value)}
        for index, value in close.dropna().items()
        if window_start <= index.date() < public_disclosure_date
    ]
    if len(sessions) < 2:
        raise RuntimeError(
            f"insufficient pre-disclosure sessions for {market_ticker}: {sessions}"
        )
    if any(date.fromisoformat(row["date"]) >= public_disclosure_date for row in sessions):
        raise RuntimeError("same-day/post-event price leaked into calibration window")
    average = sum(row["close"] for row in sessions) / len(sessions)
    if average <= 0:
        raise RuntimeError("pre-disclosure average close must be positive")
    return {
        "market_ticker": market_ticker,
        "currency": currency,
        "provider": "Yahoo Finance via yfinance",
        "source_url": f"https://finance.yahoo.com/quote/{market_ticker}/history/",
        "price_field": "unadjusted Close",
        "window_days": window_days,
        "window_start": window_start.isoformat(),
        "window_end_exclusive": public_disclosure_date.isoformat(),
        "sessions": sessions,
        "session_count": len(sessions),
        "average_close": average,
        "same_day_or_post_event_prices_used": False,
    }


def _estimate_breakdown_loa(
    workbook: Path,
    target_universe: list[str],
    average_close: float,
) -> dict[str, Any]:
    """Scale prior target LOAs while preserving their relative information.

    The constrained adjustment keeps every LOA in [0, 1].  It fails closed if
    even 100% LOA for every target cannot reconcile to the historical average.
    """
    formulas = load_workbook(workbook, data_only=False, read_only=False, keep_links=True)
    values = load_workbook(workbook, data_only=True, read_only=False, keep_links=True)
    try:
        fw = formulas["Catalyst"]
        vw = values["Catalyst"]
        groups: dict[str, tuple[float, float]] = {}
        for col in range(1, fw.max_column - 2):
            if [fw.cell(8, col + offset).value for offset in range(4)] != [
                "USD/Share", "LOA", "USD/Share", "LOA"
            ]:
                continue
            target = str(fw.cell(7, col).value or "")
            if target not in target_universe:
                continue
            un_risked = vw.cell(9, col).value
            prior_loa = vw.cell(9, col + 3).value
            if not isinstance(un_risked, (int, float)) or abs(float(un_risked)) <= 1e-12:
                raise RuntimeError(f"invalid Catalyst base value for {target}: {un_risked}")
            if not isinstance(prior_loa, (int, float)) or not 0 <= float(prior_loa) <= 1:
                raise RuntimeError(f"invalid Catalyst base LOA for {target}: {prior_loa}")
            groups[target] = (float(un_risked), float(prior_loa))
        if set(groups) != set(target_universe):
            raise RuntimeError(
                f"Catalyst breakdown targets mismatch: missing={set(target_universe)-set(groups)}"
            )
    finally:
        formulas.close()
        values.close()

    total_un_risked = sum(value for value, _ in groups.values())
    feasible_min = sum(min(0.0, value) for value, _ in groups.values())
    feasible_max = sum(max(0.0, value) for value, _ in groups.values())
    if not feasible_min - 1e-10 <= average_close <= feasible_max + 1e-10:
        raise RuntimeError(
            f"average close {average_close} is outside feasible LOA range "
            f"[{feasible_min}, {feasible_max}]"
        )
    prior_total = sum(value * loa for value, loa in groups.values())
    if prior_total > 0:
        scale = average_close / prior_total
        estimated = {
            target: min(1.0, max(0.0, prior_loa * scale))
            for target, (_, prior_loa) in groups.items()
        }
        method = "proportional scaling of existing target LOAs with [0%,100%] constraints"
    else:
        common = average_close / total_un_risked if total_un_risked else 0.0
        estimated = {
            target: min(1.0, max(0.0, common)) for target in groups
        }
        scale = None
        method = "constrained implied LOA because all prior risk-adjusted value was zero"

    # Resolve residuals created by clipping. This candidate search also handles
    # negative standalone target values without assuming every denominator is
    # positive: each iteration picks the feasible single-target LOA adjustment
    # that most reduces the absolute reconciliation error.
    for _ in range(len(groups) * 3 + 3):
        current = sum(groups[target][0] * estimated[target] for target in groups)
        residual = average_close - current
        if abs(residual) <= max(1e-10, average_close * 1e-10):
            break
        best = None
        for target, (value, _) in groups.items():
            proposed = min(1.0, max(0.0, estimated[target] + residual / value))
            improvement = value * (proposed - estimated[target])
            new_error = abs(residual - improvement)
            if best is None or new_error < best[0]:
                best = (new_error, target, proposed)
        if best is None or best[0] >= abs(residual) - 1e-12:
            break
        estimated[best[1]] = best[2]

    calibrated_total = sum(groups[target][0] * estimated[target] for target in groups)
    relative_error = abs(calibrated_total / average_close - 1)
    if relative_error >= 0.005:
        raise RuntimeError(
            f"breakdown LOA calibration error {relative_error:.4%} exceeds 0.5%"
        )
    return {
        "method": method,
        "proportional_scale": scale,
        "unrisked_total": total_un_risked,
        "feasible_market_price_min": feasible_min,
        "feasible_market_price_max": feasible_max,
        "prior_risk_adjusted_total": prior_total,
        "estimated_breakdown_total": calibrated_total,
        "relative_error": relative_error,
        "targets": {
            target: {
                "unrisked_value": groups[target][0],
                "prior_loa": groups[target][1],
                "estimated_loa": estimated[target],
                "calibrated_market_price": groups[target][0] * estimated[target],
            }
            for target in target_universe
        },
    }


def attach_pre_disclosure_calibration(
    payload: dict[str, Any],
    workbook: Path,
    market_ticker: str,
    currency: str,
    window_days: int,
) -> dict[str, Any]:
    public_date = date.fromisoformat(payload["public_disclosure_date"])
    calibration = _pre_disclosure_average_close(
        market_ticker, public_date, window_days, currency
    )
    calibration.update(
        _estimate_breakdown_loa(
            workbook, payload["target_universe"], calibration["average_close"]
        )
    )
    payload["pre_event_price_calibration"] = calibration
    payload["clinical_interpretation_price_blind"] = True
    payload["pre_event_price_calibration_used"] = True
    payload["same_day_or_post_event_price_data_used"] = False
    payload["price_data_used"] = True
    return payload


def build_prompt(
    ticker: str,
    company: str,
    event: str,
    targets: list[str],
    db_context: list[dict[str, Any]],
) -> str:
    context = json.dumps(db_context, ensure_ascii=False)
    return f"""Perform a clinical-only historical catalyst backtest for {company} ({ticker}),
event label {event}. Identify the actual completed clinical disclosure represented by the event,
read the full company clinical release and the conference abstract/poster or peer-reviewed report,
then compare the result with line-, biomarker-, population-, dose- and maturity-matched competitors.

ABSOLUTE PROHIBITION: do not search for, retrieve, mention or infer stock prices, price reactions,
returns, trading volume, market capitalization, beta, analyst price targets, financing or security
performance. Do not use finance portals. This test evaluates medical-data interpretation only.

Workbook target universe (use exact strings): {json.dumps(targets, ensure_ascii=False)}

Internal DD clinical-only peer rows (financial/price rows were removed before this prompt):
{context}

Use actual patient counts and denominators; distinguish evaluable from enrolled; preserve cutoff,
follow-up, response criteria, dose, line and subgroup. Treat cross-trial comparisons as uncertain.
For every directly affected target, provide an observed outcome plus a four-outcome probability
distribution. Convictions must sum to 1.0. market_share_change is an absolute share-point change
(0.02 means +2pp); loa_change is an absolute probability-point change. Derive both from clinical
strength, safety, durability, data maturity and matched competitor evidence—not from security data.

Output one fenced JSON object only with exactly this structure:
{{
  "ticker":"{ticker}", "event":"{event}", "event_name":"...", "event_date":"YYYY-MM-DD",
  "public_disclosure_date":"YYYY-MM-DD (earliest public release of these data)",
  "venue":"...", "data_cutoff":"...", "trial":"...", "phase":"...",
  "relevant_targets":["exact workbook target"],
  "reported_data":[{{"target":"...","population":"...","n_enrolled":0,"n_evaluable":0,
    "efficacy":"exact metrics","safety":"exact metrics","follow_up":"...","source_ids":["S1"]}}],
  "competitor_comparisons":[{{"target":"...","competitor":"...","matched_setting":"...",
    "clinical_comparison":"...","limitations":"...","source_ids":["S2"]}}],
  "target_assessments":[{{"target":"...","observed_outcome":"Increase|Remain|Decrease|Suspension",
    "data_quality":"...","clinical_interpretation":"...","outcomes":[
      {{"outcome":"Increase","market_share_change":0.0,"loa_change":0.0,"conviction":0.0,"rationale":"..."}},
      {{"outcome":"Remain","market_share_change":0.0,"loa_change":0.0,"conviction":0.0,"rationale":"..."}},
      {{"outcome":"Decrease","market_share_change":0.0,"loa_change":0.0,"conviction":0.0,"rationale":"..."}},
      {{"outcome":"Suspension","market_share_change":0.0,"loa_change":0.0,"conviction":0.0,"rationale":"..."}}
    ]}}],
  "sources":[{{"id":"S1","url":"https://...","source_kind":"company_clinical_release|conference_abstract|conference_poster|clinical_trial_registry|peer_reviewed_publication|regulator_clinical_document|competitor_clinical_release","title":"...","accessed":"YYYY-MM-DD"}}],
  "limitations":["..."]
}}
"""


def parse_json_object(text: str) -> dict[str, Any]:
    candidates = re.findall(r"```(?:json)?\s*(.*?)```", text, re.S | re.I) + [text]
    for candidate in candidates:
        match = re.search(r"\{.*\}", candidate, re.S)
        if not match:
            continue
        try:
            value = json.loads(match.group(0))
        except json.JSONDecodeError:
            continue
        if isinstance(value, dict):
            return value
    raise ValueError("research response did not contain a valid JSON object")


def research_payload(
    prompt: str, model: str, allowed_domains: list[str]
) -> dict[str, Any]:
    import openai

    os.environ["OPENAI_API_KEY"] = load_api_key()
    client = openai.OpenAI()
    started = time.time()
    response = client.responses.create(
        model=model,
        input=prompt,
        tools=[{
            "type": "web_search",
            "filters": {"allowed_domains": allowed_domains},
        }],
    )
    text = getattr(response, "output_text", None) or ""
    print(f"clinical-only research: {len(text):,} chars, {int(time.time()-started)}s")
    return parse_json_object(text)


def windows_path(path: Path) -> str:
    return subprocess.run(
        ["wslpath", "-w", str(path)], check=True, capture_output=True, text=True
    ).stdout.strip()


def audit_test_sheet(
    path: Path, payload: dict[str, Any], conviction_threshold: float
) -> dict[str, Any]:
    sheet_name = payload["sheet_name"]
    active_targets = payload["relevant_targets"]
    target_universe = payload["target_universe"]
    calibration = payload.get("pre_event_price_calibration") or {}
    average_close = _as_number(
        calibration.get("average_close"), "pre_event_price_calibration/average_close", 0.0001, 1_000_000
    )
    if calibration.get("same_day_or_post_event_prices_used") is not False:
        raise RuntimeError("price calibration is not strictly pre-disclosure")
    with zipfile.ZipFile(path) as archive:
        workbook_xml = archive.read("xl/workbook.xml").decode("utf-8", "ignore")
    calc_match = re.search(r"<(?:\w+:)?calcPr\b([^>]*)/?>", workbook_xml)
    calc_attrs = dict(re.findall(r'(\w+)="([^"]*)"', calc_match.group(1))) \
        if calc_match else {}
    if calc_attrs.get("calcMode", "auto").lower() == "manual" or \
            calc_attrs.get("calcCompleted") == "0" or \
            calc_attrs.get("calcOnSave") == "0" or \
            calc_attrs.get("forceFullCalc") == "1" or \
            "calcFeatures" in workbook_xml:
        raise RuntimeError(f"test-event left stale calculation state: {calc_attrs}")
    wb = load_workbook(path, data_only=False, read_only=False, keep_links=True)
    values_wb = load_workbook(path, data_only=True, read_only=False, keep_links=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise RuntimeError(f"missing generated sheet {sheet_name}")
        ws = wb[sheet_name]
        values_ws = values_wb[sheet_name]
        violations = []
        strikes = []
        formulas = 0
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if cell.font and cell.font.strike:
                    strikes.append(cell.coordinate)
                if isinstance(value, str) and value.startswith("="):
                    formulas += 1
                    if FORBIDDEN_FORMULA.search(value):
                        violations.append(f"{cell.coordinate}: {value}")
                elif (
                    cell.coordinate != "B5"
                    and isinstance(value, str)
                    and FORBIDDEN_TEXT.search(value)
                    and value not in {
                        str(calibration.get("provider") or ""),
                        str(calibration.get("source_url") or ""),
                    }
                ):
                    violations.append(f"{cell.coordinate}: {value}")
        if violations:
            raise RuntimeError(
                "test sheet contains forbidden external-market content: "
                + "; ".join(violations[:20])
            )
        if strikes:
            raise RuntimeError(f"test sheet contains strikethrough residue: {strikes[:20]}")
        if ws["B5"].value != "CLINICAL INTERPRETATION | PRE-DISCLOSURE PRICE-CALIBRATED":
            raise RuntimeError("pre-disclosure price-calibration banner missing")
        if abs(float(ws["C4"].value) - average_close) > 1e-10:
            raise RuntimeError(f"displayed pre-disclosure average mismatch: {ws['C4'].value}")
        if [ws.cell(7, col).value for col in range(2, 7)] != [
            "Scenario", "Base Case (USD/Share)", "Final Market Price", "Upside", "RJConv."
        ]:
            raise RuntimeError("Test main-table valuation headers do not match Catalyst")
        if not isinstance(ws["C9"].value, DataTableFormula):
            raise RuntimeError("Test B:C native What-If Data Table is missing")
        if ws["B6"].value != 1 or str(ws["C6"].value).replace(" ", "") != "=$C$8":
            raise RuntimeError("Test local What-If bridges B6/C6 are invalid")
        rjconv_col = 6
        outcome_first = 7
        headers = [ws.cell(7, outcome_first + i).value for i in range(len(active_targets))]
        if headers != active_targets:
            raise RuntimeError(f"active target headers mismatch: {headers} vs {active_targets}")
        assessment = {item["target"]: item for item in payload["target_assessments"]}
        allowed = {
            target: [
                row["outcome"] for row in assessment[target]["outcomes"]
                if row["conviction"] >= conviction_threshold
            ]
            for target in active_targets
        }
        expected_scenarios = 1
        for outcomes in allowed.values():
            expected_scenarios *= len(outcomes)
        scenario_rows = [
            row for row in range(10, 10 + expected_scenarios)
            if isinstance(ws.cell(row, 2).value, (int, float))
        ]
        if len(scenario_rows) != expected_scenarios:
            raise RuntimeError(
                f"scenario count mismatch: {len(scenario_rows)} vs {expected_scenarios}"
            )
        ordered_combinations = [
            tuple(ws.cell(row, outcome_first + i).value for i in range(len(active_targets)))
            for row in scenario_rows
        ]
        combinations = set(ordered_combinations)
        if len(combinations) != expected_scenarios:
            raise RuntimeError("filtered Cartesian scenarios are not unique")
        for combination in combinations:
            for index, outcome in enumerate(combination):
                if outcome not in allowed[active_targets[index]]:
                    raise RuntimeError(f"below-threshold outcome leaked into scenarios: {combination}")
        conviction_by_target = {
            target: {row["outcome"]: row["conviction"] for row in assessment[target]["outcomes"]}
            for target in active_targets
        }
        expected_ranked = sorted(
            product(*(allowed[target] for target in active_targets)),
            key=lambda combo: (
                -math.prod(
                    conviction_by_target[target][outcome]
                    for target, outcome in zip(active_targets, combo)
                ),
                "|".join(combo),
            ),
        )
        if ordered_combinations != expected_ranked:
            raise RuntimeError("Test scenarios are not sorted by descending RJConv.")
        if abs(float(ws.cell(9, rjconv_col).value) - 1.0) > 1e-12:
            raise RuntimeError("Test Base RJConv. must equal 100%")
        prior_probability = float("inf")
        for row, combination in zip(scenario_rows, ordered_combinations):
            expected_probability = math.prod(
                conviction_by_target[target][outcome]
                for target, outcome in zip(active_targets, combination)
            )
            formula = str(ws.cell(row, rjconv_col).value or "")
            cached = values_ws.cell(row, rjconv_col).value
            if "PRODUCT(" not in formula or "INDEX(" not in formula or "MATCH(" not in formula:
                raise RuntimeError(f"RJConv. formula is invalid at {ws.cell(row, rjconv_col).coordinate}")
            if not isinstance(cached, (int, float)) or abs(float(cached) - expected_probability) > 1e-10:
                raise RuntimeError(
                    f"RJConv. cache mismatch at {ws.cell(row, rjconv_col).coordinate}: "
                    f"{cached} vs {expected_probability}"
                )
            if expected_probability > prior_probability + 1e-12:
                raise RuntimeError("RJConv. is not descending")
            prior_probability = expected_probability
        scenario_ids = [int(ws.cell(row, 2).value) for row in scenario_rows]
        if len(scenario_ids) != len(set(scenario_ids)):
            raise RuntimeError("Test scenario IDs are not unique")
        display_targets = active_targets + [
            target for target in target_universe if target not in set(active_targets)
        ]
        group_first = outcome_first + len(active_targets)
        table_title = 11 + expected_scenarios
        table_target = table_title + 1
        table_header = table_title + 2
        table_input_first = table_title + 3
        table_input_last = table_input_first + 3
        outcome_order = list(OUTCOMES)
        market_price_cols = []
        for index, target in enumerate(display_targets):
            group_col = group_first + 4 * index
            if ws.cell(7, group_col).value != target:
                raise RuntimeError(f"target group header mismatch at {ws.cell(7, group_col).coordinate}")
            if ws.cell(7, group_col + 2).value != "Market Price" or [
                ws.cell(8, group_col + offset).value for offset in range(4)
            ] != ["USD/Share", "LOA", "USD/Share", "LOA"]:
                raise RuntimeError(f"Catalyst-style target headers missing for {target}")
            if any(
                bool(ws.column_dimensions[get_column_letter(col)].hidden)
                for col in range(group_col, group_col + 4)
            ):
                raise RuntimeError(f"test target group is hidden: {target}")
            market_price_cols.append(get_column_letter(group_col + 2))
            if not str(ws.cell(9, group_col).value or "").startswith("=") or \
                    str(ws.cell(9, group_col + 2).value or "").replace(" ", "") != \
                    f"=IFERROR(${get_column_letter(group_col)}9*${get_column_letter(group_col + 3)}9,0)":
                raise RuntimeError(f"Test Base target valuation logic mismatch for {target}")
            expected_loa = float(calibration["targets"][target]["estimated_loa"])
            actual_loa = ws.cell(9, group_col + 3).value
            if not isinstance(actual_loa, (int, float)) or abs(float(actual_loa) - expected_loa) > 1e-10:
                raise RuntimeError(
                    f"pre-disclosure implied LOA mismatch for {target}: {actual_loa} vs {expected_loa}"
                )
            expected_rows = {
                row["outcome"]: row
                for row in assessment[target]["outcomes"]
            } if target in assessment else {
                outcome: {
                    "market_share_change": 0.0,
                    "loa_change": 0.0,
                    "conviction": 1.0 if outcome == "Remain" else 0.0,
                }
                for outcome in OUTCOMES
            }
            for offset, outcome in enumerate(outcome_order):
                row = table_input_first + offset
                expected = expected_rows[outcome]
                actual = [ws.cell(row, group_col + col).value for col in range(3)]
                wanted = [
                    expected["market_share_change"], expected["loa_change"],
                    expected["conviction"],
                ]
                if any(abs(float(a) - float(b)) > 1e-10 for a, b in zip(actual, wanted)):
                    raise RuntimeError(f"clinical input mismatch for {target}/{outcome}: {actual} vs {wanted}")
            if target not in set(active_targets):
                for min_row, max_row in ((6, 9 + expected_scenarios),
                                         (table_target, table_input_last)):
                    for row in range(min_row, max_row + 1):
                        for col in range(group_col, group_col + 4):
                            cell = ws.cell(row, col)
                            if isinstance(cell, MergedCell):
                                continue
                            fill = (
                                (cell.fill.fgColor.rgb or "")[-6:].upper()
                                if cell.fill.fgColor.type == "rgb" else ""
                            )
                            font = (
                                (cell.font.color.rgb or "")[-6:].upper()
                                if cell.font.color and cell.font.color.type == "rgb" else ""
                            )
                            if cell.fill.fill_type != "solid" or fill != "E7E6E6" or font != "7F7F7F":
                                raise RuntimeError(
                                    f"inactive test target mask mismatch at {cell.coordinate}"
                                )
            for scenario_row in scenario_rows:
                value_formula = str(ws.cell(scenario_row, group_col).value or "").replace(" ", "")
                price_formula = str(ws.cell(scenario_row, group_col + 2).value or "").replace(" ", "")
                loa_formula = str(ws.cell(scenario_row, group_col + 3).value or "").replace(" ", "")
                value_required = (
                    f"$C{scenario_row}",
                    f"${get_column_letter(group_col)}$6",
                )
                if any(token not in value_formula for token in value_required):
                    raise RuntimeError(
                        f"Test scenario target value logic mismatch at "
                        f"{ws.cell(scenario_row, group_col).coordinate}: {value_formula}"
                    )
                expected_price = (
                    f"=IFERROR(${get_column_letter(group_col)}{scenario_row}*"
                    f"${get_column_letter(group_col + 3)}{scenario_row},0)"
                )
                if price_formula != expected_price:
                    raise RuntimeError(
                        f"Test scenario target market-price logic mismatch at "
                        f"{ws.cell(scenario_row, group_col + 2).coordinate}: {price_formula}"
                    )
                if target in set(active_targets):
                    active_index = active_targets.index(target)
                    outcome_col = get_column_letter(outcome_first + active_index)
                    required = (
                        f"${outcome_col}{scenario_row}",
                        f"${get_column_letter(group_col + 1)}${table_input_first}:"
                        f"${get_column_letter(group_col + 1)}${table_input_last}",
                        f"${get_column_letter(group_col + 3)}$9",
                        "MAX(0,",
                    )
                    if any(token not in loa_formula for token in required):
                        raise RuntimeError(
                            f"active Test LOA lookup mismatch at "
                            f"{ws.cell(scenario_row, group_col + 3).coordinate}: {loa_formula}"
                        )
                elif loa_formula != f"=${get_column_letter(group_col + 3)}$9":
                    raise RuntimeError(
                        f"inactive Test LOA changed at "
                        f"{ws.cell(scenario_row, group_col + 3).coordinate}: {loa_formula}"
                    )

        for scenario_row in [9] + scenario_rows:
            final_formula = str(ws.cell(scenario_row, 4).value or "").replace(" ", "")
            for market_col in market_price_cols:
                if f"${market_col}{scenario_row}" not in final_formula:
                    raise RuntimeError(
                        f"Test Final Market omits {market_col}{scenario_row}: {final_formula}"
                    )
            upside_formula = str(ws.cell(scenario_row, 5).value or "").replace(" ", "")
            if f"$D{scenario_row}" not in upside_formula or "$C$4" not in upside_formula:
                raise RuntimeError(
                    f"Test Upside formula mismatch at E{scenario_row}: {upside_formula}"
                )

        base_breakdown = sum(
            float(values_ws[f"{market_col}9"].value or 0) for market_col in market_price_cols
        )
        base_final = values_ws["D9"].value
        if not isinstance(base_final, (int, float)):
            raise RuntimeError(f"Test Base Final Market cache is not numeric: {base_final}")
        if abs(base_breakdown - float(base_final)) > 1e-8:
            raise RuntimeError(
                f"Test Base breakdown {base_breakdown} does not sum to Final Market {base_final}"
            )
        base_price_error = abs(base_breakdown / average_close - 1)
        if base_price_error >= 0.005:
            raise RuntimeError(
                f"Test Base breakdown differs from pre-disclosure average by {base_price_error:.4%}"
            )

        valuation_ws = wb["VALUATION"]
        scenario_route = str(valuation_ws["C3"].value or "").replace(" ", "")
        growth_route = str(valuation_ws["C5"].value or "").replace(" ", "")
        if "Catalyst!$B$6" not in scenario_route or f"'{sheet_name}'!$B$6" not in scenario_route:
            raise RuntimeError(f"VALUATION C3 Test routing missing: {scenario_route}")
        if "Catalyst!$C$6" not in growth_route or f"'{sheet_name}'!$C$6" not in growth_route:
            raise RuntimeError(f"VALUATION C5 Test routing missing: {growth_route}")

        # The Scenarios module and native Test Data Table form the model bridge.
        # It must use the same globally unique IDs and look up market-share
        # changes from the matching Test tab, never from Catalyst.
        sws = wb["Scenarios"]
        module_label = f"Test Scenarios - {payload['event_name']}"
        divider = next(
            (row for row in range(1, sws.max_row + 1)
             if sws.cell(row, 3).value == module_label),
            None,
        )
        if divider is None:
            raise RuntimeError(f"missing Scenarios module {module_label}")
        next_divider = next(
            (row for row in range(divider + 1, sws.max_row + 1)
             if str(sws.cell(row, 3).value or "").startswith("Test Scenarios - ")),
            sws.max_row + 1,
        )
        module_headers = [
            row for row in range(divider + 1, next_divider)
            if isinstance(sws.cell(row, 2).value, (int, float))
            and sws.cell(row, 3).value not in (None, "")
        ]
        module_ids = [int(sws.cell(row, 2).value) for row in module_headers]
        if module_ids != scenario_ids:
            raise RuntimeError(
                f"Test Scenarios IDs differ from Test tab: {module_ids} vs {scenario_ids}"
            )
        all_ids = [
            int(sws.cell(row, 2).value)
            for row in range(1, sws.max_row + 1)
            if isinstance(sws.cell(row, 2).value, (int, float))
            and sws.cell(row, 3).value not in (None, "")
        ]
        if len(all_ids) != len(set(all_ids)):
            raise RuntimeError("scenario IDs are not workbook-global unique")

        absolute_rows = [
            row for row in range(1, sws.max_row + 1)
            if sws.cell(row, 1).value == 4
            and str(sws.cell(row, 2).value or "") == "Absolute"
        ]
        if not absolute_rows:
            raise RuntimeError("Scenarios Absolute block missing")
        first_abs = min(absolute_rows)
        current_asset = None
        target_ms_rows: dict[str, int] = {}
        asset_rows: list[int] = []
        for row in absolute_rows:
            label = sws.cell(row, 3).value
            unit = sws.cell(row, 4).value
            if unit == "[%]" and current_asset:
                text = str(label or "")
                match = re.search(r'&"\s*(.*?)\s+Market Share"', text)
                if not match:
                    match = re.search(r'\)\s+(.*?)\s+Market Share$', text)
                indication = match.group(1).strip() if match else "All"
                name = current_asset if indication in ("", "All") else f"{current_asset} - {indication}"
                target_ms_rows[name] = row
            else:
                asset_rows.append(row)
                raw = str(label or "")
                current_asset = raw.split(" (")[0].strip()
        for scenario_index, header in enumerate(module_headers):
            for asset_row in asset_rows:
                destination = header + (asset_row - first_abs + 1)
                if sws.cell(destination, 25).value not in (None, ""):
                    raise RuntimeError(f"Test scenario asset-title Y residue at Y{destination}")
            for target_index, target_name in enumerate(display_targets):
                base_row = target_ms_rows.get(target_name)
                if base_row is None:
                    raise RuntimeError(f"Test target missing from Absolute: {target_name}")
                destination = header + (base_row - first_abs + 1)
                formula = str(sws.cell(destination, 25).value or "").replace(" ", "")
                if target_name not in set(active_targets):
                    if formula != f"=$Y${base_row}":
                        raise RuntimeError(
                            f"inactive Test market share changed at Y{destination}: {formula}"
                        )
                else:
                    active_index = active_targets.index(target_name)
                    outcome_col = get_column_letter(outcome_first + active_index)
                    ms_col = get_column_letter(group_first + 4 * target_index)
                    main_row = 10 + scenario_index
                    required = (
                        f"'{sheet_name}'!${outcome_col}${main_row}",
                        f"'{sheet_name}'!${ms_col}${table_input_first}:${ms_col}${table_input_last}",
                        f"$Y${base_row}",
                        "MAX(0,",
                    )
                    if any(token not in formula for token in required):
                        raise RuntimeError(
                            f"active Test market-share lookup mismatch at Y{destination}: {formula}"
                        )
        return {
            "sheet": sheet_name,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "formulas": formulas,
            "target_universe": len(target_universe),
            "active_targets": len(active_targets),
            "scenario_count": expected_scenarios,
            "scenario_ids": scenario_ids,
            "scenarios_module": {
                "label": module_label,
                "divider_row": divider,
                "first_id": scenario_ids[0],
                "last_id": scenario_ids[-1],
            },
            "research_clinical_only": True,
            "clinical_interpretation_price_blind": True,
            "pre_event_price_calibration_used": True,
            "same_day_or_post_event_price_data_used": False,
            "average_close": average_close,
            "base_breakdown_total": base_breakdown,
            "base_price_relative_error": base_price_error,
            "model_implied_valuation": True,
            "observed_market_data_used": True,
            "native_data_table": f"B8:C{9 + expected_scenarios}",
            "calculation_state": "automatic_non_stale",
        }
    finally:
        wb.close()
        values_wb.close()


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Clinical interpretation plus pre-disclosure price-calibrated Catalyst backtest"
    )
    parser.add_argument("--ticker", required=True)
    parser.add_argument("--event", required=True)
    parser.add_argument("--company-name")
    parser.add_argument("--path")
    parser.add_argument("--research-file")
    parser.add_argument("--research-only", action="store_true")
    parser.add_argument("--model", default="gpt-5.6-sol")
    parser.add_argument("--db", default=str(REPO / "datastore" / "dd.duckdb"))
    parser.add_argument("--conviction-threshold", type=float, default=0.10)
    parser.add_argument(
        "--market-ticker",
        help="Listing used only for the pre-disclosure raw-close calibration; defaults to TICKER",
    )
    parser.add_argument("--market-currency", default="USD")
    parser.add_argument("--price-window-days", type=int, default=7)
    parser.add_argument("--no-backup", action="store_true")
    args = parser.parse_args()

    ticker = args.ticker.upper()
    workbook = Path(args.path) if args.path else default_workbook(ticker)
    if not workbook.exists():
        raise SystemExit(f"workbook not found: {workbook}")
    targets = read_assets(workbook)
    if not targets:
        raise SystemExit("no Catalyst target universe found in Scenarios Absolute")
    db_rows = clinical_db_context(ticker, targets, Path(args.db))
    if args.research_file:
        payload = json.loads(Path(args.research_file).read_text(encoding="utf-8"))
    else:
        company_name = resolve_company_name(ticker, args.company_name)
        payload = research_payload(
            build_prompt(ticker, company_name, args.event, targets, db_rows),
            args.model,
            clinical_allowed_domains(ticker),
        )
    # A previously generated runtime artifact may be supplied on a rerun. Strip
    # its old market calibration before the clinical-only validator runs, then
    # fetch a fresh, identically bounded pre-disclosure window after judgement.
    for runtime_key in (
        "pre_event_price_calibration",
        "clinical_interpretation_price_blind",
        "pre_event_price_calibration_used",
        "same_day_or_post_event_price_data_used",
    ):
        payload.pop(runtime_key, None)
    if payload.get("price_data_used") is True:
        payload["price_data_used"] = False
    payload["internal_clinical_context_rows"] = db_rows
    payload = validate_payload(payload, ticker, args.event, targets)
    payload = attach_pre_disclosure_calibration(
        payload,
        workbook,
        (args.market_ticker or ticker).upper(),
        args.market_currency.upper(),
        args.price_window_days,
    )
    stored_facts = upsert_research_facts(_facts_from_test_payload(payload), ticker)
    print(f"database scan: stored/refreshed {len(stored_facts)} validated Test facts")

    artifact_dir = REPO / "artifacts" / ticker
    artifact_dir.mkdir(parents=True, exist_ok=True)
    slug = event_slug(args.event)
    research_path = artifact_dir / f"{ticker}_test_{slug}_clinical.json"
    research_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"validated clinical interpretation + pre-disclosure calibration → {research_path}")
    if args.research_only:
        return

    if not args.no_backup:
        backup = workbook.with_name(
            f"{workbook.stem}_pre_test_{slug}_{datetime.now():%Y%m%d_%H%M%S}{workbook.suffix}"
        )
        shutil.copy2(workbook, backup)
        print(f"backup → {backup}")
    powershell = Path("/mnt/c/Windows/System32/WindowsPowerShell/v1.0/powershell.exe")
    script = REPO / "tools" / "build_test_catalyst_event.ps1"
    subprocess.run([
        str(powershell), "-NoProfile", "-ExecutionPolicy", "Bypass", "-File",
        windows_path(script), "-Path", windows_path(workbook),
        "-ResearchPath", windows_path(research_path),
        "-ConvictionThreshold", str(args.conviction_threshold),
    ], check=True)
    subprocess.run([
        sys.executable,
        str(REPO / "tools" / "normalize_calc_state.py"),
        "--path", str(workbook),
        "--no-backup",
    ], check=True)
    audit = audit_test_sheet(workbook, payload, args.conviction_threshold)
    with zipfile.ZipFile(workbook) as archive:
        if archive.testzip() is not None:
            raise RuntimeError("generated workbook ZIP integrity failure")
    manifest = {
        "ticker": ticker,
        "event": args.event,
        "sheet_name": payload["sheet_name"],
        "research_clinical_only": True,
        "clinical_interpretation_price_blind": True,
        "pre_event_price_calibration_used": True,
        "same_day_or_post_event_price_data_used": False,
        "observed_market_data_used": True,
        "model_implied_valuation": True,
        "clinical_only": False,
        "price_data_used": True,
        "price_calibration": payload["pre_event_price_calibration"],
        "blind_snapshot": {
            "frozen_before_post_release_fetch": True,
            "frozen_at": datetime.now().astimezone().isoformat(),
            "clinical_artifact_sha256": hashlib.sha256(research_path.read_bytes()).hexdigest(),
            "blind_prediction_sha256": canonical_sha256(blind_prediction_payload(payload)),
            "clinical_interpretation_price_blind": True,
            "blind_stage_post_release_market_data_used": False,
        },
        "post_release_scoring_market_data_used": False,
        "post_release_scoring": {
            "status": "pending",
            "requires_independent_agent": True,
            "required_close_count": 3,
        },
        "research": str(research_path),
        "workbook": str(workbook),
        "audit": audit,
    }
    manifest_path = artifact_dir / f"{ticker}_test_{slug}_manifest.json"
    manifest_path.write_text(json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8")
    print(
        f"price-calibrated clinical Catalyst test complete → {payload['sheet_name']}; "
        f"manifest → {manifest_path}"
    )


if __name__ == "__main__":
    main()
