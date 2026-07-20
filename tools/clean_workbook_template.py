#!/usr/bin/env python3
"""Clean stale template labels after a ticker-specific DCF build.

The model template carries labels from the prior company in several support
sheets.  This script only patches text/header cells; it does not rewrite the
model structure.
"""

from __future__ import annotations

import argparse
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple

import openpyxl

REPO = Path(__file__).resolve().parents[1]
if str(REPO) not in sys.path:
    sys.path.insert(0, str(REPO))

from core.excel_writer import _apply_xlsx_patches
from tools.workbook_audit import DEFAULT_FORBIDDEN, FORBIDDEN_BY_TICKER


HIST_COLS = "FGHIJ"
FORECAST_COL = "K"
# RIS (rows 8-11/26-29/37-40), Schedules (rows 7-10) and Catalyst (rows 51-54)
# each expose exactly this many drug slots.  We do not insert Excel rows, so
# pipelines with more assets are truncated to this capacity (with a warning).
TEMPLATE_DRUG_SLOTS = 4
DATA_CENTER_SHEETS = {
    "Peer View",
    "Peer Views",
    "TAM Solid",
    "TAM Solid+MM",
    "TAM Blood",
}


def _clean_code(asset: str) -> str:
    if not asset:
        return ""
    return str(asset).split(" ", 1)[0].strip()


def _read_assets(path: Path) -> List[str]:
    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    try:
        ws = wb["Pipeline"]
        assets: List[str] = []
        for row in range(1, min(ws.max_row, 120) + 1):
            a = ws.cell(row, 1).value
            c = ws.cell(row, 3).value
            d = ws.cell(row, 4).value
            if a == "X" and (c is None or c == "") and isinstance(d, str):
                label = d.strip()
                if (
                    not label
                    or label.startswith("=")
                    or label in {"Referred Table", "Fiscal Year"}
                    or label.startswith("Referred")
                ):
                    continue
                if label not in assets:
                    assets.append(label)
        if len(assets) > TEMPLATE_DRUG_SLOTS:
            dropped = assets[TEMPLATE_DRUG_SLOTS:]
            print(
                f"WARNING: Pipeline has {len(assets)} assets but the template "
                f"exposes only {TEMPLATE_DRUG_SLOTS} drug slots; dropping "
                f"{dropped!r} from RIS/Schedules/Catalyst labels.",
                file=sys.stderr,
            )
        return assets[:TEMPLATE_DRUG_SLOTS]
    finally:
        wb.close()


def _read_catalyst_loa_labels(path: Path, n: int = 4) -> List[str]:
    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    try:
        ws = wb["Catalyst"]
        labels: List[str] = []
        for cell in ("J9", "N9", "R9", "V9"):
            value = ws[cell].value
            if isinstance(value, (int, float)):
                labels.append(f"{value:.0%} LOA")
            elif isinstance(value, str) and value.startswith("="):
                labels.append("Formula LOA")
            elif value not in (None, ""):
                labels.append(f"{value} LOA")
            else:
                labels.append("0% LOA")
        while len(labels) < n:
            labels.append("0% LOA")
        return labels[:n]
    finally:
        wb.close()


def _read_note_rows(path: Path) -> Tuple[List[str], List[str], List[str]]:
    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    try:
        ws = wb["FY DATA K USD"]
        rd: List[str] = []
        ga: List[str] = []
        ppe: List[str] = []
        for row in range(28, 37):
            name = ws.cell(row, 4).value
            if isinstance(name, str) and name and name != "Reserved":
                rd.append(name)
        for row in range(42, 47):
            name = ws.cell(row, 4).value
            if isinstance(name, str) and name and name != "Reserved":
                ga.append(name)
        for row in range(93, 98):
            name = ws.cell(row, 4).value
            if isinstance(name, str) and name and name != "Reserved":
                ppe.append(name)
        return rd, ga, ppe
    finally:
        wb.close()


# RIS drug-row layout (revenue-timeline / sales / COGS blocks).
_RIS_DRUG_ROWS = [8, 9, 10, 11, 26, 27, 28, 29, 37, 38, 39, 40]


def _read_ris_drug_labels(path: Path) -> Dict[int, Any]:
    """Current RIS col-D labels on the drug rows.

    adapt_ris runs earlier in the pipeline and, for a pipeline with more than
    TEMPLATE_DRUG_SLOTS assets, sets rows 11/29/40 to 'Other Pipeline …'
    aggregate labels (with a tail-aggregate formula).  We must not clobber those
    with the concrete 4th asset, otherwise fix_statement_logic — which keys its
    stage-load aggregate on D11=='Other Pipeline Active Stage Load' — drops the
    tail programs' R&D/G&A burden.
    """
    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    try:
        ws = wb["RIS"]
        return {r: ws.cell(r, 4).value for r in _RIS_DRUG_ROWS}
    finally:
        wb.close()


def _short_asset_labels(assets: List[str], n: int) -> List[str]:
    labels = list(assets[:n])
    while len(labels) < n:
        labels.append("Reserved")
    return labels


def _build_patches(path: Path, ticker: str, years: List[int]) -> Dict[str, List[Tuple[int, int, Any]]]:
    assets = _read_assets(path)
    rd_rows, ga_rows, ppe_rows = _read_note_rows(path)
    ris_labels = _read_ris_drug_labels(path)

    def _keep_aggregate(row: int) -> bool:
        """Preserve adapt_ris's 'Other Pipeline …' aggregate label on this RIS
        drug row (only present when the pipeline has >TEMPLATE_DRUG_SLOTS
        assets) instead of overwriting it with the concrete 4th asset."""
        v = ris_labels.get(row)
        return isinstance(v, str) and v.startswith("Other Pipeline")

    patches: Dict[str, List[Tuple[int, int, Any]]] = {}

    def add(sheet: str, row: int, col: int, value: Any) -> None:
        patches.setdefault(sheet, []).append((row, col, value))

    # Global ticker labels.
    add("WELCOME!", 18, 2, ("text", f"{ticker} US Equity"))
    if "Playgroud":
        add("Playgroud", 5, 2, ("text", f"{ticker} pipeline review"))

    # Historical headers for schedule-style tabs: F:J are completed FYs;
    # K remains the first forecast year.
    if years:
        header_years = years[-5:]
        for sheet, row in (("RIS", 5), ("Schedules", 4)):
            for idx, year in enumerate(header_years):
                add(sheet, row, 6 + idx, year)
            add(sheet, row, 11, header_years[-1] + 1)

    # Current assets in RIS and Schedules.
    four_assets = _short_asset_labels(assets, TEMPLATE_DRUG_SLOTS)
    for idx, asset in enumerate(four_assets):
        row = 8 + idx
        if _keep_aggregate(row):
            continue
        add("RIS", row, 4, ("text", asset))
    for idx, asset in enumerate(four_assets):
        row = 26 + idx
        if _keep_aggregate(row):
            continue
        label = f"{asset} Revenue" if asset != "Reserved" else "Reserved"
        add("RIS", row, 4, ("text", label))
    for idx, asset in enumerate(four_assets):
        row = 37 + idx
        if _keep_aggregate(row):
            continue
        label = f"{asset} COGS" if asset != "Reserved" else "Reserved"
        add("RIS", row, 4, ("text", label))

    for idx, asset in enumerate(four_assets):
        add("Schedules", 7 + idx, 4, ("text", asset))

    # RIS operating expense line labels.  The row layout has gaps that contain
    # ratio/helper rows, so only patch rows that carry ISN labels.
    rd_slots = [47, 48, 49, 50, 52, 54, 55, 56, 58]
    ga_slots = [62, 63, 64, 66, 67]
    for row, name in zip(rd_slots, rd_rows + ["Reserved"] * len(rd_slots)):
        add("RIS", row, 4, ("text", name))
    for row, name in zip(ga_slots, ga_rows + ["Reserved"] * len(ga_slots)):
        add("RIS", row, 4, ("text", name))

    # Schedules PP&E rows are a 5-row carrying-amount block for IFRS filers.
    ppe_slots = [20, 22, 24, 26, 29]
    for row, name in zip(ppe_slots, ppe_rows + ["Reserved"] * len(ppe_slots)):
        add("Schedules", row, 4, ("text", name))
    add("Schedules", 28, 4, ("text", "PP&E [Subtotal]"))
    add("Schedules", 30, 4, ("text", "PP&E [Net]"))

    # Catalyst support table labels.
    asset_codes = [_clean_code(a) for a in assets]
    while len(asset_codes) < TEMPLATE_DRUG_SLOTS:
        asset_codes.append("Reserved")
    loa_labels = _read_catalyst_loa_labels(path, TEMPLATE_DRUG_SLOTS)
    for i, code in enumerate(asset_codes[:TEMPLATE_DRUG_SLOTS], start=51):
        label = f"{code}, {loa_labels[i - 51]}" if code != "Reserved" else "Reserved, 0% LOA"
        add("Catalyst", i, 6, ("text", label))

    # Remove template-only coverage/market-reaction examples.  They are not
    # model inputs for the current issuer and can leak unrelated peer tickers
    # (e.g. IOBT/NUVL) into an otherwise clean new-ticker workbook.  Preserve
    # F:G, where the live Catalyst drug/LOA support table sits.
    for row in range(51, 59):
        for col in range(2, 5):
            add("Catalyst", row, col, ("text", ""))
    for row in range(62, 65):
        for col in range(2, 5):
            add("Catalyst", row, col, ("text", ""))

    # Data-center helper column repair.  Some legacy workbooks carry #REF!
    # inside TAM Blood column-B duplicate counters after row deletions.  These
    # are display/helper formulas only; rebuild them to match TAM Solid+MM.
    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    try:
        if "TAM Blood" in wb.sheetnames:
            ws = wb["TAM Blood"]
            for row in range(1, ws.max_row + 1):
                val = ws.cell(row, 2).value
                if not (isinstance(val, str) and val.startswith("=") and "COUNTIFS" in val):
                    continue
                formula = (
                    f'=IF(AND($C{row}&lt;&gt;"",'
                    f'COUNTIFS($D:$D,$D{row},$C:$C,$C{row})&gt;1),'
                    f'COUNTIFS($D$1:D{row},$D{row},$C$1:C{row},$C{row}),"")'
                )
                add("TAM Blood", row, 2, formula)
    finally:
        wb.close()

    return patches


def _validate_no_stale(path: Path, ticker: str) -> List[str]:
    own = {
        term.lower()
        for term in FORBIDDEN_BY_TICKER.get(ticker.upper(), [ticker])
    }
    stale_terms = [
        term for term in DEFAULT_FORBIDDEN
        if term.lower() not in own and term.upper() != ticker.upper()
    ]
    found: List[str] = []
    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    try:
        for ws in wb.worksheets:
            if ws.title in DATA_CENTER_SHEETS:
                continue
            for row in ws.iter_rows():
                for cell in row:
                    val = cell.value
                    if not isinstance(val, str):
                        continue
                    for term in stale_terms:
                        if re.search(re.escape(term), val, flags=re.IGNORECASE):
                            found.append(f"{ws.title}!{cell.coordinate}: {term} in {val!r}")
                            break
    finally:
        wb.close()
    return found


def main() -> None:
    parser = argparse.ArgumentParser(description="Clean stale DCF template labels")
    parser.add_argument("--ticker", required=True)
    parser.add_argument("--path", required=True)
    parser.add_argument("--years", nargs="*", type=int, default=None)
    args = parser.parse_args()

    path = Path(args.path)
    if not path.exists():
        raise FileNotFoundError(path)

    # Match build_dcf.default_years(): the last five completed fiscal years.
    last_full_year = datetime.now().year - 1
    years = args.years or list(range(last_full_year - 4, last_full_year + 1))
    backup = path.with_name(f"{path.stem}_pre_templateclean_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
    shutil.copy2(path, backup)

    patches = _build_patches(path, args.ticker.upper(), years)
    _apply_xlsx_patches(path, patches)

    found = _validate_no_stale(path, args.ticker.upper())
    print(f"Template labels cleaned. backup: {backup.name}")
    if found:
        print("Remaining stale labels:")
        for item in found[:80]:
            print(f"  {item}")
        raise SystemExit(1)


if __name__ == "__main__":
    main()
