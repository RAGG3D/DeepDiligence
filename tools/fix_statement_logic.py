#!/usr/bin/env python3
"""Patch ticker-specific statement logic without rewriting the workbook.

This keeps the model auditable after the template has been populated:
  * RIS R&D/G&A totals are calculated from their visible FY DATA note rows.
  * Non-FY DATA placeholder opex rows are cleared and hidden.
  * RBS accumulated deficit rolls forward from the explicit Net Income row.
  * RBS detail rows that do not exist in the company's FY DATA are cleared and
    hidden while totals continue to use only tagged rows.

The script edits OOXML directly and does not save with openpyxl.
"""

from __future__ import annotations

import argparse
import html
import re
import shutil
import time
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl


NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
RIS_FORECAST_START = "K"
RIS_FORECAST_END = "Y"
RBS_FORECAST_START = "K"
RBS_FORECAST_END = "W"
SCHEDULES_FORECAST_START = "K"
SCHEDULES_FORECAST_END = "X"
BASE_OPEX_GROWTH = 0.03


def _default_path(ticker: str) -> Path:
    return Path(f"/mnt/c/Users/yzsun/Desktop/DD/{ticker}/DCF {ticker}.xlsx")


def _sheet_zip_path(parts: dict[str, bytes], sheet_name: str) -> str:
    wb = ET.fromstring(parts["xl/workbook.xml"])
    rels = ET.fromstring(parts["xl/_rels/workbook.xml.rels"])
    relmap = {rel.get("Id"): rel.get("Target") for rel in rels}
    for sheet in wb.findall(f".//{{{NS_MAIN}}}sheet"):
        if html.unescape(sheet.get("name", "")) == sheet_name:
            target = relmap[sheet.get(f"{{{NS_R}}}id")]
            target = target.lstrip("/")
            return target if target.startswith("xl/") else "xl/" + target
    raise RuntimeError(f"Sheet not found: {sheet_name}")


def _read_parts(path: Path) -> tuple[dict[str, bytes], list[str]]:
    with zipfile.ZipFile(path, "r") as zf:
        bad = zf.testzip()
        if bad:
            raise RuntimeError(f"zip CRC failure in {bad}")
        return {name: zf.read(name) for name in zf.namelist()}, zf.namelist()


def _write_parts(path: Path, parts: dict[str, bytes], order: list[str]) -> None:
    tmp = path.with_suffix(".~statement_logic.xlsx")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        written: set[str] = set()
        for name in order:
            if name == "xl/calcChain.xml":
                continue
            if name in parts:
                zout.writestr(name, parts[name])
                written.add(name)
        for name, data in parts.items():
            if name not in written and name != "xl/calcChain.xml":
                zout.writestr(name, data)
    tmp.replace(path)


def _cell_col(addr: str) -> int:
    letters = re.match(r"([A-Z]+)", addr).group(1)
    out = 0
    for ch in letters:
        out = out * 26 + ord(ch) - 64
    return out


def _col_letter(index: int) -> str:
    out = ""
    while index:
        index, rem = divmod(index - 1, 26)
        out = chr(65 + rem) + out
    return out


def _cell_style(xml: str, addr: str, default: str = "0") -> str:
    m = re.search(
        rf'<c\b(?=[^>]*\br="{addr}")[^>]*/>|'
        rf'<c\b(?=[^>]*\br="{addr}")[^>]*>.*?</c>',
        xml,
        re.S,
    )
    if not m:
        return default
    sm = re.search(r'\bs="(\d+)"', m.group(0))
    return sm.group(1) if sm else default


def _replace_cell(xml: str, addr: str, cell_xml: str) -> tuple[str, bool]:
    m = re.search(
        rf'<c\b(?=[^>]*\br="{addr}")[^>]*/>|'
        rf'<c\b(?=[^>]*\br="{addr}")[^>]*>.*?</c>',
        xml,
        re.S,
    )
    if m:
        return xml[:m.start()] + cell_xml + xml[m.end():], True

    row_num = int(re.search(r"\d+", addr).group(0))
    row_m = re.search(rf'(<row\b(?=[^>]*\br="{row_num}")[^>]*>)(.*?)(</row>)', xml, re.S)
    if not row_m:
        return xml, False
    body = row_m.group(2)
    insert_at = len(body)
    target_col = _cell_col(addr)
    for cm in re.finditer(r'<c\b[^>]*\br="([A-Z]+\d+)"', body):
        if _cell_col(cm.group(1)) > target_col:
            insert_at = cm.start()
            break
    new_body = body[:insert_at] + cell_xml + body[insert_at:]
    return xml[:row_m.start(2)] + new_body + xml[row_m.end(2):], True


def _text_cell(addr: str, text: str, style: str) -> str:
    return f'<c r="{addr}" s="{style}" t="inlineStr"><is><t>{html.escape(text, quote=False)}</t></is></c>'


def _formula_cell(addr: str, formula: str, style: str) -> str:
    return f'<c r="{addr}" s="{style}"><f>{html.escape(formula, quote=False)}</f></c>'


def _blank_cell(addr: str, style: str) -> str:
    return f'<c r="{addr}" s="{style}"/>'


def _patch_text(xml: str, addr: str, text: str, default_style: str = "0") -> tuple[str, bool]:
    return _replace_cell(xml, addr, _text_cell(addr, text, _cell_style(xml, addr, default_style)))


def _patch_formula(xml: str, addr: str, formula: str, default_style: str = "0") -> tuple[str, bool]:
    return _replace_cell(xml, addr, _formula_cell(addr, formula, _cell_style(xml, addr, default_style)))


def _patch_blank(xml: str, addr: str, default_style: str = "0") -> tuple[str, bool]:
    return _replace_cell(xml, addr, _blank_cell(addr, _cell_style(xml, addr, default_style)))


def _set_row_hidden(xml: str, row: int, hidden: bool) -> str:
    m = re.search(rf'<row\b(?=[^>]*\br="{row}")[^>]*(?:/>|>.*?</row>)', xml, re.S)
    if not m:
        return xml
    row_xml = m.group(0)
    if hidden:
        if 'hidden="' in row_xml:
            row_xml = re.sub(r'\bhidden="[^"]*"', 'hidden="1"', row_xml, count=1)
        else:
            row_xml = row_xml.replace("<row ", '<row hidden="1" ', 1)
    else:
        row_xml = re.sub(r'\s+hidden="[^"]*"', "", row_xml, count=1)
    return xml[:m.start()] + row_xml + xml[m.end():]


def _clear_row_cells(xml: str, row: int, start_col: str = "B", end_col: str = "Y") -> tuple[str, int]:
    changed = 0
    for col_idx in range(_cell_col(start_col + "1"), _cell_col(end_col + "1") + 1):
        addr = f"{_col_letter(col_idx)}{row}"
        before = xml
        xml, ok = _patch_blank(xml, addr)
        changed += int(ok and before != xml)
    return xml, changed


def _duplicate_counter_formula(row: int) -> str:
    return f'IF(AND($C{row}<>"",COUNTIFS($D:$D,$D{row},$C:$C,$C{row})>1),COUNTIFS($D$1:D{row},$D{row},$C$1:C{row},$C{row}),"")'


def _fy_note_categories(path: Path, heading: str) -> list[str]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        ws = wb["FY DATA K USD"] if "FY DATA K USD" in wb.sheetnames else wb["FY DATA"]
        start = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 4).value == heading:
                start = r + 1
                break
        if start is None:
            return []
        out: list[str] = []
        for r in range(start, ws.max_row + 1):
            c = ws.cell(r, 3).value
            d = ws.cell(r, 4).value
            if d is None:
                continue
            text = str(d).strip()
            low = text.lower()
            if low.startswith("check - "):
                break
            if low.startswith("total "):
                break
            if c == "ISN" and text and text.lower() != "reserved":
                out.append(text)
        return out
    finally:
        wb.close()


def _fy_bs_items(path: Path) -> set[tuple[str, str]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        ws = wb["FY DATA K USD"] if "FY DATA K USD" in wb.sheetnames else wb["FY DATA"]
        out: set[tuple[str, str]] = set()
        for r in range(1, ws.max_row + 1):
            c = ws.cell(r, 3).value
            d = ws.cell(r, 4).value
            if c == "BS" and d:
                out.add((str(c).strip(), str(d).strip()))
        return out
    finally:
        wb.close()


def _stage_load(col: str) -> str:
    """Stage-weighted active clinical burden for RIS Pipeline Timeline rows."""
    # Rows 8:10 are individual assets. Row 11 is either a regular fourth asset
    # or an aggregate "Other Pipeline Active Stage Load" row written by
    # generate/adapt_ris.py when a company has >4 clinical assets.  The aggregate
    # row is already a sum of pre-approval stage loads and can exceed 5, so do
    # not apply the <5 filter to it when that label is present.
    return (
        f'SUMPRODUCT({col}$8:{col}$10,--({col}$8:{col}$10>0),--({col}$8:{col}$10<5))'
        f'+IF($D$11="Other Pipeline Active Stage Load",{col}$11,'
        f'IF(AND({col}$11>0,{col}$11<5),{col}$11,0))'
    )


def _opex_sensitivities(name: str, section: str) -> tuple[float, float, float]:
    """Return (stage-up, stage-down, revenue-increment coefficient)."""
    low = name.lower()
    if section == "rd":
        if any(x in low for x in ("research consumables", "external", "development", "project", "material")):
            return 0.12, 0.06, 0.0
        if "personnel" in low or "employee" in low:
            return 0.06, 0.03, 0.0
        if any(x in low for x in ("depreciation", "amortization", "facility")):
            return 0.04, 0.02, 0.0
        if any(x in low for x in ("intellectual property", "royalties", "license")):
            return 0.03, 0.015, 0.0
        return 0.05, 0.025, 0.0

    if "personnel" in low or "employee" in low:
        return 0.04, 0.02, 0.08
    if any(x in low for x in ("administrative", "professional", "other")):
        return 0.03, 0.015, 0.12
    if any(x in low for x in ("depreciation", "amortization", "facility")):
        return 0.02, 0.01, 0.02
    return 0.03, 0.015, 0.05


def _opex_forecast_formula(row: int, name: str, section: str, col: str, prev: str) -> str:
    up, down, revenue_coeff = _opex_sensitivities(name, section)
    delta = f"({_stage_load(col)}-{_stage_load(prev)})"
    stage_factor = f"(1+{up:.4g}*MAX(0,{delta})-{down:.4g}*MAX(0,-{delta}))"
    formula = f"{prev}{row}*(1+{BASE_OPEX_GROWTH:.4g})*{stage_factor}"
    if revenue_coeff:
        # Commercial organization builds with launch/revenue, but only on new
        # revenue added in the year so it does not double-count the prior base.
        formula += f"+{revenue_coeff:.4g}*MAX(0,{col}$30-{prev}$30)"
    return f"IFERROR(MAX(0,{formula}),0)"


def _patch_ris_opex(xml: str, rd_categories: list[str], ga_categories: list[str]) -> tuple[str, int]:
    changed = 0
    row_style = {
        "b": _cell_style(xml, "B47", "0"),
        "c": _cell_style(xml, "C47", "0"),
        "d": _cell_style(xml, "D47", "0"),
        "e": _cell_style(xml, "E47", "0"),
        "data": _cell_style(xml, "F47", "0"),
        "total_data": _cell_style(xml, "F46", "0"),
    }

    def write_category_row(row: int, name: str, section: str) -> None:
        nonlocal xml, changed
        xml = _set_row_hidden(xml, row, False)
        for addr, val, style in [
            (f"B{row}", _duplicate_counter_formula(row), row_style["b"]),
            (f"C{row}", "ISN", row_style["c"]),
            (f"D{row}", name, row_style["d"]),
            (f"E{row}", "[MM USD]", row_style["e"]),
        ]:
            before = xml
            if addr[0] == "B":
                xml, ok = _replace_cell(xml, addr, _formula_cell(addr, val, style))
            else:
                xml, ok = _replace_cell(xml, addr, _text_cell(addr, val, style))
            changed += int(ok and before != xml)
        for col_idx in range(_cell_col("F1"), _cell_col("J1") + 1):
            col = _col_letter(col_idx)
            formula = (
                f"SUMIFS('FY DATA'!{col}:{col},'FY DATA'!$D:$D,$D{row},"
                f"'FY DATA'!$C:$C,RIS!$C{row},'FY DATA'!$B:$B,RIS!$B{row})"
            )
            before = xml
            xml, ok = _replace_cell(xml, f"{col}{row}", _formula_cell(f"{col}{row}", formula, row_style["data"]))
            changed += int(ok and before != xml)
        for col_idx in range(_cell_col(RIS_FORECAST_START + "1"), _cell_col(RIS_FORECAST_END + "1") + 1):
            col = _col_letter(col_idx)
            prev = _col_letter(col_idx - 1)
            before = xml
            formula = _opex_forecast_formula(row, name, section, col, prev)
            xml, ok = _replace_cell(xml, f"{col}{row}", _formula_cell(f"{col}{row}", formula, row_style["data"]))
            changed += int(ok and before != xml)

    def clear_block(rows: range) -> None:
        nonlocal xml, changed
        for row in rows:
            before = xml
            xml, n = _clear_row_cells(xml, row, "B", "Y")
            changed += n
            xml = _set_row_hidden(xml, row, True)
            changed += int(xml != before)

    def write_total(row: int, start: int, end: int) -> None:
        nonlocal xml, changed
        xml = _set_row_hidden(xml, row, False)
        for col_idx in range(_cell_col("F1"), _cell_col("Y1") + 1):
            col = _col_letter(col_idx)
            before = xml
            xml, ok = _replace_cell(
                xml,
                f"{col}{row}",
                _formula_cell(f"{col}{row}", f'SUMIF($C${start}:$C${end},"ISN",{col}${start}:{col}${end})', row_style["total_data"]),
            )
            changed += int(ok and before != xml)

    clear_block(range(47, 61))
    for row, name in zip(range(47, 61), rd_categories):
        write_category_row(row, name, "rd")
    write_total(46, 47, 60)

    clear_block(range(62, 71))
    for row, name in zip(range(62, 71), ga_categories):
        write_category_row(row, name, "ga")
    write_total(61, 62, 70)

    for col_idx in range(_cell_col("F1"), _cell_col("Y1") + 1):
        col = _col_letter(col_idx)
        before = xml
        xml, ok = _replace_cell(xml, f"{col}71", _formula_cell(f"{col}71", f"SUM({col}46,{col}61)", row_style["total_data"]))
        changed += int(ok and before != xml)

    return xml, changed


def _patch_schedules(xml: str) -> tuple[str, int]:
    changed = 0

    def patch(addr: str, formula: str) -> None:
        nonlocal xml, changed
        before = xml
        xml, ok = _patch_formula(xml, addr, formula)
        changed += int(ok and before != xml)

    for col_idx in range(_cell_col(SCHEDULES_FORECAST_START + "1"), _cell_col(SCHEDULES_FORECAST_END + "1") + 1):
        col = _col_letter(col_idx)
        prev = _col_letter(col_idx - 1)
        if col == SCHEDULES_FORECAST_START:
            patch(f"{col}21", "IFERROR(AVERAGE(G21:J21),0)")
            patch(f"{col}23", "IFERROR(AVERAGE(G23:J23),0)")
            patch(f"{col}25", "IFERROR(AVERAGE(G25:J25),0)")
            patch(f"{col}27", "IFERROR(AVERAGE(G27:J27),0)")
            patch(f"{col}35", "IFERROR(AVERAGE(G35:J35),0)")
        else:
            for row in (21, 23, 25, 27, 35):
                patch(f"{col}{row}", f"IFERROR({prev}{row},0)")

        patch(f"{col}20", f"IFERROR(MAX({prev}20,{col}$15*{col}21),0)")
        patch(f"{col}22", f"IFERROR({col}$20*{col}23,0)")
        patch(f"{col}24", f"IFERROR({col}$20*{col}25,0)")
        patch(f"{col}26", f"IFERROR({col}$15*{col}27,0)")
        patch(f"{col}28", f'SUMIF($C$20:$C$26,"BSN",{col}20:{col}26)')
        patch(f"{col}29", f"IFERROR(MAX(0,{prev}29*(1+0.02)),0)")
        patch(f"{col}30", f"IFERROR(SUM({col}28:{col}29),0)")
        patch(f"{col}34", f"IFERROR({col}35*AVERAGE({col}30,{prev}30),0)")
        patch(f"{col}37", f"IFERROR(SUM({col}34,{col}36),0)")
        patch(f"{col}39", f"IFERROR(SUM({col}37),0)")
        patch(f"{col}42", f"IFERROR({col}34,0)")
        patch(f"{col}43", f"IFERROR(MAX(0,{col}30-{prev}30),0)")
        patch(f"{col}44", f"IFERROR(SUM({col}42:{col}43),0)")

    return xml, changed


def _read_rbs_detail_rows(path: Path) -> list[tuple[int, str, str]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        ws = wb["RBS"]
        rows = []
        for r in list(range(11, 28)) + list(range(41, 45)):
            c = ws.cell(r, 3).value
            d = ws.cell(r, 4).value
            if c == "BS" and d:
                rows.append((r, str(c).strip(), str(d).strip()))
        return rows
    finally:
        wb.close()


def _cash_plug_formula(col: str, cash_row: int,
                       financial_asset_rows: list[int]) -> str:
    """Return the accounting-equation cash plug net of non-cash FA rows."""
    other_refs = [f"{col}{row}" for row in financial_asset_rows if row != cash_row]
    formula = f"{col}66-{col}63"
    if other_refs:
        formula += f"-SUM({','.join(other_refs)})"
    return formula


def _read_rbs_financial_asset_layout(path: Path) -> tuple[int, list[int]]:
    """Discover the cash row and all financial-asset detail rows by labels."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        ws = wb["RBS"]
        cash_row = 33
        start = None
        end = None
        for row in range(1, ws.max_row + 1):
            label = str(ws.cell(row, 4).value or "").strip()
            if label == "Financial Assets":
                start = row
            elif label == "Total Financial Assets" and start is not None and end is None:
                end = row
            elif label == "Cash And Cash Equivalents":
                cash_row = row
        if start is not None and end is not None:
            rows = [
                row for row in range(start + 1, end)
                if str(ws.cell(row, 3).value or "").strip() == "BS"
                and str(ws.cell(row, 4).value or "").strip()
            ]
        else:
            rows = [cash_row]
        if cash_row not in rows:
            rows.append(cash_row)
        return cash_row, rows
    finally:
        wb.close()


def _patch_rbs(xml: str, existing_bs: set[tuple[str, str]],
               rbs_rows: list[tuple[int, str, str]], cash_row: int,
               financial_asset_rows: list[int]) -> tuple[str, int]:
    changed = 0
    for row, c, d in rbs_rows:
        if (c, d) in existing_bs:
            xml = _set_row_hidden(xml, row, False)
            continue
        before = xml
        xml, n = _clear_row_cells(xml, row, "B", "Y")
        changed += n
        xml = _set_row_hidden(xml, row, True)
        changed += int(xml != before)

    def avg_ratio_to_ris(row: int, ris_row: int) -> str:
        terms = [f"IFERROR(${c}${row}/RIS!${c}${ris_row},0)" for c in "FGHIJ"]
        return "AVERAGE(" + ",".join(terms) + ")"

    def avg_ratio_same_sheet(num_row: int, denom_row: int) -> str:
        terms = [f"IFERROR(${c}${num_row}/${c}${denom_row},0)" for c in "FGHIJ"]
        return "AVERAGE(" + ",".join(terms) + ")"

    for col_idx in range(_cell_col(RBS_FORECAST_START + "1"), _cell_col(RBS_FORECAST_END + "1") + 1):
        col = _col_letter(col_idx)
        prev = _col_letter(col_idx - 1)
        patches = {
            f"{col}11": f"Schedules!{col}30",
            f"{col}12": f"IFERROR(MAX(0,{prev}12*0.5),0)",
            f"{col}13": f"IFERROR(MAX(0,RIS!{col}71*{avg_ratio_to_ris(13, 71)}),0)",
            f"{col}14": f"IFERROR(MAX(0,RIS!{col}46*{avg_ratio_to_ris(14, 46)}),0)",
            f"{col}15": f"IFERROR(Schedules!{col}26,0)",
            f"{col}16": f"IFERROR(MAX(0,{prev}16*(1+0.03)),0)",
            f"{col}20": f"IFERROR(MAX(0,RIS!{col}71*{avg_ratio_to_ris(20, 71)}),0)",
            f"{col}21": f"IFERROR(MAX(0,RIS!{col}71*{avg_ratio_to_ris(21, 71)}),0)",
            f"{col}22": f"IFERROR(MAX(0,{col}15*{avg_ratio_same_sheet(22, 15)}),0)",
            f"{col}23": f"IFERROR(MAX(0,{prev}23*0.5),0)",
            f"{col}24": f"IFERROR(MAX(0,{prev}24*0.8),0)",
            f"{col}25": f"IFERROR(MAX(0,{col}15*{avg_ratio_same_sheet(25, 15)}),0)",
            f"{col}26": f"IFERROR(MAX(0,{prev}26*0.75),0)",
            f"{col}27": f"IFERROR(MAX(0,{prev}27*(1+0.03)),0)",
            f"{col}41": f"IFERROR({prev}41,0)",
        }
        # Preserve non-cash financial assets at their historical average, then
        # make the dynamically located cash row the residual accounting plug.
        # Explicitly rewriting these rows also repairs a workbook previously
        # touched by the old hard-coded row-33 cash logic.
        for row in financial_asset_rows:
            if row != cash_row:
                patches[f"{col}{row}"] = (
                    f"IFERROR(MAX(0,AVERAGE($F{row}:$J{row})),0)"
                )
        patches[f"{col}{cash_row}"] = _cash_plug_formula(
            col, cash_row, financial_asset_rows
        )
        for addr, formula in patches.items():
            before = xml
            xml, ok = _patch_formula(xml, addr, formula)
            changed += int(ok and before != xml)

    # Net income is row 6. Accumulated deficit should visibly roll from it,
    # instead of bypassing the RBS operating section with a direct RIS row link.
    # The four equity lines (rows 41-44) are emitted in the issuer's filing
    # order, so the Accumulated Deficit row is NOT fixed: MOLN/CMPX carry it on
    # row 44, but BHVN/KYMR carry it on row 43 with AOCI on 44.  Locate it
    # dynamically from the RBS detail scan and roll net income into THAT row
    # only, leaving the true AOCI/APIC rows on their generator formulas.
    deficit_row = 44
    for r, _c, d in rbs_rows:
        if 41 <= r <= 44 and d.strip().lower() == "accumulated deficit":
            deficit_row = r
            break
    for col_idx in range(_cell_col(RBS_FORECAST_START + "1"), _cell_col(RBS_FORECAST_END + "1") + 1):
        col = _col_letter(col_idx)
        prev = _col_letter(col_idx - 1)
        before = xml
        xml, ok = _patch_formula(xml, f"{col}{deficit_row}", f"{prev}{deficit_row}+{col}6")
        changed += int(ok and before != xml)

    # Keep totals conditional on active source tags.
    for col_idx in range(_cell_col("F1"), _cell_col("Y1") + 1):
        col = _col_letter(col_idx)
        patches = {
            # Keep the visible operating totals and their summary helper rows
            # on one definition.  The old R28 range started at row 20 and
            # omitted Accounts Payable (R19), while R16/R27 retained unrelated
            # template roll-forward formulas; that made R47 and R67 disagree.
            f"{col}16": (
                f'SUMIF($C$11:$C$15,"SCHE",{col}$11:{col}$15)'
                f'+SUMIF($C$11:$C$15,"BS",{col}$11:{col}$15)'
            ),
            f"{col}17": f"{col}16",
            f"{col}27": f'SUMIF($C$19:$C$26,"BS",{col}$19:{col}$26)',
            f"{col}28": f"{col}27",
            f"{col}45": f"SUM({col}41:{col}44)",
            f"{col}61": f"{col}17",
            f"{col}62": f"{col}28",
            f"{col}66": f"{col}45",
        }
        for addr, formula in patches.items():
            before = xml
            xml, ok = _patch_formula(xml, addr, formula)
            changed += int(ok and before != xml)
    return xml, changed


def _mark_full_calc(parts: dict[str, bytes]) -> None:
    wb_xml = parts["xl/workbook.xml"].decode("utf-8", "ignore")
    if "fullCalcOnLoad" not in wb_xml:
        wb_xml = wb_xml.replace("<calcPr", '<calcPr fullCalcOnLoad="1"', 1)
    parts["xl/workbook.xml"] = wb_xml.encode("utf-8")
    ct = parts.get("[Content_Types].xml", b"").decode("utf-8", "ignore")
    if ct:
        ct = re.sub(r'<Override[^>]*/xl/calcChain\.xml[^>]*/>', "", ct)
        parts["[Content_Types].xml"] = ct.encode("utf-8")
    rels = parts.get("xl/_rels/workbook.xml.rels", b"").decode("utf-8", "ignore")
    if rels:
        rels = re.sub(r'<Relationship[^>]*calcChain[^>]*/>', "", rels)
        parts["xl/_rels/workbook.xml.rels"] = rels.encode("utf-8")
    parts.pop("xl/calcChain.xml", None)


def fix_statement_logic(path: Path, make_backup: bool = True) -> dict[str, int]:
    rd_categories = _fy_note_categories(path, "Research And Development Expenses")
    ga_categories = _fy_note_categories(path, "General And Administrative Expenses")
    existing_bs = _fy_bs_items(path)
    rbs_rows = _read_rbs_detail_rows(path)
    cash_row, financial_asset_rows = _read_rbs_financial_asset_layout(path)

    parts, order = _read_parts(path)
    report: dict[str, int] = {}

    ris_path = _sheet_zip_path(parts, "RIS")
    ris_xml = parts[ris_path].decode("utf-8", "ignore")
    ris_xml, report["RIS"] = _patch_ris_opex(ris_xml, rd_categories, ga_categories)
    parts[ris_path] = ris_xml.encode("utf-8")

    schedules_path = _sheet_zip_path(parts, "Schedules")
    schedules_xml = parts[schedules_path].decode("utf-8", "ignore")
    schedules_xml, report["Schedules"] = _patch_schedules(schedules_xml)
    parts[schedules_path] = schedules_xml.encode("utf-8")

    rbs_path = _sheet_zip_path(parts, "RBS")
    rbs_xml = parts[rbs_path].decode("utf-8", "ignore")
    rbs_xml, report["RBS"] = _patch_rbs(
        rbs_xml, existing_bs, rbs_rows, cash_row, financial_asset_rows
    )
    parts[rbs_path] = rbs_xml.encode("utf-8")

    _mark_full_calc(parts)
    if make_backup:
        backup = path.with_name(f"{path.stem}_pre_statement_logic_{time.strftime('%Y%m%d_%H%M%S')}.xlsx")
        shutil.copy2(path, backup)
        print(f"Backup: {backup}")
    _write_parts(path, parts, order)
    print(
        "Statement logic fixed: "
        f"R&D rows={len(rd_categories)}, G&A rows={len(ga_categories)}, "
        f"patches={report}"
    )
    return report


def main() -> None:
    parser = argparse.ArgumentParser(description="Patch RIS/RBS ticker-specific statement logic")
    parser.add_argument("--ticker")
    parser.add_argument("--path")
    parser.add_argument("--no-backup", action="store_true")
    args = parser.parse_args()
    path = Path(args.path) if args.path else _default_path(args.ticker)
    if not path.exists():
        raise FileNotFoundError(path)
    fix_statement_logic(path, make_backup=not args.no_backup)


if __name__ == "__main__":
    main()
