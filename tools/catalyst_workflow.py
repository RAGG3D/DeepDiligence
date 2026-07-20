#!/usr/bin/env python3
"""Catalyst run/post-catalyst lifecycle with durable snapshots.

Commands:
  run   research artifact -> event metadata + grey unrelated target columns
  refresh rebuild conviction-filtered combinations after analyst input changes
  post  snapshot workbook into durable DuckDB, add price reaction/interpretation,
        then and only then clear analyst inputs and restore original colours
  clean restore a neutral Catalyst framework before a new-ticker delivery

The workbook is edited at OOXML level; Excel data tables/charts are preserved.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
import shutil
import subprocess
import sys
import zipfile
from copy import deepcopy
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.formula import DataTableFormula

REPO = Path(__file__).resolve().parents[1]
if str(REPO) not in sys.path:
    sys.path.insert(0, str(REPO))

from datastore.catalyst_store import latest_open_run, save_snapshot, start_run

try:
    import yfinance as yf
except Exception:  # pragma: no cover
    yf = None

NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
RNS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
ET.register_namespace("", NS)
ET.register_namespace("r", RNS)
# Excel places these prefixes in mc:Ignorable attribute values. ElementTree
# must retain their canonical names when serializing; auto-renaming them to
# ns1/ns2 leaves the Ignorable QName list unresolved and desktop Excel refuses
# to open the otherwise well-formed package.
ET.register_namespace("mc", "http://schemas.openxmlformats.org/markup-compatibility/2006")
ET.register_namespace("x14ac", "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac")
ET.register_namespace("x16r2", "http://schemas.microsoft.com/office/spreadsheetml/2015/02/main")
ET.register_namespace("xr", "http://schemas.microsoft.com/office/spreadsheetml/2014/revision")
ET.register_namespace("xr2", "http://schemas.microsoft.com/office/spreadsheetml/2015/revision2")
ET.register_namespace("xr3", "http://schemas.microsoft.com/office/spreadsheetml/2016/revision3")
MCNS = "http://schemas.openxmlformats.org/markup-compatibility/2006"


def _excel_xml(root: ET.Element) -> bytes:
    """Serialize without unresolved prefixes in mc:Ignorable.

    ElementTree drops namespace declarations used only inside the whitespace-
    separated QName value. The workbook parts we edit actively use x14ac/xr,
    while x16r2/xr2/xr3 are declaration-only. Keeping the latter tokens after
    their declarations disappear makes desktop Excel reject the package.
    """
    ignorable = f"{{{MCNS}}}Ignorable"
    if ignorable in root.attrib:
        root.set(ignorable, "x14ac xr")
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _default_workbook(ticker: str) -> Path:
    return Path(f"/mnt/c/Users/yzsun/Desktop/DD/{ticker}/DCF {ticker}.xlsx")


def _artifact_dir(ticker: str) -> Path:
    path = REPO / "artifacts" / ticker.upper()
    path.mkdir(parents=True, exist_ok=True)
    return path


def _sheet_path(parts: Dict[str, bytes], name: str) -> str:
    wb = ET.fromstring(parts["xl/workbook.xml"])
    rels = ET.fromstring(parts["xl/_rels/workbook.xml.rels"])
    relmap = {r.get("Id"): r.get("Target") for r in rels}
    for sheet in wb.findall(f".//{{{NS}}}sheet"):
        if sheet.get("name") == name:
            target = relmap[sheet.get(f"{{{RNS}}}id")].lstrip("/")
            return target if target.startswith("xl/") else "xl/" + target
    raise RuntimeError(f"worksheet not found: {name}")


def _read_parts(path: Path) -> Tuple[Dict[str, bytes], List[str]]:
    with zipfile.ZipFile(path) as zf:
        return {n: zf.read(n) for n in zf.namelist()}, zf.namelist()


def _write_parts(path: Path, parts: Dict[str, bytes], order: List[str]) -> None:
    tmp = path.with_suffix(".~catalyst.xlsx")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as out:
        done = set()
        for name in order:
            if name in parts:
                out.writestr(name, parts[name]); done.add(name)
        for name, payload in parts.items():
            if name not in done:
                out.writestr(name, payload)
    tmp.replace(path)


def _cells(root: ET.Element) -> Dict[str, ET.Element]:
    return {c.get("r"): c for c in root.findall(f".//{{{NS}}}c") if c.get("r")}


def _column_number(addr: str) -> int:
    value = 0
    for ch in re.match(r"[A-Z]+", addr).group(0):
        value = value * 26 + ord(ch) - 64
    return value


def _ensure_cell(root: ET.Element, addr: str, style: str = "0") -> ET.Element:
    existing = _cells(root).get(addr)
    if existing is not None:
        return existing
    row_num = int(re.search(r"\d+", addr).group(0))
    sheet_data = root.find(f"{{{NS}}}sheetData")
    if sheet_data is None:
        raise RuntimeError("Catalyst sheetData missing")
    row = next((r for r in sheet_data.findall(f"{{{NS}}}row") if int(r.get("r", "0")) == row_num), None)
    if row is None:
        row = ET.Element(f"{{{NS}}}row", {"r": str(row_num)})
        inserted = False
        for idx, candidate in enumerate(list(sheet_data)):
            if candidate.tag == f"{{{NS}}}row" and int(candidate.get("r", "0")) > row_num:
                sheet_data.insert(idx, row); inserted = True; break
        if not inserted:
            sheet_data.append(row)
    cell = ET.Element(f"{{{NS}}}c", {"r": addr, "s": style})
    target_col = _column_number(addr)
    inserted = False
    for idx, candidate in enumerate(list(row)):
        if candidate.tag == f"{{{NS}}}c" and _column_number(candidate.get("r")) > target_col:
            row.insert(idx, cell); inserted = True; break
    if not inserted:
        row.append(cell)
    return cell


def _set_text(cell: ET.Element, value: str) -> None:
    for child in list(cell):
        cell.remove(child)
    cell.set("t", "inlineStr")
    inline = ET.SubElement(cell, f"{{{NS}}}is")
    text = ET.SubElement(inline, f"{{{NS}}}t")
    text.text = value


def _clear(cell: ET.Element) -> None:
    for child in list(cell):
        cell.remove(child)
    cell.attrib.pop("t", None)


def _gray_style_map(styles_root: ET.Element, source_ids: Iterable[int]) -> Dict[int, int]:
    """Return XF mappings for a visible grey-background/grey-data mask.

    The mask is deliberately visual only: formulas and values remain present so
    every drug x indication still contributes to the final market price.
    """
    fonts = styles_root.find(f"{{{NS}}}fonts")
    fills = styles_root.find(f"{{{NS}}}fills")
    xfs = styles_root.find(f"{{{NS}}}cellXfs")
    if fonts is None or fills is None or xfs is None:
        raise RuntimeError("styles.xml missing fonts/fills/cellXfs")

    gray_fill = ET.Element(f"{{{NS}}}fill")
    pattern = ET.SubElement(gray_fill, f"{{{NS}}}patternFill", {"patternType": "solid"})
    ET.SubElement(pattern, f"{{{NS}}}fgColor", {"rgb": "FFE7E6E6"})
    ET.SubElement(pattern, f"{{{NS}}}bgColor", {"indexed": "64"})
    fill_lookup = {ET.tostring(item): i for i, item in enumerate(list(fills))}
    fill_key = ET.tostring(gray_fill)
    fill_id = fill_lookup.get(fill_key)
    if fill_id is None:
        fills.append(gray_fill)
        fill_id = len(list(fills)) - 1
    fills.set("count", str(len(list(fills))))

    font_lookup = {ET.tostring(item): i for i, item in enumerate(list(fonts))}
    gray_font_by_source: Dict[int, int] = {}
    source = list(xfs)
    mapping: Dict[int, int] = {}
    for sid in sorted(set(source_ids)):
        if sid >= len(source):
            continue
        source_font_id = int(source[sid].get("fontId", "0"))
        if source_font_id >= len(fonts):
            raise RuntimeError(f"source font {source_font_id} missing for style {sid}")
        gray_font_id = gray_font_by_source.get(source_font_id)
        if gray_font_id is None:
            font = deepcopy(list(fonts)[source_font_id])
            color = font.find(f"{{{NS}}}color")
            if color is None:
                color = ET.Element(f"{{{NS}}}color")
                children = list(font)
                insert_at = next(
                    (i for i, child in enumerate(children)
                     if child.tag in {f"{{{NS}}}sz", f"{{{NS}}}u",
                                      f"{{{NS}}}vertAlign", f"{{{NS}}}scheme"}),
                    len(children),
                )
                font.insert(insert_at, color)
            color.attrib.clear()
            color.set("rgb", "FF7F7F7F")
            font_key = ET.tostring(font)
            gray_font_id = font_lookup.get(font_key)
            if gray_font_id is None:
                fonts.append(font)
                gray_font_id = len(list(fonts)) - 1
                font_lookup[font_key] = gray_font_id
            gray_font_by_source[source_font_id] = gray_font_id
        clone = ET.fromstring(ET.tostring(source[sid], encoding="utf-8"))
        clone.set("fillId", str(fill_id)); clone.set("applyFill", "1")
        clone.set("fontId", str(gray_font_id)); clone.set("applyFont", "1")
        xfs.append(clone)
        mapping[sid] = len(list(xfs)) - 1
    fonts.set("count", str(len(list(fonts))))
    xfs.set("count", str(len(list(xfs))))
    return mapping


def _addresses(cell_range: str) -> Iterable[str]:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            letters = ""
            n = col
            while n:
                n, rem = divmod(n - 1, 26)
                letters = chr(65 + rem) + letters
            yield f"{letters}{row}"


def _target_match(name: str, relevant: Iterable[str]) -> bool:
    norm = re.sub(r"[^a-z0-9]", "", name.lower())
    for item in relevant:
        other = re.sub(r"[^a-z0-9]", "", str(item).lower())
        if other and (other == norm or other in norm or norm in other):
            return True
    return False


def _patch_run(path: Path, manifest: Dict[str, Any], research: Dict[str, Any]) -> Dict[str, str]:
    parts, order = _read_parts(path)
    sheet_path = _sheet_path(parts, "Catalyst")
    sheet = ET.fromstring(parts[sheet_path])
    styles = ET.fromstring(parts["xl/styles.xml"])
    cells = _cells(sheet)
    relevant = research.get("relevant_targets") or []
    if not relevant:
        raise ValueError("research JSON must identify at least one relevant_targets entry")

    original: Dict[str, str] = {}
    gray_addrs: List[str] = []
    for target in manifest.get("targets", []):
        if _target_match(target["name"], relevant):
            continue
        ranges = target.get("display_ranges") or ([target.get("main_range")] if target.get("main_range") else [])
        for cell_range in ranges:
            for addr in _addresses(cell_range):
                if addr in original:
                    continue
                cell = cells.get(addr)
                if cell is None:
                    # Excel may omit a formatted-but-empty cell from sheet XML.
                    # Materialize it so the visible grey mask is continuous
                    # across the complete four-column target block.
                    cell = _ensure_cell(sheet, addr)
                    cells[addr] = cell
                original[addr] = cell.get("s", "0")
                gray_addrs.append(addr)
    mapping = _gray_style_map(styles, [int(original[a]) for a in gray_addrs]) if gray_addrs else {}
    for addr in gray_addrs:
        cells[addr].set("s", str(mapping[int(original[addr])]))

    metadata = manifest.get("event_metadata", {"name": "C2", "disclosure": "C3", "source": "C4"})
    values = {
        metadata["name"]: research.get("event_name") or "",
        metadata["disclosure"]: research.get("expected_disclosure") or "",
        metadata["source"]: " | ".join(research.get("sources") or []),
    }
    for addr, value in values.items():
        cell = cells.get(addr)
        if cell is None:
            cell = _ensure_cell(sheet, addr)
        _set_text(cell, str(value))
    parts[sheet_path] = _excel_xml(sheet)
    parts["xl/styles.xml"] = _excel_xml(styles)
    _write_parts(path, parts, order)
    return original


def _restore_original_fills(parts: Dict[str, bytes], sheet: ET.Element,
                            original_styles: Dict[str, str], backup: Path) -> None:
    """Restore pre-run fill and font colours without unstable Excel style IDs.

    Excel may renumber/deduplicate XFs whenever the analyst saves the workbook.
    The pre-run backup remains authoritative for the desired appearance. For
    each grey-masked cell, clone its *current* normalized XF and replace fillId
    and fontId with the matching pre-run values, preserving number formats,
    borders and alignment.
    """
    backup_parts, _ = _read_parts(backup)
    old_styles = ET.fromstring(backup_parts["xl/styles.xml"])
    new_styles = ET.fromstring(parts["xl/styles.xml"])
    old_fonts = old_styles.find(f"{{{NS}}}fonts")
    old_fills = old_styles.find(f"{{{NS}}}fills")
    old_xfs = old_styles.find(f"{{{NS}}}cellXfs")
    new_fonts = new_styles.find(f"{{{NS}}}fonts")
    new_fills = new_styles.find(f"{{{NS}}}fills")
    new_xfs = new_styles.find(f"{{{NS}}}cellXfs")
    if any(x is None for x in (
        old_fonts, old_fills, old_xfs, new_fonts, new_fills, new_xfs
    )):
        raise RuntimeError("styles.xml missing fonts/fills/cellXfs during Catalyst reset")

    font_lookup = {ET.tostring(font): i for i, font in enumerate(list(new_fonts))}
    fill_lookup = {ET.tostring(fill): i for i, fill in enumerate(list(new_fills))}
    desired_appearance: Dict[str, Tuple[int, int]] = {}
    for addr, old_sid_text in original_styles.items():
        old_sid = int(old_sid_text)
        if old_sid >= len(old_xfs):
            raise RuntimeError(f"pre-run style {old_sid} missing for {addr}")
        font_id = int(old_xfs[old_sid].get("fontId", "0"))
        fill_id = int(old_xfs[old_sid].get("fillId", "0"))
        if font_id >= len(old_fonts):
            raise RuntimeError(f"pre-run font {font_id} missing for {addr}")
        if fill_id >= len(old_fills):
            raise RuntimeError(f"pre-run fill {fill_id} missing for {addr}")
        source_font = old_fonts[font_id]
        source_fill = old_fills[fill_id]
        font_key = ET.tostring(source_font)
        mapped_font = font_lookup.get(font_key)
        if mapped_font is None:
            new_fonts.append(deepcopy(source_font))
            mapped_font = len(list(new_fonts)) - 1
            font_lookup[font_key] = mapped_font
        fill_key = ET.tostring(source_fill)
        mapped_fill = fill_lookup.get(fill_key)
        if mapped_fill is None:
            new_fills.append(deepcopy(source_fill))
            mapped_fill = len(list(new_fills)) - 1
            fill_lookup[fill_key] = mapped_fill
        desired_appearance[addr] = (mapped_font, mapped_fill)
    new_fonts.set("count", str(len(list(new_fonts))))
    new_fills.set("count", str(len(list(new_fills))))

    cells = _cells(sheet)
    xf_map: Dict[Tuple[int, int, int], int] = {}
    for addr, (font_id, fill_id) in desired_appearance.items():
        cell = cells.get(addr)
        if cell is None:
            continue
        current_sid = int(cell.get("s", "0"))
        key = (current_sid, font_id, fill_id)
        mapped_sid = xf_map.get(key)
        if mapped_sid is None:
            if current_sid >= len(new_xfs):
                raise RuntimeError(f"current style {current_sid} missing for {addr}")
            xf = deepcopy(new_xfs[current_sid])
            xf.set("fillId", str(fill_id)); xf.set("applyFill", "1")
            xf.set("fontId", str(font_id)); xf.set("applyFont", "1")
            new_xfs.append(xf)
            mapped_sid = len(list(new_xfs)) - 1
            xf_map[key] = mapped_sid
        cell.set("s", str(mapped_sid))
    new_xfs.set("count", str(len(list(new_xfs))))
    parts["xl/styles.xml"] = _excel_xml(new_styles)


def _patch_reset(path: Path, manifest: Dict[str, Any], original_styles: Dict[str, str],
                 backup: Optional[Path] = None) -> None:
    parts, order = _read_parts(path)
    sheet_path = _sheet_path(parts, "Catalyst")
    sheet = ET.fromstring(parts[sheet_path])
    cells = _cells(sheet)
    if original_styles and backup and backup.exists():
        _restore_original_fills(parts, sheet, original_styles, backup)
    else:
        for addr, style in original_styles.items():
            if addr in cells:
                cells[addr].set("s", style)
    for addr in manifest.get("manual_cells", []):
        if addr in cells:
            _clear(cells[addr])
    for addr, value in manifest.get("neutral_defaults", {}).items():
        if addr not in cells:
            continue
        _clear(cells[addr])
        cells[addr].set("t", "n")
        v = ET.SubElement(cells[addr], f"{{{NS}}}v")
        v.text = str(value)
    for addr in manifest.get("event_metadata", {"name": "C2", "disclosure": "C3", "source": "C4"}).values():
        cell = cells.get(addr)
        if cell is None:
            cell = _ensure_cell(sheet, addr)
        _set_text(cell, "")
    parts[sheet_path] = _excel_xml(sheet)
    _write_parts(path, parts, order)


def _snapshot_cells(path: Path) -> Dict[str, Any]:
    # Normal mode is required for merged ranges and for styled empty cells;
    # cached values can stay read-only to keep the second workbook lightweight.
    formulas = load_workbook(path, read_only=False, data_only=False)
    values = load_workbook(path, read_only=True, data_only=True)
    try:
        wf = formulas["Catalyst"]; wv = values["Catalyst"]
        cells = []
        for row in wf.iter_rows():
            for cell in row:
                value = getattr(cell, "value", None)
                if value is None and not getattr(cell, "has_style", False):
                    continue
                coordinate = getattr(cell, "coordinate", None)
                if not coordinate:
                    continue
                formula_or_input: Any = value
                if isinstance(formula_or_input, DataTableFormula):
                    formula_or_input = {
                        "type": "dataTable",
                        **{
                            key: getattr(formula_or_input, key, None)
                            for key in ("ref", "ca", "dt2D", "dtr", "r1", "r2", "del1", "del2")
                            if getattr(formula_or_input, key, None) is not None
                        },
                    }
                cells.append({
                    "address": coordinate, "formula_or_input": formula_or_input,
                    "cached_value": wv[coordinate].value,
                    "style_id": getattr(cell, "style_id", 0),
                    "number_format": getattr(cell, "number_format", "General"),
                })
        return {"max_row": wf.max_row, "max_column": wf.max_column, "merged_ranges": [str(x) for x in wf.merged_cells], "cells": cells}
    finally:
        formulas.close(); values.close()


def _price_reaction(ticker: str, event_date: str) -> Dict[str, Any]:
    if not event_date or yf is None:
        return {"status": "unavailable", "reason": "event date or yfinance unavailable"}
    d = date.fromisoformat(event_date)
    data = yf.download(ticker, start=(d - timedelta(days=10)).isoformat(),
                       end=(d + timedelta(days=12)).isoformat(), auto_adjust=False,
                       progress=False, threads=False)
    if data is None or data.empty:
        return {"status": "unavailable", "reason": "no price history"}
    close = data["Close"]
    if hasattr(close, "columns"):
        close = close.iloc[:, 0]
    rows = [(idx.date(), float(v)) for idx, v in close.dropna().items()]
    pre = [x for x in rows if x[0] < d]
    post = [x for x in rows if x[0] >= d]
    if not pre or not post:
        return {"status": "unavailable", "reason": "insufficient pre/post sessions"}
    base = pre[-1]
    result: Dict[str, Any] = {"status": "ok", "pre_session": base[0].isoformat(), "pre_close": base[1], "sessions": []}
    for session, px in post[:3]:
        result["sessions"].append({"date": session.isoformat(), "close": px, "return_vs_pre": px / base[1] - 1})
    return result


def _load_json(path: Path) -> Dict[str, Any]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        raise ValueError(f"expected JSON object: {path}")
    return data


def _manifest(ticker: str, explicit: Optional[str]) -> Tuple[Path, Dict[str, Any]]:
    path = Path(explicit) if explicit else _artifact_dir(ticker) / f"{ticker.upper()}_catalyst_manifest.json"
    if not path.exists():
        raise FileNotFoundError(f"Catalyst manifest missing: {path}; framework generation must create it")
    return path, _load_json(path)


def command_run(args) -> None:
    ticker = args.ticker.upper(); workbook = Path(args.path) if args.path else _default_workbook(ticker)
    research = _load_json(Path(args.research))
    if latest_open_run(ticker):
        raise SystemExit(f"{ticker} already has an open catalyst run; close it with post first")
    sources = research.get("sources") or []
    if not sources or not all(str(x).startswith(("http://", "https://")) for x in sources):
        raise SystemExit("catalyst run requires cited official/conference source URLs")
    if research.get("relevant_targets"):
        subprocess.run([
            sys.executable,
            str(REPO / "generate" / "build_catalyst_framework.py"),
            "--ticker", ticker,
            "--path", str(workbook),
            "--research", str(Path(args.research)),
            "--conviction-threshold", str(args.conviction_threshold),
            "--no-backup",
        ], check=True)
    _, manifest = _manifest(ticker, args.manifest)
    backup = workbook.with_name(f"{workbook.stem}_pre_catalyst_run_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
    shutil.copy2(workbook, backup)
    original = _patch_run(workbook, manifest, research)
    subprocess.run([
        sys.executable,
        str(REPO / "tools" / "normalize_calc_state.py"),
        "--path", str(workbook),
        "--no-backup",
    ], check=True)
    run_id = start_run(ticker, research, str(workbook))
    state = {
        "run_id": run_id, "workbook": str(workbook),
        "original_styles": original, "research": research, "backup": str(backup),
    }
    state_path = _artifact_dir(ticker) / f"{ticker}_catalyst_active_state.json"
    state_path.write_text(json.dumps(state, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Catalyst run opened {run_id}; unrelated targets greyed; state → {state_path}")


def command_post(args) -> None:
    ticker = args.ticker.upper()
    state_path = _artifact_dir(ticker) / f"{ticker}_catalyst_active_state.json"
    if not state_path.exists():
        raise SystemExit("no active Catalyst state; refusing to clear analyst inputs")
    state = _load_json(state_path)
    workbook = Path(args.path) if args.path else Path(state.get("workbook") or _default_workbook(ticker))
    _, manifest = _manifest(ticker, args.manifest)
    interpretation = _load_json(Path(args.interpretation))
    if not interpretation.get("summary") or not interpretation.get("sources"):
        raise SystemExit("post-catalyst interpretation must contain summary and cited sources")
    blob = workbook.read_bytes(); digest = hashlib.sha256(blob).hexdigest()
    parts, _ = _read_parts(workbook); sheet_xml = parts[_sheet_path(parts, "Catalyst")]
    research = state.get("research", {})
    reaction = _price_reaction(ticker, args.event_date or research.get("event_date", ""))
    snapshot_id = save_snapshot(
        state["run_id"], ticker, digest, blob, sheet_xml, _snapshot_cells(workbook),
        reaction, interpretation,
    )
    # The database transaction above must commit before any user input is cleared.
    backup = workbook.with_name(f"{workbook.stem}_post_catalyst_snapshot_{snapshot_id[:8]}.xlsx")
    shutil.copy2(workbook, backup)
    _patch_reset(workbook, manifest, state.get("original_styles", {}), Path(state.get("backup", "")))
    state_path.unlink()
    print(f"Snapshot {snapshot_id} committed; Catalyst reset to neutral framework → {workbook}")


def command_refresh(args) -> None:
    """Rebuild an open run after the analyst changes Table-3 Conv./MS/LOA inputs."""
    ticker = args.ticker.upper()
    state_path = _artifact_dir(ticker) / f"{ticker}_catalyst_active_state.json"
    if not state_path.exists():
        raise SystemExit("no active Catalyst state; start the run before refreshing combinations")
    state = _load_json(state_path)
    workbook = Path(args.path) if args.path else Path(
        state.get("workbook") or _default_workbook(ticker)
    )
    subprocess.run([
        sys.executable,
        str(REPO / "generate" / "build_catalyst_framework.py"),
        "--ticker", ticker,
        "--path", str(workbook),
        "--conviction-threshold", str(args.conviction_threshold),
        "--no-backup",
    ], check=True)
    print(f"Catalyst active combinations refreshed from current Table-3 inputs → {workbook}")


def command_clean(args) -> None:
    ticker = args.ticker.upper(); workbook = Path(args.path) if args.path else _default_workbook(ticker)
    _, manifest = _manifest(ticker, args.manifest)
    state_path = _artifact_dir(ticker) / f"{ticker}_catalyst_active_state.json"
    original: Dict[str, str] = {}
    if state_path.exists():
        state = _load_json(state_path)
        original = state.get("original_styles", {})
        if not args.force:
            raise SystemExit("active catalyst run exists; use post or --force only for a new-model clean")
    backup = Path(state.get("backup", "")) if state_path.exists() else None
    _patch_reset(workbook, manifest, original, backup)
    if args.force and state_path.exists():
        state_path.unlink()
    print(f"Catalyst neutral-input clean complete → {workbook}")


def main() -> None:
    ap = argparse.ArgumentParser()
    sub = ap.add_subparsers(dest="command", required=True)
    for name in ("run", "refresh", "post", "clean"):
        p = sub.add_parser(name)
        p.add_argument("--ticker", required=True); p.add_argument("--path"); p.add_argument("--manifest")
    sub.choices["run"].add_argument("--research", required=True)
    sub.choices["run"].add_argument("--conviction-threshold", type=float, default=0.10)
    sub.choices["refresh"].add_argument("--conviction-threshold", type=float, default=0.10)
    sub.choices["post"].add_argument("--interpretation", required=True)
    sub.choices["post"].add_argument("--event-date")
    sub.choices["clean"].add_argument("--force", action="store_true")
    args = ap.parse_args()
    {
        "run": command_run,
        "refresh": command_refresh,
        "post": command_post,
        "clean": command_clean,
    }[args.command](args)


if __name__ == "__main__":
    main()
