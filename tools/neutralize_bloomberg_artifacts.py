#!/usr/bin/env python3
"""Replace Bloomberg add-in artifacts with static public-data values.

The DCF template contains Bloomberg Desktop API formulas (`_xll.bdp`,
`_xll.bdh`, `_xll.bql`, `_xll.bds`) across WELCOME!, Earnings, BBG DAPI,
Historical Events, Peer View(s), and Catalyst helper areas.  Machines without
Bloomberg show those as #NAME?/add-in errors.  This tool keeps the workbook as
an xlsx package and patches the XML directly:

* writes current public market data into WELCOME! and BBG DAPI;
* freezes any remaining Bloomberg add-in formulas to cached values or blanks;
* removes workbook comments/VML comment drawings;
* removes calcChain and forces recalculation on open.
"""

from __future__ import annotations

import argparse
import html
import json
import math
import re
import shutil
import sys
import time
import warnings
import zipfile
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import openpyxl
import yfinance as yf
from lxml import etree
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from core.excel_writer import _normalize_formula_xml, _strip_formula_cache


ERROR_VALUES = {
    "#NULL!",
    "#DIV/0!",
    "#VALUE!",
    "#REF!",
    "#NAME?",
    "#NUM!",
    "#N/A",
    "#N/A N/A",
    "#GETTING_DATA",
}

COMMENT_REL_TYPES = (
    "/comments",
    "/threadedComments",
    "/vmlDrawing",
)


def _default_path(ticker: str) -> Path:
    return Path(f"/mnt/c/Users/yzsun/Desktop/DD/{ticker}/DCF {ticker}.xlsx")


def _sheet_zip_paths(parts: dict[str, bytes]) -> dict[str, str]:
    wb_xml = parts["xl/workbook.xml"].decode("utf-8", "ignore")
    rels_xml = parts["xl/_rels/workbook.xml.rels"].decode("utf-8", "ignore")

    rid_to_target: dict[str, str] = {}
    for rel in re.finditer(r"<Relationship\b[^>]+>", rels_xml):
        tag = rel.group(0)
        rid_m = re.search(r'Id="([^"]+)"', tag)
        target_m = re.search(r'Target="([^"]+)"', tag)
        type_m = re.search(r'Type="([^"]+)"', tag)
        if not rid_m or not target_m or not type_m:
            continue
        if "worksheet" not in type_m.group(1):
            continue
        target = target_m.group(1).lstrip("/")
        if not target.startswith("xl/"):
            target = "xl/" + target
        rid_to_target[rid_m.group(1)] = target

    out: dict[str, str] = {}
    for sheet in re.finditer(r"<sheet\b[^>]+>", wb_xml):
        tag = sheet.group(0)
        name_m = re.search(r'name="([^"]+)"', tag)
        rid_m = re.search(r'r:id="([^"]+)"', tag)
        if name_m and rid_m and rid_m.group(1) in rid_to_target:
            out[html.unescape(name_m.group(1))] = rid_to_target[rid_m.group(1)]
    return out


def _read_parts(path: Path) -> dict[str, bytes]:
    with zipfile.ZipFile(path, "r") as zf:
        bad = zf.testzip()
        if bad:
            raise RuntimeError(f"zip CRC failure in {bad}")
        return {name: zf.read(name) for name in zf.namelist()}


def _write_parts(path: Path, parts: dict[str, bytes], original_order: list[str]) -> None:
    tmp = path.with_suffix(".~bbgclean.xlsx")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        written = set()
        for name in original_order:
            if name not in parts:
                continue
            zout.writestr(name, parts[name])
            written.add(name)
        for name, data in parts.items():
            if name not in written:
                zout.writestr(name, data)
    tmp.replace(path)


def _cell_addr(addr: str) -> tuple[int, int]:
    col = "".join(ch for ch in addr if ch.isalpha())
    row = int("".join(ch for ch in addr if ch.isdigit()))
    col_idx = 0
    for ch in col:
        col_idx = col_idx * 26 + ord(ch.upper()) - 64
    return row, col_idx


def _col_idx(addr: str) -> int:
    return _cell_addr(addr)[1]


def _is_bad_cached_value(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() in ERROR_VALUES or value.startswith("#")
    if isinstance(value, float):
        return math.isnan(value) or math.isinf(value)
    return False


def _format_number(value: float | int) -> str:
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return f"{value:.12g}" if isinstance(value, float) else str(value)


def _xml_text(value: str) -> str:
    return html.escape(value, quote=False)


def _value_cell(addr: str, value: Any, style_attrs: str = "") -> str:
    attrs = f' r="{addr}"{style_attrs}'
    if value is None or value == "":
        return f"<c{attrs}/>"
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return f"<c{attrs}><v>{_format_number(value)}</v></c>"
    if isinstance(value, (datetime, date)):
        value = value.strftime("%Y-%m-%d")
    return f'<c{attrs} t="inlineStr"><is><t>{_xml_text(str(value))}</t></is></c>'


def _style_attrs(open_tag: str) -> str:
    style = re.search(r'\s+s="[^"]+"', open_tag)
    return style.group(0) if style else ""


def _cell_span(xml: str, addr: str) -> tuple[int, int, str] | None:
    """Return (start, end, open_tag) for a cell without crossing cell bounds."""
    search = f'r="{addr}"'
    start = 0
    while True:
        pos = xml.find(search, start)
        if pos == -1:
            return None
        lt = xml.rfind("<", 0, pos)
        if lt == -1:
            start = pos + 1
            continue
        if xml[lt + 1] != "c" or xml[lt + 2] not in (" ", "\t", "\n", "/", ">"):
            start = pos + 1
            continue
        tag_end = xml.index(">", lt) + 1
        if xml[tag_end - 2 : tag_end] == "/>":
            return lt, tag_end, xml[lt : tag_end - 2]
        cell_end = xml.index("</c>", tag_end) + 4
        return lt, cell_end, xml[lt : tag_end - 1]


def _replace_cell(xml: str, addr: str, value: Any) -> str:
    row_num, col_idx = _cell_addr(addr)
    span = _cell_span(xml, addr)
    if span:
        start, end, open_tag = span
        new_cell = _value_cell(addr, value, _style_attrs(open_tag))
        return xml[:start] + new_cell + xml[end:]

    row_start = rf'<row\b(?=[^>]*\br="{row_num}")[^>]*'
    row_pattern = re.compile(rf'({row_start})(?<!/)>(.*?)(</row>)', re.S)
    row_match = row_pattern.search(xml)
    if not row_match:
        self_closing = re.search(rf'({row_start})/>', xml)
        if not self_closing:
            return xml
        new_row = f"{self_closing.group(1)}>{_value_cell(addr, value)}</row>"
        return xml[: self_closing.start()] + new_row + xml[self_closing.end() :]

    row_body = row_match.group(2)
    cells = list(re.finditer(r'<c\b[^>]*\br="([A-Z]+\d+)"', row_body))
    insert_at = len(row_body)
    for c_match in cells:
        if _col_idx(c_match.group(1)) > col_idx:
            insert_at = c_match.start()
            break
    new_body = row_body[:insert_at] + _value_cell(addr, value) + row_body[insert_at:]
    return xml[: row_match.start(2)] + new_body + xml[row_match.end(2) :]


def _cached_values(path: Path) -> dict[str, dict[str, Any]]:
    out: dict[str, dict[str, Any]] = {}
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(path, read_only=False, data_only=True, keep_links=True)
    for ws in wb.worksheets:
        vals: dict[str, Any] = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    vals[cell.coordinate] = cell.value
        out[ws.title] = vals
    wb.close()
    return out


def _last_history_row(hist):
    if hist.empty:
        raise RuntimeError("No yfinance price history returned")
    return hist.iloc[-1]


def _load_company_facts(ticker: str) -> dict[str, Any]:
    """Load researched, primary-source company facts used by market-data QA.

    Price vendors key history to a security/ticker, not necessarily to a continuous
    operating company.  Reverse mergers and shell/predecessor histories therefore
    require an explicit identity start date before a "lifetime" price statistic is
    meaningful.  The research workflow writes this file after checking SEC/company
    sources; absence is allowed, but is reported loudly by :func:`_market_data`.
    """
    path = ROOT / "artifacts" / ticker.upper() / f"{ticker.upper()}_company_facts.json"
    if not path.exists():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        print(f"  ! invalid company-facts file {path}: {exc}")
        return {}
    return data if isinstance(data, dict) else {}


def _completed_daily_history(hist):
    """Exclude a still-open U.S. daily bar so all quote fields share one timestamp."""
    if hist.empty:
        return hist
    now_et = datetime.now(ZoneInfo("America/New_York"))
    # Yahoo's daily bar is live during the session.  Do not stamp a model with a
    # mixture of an intraday last price and a partial daily OHLC bar.
    if now_et.weekday() < 5 and now_et.time() < datetime.strptime("16:15", "%H:%M").time():
        idx = hist.index
        local_dates = idx.tz_convert("America/New_York").date if getattr(idx, "tz", None) else idx.date
        hist = hist[local_dates < now_et.date()]
    return hist


def _scope_company_history(hist, facts: dict[str, Any], ticker: str):
    """Apply the researched current-company identity boundary to vendor history."""
    start = facts.get("price_history_start_date") or facts.get("current_company_identity_start_date")
    if not start:
        print(f"  !! {ticker}: no researched price_history_start_date; lifetime statistics "
              "retain the vendor's full security history and require manual review")
        return hist, ""
    try:
        start_dt = datetime.fromisoformat(str(start)).replace(tzinfo=None)
    except ValueError:
        print(f"  !! {ticker}: invalid price_history_start_date={start!r}; using full history")
        return hist, ""
    idx = hist.index.tz_localize(None) if getattr(hist.index, "tz", None) else hist.index
    scoped = hist[idx >= start_dt]
    if scoped.empty:
        raise RuntimeError(f"{ticker}: no price history on/after researched identity date {start}")
    return scoped, str(start)


def _year_end_series(hist, shares, through_year: int) -> tuple[dict[int, float], dict[int, float]]:
    price_by_year: dict[int, float] = {}
    shares_by_year: dict[int, float] = {}

    hist_sorted = hist.sort_index()
    for year in range(2001, through_year + 1):
        cut = hist_sorted[hist_sorted.index.tz_localize(None) <= datetime(year, 12, 31, 23, 59, 59)]
        if not cut.empty and cut.index.min().year <= year:
            price_by_year[year] = round(float(cut.iloc[-1]["Close"]), 4)
        else:
            price_by_year[year] = 0.0

        if shares is not None and len(shares):
            s = shares.sort_index()
            s_cut = s[s.index.tz_localize(None) <= datetime(year, 12, 31, 23, 59, 59)]
            shares_by_year[year] = round(float(s_cut.iloc[-1]) / 1e6, 4) if len(s_cut) else 0.0
        else:
            shares_by_year[year] = 0.0

    return price_by_year, shares_by_year


# Intervals used by the BBG DAPI Interval Reference Table (col M) → trailing window.
# "100Y" is the template's stand-in for full/lifetime history.
_INTERVAL_DAYS = {"365D": 365, "1Y": 365, "5Y": 365 * 5}


def _px_interval_stats(hist) -> dict[str, dict[str, float]]:
    """Interval statistics of the Close price for the BBG DAPI Interval table.

    Returns {interval_label: {stat_key: value}} for the intervals the template's
    PX_LAST rows use (365D / 1Y / 5Y / YTD / 100Y).  All figures are THIS ticker's
    own price history — never a reference company's cached statistics.
    """
    close = hist["Close"].dropna()
    if close.empty:
        return {}
    idx = close.index
    try:
        idx = idx.tz_localize(None)
    except (TypeError, AttributeError):
        pass
    close = close.copy()
    close.index = idx
    close = close.sort_index()
    end = close.index.max()
    year_start = datetime(datetime.now().year, 1, 1)

    windows: dict[str, Any] = {
        interval: close[close.index >= end - timedelta(days=days)]
        for interval, days in _INTERVAL_DAYS.items()
    }
    windows["YTD"] = close[close.index >= year_start]
    windows["100Y"] = close  # full/lifetime history

    out: dict[str, dict[str, float]] = {}
    for label, series in windows.items():
        if series.empty:
            series = close  # fall back to full history rather than emit nothing
        out[label] = {
            "INTERVAL_HIGH": round(float(series.max()), 4),
            "INTERVAL_MEDIAN": round(float(series.median()), 4),
            "INTERVAL_AVG": round(float(series.mean()), 4),
            "INTERVAL_LOW": round(float(series.min()), 4),
            "P25": round(float(series.quantile(0.25)), 4),
            "P75": round(float(series.quantile(0.75)), 4),
        }
    return out


def _interval_table_values(bbg_cached: dict[str, Any],
                           px_stats: dict[str, dict[str, float]]) -> dict[str, Any]:
    """Refill the BBG DAPI Interval Reference Table's S (OVERRIDE) and T (FINAL).

    The template ships this table (K4:T113) with the reference company's price
    stats in S and #NAME?-cached _xll formulas in R, so once R is frozen blank the
    FINAL column T (=IFERROR(IF(LEFT(R,1)="#",S,R),S)) evaluates to 0 for every
    row — zeroing the VALUATION football field, which reads it via SUMIFS on $T:$T.
    Recompute the active ticker's PX_LAST interval stats and write BOTH S and T
    (writing S alone is insufficient: the frozen R is blank, so T falls back to R,
    not S).  Rows with no free public source (target price / multiples) get 0 —
    never the reference company's stale value.  Row identity is read from the
    sheet's own K (stat) / L (field) / M (interval) / Q (percentile) labels.
    """
    out: dict[str, Any] = {}
    for r in range(4, 114):
        field = bbg_cached.get(f"L{r}")
        stat = bbg_cached.get(f"K{r}")
        interval = bbg_cached.get(f"M{r}")
        if not (field and stat and interval):
            continue
        value: Any = 0
        if str(field) == "PX_LAST":
            wstats = px_stats.get(str(interval), {})
            if str(stat) == "INTERVAL_PERCENTILE":
                try:
                    key = "P25" if float(bbg_cached.get(f"Q{r}")) < 0.5 else "P75"
                except (TypeError, ValueError):
                    key = None
                value = wstats.get(key, 0) if key else 0
            else:
                value = wstats.get(str(stat), 0)
        out[f"S{r}"] = value
        out[f"T{r}"] = value
    return out


def _market_data(ticker: str) -> dict[str, Any]:
    yf_ticker = yf.Ticker(ticker)
    info = yf_ticker.info
    facts = _load_company_facts(ticker)
    hist = yf_ticker.history(period="max", auto_adjust=False, actions=True)
    hist = _completed_daily_history(hist)
    hist, identity_start = _scope_company_history(hist, facts, ticker)
    last = _last_history_row(hist)

    # One completed daily bar is the atomic quote snapshot.  Do not mix an
    # intraday info.currentPrice with a different date's OHLC values.
    price = float(last["Close"])
    open_px = float(last["Open"])
    high_px = float(last["High"])
    low_px = float(last["Low"])
    price_as_of = last.name.tz_localize(None).date().isoformat() \
        if getattr(last.name, "tzinfo", None) else last.name.date().isoformat()
    share_fact = facts.get("shares_outstanding")
    if isinstance(share_fact, dict):
        share_fact = share_fact.get("value")
    shares = share_fact or info.get("sharesOutstanding") or info.get("impliedSharesOutstanding")
    shares_m = (float(shares) / 1e6) if shares else 0.0
    # Keep market cap mathematically consistent with the quote and share count
    # shown in the same workbook snapshot.
    mktcap_m = price * shares_m

    try:
        shares_full = yf_ticker.get_shares_full(start="2000-01-01")
    except Exception:
        shares_full = None

    price_by_year, shares_by_year = _year_end_series(hist, shares_full, datetime.now().year)
    if shares_m:
        shares_by_year[datetime.now().year] = round(shares_m, 4)

    try:
        tnx_hist = yf.Ticker("^TNX").history(period="10d")
        tnx_hist = _completed_daily_history(tnx_hist)
        tnx = tnx_hist["Close"].iloc[-1]
        risk_free = round(float(tnx), 4)
    except Exception:
        risk_free = 4.5

    hist_high = float(hist["High"].max())
    hist_low = float(hist["Low"].min())
    px_interval_stats = _px_interval_stats(hist)

    country = facts.get("country") or info.get("country") or ""
    fin_ccy = facts.get("reporting_currency") or info.get("financialCurrency") or info.get("currency") or "USD"
    listing = identity_start
    ts = info.get("firstTradeDateEpochUtc") or info.get("firstTradeDateMilliseconds")
    if ts and not listing:
        try:
            listing = datetime.utcfromtimestamp(float(ts) / (1000 if ts > 1e11 else 1)).strftime("%Y-%m-%d")
        except Exception:
            listing = ""

    # Beta / headcount are often missing for small/foreign biotechs. Fall back to
    # ticker-neutral defaults (never a reference company's figure) and warn.
    beta = info.get("beta")
    if not beta:
        print(f"  ! beta unavailable for {ticker}; using generic market beta 1.0")
        beta = 1.0            # market beta default (not a specific company's value)
    employees = info.get("fullTimeEmployees")
    if not employees:
        print(f"  ! employee count unavailable for {ticker}; leaving blank")
        employees = 0

    return {
        "ticker_bbg": f"{ticker} US Equity",
        # ── company identity (derived per-ticker; never hardcode a reference co.) ──
        "company": facts.get("legal_name") or info.get("longName") or info.get("shortName") or ticker,
        "country": country,
        "exchange": facts.get("exchange") or info.get("fullExchangeName") or info.get("exchange") or "",
        "security_type": {"ADR": "American Depositary Receipt"}.get(
            info.get("quoteType", ""), "Common Stock"),
        "fin_currency": fin_ccy,
        "listing_date": listing,
        # US filers report US GAAP; foreign private issuers (20-F) report IFRS.
        "accounting_std": facts.get("accounting_standard") or
                          ("US GAAP" if country in ("United States", "USA", "US", "") else "IFRS"),
        "price_as_of": price_as_of,
        "price": round(float(price), 4),
        "open": round(float(open_px), 4),
        "high": round(float(high_px), 4),
        "low": round(float(low_px), 4),
        "shares_m": round(shares_m, 4),
        "mktcap_m": round(float(mktcap_m), 4),
        "beta": beta,
        "target_median": info.get("targetMedianPrice") or info.get("targetMeanPrice") or 0,
        "target_mean": info.get("targetMeanPrice") or info.get("targetMedianPrice") or 0,
        "target_high": info.get("targetHighPrice") or 0,
        "target_low": info.get("targetLowPrice") or 0,
        "lifetime_high": round(hist_high, 4),
        "lifetime_low": round(hist_low, 4),
        "risk_free": risk_free,
        "employees": employees,
        "currency": info.get("currency") or "USD",
        "price_by_year": price_by_year,
        "shares_by_year": shares_by_year,
        "px_interval_stats": px_interval_stats,
    }


def _earnings_from_fydata(path: Path, fin_ccy: str) -> dict[str, Any]:
    """Build the Earnings P&L from THIS ticker's own FY DATA (never a reference co).

    Earnings is a display-only snapshot (no other tab references it). Rows map to
    FY DATA K USD line items; values shown in millions of the reporting currency.
    Net income uses the operating-income approximation (Revenue − R&D − G&A) since
    the IS total is a formula cell; EPS = net income / weighted shares.
    """
    try:
        wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
        ws = wb["FY DATA K USD"]
    except Exception:
        return {}
    base = next((int(ws.cell(4, c).value) for c in range(6, 12)
                 if isinstance(ws.cell(4, c).value, int) and 2000 <= ws.cell(4, c).value <= 2100), None)
    if base is None:
        wb.close(); return {}
    year_col = {base + (c - 6): c for c in range(6, 12)}

    def find(tag, *keys):
        for r in range(1, ws.max_row + 1):
            if str(ws.cell(r, 3).value) == tag:
                d = str(ws.cell(r, 4).value or "")
                if any(k.lower() in d.lower() for k in keys):
                    return r
        return None

    r_rev = find("IS", "Revenue", "Collaboration Revenues")
    r_rd = find("IS", "Research And Development")
    r_ga = find("IS", "General And Administrative")
    r_sh = find("IS", "Weighted-Average Number Of Common Shares", "Shares")
    r_cash = find("BS", "Cash And Cash Equivalents")

    def val(r, y):
        if not r or y not in year_col:
            return 0.0
        v = ws.cell(r, year_col[y]).value
        return float(v) if isinstance(v, (int, float)) else 0.0

    cur = datetime.now().year
    years = [y for y in sorted(year_col) if y <= cur
             and (val(r_rev, y) or val(r_rd, y) or val(r_ga, y))][-4:]
    out: dict[str, Any] = {}
    if years:
        out = {"B7": f"In Millions of {fin_ccy}", "B6": f"{years[-1]} FY",
               "G7": f"{years[-1]} FY reported", "H7": f"{years[-1] + 1} latest public"}
        for col, y in zip(["C", "D", "E", "F"][-len(years):], years):
            rev, rd, ga = val(r_rev, y), val(r_rd, y), val(r_ga, y)
            sh, cash = val(r_sh, y), val(r_cash, y)
            opex = rd + ga
            ni = rev - opex
            out[f"{col}7"] = f"{y} FY"
            out[f"{col}9"] = round((ni * 1000 / sh), 2) if sh else 0.0   # EPS
            out[f"{col}10"] = round(rev / 1000, 1)   # Revenue (MM)
            out[f"{col}11"] = 0                       # CapEx (not derived from FY DATA)
            out[f"{col}12"] = round(opex / 1000, 1)  # Total OpEx (MM)
            out[f"{col}13"] = round(rd / 1000, 1)    # R&D (MM)
            out[f"{col}14"] = round(ga / 1000, 1)    # SG&A (MM)
            out[f"{col}15"] = round(ni / 1000, 1)    # Net income (MM, operating approx.)
            out[f"{col}16"] = round(cash / 1000, 1)  # Cash (MM)
    wb.close()
    return out


def _explicit_values(ticker: str, md: dict[str, Any],
                     earnings: dict[str, Any] | None = None) -> dict[str, dict[str, Any]]:
    target_low = float(md["target_low"] or 0)
    target_high = float(md["target_high"] or 0)
    target_range = f"{target_low:.2f} to {target_high:.2f}" if target_low and target_high else ""
    target_mean = float(md["target_mean"] or md["target_median"] or 0)
    next_fye = f"31/12/{datetime.now().year}"

    values: dict[str, dict[str, Any]] = {
        "WELCOME!": {
            "B18": md["ticker_bbg"],
            "C19": md["company"],
            "E19": md["company"],
            "F19": md["company"],
            "E20": "Date",
            "C21": ticker,
            "F21": ticker,
            "C22": md["country"],
            "F22": md["country"],
            "C25": md["price"],
            "E25": "Last Price",
            "F25": md["price"],
            "C26": round(target_mean, 4),
            "F26": round(target_mean, 4),
            "C27": target_range,
            "F27": target_range,
            "C28": md["mktcap_m"],
            "F28": md["mktcap_m"],
            "C29": md["currency"],          # trading currency
            "F29": md["currency"],
            "C30": md["exchange"],
            "E30": "Primary Exchange Name",
            "F30": md["exchange"],
            "C31": md["security_type"],
            "E31": "Security Type",
            "F31": md["security_type"],
            "C32": md["listing_date"],
            "E32": "Listing Date",
            "F32": md["listing_date"],
            "C33": md["employees"],
            "E33": "Number of Employees",
            "F33": md["employees"],
            "C34": 0,
            "F34": 0,
            "C35": md["fin_currency"],      # reporting currency
            "F35": md["fin_currency"],
            "C36": "31 December",           # FYE not exposed by yfinance; common default
            "E36": "Fiscal Year End",
            "F36": "31 December",
            "C37": md["accounting_std"],
            "E37": "Accounting Standard",
            "F37": md["accounting_std"],
        },
        "BBG DAPI": {
            "C5": md["ticker_bbg"],
            "E5": md["ticker_bbg"],
            "E6": md["ticker_bbg"],
            "E7": md["ticker_bbg"],
            "E8": md["ticker_bbg"],
            "E9": "USGG10YR Index",
            "G4": next_fye,
            "H4": next_fye,
            "I4": next_fye,
            "G5": md["price"],
            "H5": md["price"],
            "I5": md["price"],
            "G6": md["open"],
            "H6": md["open"],
            "I6": md["open"],
            "G7": md["high"],
            "H7": md["high"],
            "I7": md["high"],
            "G8": md["low"],
            "H8": md["low"],
            "I8": md["low"],
            "G9": md["risk_free"],
            "H9": md["risk_free"],
            "I9": md["risk_free"],
            "G10": md["beta"],
            "H10": md["beta"],
            "I10": md["beta"],
            "H11": 5.5,      # generic equity-risk-premium default (not a reference co.'s)
            "I11": 5.5,
            "G12": md["shares_m"],
            "H12": md["shares_m"],
            "I12": md["shares_m"],
            "G13": round(float(md["target_median"] or target_mean), 4),
            "H13": round(float(md["target_median"] or target_mean), 4),
            "I13": round(float(md["target_median"] or target_mean), 4),
            "G14": round(target_high, 4),
            "H14": round(target_high, 4),
            "I14": round(target_high, 4),
            "G15": round(target_low, 4),
            "H15": round(target_low, 4),
            "I15": round(target_low, 4),
            "G16": md["lifetime_high"],
            "H16": md["lifetime_high"],
            "I16": md["lifetime_high"],
            "G17": md["lifetime_low"],
            "H17": md["lifetime_low"],
            "I17": md["lifetime_low"],
            "H18": 0,         # implied CDS spread — most pre-commercial biotechs carry no debt
            "I18": 0,
            "G19": md["mktcap_m"],
            "H19": md["mktcap_m"],
            "I19": md["mktcap_m"],
        },
        # Earnings P&L is derived from THIS ticker's own FY DATA (see _earnings_from_fydata),
        # never a reference company's income statement.
        "Earnings": {"B4": md["ticker_bbg"], **(earnings or {})},
    }

    # Blank the reference company's stale analyst-recommendation block (V4:AG11)
    # and seed the AH FINAL target-price column with THIS ticker's own consensus,
    # so VALUATION G14/J14 =PERCENTILE.INC('BBG DAPI'!$AH:$AH,0.25/0.75) computes
    # off the active ticker instead of the template company's analyst price targets.
    # (In the template AH holds the only numeric cells the percentile sees: rows
    # 4-11; AH12+ evaluate to blank/text.)
    bbg = values["BBG DAPI"]
    for row in range(4, 12):
        for col_idx in range(22, 34):        # columns V(22) .. AG(33)
            bbg[f"{get_column_letter(col_idx)}{row}"] = ""
    bbg["AH4"] = md["target_low"]
    bbg["AH5"] = md["target_median"]
    bbg["AH6"] = md["target_high"]
    for row in range(7, 12):
        bbg[f"AH{row}"] = ""

    for year in range(2001, datetime.now().year + 1):
        row = year - 1997
        values["BBG DAPI"][f"AN{row}"] = year
        values["BBG DAPI"][f"AO{row}"] = md["price_by_year"].get(year, 0.0)
        values["BBG DAPI"][f"AU{row}"] = year
        values["BBG DAPI"][f"AV{row}"] = md["shares_by_year"].get(year, 0.0)

    return values


def _contains_bbg_formula(cell_xml: str) -> bool:
    f_match = re.search(r"<f\b[^>]*>(.*?)</f>", cell_xml, re.S | re.I)
    if not f_match:
        return False
    formula = html.unescape(f_match.group(1)).lower()
    return "_xll." in formula or "_xll" in formula or "bql(" in formula


def _freeze_bbg_formulas(
    sheet_name: str,
    xml: str,
    cached: dict[str, dict[str, Any]],
    explicit: dict[str, Any],
) -> tuple[str, int]:
    count = 0
    sheet_cached = cached.get(sheet_name, {})

    def repl(match: re.Match[str]) -> str:
        nonlocal count
        cell_xml = match.group(0)
        if not _contains_bbg_formula(cell_xml):
            return cell_xml
        addr_m = re.search(r'\br="([A-Z]+\d+)"', cell_xml)
        if not addr_m:
            return cell_xml
        addr = addr_m.group(1)
        open_tag = re.match(r"(<c\b[^>]*)(?:>|/>)", cell_xml, re.S).group(1)
        value = explicit.get(addr)
        if value is None:
            cached_value = sheet_cached.get(addr)
            value = "" if _is_bad_cached_value(cached_value) else cached_value
        count += 1
        return _value_cell(addr, value, _style_attrs(open_tag))

    xml = re.sub(r"<c\b[^>]*/>|<c\b[^>]*>.*?</c>", repl, xml, flags=re.S)
    return xml, count


def _apply_explicit_values(xml: str, values: dict[str, Any]) -> str:
    for addr, value in sorted(values.items(), key=lambda kv: (_cell_addr(kv[0])[0], _cell_addr(kv[0])[1])):
        xml = _replace_cell(xml, addr, value)
    return xml


def _clear_static_error_literals(xml: str, sheet_cached: dict[str, Any]) -> tuple[str, int]:
    """Blank literal error markers while preserving formulas for hardening."""
    cleared = 0
    for addr, value in sheet_cached.items():
        if not _is_bad_cached_value(value):
            continue
        span = _cell_span(xml, addr)
        if not span:
            continue
        cell_xml = xml[span[0] : span[1]]
        if "<f" in cell_xml:
            continue
        xml = _replace_cell(xml, addr, "")
        cleared += 1
    return xml, cleared


def _remove_comment_parts(parts: dict[str, bytes]) -> tuple[dict[str, bytes], int]:
    removed = 0
    new_parts: dict[str, bytes] = {}
    for name, data in parts.items():
        lower = name.lower()
        if (
            lower.startswith("xl/comments")
            or lower.startswith("xl/threadedcomments/")
            or lower.startswith("xl/drawings/vmldrawing")
            or lower.startswith("xl/persons/")
        ):
            removed += 1
            continue
        new_parts[name] = data

    for name, data in list(new_parts.items()):
        if name.startswith("xl/worksheets/") and name.endswith(".xml"):
            xml = data.decode("utf-8", "ignore")
            xml = re.sub(r"<legacyDrawing\b[^>]*/>", "", xml)
            xml = re.sub(r"<legacyDrawingHF\b[^>]*/>", "", xml)
            new_parts[name] = xml.encode("utf-8")

        if name.startswith("xl/worksheets/_rels/") and name.endswith(".rels"):
            xml = data.decode("utf-8", "ignore")
            before = xml
            xml = re.sub(
                r"<Relationship\b(?=[^>]*Type=\"[^\"]*(?:comments|threadedComments|vmlDrawing)\")[^>]*/>",
                "",
                xml,
            )
            if xml != before:
                removed += 1
            new_parts[name] = xml.encode("utf-8")

    if "[Content_Types].xml" in new_parts:
        ct = new_parts["[Content_Types].xml"].decode("utf-8", "ignore")
        ct = re.sub(r'<Override\b[^>]*(?:comments|threadedComments|persons)[^>]*/>', "", ct)
        new_parts["[Content_Types].xml"] = ct.encode("utf-8")

    return new_parts, removed


def _remove_calc_chain_and_force_recalc(parts: dict[str, bytes]) -> dict[str, bytes]:
    parts.pop("xl/calcChain.xml", None)
    if "[Content_Types].xml" in parts:
        ct = parts["[Content_Types].xml"].decode("utf-8", "ignore")
        ct = re.sub(r'<Override[^>]*/xl/calcChain\.xml[^>]*/>', "", ct)
        parts["[Content_Types].xml"] = ct.encode("utf-8")
    if "xl/_rels/workbook.xml.rels" in parts:
        rels = parts["xl/_rels/workbook.xml.rels"].decode("utf-8", "ignore")
        rels = re.sub(r"<Relationship[^>]*calcChain[^>]*/>", "", rels)
        parts["xl/_rels/workbook.xml.rels"] = rels.encode("utf-8")
    if "xl/workbook.xml" in parts:
        wb = parts["xl/workbook.xml"].decode("utf-8", "ignore")
        if "<calcPr" in wb:
            wb = re.sub(r'\s+fullCalcOnLoad="[^"]*"', "", wb, count=1)
            wb = re.sub(r"<calcPr\b", '<calcPr fullCalcOnLoad="1"', wb, count=1)
        else:
            wb = wb.replace("</workbook>", '<calcPr fullCalcOnLoad="1"/></workbook>')
        parts["xl/workbook.xml"] = wb.encode("utf-8")
    return parts


def _validate_xml_parts(path: Path) -> None:
    with zipfile.ZipFile(path, "r") as zf:
        bad = zf.testzip()
        if bad:
            raise RuntimeError(f"post-write zip CRC failure in {bad}")
        for name in zf.namelist():
            if name.endswith(".xml") or name.endswith(".rels"):
                etree.fromstring(zf.read(name))
    with warnings.catch_warnings():
        warnings.simplefilter("error")
        wb = openpyxl.load_workbook(path, read_only=False, data_only=False, keep_links=True)
        wb.close()


def neutralize(path: Path, ticker: str, backup: bool = True) -> dict[str, int]:
    if not path.exists():
        raise FileNotFoundError(path)
    if backup:
        dst = path.with_name(f"{path.stem}_pre_bbgclean_{time.strftime('%Y%m%d_%H%M%S')}.xlsx")
        shutil.copy2(path, dst)
        print(f"Backup: {dst}")

    original_parts = _read_parts(path)
    original_order = list(original_parts)
    cached = _cached_values(path)
    md = _market_data(ticker)
    earnings = _earnings_from_fydata(path, md["fin_currency"])
    explicit = _explicit_values(ticker, md, earnings)

    # Refill the BBG DAPI Interval Reference Table (K4:T113) for THIS ticker so the
    # FINAL column T (read by the VALUATION football field) is not zeroed and S no
    # longer carries the reference company's price statistics. Row identity comes
    # from the sheet's own cached K/L/M/Q labels.
    interval_vals = _interval_table_values(
        cached.get("BBG DAPI", {}), md.get("px_interval_stats", {}))
    if interval_vals:
        explicit.setdefault("BBG DAPI", {}).update(interval_vals)

    parts = dict(original_parts)
    sheet_paths = _sheet_zip_paths(parts)
    frozen = 0
    static_errors_cleared = 0
    patched_sheets = 0
    for sheet_name, zip_path in sheet_paths.items():
        if zip_path not in parts:
            continue
        xml = parts[zip_path].decode("utf-8", "ignore")
        sheet_explicit = explicit.get(sheet_name, {})
        xml, n = _freeze_bbg_formulas(sheet_name, xml, cached, sheet_explicit)
        if sheet_explicit:
            xml = _apply_explicit_values(xml, sheet_explicit)
        xml, cleared = _clear_static_error_literals(xml, cached.get(sheet_name, {}))
        static_errors_cleared += cleared
        xml = _strip_formula_cache(_normalize_formula_xml(xml))
        if n or sheet_explicit or cleared:
            patched_sheets += 1
        frozen += n
        parts[zip_path] = xml.encode("utf-8")

    parts, comment_removed = _remove_comment_parts(parts)
    parts = _remove_calc_chain_and_force_recalc(parts)

    _write_parts(path, parts, original_order)
    _validate_xml_parts(path)

    return {
        "sheets_patched": patched_sheets,
        "bbg_formulas_frozen": frozen,
        "static_error_literals_cleared": static_errors_cleared,
        "comment_artifacts_removed": comment_removed,
    }


def main() -> int:
    ap = argparse.ArgumentParser(description="Freeze Bloomberg add-in formulas and remove comments from a DCF workbook")
    ap.add_argument("--ticker", required=True)
    ap.add_argument("--path")
    ap.add_argument("--no-backup", action="store_true")
    args = ap.parse_args()

    ticker = args.ticker.upper()
    path = Path(args.path) if args.path else _default_path(ticker)
    report = neutralize(path, ticker, backup=not args.no_backup)
    print(f"Neutralized Bloomberg/comment artifacts: {path}")
    for key, value in report.items():
        print(f"  {key}: {value}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
