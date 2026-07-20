#!/usr/bin/env python3
"""Audit generated DCF workbooks for stale ticker/template residue."""
import argparse
import csv
import json
import re
import sys
import zipfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.formula import DataTableFormula

REPO = Path(__file__).resolve().parents[1]


# Static reference issuers whose data must never leak into another ticker's
# model, grouped by ticker so a ticker's OWN terms are excluded when auditing
# itself (otherwise MOLN's own MP0533/MP0712 would false-positive in the MOLN
# model). This is a floor that is ALWAYS applied and the fallback used when the
# datastore is unreadable. BCYC (template base) and HOWL (Bloomberg example)
# never enter the datastore, so they must live here; CMPX/MOLN are kept as a
# belt-and-braces fallback (e.g. MOLN's own assets are known to be absent from
# dd.duckdb), while every built ticker is additionally derived dynamically below.
STATIC_FORBIDDEN_BY_TICKER = {
    # Base template issuer (Bicycle Therapeutics) + its projects/drugs
    # NB: "Churchill" is NOT listed — it's a generic template analyst-price-target
    # label ("Churchill Analyst Price Target"/"Churchill Base"), present in every
    # model by design, not Bicycle-specific data.
    "BCYC": ["BCYC", "Bicycle", "BT1718", "BT5528", "BT8009", "BT7480", "Zelenectide"],
    # Reference build (Compass Therapeutics) + drugs
    "CMPX": ["CMPX", "Compass Therapeutics", "CTX-009", "CTX-471", "CTX-8371", "CTX-10726"],
    # Prior build (Molecular Partners) + drugs
    "MOLN": ["MOLN", "Molecular Partners", "MP0533", "MP0712", "MP0317", "MP0310", "MP0250"],
    # Bloomberg example issuer left in the template's Earnings BQL formulas
    "HOWL": ["HOWL", "Werewolf"],
}

_EXPORT_DIR = Path(__file__).resolve().parent.parent / "datastore" / "export"
# 4-6 char alpha symbols only: keeps build/biotech tickers (BHVN, CMPX, IOBT)
# and drops Bloomberg-style peer codes ("4503 JP Equity", "Les Lab", "Private")
# and short big-pharma symbols (JNJ/LLY/PFE) whose substrings would false-match.
_TICKER_RE = re.compile(r"[A-Za-z]{4,6}")
# A drug string is a proprietary development code (residue we CAN safely forbid)
# when a letter sits next to a digit: BHV-7000, CTX-471, MP0533, IOB-022, AFM13.
# Plain-word INN/brand names (Zanidatamab, Keytruda) are shared comparators that
# legitimately appear in a same-indication model, so they are NOT auto-forbidden.
_DEVCODE_RE = re.compile(r"[A-Za-z]-?\d")
_PAREN_SUFFIX = re.compile(r"\s*\(.*\)\s*$")


def _norm_drug(name: str) -> str:
    return _PAREN_SUFFIX.sub("", (name or "").strip()).strip().lower()


def _marketed_drug_names(export_dir: Path) -> set[str]:
    """Normalized brand + molecule (INN) names of marketed reference drugs from
    the drug table. These are shared comparators, so they must never become
    forbidden terms (else Cabometyx/cabozantinib would flag every oncology model)."""
    names: set[str] = set()
    try:
        with (export_dir / "drug.csv").open(newline="", encoding="utf-8") as fh:
            for row in csv.DictReader(fh):
                for col in ("drug_name", "molecule"):
                    nm = _norm_drug(row.get(col, ""))
                    if nm:
                        names.add(nm)
    except (OSError, csv.Error, KeyError):
        pass
    return names


def _datastore_forbidden_groups(export_dir: Path) -> dict[str, list[str]]:
    """Derive ticker -> [symbol + proprietary development codes] from peer_drug so
    every ticker the pipeline has ever built (its assets upserted per G3) is
    automatically screened in future builds. Marketed comparators and non-code
    plain names are dropped to avoid false positives on legitimate models."""
    marketed = _marketed_drug_names(export_dir)
    groups: dict[str, set[str]] = {}
    try:
        with (export_dir / "peer_drug.csv").open(newline="", encoding="utf-8") as fh:
            for row in csv.DictReader(fh):
                ticker = (row.get("ticker") or "").strip()
                if not _TICKER_RE.fullmatch(ticker):
                    continue
                tkr = ticker.upper()
                drug = (row.get("drug") or "").strip()
                if len(drug) < 4 or _norm_drug(drug) in marketed:
                    continue
                if not _DEVCODE_RE.search(drug):
                    continue
                groups.setdefault(tkr, {tkr}).add(drug)
    except (OSError, csv.Error, KeyError):
        return {}
    # Keep only groups that contributed at least one development code (a bare
    # symbol-only group adds no residue-detection value).
    return {t: sorted(v) for t, v in groups.items() if len(v) > 1}


def _build_forbidden_by_ticker() -> dict[str, list[str]]:
    merged = {t: set(v) for t, v in STATIC_FORBIDDEN_BY_TICKER.items()}
    for tkr, terms in _datastore_forbidden_groups(_EXPORT_DIR).items():
        merged.setdefault(tkr, set()).update(terms)
    return {t: sorted(v) for t, v in merged.items()}


FORBIDDEN_BY_TICKER = _build_forbidden_by_ticker()
DEFAULT_FORBIDDEN = sorted({t for group in FORBIDDEN_BY_TICKER.values() for t in group})

DATA_CENTER_SHEETS = {
    "Peer View",
    "Peer Views",
    "TAM Solid",
    "TAM Solid+MM",
    "TAM Blood",
}


def audit_workbook(path: Path, ticker: str, forbidden: list[str],
                   include_datacenter: bool = False,
                   strict_research: bool = False) -> int:
    ticker = ticker.upper()
    # Exclude the current ticker's OWN reference group (its symbol, name, drugs)
    # so a reference ticker's legitimate data isn't flagged when auditing itself.
    own = {t.lower() for t in FORBIDDEN_BY_TICKER.get(ticker, [ticker])}
    terms = [t for t in forbidden if t.lower() not in own and t.upper() != ticker]
    hits = []
    # Excel's modern "stale values" display uses a visual strikethrough that is
    # not stored as Font.Strike. It is triggered by a workbook saved in Manual
    # or otherwise incomplete calculation state, so inspect calcPr directly.
    with zipfile.ZipFile(path) as archive:
        workbook_xml = archive.read("xl/workbook.xml").decode("utf-8", "ignore")
        styles_xml = archive.read("xl/styles.xml").decode("utf-8", "ignore")
    calc_match = re.search(r"<(?:\w+:)?calcPr\b([^>]*)/?>", workbook_xml)
    calc_attrs = dict(re.findall(r'(\w+)="([^"]*)"', calc_match.group(1))) \
        if calc_match else {}
    if calc_attrs.get("calcMode", "auto").lower() == "manual":
        hits.append(("Workbook", "calcPr", "manual calculation causes stale-value strikethrough",
                     str(calc_attrs)))
    for attr in ("calcCompleted", "calcOnSave"):
        if calc_attrs.get(attr) == "0":
            hits.append(("Workbook", "calcPr", f"{attr}=0 stale calculation state",
                         str(calc_attrs)))
    if calc_attrs.get("forceFullCalc") == "1":
        hits.append(("Workbook", "calcPr", "forceFullCalc=1 incomplete calculation state",
                     str(calc_attrs)))
    if "calcFeatures" in workbook_xml:
        hits.append(("Workbook", "extLst", "stale calcFeatures extension", "calcFeatures"))
    if re.search(r"<(?:\w+:)?strike(?:\s|/|>)", styles_xml):
        hits.append(("Workbook", "styles.xml", "stored font strikethrough", "<strike>"))

    wb = load_workbook(path, data_only=False, read_only=not strict_research)
    for ws in wb.worksheets:
        if not include_datacenter and ws.title in DATA_CENTER_SHEETS:
            continue
        for row in ws.iter_rows():
            for cell in row:
                if cell.font and cell.font.strike:
                    hits.append((ws.title, cell.coordinate, "stored font strikethrough",
                                 str(cell.value or "")[:120]))
                v = cell.value
                if v is None:
                    continue
                text = str(v)
                if text.startswith("=") and ("#REF!" in text or "#NAME?" in text):
                    hits.append((ws.title, cell.coordinate, "broken formula reference", text[:120]))
                for term in terms:
                    if term.lower() in text.lower():
                        hits.append((ws.title, cell.coordinate, term, text[:120]))

    if strict_research:
        artifact_dir = REPO / "artifacts" / ticker
        manifest_path = artifact_dir / f"{ticker}_catalyst_manifest.json"
        event_path = artifact_dir / f"{ticker}_historical_events_enriched.json"
        if not manifest_path.exists():
            hits.append(("Catalyst", "manifest", "missing", str(manifest_path)))
        else:
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            expected = len(manifest.get("targets") or [])
            framework_version = int(manifest.get("framework_version") or 0)
            ws = wb["Catalyst"]
            if framework_version >= 3:
                layout = manifest.get("layout") or {}
                target_row = int(layout.get("table3_target_row") or 0)
                header_row = int(layout.get("table3_header_row") or 0)
                table3 = 0
                for target in manifest.get("targets") or []:
                    ms_col = column_index_from_string(
                        target["market_share_change_col"]
                    )
                    if str(ws.cell(target_row, ms_col).value or "").strip() == target["name"] and \
                            str(ws.cell(header_row, ms_col).value or "").strip() == "Market Share Change" and \
                            str(ws.cell(header_row, ms_col + 1).value or "").strip() == "LOA Change" and \
                            str(ws.cell(header_row, ms_col + 2).value or "").strip() == "Conv.":
                        table3 += 1
                first = int(layout.get("main_scenario_first") or 10)
                last = int(layout.get("main_scenario_last") or first - 1)
                scenario_ids = [
                    ws.cell(row, 2).value for row in range(first, last + 1)
                    if isinstance(ws.cell(row, 2).value, (int, float))
                ]
                # Table 2 was removed in framework v3.  Its legacy standalone
                # Scenario header must not survive anywhere in row 8.
                if any(str(ws.cell(8, col).value or "").strip() == "Scenario"
                       for col in range(27, ws.max_column + 1)):
                    hits.append(("Catalyst", "row 8", "legacy Table 2 remains", "Scenario header"))
            else:
                table3 = 0
                for col in range(45, ws.max_column - 2):
                    if str(ws.cell(8, col + 1).value or "").strip() == "Conv." and \
                            str(ws.cell(8, col + 2).value or "").strip() == "Market Share Change":
                        table3 += 1
                scenario_ids = [
                    ws.cell(row, 27).value for row in range(10, min(ws.max_row, 200) + 1)
                    if isinstance(ws.cell(row, 27).value, (int, float))
                ]
            expected_scenarios = (
                int((manifest.get("layout") or {}).get("scenario_count") or 0)
                if framework_version >= 4 else expected * 4
            )
            if not expected or table3 != expected or len(scenario_ids) != expected_scenarios:
                hits.append((
                    "Catalyst", "framework", "target-count mismatch",
                    f"manifest={expected}, table3={table3}, "
                    f"scenario_ids={len(scenario_ids)}, expected_scenarios={expected_scenarios}",
                ))
            if framework_version >= 4:
                active = manifest.get("active_targets") or []
                base_col = int(layout.get("base_col") or 0)
                final_col = int(layout.get("final_market_col") or 0)
                outcome_first = int(layout.get("outcome_first_col") or 3)
                if [ws.cell(7, outcome_first + i).value for i in range(len(active))] != active:
                    hits.append(("Catalyst", ws.cell(7, outcome_first).coordinate,
                                 "active-target outcome headers mismatch", str(active)))
                if final_col != base_col + 1 or ws.cell(7, final_col).value != "Final Market Price":
                    hits.append(("Catalyst", ws.cell(7, final_col).coordinate,
                                 "Final Market is not immediately after Base Case", ""))
                group_first = int(layout.get("target_group_first_col") or 0)
                for i, target in enumerate(manifest.get("targets") or []):
                    group_col = group_first + 4 * i
                    hidden = [
                        ws.column_dimensions[get_column_letter(col)].hidden
                        for col in range(group_col, group_col + 4)
                    ]
                    if any(bool(value) for value in hidden):
                        hits.append(("Catalyst", ws.cell(7, group_col).coordinate,
                                     "target breakdown group is hidden", target["name"]))
            if framework_version >= 5:
                embedded = ws["C9"].value
                if not isinstance(embedded, DataTableFormula):
                    hits.append(("Catalyst", "C9", "embedded Catalyst data table missing", ""))
                elif getattr(embedded, "r1", None) != "C6" or getattr(embedded, "r2", None) != "B6":
                    hits.append(("Catalyst", "C9", "embedded input bridge mismatch", str(vars(embedded))))
                if ws["C8"].value != 0.03:
                    hits.append(("Catalyst", "C8", "embedded terminal growth is not 3%", str(ws["C8"].value)))
                c3 = str(wb["VALUATION"]["C3"].value or "").replace("$", "").upper()
                c5 = str(wb["VALUATION"]["C5"].value or "").replace("$", "").upper()
                if "CATALYST!B6" not in c3 or "CATALYST!C6" not in c5:
                    hits.append(("VALUATION", "C3/C5", "embedded input bridges missing", f"{c3}; {c5}"))
                legacy = [
                    wb["VALUATION"].cell(row, col).value
                    for row in range(4, 201) for col in (15, 16)
                    if wb["VALUATION"].cell(row, col).value not in (None, "")
                ]
                if legacy:
                    hits.append(("VALUATION", "O4:P200", "legacy Catalyst price table remains", str(legacy[:10])))
            else:
                val_ids = [wb["VALUATION"].cell(row, 15).value for row in range(5, 5 + len(scenario_ids))]
                if val_ids != scenario_ids:
                    hits.append(("VALUATION", "O5", "scenario ID mismatch", f"{val_ids} vs {scenario_ids}"))
            sws = wb["Scenarios"]
            divider = next((row for row in range(1, sws.max_row + 1)
                            if sws.cell(row, 3).value == "Catalyst Scenarios"), None)
            if divider is None:
                hits.append(("Scenarios", "Catalyst Scenarios", "missing divider", ""))
            else:
                actual_end = next(
                    (row for row in range(divider + 1, sws.max_row + 1)
                     if str(sws.cell(row, 3).value or "").startswith("Test Scenarios - ")),
                    sws.max_row + 1,
                )
                title_ids = [
                    sws.cell(row, 2).value
                    for row in range(divider + 1, actual_end)
                    if isinstance(sws.cell(row, 2).value, (int, float)) and
                    sws.cell(row, 3).value not in (None, "")
                ]
                numeric_only = [
                    row for row in range(divider + 1, actual_end)
                    if isinstance(sws.cell(row, 2).value, (int, float)) and
                    sws.cell(row, 3).value in (None, "")
                ]
                if title_ids != scenario_ids or len(title_ids) != len(set(title_ids)):
                    hits.append(("Scenarios", f"B{divider + 1}",
                                 "Catalyst title count/ID mismatch",
                                 f"titles={title_ids} vs main={scenario_ids}"))
                if numeric_only:
                    hits.append(("Scenarios", f"B{numeric_only[0]}",
                                 "numeric-only Catalyst header",
                                 f"rows={numeric_only[:20]}"))
                if framework_version >= 4 and title_ids:
                    absolute_asset_rows = [
                        row for row in range(1, sws.max_row + 1)
                        if sws.cell(row, 1).value == 4 and
                        str(sws.cell(row, 2).value or "").strip() == "Absolute" and
                        str(sws.cell(row, 4).value or "").strip() != "[%]"
                    ]
                    absolute_rows = [
                        row for row in range(1, sws.max_row + 1)
                        if sws.cell(row, 1).value == 4 and
                        str(sws.cell(row, 2).value or "").strip() == "Absolute"
                    ]
                    if absolute_rows:
                        first_abs = min(absolute_rows)
                        current_asset = None
                        target_ms_rows: dict[str, int] = {}
                        for row in absolute_rows:
                            value = sws.cell(row, 3).value
                            unit = sws.cell(row, 4).value
                            if unit == "[%]" and current_asset:
                                text = str(value or "")
                                match = re.search(r'&"\s*(.*?)\s+Market Share"', text)
                                if not match:
                                    match = re.search(r'\)\s+(.*?)\s+Market Share$', text)
                                indication = match.group(1).strip() if match else "All"
                                name = (
                                    current_asset if indication in ("", "All")
                                    else f"{current_asset} - {indication}"
                                )
                                target_ms_rows[name] = row
                            elif isinstance(value, str) and not value.startswith("=") and \
                                    "TAM" not in value:
                                current_asset = value.split(" (")[0].strip()
                        title_rows = [
                            row for row in range(divider + 1, actual_end)
                            if isinstance(sws.cell(row, 2).value, (int, float)) and
                            sws.cell(row, 3).value not in (None, "")
                        ]
                        residue = [
                            f"Y{header + (asset_row - first_abs + 1)}"
                            for header in title_rows
                            for asset_row in absolute_asset_rows
                            if sws.cell(header + (asset_row - first_abs + 1), 25).value not in (None, "")
                        ]
                        if residue:
                            hits.append(("Scenarios", residue[0],
                                         "Catalyst asset-title Y residue", str(residue[:20])))
                        active_targets = manifest.get("active_targets") or []
                        active_set = set(active_targets)
                        table_first = int(layout["table3_input_first"])
                        table_last = int(layout["table3_input_last"])
                        main_first = int(layout["main_scenario_first"])
                        for scenario_index, header in enumerate(title_rows):
                            main_row = main_first + scenario_index
                            for target in manifest.get("targets") or []:
                                base_ms_row = target_ms_rows.get(target["name"])
                                if base_ms_row is None:
                                    hits.append(("Scenarios", f"Y{header}",
                                                 "Catalyst target missing from Absolute",
                                                 target["name"]))
                                    continue
                                destination = header + (base_ms_row - first_abs + 1)
                                formula = str(sws.cell(destination, 25).value or "").replace(" ", "")
                                if target["name"] not in active_set:
                                    if formula != f"=$Y${base_ms_row}":
                                        hits.append(("Scenarios", f"Y{destination}",
                                                     "non-catalyst market share must remain Absolute",
                                                     formula[:120]))
                                else:
                                    active_index = active_targets.index(target["name"])
                                    outcome_col = get_column_letter(
                                        int(layout.get("outcome_first_col") or 3) + active_index
                                    )
                                    ms_col = target["market_share_change_col"]
                                    required = (
                                        f"$Y${base_ms_row}",
                                        f"Catalyst!${outcome_col}${main_row}",
                                        f"Catalyst!${ms_col}${table_first}:${ms_col}${table_last}",
                                        "MAX(0,",
                                    )
                                    if any(token not in formula for token in required):
                                        hits.append(("Scenarios", f"Y{destination}",
                                                     "active Catalyst market-share lookup mismatch",
                                                     formula[:120]))
        if not event_path.exists():
            hits.append(("Historical Events", "artifact", "missing", str(event_path)))
            allowed_event_urls: set[str] = set()
        else:
            data = json.loads(event_path.read_text(encoding="utf-8"))
            if int(data.get("official_release_count") or 0) <= 0:
                hits.append(("Historical Events", "artifact", "zero official releases", str(event_path)))
            allowed_event_urls = {
                str(e.get(key))
                for e in data.get("events", [])
                for key in ("url", "source_url", "abstract_url")
                if e.get(key)
            }
        hws = wb["Historical Events"]
        for evt_col, cat_col in ((10, 11), (21, 22), (32, 33), (43, 44)):
            for row in range(9, 375):
                event_cell = hws.cell(row, evt_col)
                if event_cell.hyperlink is not None and allowed_event_urls and \
                        event_cell.hyperlink.target not in allowed_event_urls:
                    hits.append(("Historical Events", event_cell.coordinate,
                                 "hyperlink absent from current-ticker event artifact",
                                 event_cell.hyperlink.target[:120]))
                if hws.cell(row, cat_col).value != "Clinical Data":
                    continue
                text = str(event_cell.value or "")
                if not text.startswith("["):
                    hits.append(("Historical Events", event_cell.coordinate,
                                 "clinical venue prefix missing", text[:120]))
                if event_cell.hyperlink is None:
                    hits.append(("Historical Events", event_cell.coordinate,
                                 "clinical source hyperlink missing", text[:120]))

    wb.close()
    if hits:
        print("Forbidden/stale workbook terms found:")
        for sheet, coord, term, text in hits[:200]:
            print(f"  {sheet}!{coord}: {term} in {text!r}")
        if len(hits) > 200:
            print(f"  ... {len(hits) - 200} more")
        return 1

    print(f"Workbook audit OK: no stale forbidden terms"
          f"{' and strict research gates passed' if strict_research else ''} in {path}")
    return 0


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--ticker", required=True)
    ap.add_argument("--path", help="Workbook path; default DD/{ticker}/DCF {ticker}.xlsx")
    ap.add_argument("--forbidden", nargs="*", default=DEFAULT_FORBIDDEN)
    ap.add_argument("--include-datacenter", action="store_true",
                    help="Also scan Peer/TAM data-center tabs")
    ap.add_argument("--strict-research", action="store_true",
                    help="Require full Catalyst/event artifacts, counts and clinical hyperlinks")
    args = ap.parse_args()
    path = Path(args.path) if args.path else Path(
        f"/mnt/c/Users/yzsun/Desktop/DD/{args.ticker}/DCF {args.ticker}.xlsx"
    )
    return audit_workbook(path, args.ticker, args.forbidden, args.include_datacenter,
                          args.strict_research)


if __name__ == "__main__":
    sys.exit(main())
