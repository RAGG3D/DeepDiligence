#!/usr/bin/env python3
"""Normalize generated workbook formats that depend on style IDs."""

from __future__ import annotations

import argparse
import html
import re
import shutil
import time
import warnings
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

warnings.filterwarnings("ignore")
import openpyxl  # noqa: E402
from openpyxl.utils import column_index_from_string, get_column_letter  # noqa: E402


NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


def _default_path(ticker: str) -> Path:
    return Path(f"/mnt/c/Users/yzsun/Desktop/DD/{ticker}/DCF {ticker}.xlsx")


def _sheet_zip_path(parts: dict[str, bytes], sheet_name: str) -> str | None:
    wb = ET.fromstring(parts["xl/workbook.xml"])
    rels = ET.fromstring(parts["xl/_rels/workbook.xml.rels"])
    relmap = {rel.get("Id"): rel.get("Target") for rel in rels}
    for sheet in wb.findall(f".//{{{NS_MAIN}}}sheet"):
        if html.unescape(sheet.get("name", "")) != sheet_name:
            continue
        target = relmap.get(sheet.get(f"{{{NS_R}}}id"))
        if not target:
            return None
        target = target.lstrip("/")
        return target if target.startswith("xl/") else "xl/" + target
    return None


def _read_parts(path: Path) -> tuple[dict[str, bytes], list[str]]:
    with zipfile.ZipFile(path, "r") as zf:
        return {name: zf.read(name) for name in zf.namelist()}, zf.namelist()


def _write_parts(path: Path, parts: dict[str, bytes], order: list[str]) -> None:
    tmp = path.with_suffix(".~formats.xlsx")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        written = set()
        for name in order:
            if name in parts:
                zout.writestr(name, parts[name])
                written.add(name)
        for name, data in parts.items():
            if name not in written:
                zout.writestr(name, data)
    tmp.replace(path)


def _split_cell_xfs(styles_xml: str) -> list[str]:
    match = re.search(r"<cellXfs\b[^>]*>(.*?)</cellXfs>", styles_xml, re.S)
    if not match:
        return []
    return re.findall(r"<xf\b[^>]*/>|<xf\b[^>]*>.*?</xf>", match.group(1), re.S)


def _stage_style_id(styles_xml: str, pipeline_xml: str) -> str:
    numfmts = {
        m.group(1): html.unescape(m.group(2))
        for m in re.finditer(r'<numFmt numFmtId="(\d+)" formatCode="([^"]*)"', styles_xml)
    }
    xfs = _split_cell_xfs(styles_xml)
    stage_ids = []
    for idx, xf in enumerate(xfs):
        num = re.search(r'numFmtId="(\d+)"', xf)
        if num and "Stage" in numfmts.get(num.group(1), ""):
            stage_ids.append((str(idx), xf))
    if not stage_ids:
        raise RuntimeError("No cell style with a custom Stage number format found")

    used_pipeline_styles = set(re.findall(r'\bs="(\d+)"', pipeline_xml))
    candidates = [item for item in stage_ids if item[0] in used_pipeline_styles] or stage_ids

    def score(item: tuple[str, str]) -> tuple[int, int]:
        _style_id, xf = item
        fill_m = re.search(r'fillId="(\d+)"', xf)
        border_m = re.search(r'borderId="(\d+)"', xf)
        fill = int(fill_m.group(1)) if fill_m else 0
        border = int(border_m.group(1)) if border_m else 0
        return (1 if fill > 0 else 0, border)

    return max(candidates, key=score)[0]


def _stage_rows(path: Path) -> list[int]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        ws = wb["Pipeline"]
        rows: list[int] = []
        for r in range(9, min(ws.max_row, 220) + 1):
            a = ws.cell(r, 1).value
            c = ws.cell(r, 3).value
            d = ws.cell(r, 4).value
            if a == "X" and (c is None or c == "") and isinstance(d, str) and not d.startswith("="):
                rows.append(r)
        return rows
    finally:
        wb.close()


def _cell_col(addr: str) -> int:
    col = re.match(r"([A-Z]+)", addr).group(1)
    out = 0
    for ch in col:
        out = out * 26 + ord(ch) - 64
    return out


def _set_cell_style(xml: str, addr: str, style_id: str) -> str:
    m = re.search(
        rf'<c\b(?=[^>]*\br="{addr}")[^>]*/>|'
        rf'<c\b(?=[^>]*\br="{addr}")[^>]*>.*?</c>',
        xml,
        re.S,
    )
    if m:
        cell = m.group(0)
        if re.search(r'\bs="\d+"', cell):
            cell2 = re.sub(r'\bs="\d+"', f's="{style_id}"', cell, count=1)
        else:
            cell2 = cell.replace("<c ", f'<c s="{style_id}" ', 1)
        return xml[:m.start()] + cell2 + xml[m.end():]

    row_num = int(re.search(r"\d+", addr).group(0))
    row_m = re.search(rf'(<row\b(?=[^>]*\br="{row_num}")[^>]*>)(.*?)(</row>)', xml, re.S)
    if not row_m:
        return xml
    body = row_m.group(2)
    insert_at = len(body)
    target_col = _cell_col(addr)
    for cm in re.finditer(r'<c\b[^>]*\br="([A-Z]+\d+)"', body):
        if _cell_col(cm.group(1)) > target_col:
            insert_at = cm.start()
            break
    new_body = body[:insert_at] + f'<c r="{addr}" s="{style_id}"/>' + body[insert_at:]
    return xml[:row_m.start(2)] + new_body + xml[row_m.end(2):]


def _cell_xml(xml: str, addr: str) -> str | None:
    m = re.search(
        rf'<c\b(?=[^>]*\br="{addr}")[^>]*/>|'
        rf'<c\b(?=[^>]*\br="{addr}")[^>]*>.*?</c>',
        xml,
        re.S,
    )
    return m.group(0) if m else None


def _cell_style(xml: str, addr: str) -> str | None:
    cell = _cell_xml(xml, addr)
    if not cell:
        return None
    m = re.search(r'\bs="(\d+)"', cell)
    return m.group(1) if m else None


def _shared_strings(parts: dict[str, bytes]) -> list[str]:
    raw = parts.get("xl/sharedStrings.xml")
    if not raw:
        return []
    root = ET.fromstring(raw)
    return ["".join(t.text or "" for t in si.iter(f"{{{NS_MAIN}}}t"))
            for si in root.findall(f"{{{NS_MAIN}}}si")]


def _read_cell_text(xml: str, addr: str, shared: list[str]) -> str | None:
    cell = _cell_xml(xml, addr)
    if not cell:
        return None
    ctype_m = re.search(r'\bt="([^"]+)"', cell)
    ctype = ctype_m.group(1) if ctype_m else "n"
    if ctype == "inlineStr":
        tm = re.search(r"<t[^>]*>(.*?)</t>", cell, re.S)
        return html.unescape(tm.group(1)) if tm else None
    vm = re.search(r"<v>(.*?)</v>", cell, re.S)
    if not vm:
        return None
    if ctype == "s":
        try:
            return shared[int(vm.group(1))]
        except (ValueError, IndexError):
            return None
    if ctype == "str":
        return html.unescape(vm.group(1))
    return None


def _fiscal_year_label(parts: dict[str, bytes]) -> str | None:
    """Read the ticker's fiscal-year-end header from the FY DATA sheets so the
    FSA statement mirrors the actual fiscal year end rather than assuming a
    calendar-year (December 31) close."""
    shared = _shared_strings(parts)
    for sheet_name in ("FY DATA K USD", "FY DATA"):
        sheet_path = _sheet_zip_path(parts, sheet_name)
        if not sheet_path:
            continue
        xml = parts[sheet_path].decode("utf-8", "ignore")
        text = _read_cell_text(xml, "D3", shared)
        if text and "Fiscal Year Ended" in text:
            return text
    return None


def _set_text_cell(xml: str, addr: str, text: str, style_id: str | None = None) -> str:
    escaped = html.escape(text, quote=False)
    cell = f'<c r="{addr}"'
    if style_id is not None:
        cell += f' s="{style_id}"'
    cell += f' t="inlineStr"><is><t>{escaped}</t></is></c>'

    m = re.search(
        rf'<c\b(?=[^>]*\br="{addr}")[^>]*/>|'
        rf'<c\b(?=[^>]*\br="{addr}")[^>]*>.*?</c>',
        xml,
        re.S,
    )
    if m:
        return xml[:m.start()] + cell + xml[m.end():]

    row_num = int(re.search(r"\d+", addr).group(0))
    row_m = re.search(rf'(<row\b(?=[^>]*\br="{row_num}")[^>]*>)(.*?)(</row>)', xml, re.S)
    if not row_m:
        return xml
    body = row_m.group(2)
    insert_at = len(body)
    target_col = _cell_col(addr)
    for cm in re.finditer(r'<c\b[^>]*\br="([A-Z]+\d+)"', body):
        if _cell_col(cm.group(1)) > target_col:
            insert_at = cm.start()
            break
    new_body = body[:insert_at] + cell + body[insert_at:]
    return xml[:row_m.start(2)] + new_body + xml[row_m.end(2):]


def _style_fill_ids(styles_xml: str) -> dict[str, int]:
    root = ET.fromstring(styles_xml)
    xfs = root.find(f"{{{NS_MAIN}}}cellXfs")
    out: dict[str, int] = {}
    if xfs is None:
        return out
    for idx, xf in enumerate(list(xfs)):
        try:
            out[str(idx)] = int(xf.get("fillId", "0"))
        except ValueError:
            out[str(idx)] = 0
    return out


def _copy_style_if_missing_between(xml: str, row: int, start_col: str, end_col: str,
                                   fill_ids: dict[str, int]) -> tuple[str, int]:
    """Fill missing spacer cells in a header row with the nearest filled style."""
    changed = 0
    start = column_index_from_string(start_col)
    end = column_index_from_string(end_col)
    for col_idx in range(start, end + 1):
        addr = f"{get_column_letter(col_idx)}{row}"
        style = _cell_style(xml, addr)
        if style and fill_ids.get(style, 0) > 0:
            continue

        left_style = None
        for lidx in range(col_idx - 1, start - 1, -1):
            cand = _cell_style(xml, f"{get_column_letter(lidx)}{row}")
            if cand and fill_ids.get(cand, 0) > 0:
                left_style = cand
                break
        right_style = None
        for ridx in range(col_idx + 1, end + 1):
            cand = _cell_style(xml, f"{get_column_letter(ridx)}{row}")
            if cand and fill_ids.get(cand, 0) > 0:
                right_style = cand
                break
        replacement = left_style or right_style
        if not replacement:
            continue
        before = xml
        xml = _set_cell_style(xml, addr, replacement)
        changed += int(xml != before)
    return xml, changed


def _numfmt_map(styles_root: ET.Element) -> dict[str, str]:
    out: dict[str, str] = {}
    numfmts = styles_root.find(f"{{{NS_MAIN}}}numFmts")
    if numfmts is None:
        return out
    for numfmt in numfmts.findall(f"{{{NS_MAIN}}}numFmt"):
        out[numfmt.get("numFmtId", "")] = html.unescape(numfmt.get("formatCode", ""))
    return out


def _style_numfmt_ids(styles_root: ET.Element) -> dict[str, str]:
    out: dict[str, str] = {}
    xfs = styles_root.find(f"{{{NS_MAIN}}}cellXfs")
    if xfs is None:
        return out
    for idx, xf in enumerate(list(xfs)):
        out[str(idx)] = xf.get("numFmtId", "0")
    return out


def _is_hidden_text_numfmt(fmt: str) -> bool:
    sections = fmt.split(";")
    return len(sections) >= 3 and sections[-1] == ""


def _ensure_general_text_style(styles_xml: str, source_style_id: str,
                               cache: dict[str, str]) -> tuple[str, str]:
    if source_style_id in cache:
        return styles_xml, cache[source_style_id]

    ET.register_namespace("", NS_MAIN)
    root = ET.fromstring(styles_xml)
    cellxfs = root.find(f"{{{NS_MAIN}}}cellXfs")
    if cellxfs is None:
        return styles_xml, source_style_id
    source_idx = int(source_style_id)
    if source_idx >= len(list(cellxfs)):
        return styles_xml, source_style_id

    source_xf = list(cellxfs)[source_idx]
    new_xf = ET.fromstring(ET.tostring(source_xf))
    new_xf.set("numFmtId", "0")
    new_xf.attrib.pop("applyNumberFormat", None)
    cellxfs.append(new_xf)
    new_id = str(len(list(cellxfs)) - 1)
    cellxfs.set("count", str(len(list(cellxfs))))
    cache[source_style_id] = new_id
    return ET.tostring(root, encoding="unicode"), new_id


def _set_col_width_visible(xml: str, col_idx: int, width: float) -> str:
    cols_m = re.search(r"<cols\b[^>]*>.*?</cols>", xml, re.S)
    col_tag = f'<col min="{col_idx}" max="{col_idx}" width="{width}" customWidth="1"/>'
    if not cols_m:
        return xml.replace("<sheetData", f"<cols>{col_tag}</cols><sheetData", 1)
    cols = cols_m.group(0)
    target = None
    for m in re.finditer(r"<col\b[^>]*/>", cols):
        min_m = re.search(r'\bmin="(\d+)"', m.group(0))
        max_m = re.search(r'\bmax="(\d+)"', m.group(0))
        if min_m and max_m and int(min_m.group(1)) <= col_idx <= int(max_m.group(1)):
            target = m
            break
    if target:
        tag = target.group(0)
        tag = re.sub(r'\bhidden="[^"]*"', "", tag)
        tag = re.sub(r'\bwidth="[^"]*"', f'width="{width}"', tag)
        if 'width="' not in tag:
            tag = tag.replace("/>", f' width="{width}"/>')
        tag = re.sub(r'\bcustomWidth="[^"]*"', 'customWidth="1"', tag)
        if 'customWidth="' not in tag:
            tag = tag.replace("/>", ' customWidth="1"/>')
        cols = cols[:target.start()] + tag + cols[target.end():]
    else:
        cols = cols.replace("</cols>", f"{col_tag}</cols>")
    return xml[:cols_m.start()] + cols + xml[cols_m.end():]


def normalize_pipeline_stage_formats(path: Path) -> int:
    parts, order = _read_parts(path)
    sheet_path = _sheet_zip_path(parts, "Pipeline")
    if not sheet_path:
        raise RuntimeError("Pipeline sheet not found")
    pipeline_xml = parts[sheet_path].decode("utf-8", "ignore")
    style_id = _stage_style_id(parts["xl/styles.xml"].decode("utf-8", "ignore"), pipeline_xml)
    rows = _stage_rows(path)
    changed = 0
    for row in rows:
        for col_idx in range(6, 35):
            before = pipeline_xml
            pipeline_xml = _set_cell_style(pipeline_xml, f"{get_column_letter(col_idx)}{row}", style_id)
            changed += int(pipeline_xml != before)
    parts[sheet_path] = pipeline_xml.encode("utf-8")
    _write_parts(path, parts, order)
    print(f"Pipeline Stage format normalized: rows={rows}, style={style_id}, cells_changed={changed}")
    return changed


def normalize_header_and_text_formats(path: Path) -> int:
    parts, order = _read_parts(path)
    styles_xml = parts["xl/styles.xml"].decode("utf-8", "ignore")
    fill_ids = _style_fill_ids(styles_xml)
    changed = 0

    for sheet_name, row_specs in {
        "Peer View": [(5, "E", "V")],
        "Peer Views": [(12, "E", "V"), (13, "E", "V"), (14, "E", "V")],
    }.items():
        sheet_path = _sheet_zip_path(parts, sheet_name)
        if not sheet_path:
            continue
        xml = parts[sheet_path].decode("utf-8", "ignore")
        for row, start_col, end_col in row_specs:
            xml, n = _copy_style_if_missing_between(xml, row, start_col, end_col, fill_ids)
            changed += n
        parts[sheet_path] = xml.encode("utf-8")

    fsa_path = _sheet_zip_path(parts, "FSA")
    if fsa_path:
        xml = parts[fsa_path].decode("utf-8", "ignore")
        style = _cell_style(xml, "D3")
        fy_label = _fiscal_year_label(parts) or "Fiscal Year Ended December 31."
        before = xml
        xml = _set_text_cell(xml, "D3", fy_label, style)
        changed += int(xml != before)
        parts[fsa_path] = xml.encode("utf-8")

    for fy_sheet in ("FY DATA", "FY DATA K USD"):
        fy_path = _sheet_zip_path(parts, fy_sheet)
        if not fy_path:
            continue
        xml = parts[fy_path].decode("utf-8", "ignore")
        root = ET.fromstring(styles_xml)
        numfmts = _numfmt_map(root)
        style_numfmts = _style_numfmt_ids(root)
        style_cache: dict[str, str] = {}
        xml = _set_col_width_visible(xml, 4, 68)
        for m in list(re.finditer(r'<c\b(?=[^>]*\br="D\d+")[^>]*(?:/>|>.*?</c>)', xml, re.S)):
            cell = m.group(0)
            sm = re.search(r'\bs="(\d+)"', cell)
            if not sm:
                continue
            style_id = sm.group(1)
            fmt = numfmts.get(style_numfmts.get(style_id, ""), "")
            if fmt == "General":
                continue
            if _is_hidden_text_numfmt(fmt) or '<f>' in cell or 't="s"' in cell or 't="inlineStr"' in cell or "<is>" in cell:
                styles_xml, new_style = _ensure_general_text_style(styles_xml, style_id, style_cache)
                before = xml
                addr = re.search(r'\br="([A-Z]+\d+)"', cell).group(1)
                xml = _set_cell_style(xml, addr, new_style)
                changed += int(xml != before)
        parts[fy_path] = xml.encode("utf-8")
        parts["xl/styles.xml"] = styles_xml.encode("utf-8")

    _write_parts(path, parts, order)
    print(f"Header/text formats normalized: cells_changed={changed}")
    return changed


def main() -> None:
    parser = argparse.ArgumentParser(description="Normalize generated DCF workbook formats")
    parser.add_argument("--ticker")
    parser.add_argument("--path")
    parser.add_argument("--no-backup", action="store_true")
    args = parser.parse_args()
    path = Path(args.path) if args.path else _default_path(args.ticker)
    if not path.exists():
        raise FileNotFoundError(path)
    if not args.no_backup:
        backup = path.with_name(f"{path.stem}_pre_format_normalize_{time.strftime('%Y%m%d_%H%M%S')}.xlsx")
        shutil.copy2(path, backup)
        print(f"Backup: {backup}")
    normalize_pipeline_stage_formats(path)
    normalize_header_and_text_formats(path)


if __name__ == "__main__":
    main()
