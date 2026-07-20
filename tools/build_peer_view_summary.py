#!/usr/bin/env python3
"""Build the ticker-specific Peer View tab from Pipeline + datastore peers.

Peer View (singular) is the delivered current-ticker drug-vs-peer comparison
page, filtered by the indications that actually appear in Pipeline.  Peer Views
(plural) is a build-time raw data-center sheet; it is rendered when present but
final models may delete it.
"""

from __future__ import annotations

import argparse
import csv
import html
import json
import re
import shutil
import subprocess
import sys
import tempfile
import time
import zipfile
from collections import defaultdict
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

import openpyxl


NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
REPO = Path(__file__).resolve().parents[1]
if str(REPO) not in sys.path:
    sys.path.insert(0, str(REPO))
EXPORT = REPO / "datastore" / "export"
ARTIFACTS = REPO / "artifacts"


SECTION_ALIASES = {
    "BTC": "Biliary Tract Cancer (BTC)",
    "SCLC": "SCLC 3L+",
}


def _default_path(ticker: str) -> Path:
    return Path(f"/mnt/c/Users/yzsun/Desktop/DD/{ticker}/DCF {ticker}.xlsx")


def _sheet_zip_path(parts: dict[str, bytes], sheet_name: str) -> str:
    wb = ET.fromstring(parts["xl/workbook.xml"])
    rels = ET.fromstring(parts["xl/_rels/workbook.xml.rels"])
    relmap = {rel.get("Id"): rel.get("Target") for rel in rels}
    for sheet in wb.findall(f".//{{{NS_MAIN}}}sheet"):
        if html.unescape(sheet.get("name", "")) == sheet_name:
            target = relmap[sheet.get(f"{{{NS_R}}}id")]
            target = target.lstrip("/")
            return target if target.startswith("xl/") else "xl/" + target
    raise RuntimeError(f"Sheet not found: {sheet_name}")


def _optional_sheet_zip_path(parts: dict[str, bytes], sheet_name: str) -> str | None:
    try:
        return _sheet_zip_path(parts, sheet_name)
    except RuntimeError:
        return None


def _read_parts(path: Path) -> tuple[dict[str, bytes], list[str]]:
    with zipfile.ZipFile(path, "r") as zf:
        return {name: zf.read(name) for name in zf.namelist()}, zf.namelist()


def _write_parts(path: Path, parts: dict[str, bytes], order: list[str]) -> None:
    tmp = path.with_suffix(".~peer_view.xlsx")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        written = set()
        for name in order:
            if name in parts:
                zout.writestr(name, parts[name])
                written.add(name)
        for name, data in parts.items():
            if name not in written:
                zout.writestr(name, data)
    tmp.replace(path)


def _col_num(col: str) -> int:
    out = 0
    for ch in col:
        out = out * 26 + ord(ch) - 64
    return out


def _col_letter(index: int) -> str:
    out = ""
    while index:
        index, rem = divmod(index - 1, 26)
        out = chr(65 + rem) + out
    return out


def _addr_col(addr: str) -> int:
    return _col_num(re.match(r"([A-Z]+)", addr).group(1))


def _cell_style(xml: str, addr: str, default: str = "0") -> str:
    m = re.search(
        rf'<c\b(?=[^>]*\br="{addr}")[^>]*/>|'
        rf'<c\b(?=[^>]*\br="{addr}")[^>]*>.*?</c>',
        xml,
        re.S,
    )
    if not m:
        return default
    sm = re.search(r'\bs="(\d+)"', m.group(0))
    return sm.group(1) if sm else default


def _ensure_row(xml: str, row: int) -> str:
    if re.search(rf'<row\b(?=[^>]*\br="{row}")', xml):
        return xml
    row_xml = f'<row r="{row}"></row>'
    sheet_data_end = xml.find("</sheetData>")
    if sheet_data_end == -1:
        return xml
    rows = list(re.finditer(r'<row\b(?=[^>]*\br="(\d+)")[^>]*(?:/>|>.*?</row>)', xml, re.S))
    insert_at = sheet_data_end
    for m in rows:
        rm = re.search(r'\br="(\d+)"', m.group(0))
        if rm and int(rm.group(1)) > row:
            insert_at = m.start()
            break
    return xml[:insert_at] + row_xml + xml[insert_at:]


def _replace_cell(xml: str, addr: str, cell_xml: str) -> str:
    row = int(re.search(r"\d+", addr).group(0))
    xml = _ensure_row(xml, row)
    m = re.search(
        rf'<c\b(?=[^>]*\br="{addr}")[^>]*/>|'
        rf'<c\b(?=[^>]*\br="{addr}")[^>]*>.*?</c>',
        xml,
        re.S,
    )
    if m:
        return xml[:m.start()] + cell_xml + xml[m.end():]
    row_m = re.search(rf'(<row\b(?=[^>]*\br="{row}")[^>]*>)(.*?)(</row>)', xml, re.S)
    if not row_m:
        return xml
    body = row_m.group(2)
    insert_at = len(body)
    target_col = _addr_col(addr)
    for cm in re.finditer(r'<c\b[^>]*\br="([A-Z]+\d+)"', body):
        if _addr_col(cm.group(1)) > target_col:
            insert_at = cm.start()
            break
    new_body = body[:insert_at] + cell_xml + body[insert_at:]
    return xml[:row_m.start(2)] + new_body + xml[row_m.end(2):]


def _text_cell(addr: str, text: Any, style: str) -> str:
    return f'<c r="{addr}" s="{style}" t="inlineStr"><is><t>{html.escape("" if text is None else str(text), quote=False)}</t></is></c>'


def _num_cell(addr: str, value: float, style: str) -> str:
    return f'<c r="{addr}" s="{style}"><v>{value}</v></c>'


def _blank_cell(addr: str, style: str) -> str:
    return f'<c r="{addr}" s="{style}"/>'


def _shared_string_adder(parts: dict[str, bytes]):
    ET.register_namespace("", NS_MAIN)
    path = "xl/sharedStrings.xml"
    if path not in parts:
        root = ET.Element(f"{{{NS_MAIN}}}sst", {
            "count": "0",
            "uniqueCount": "0",
        })
    else:
        root = ET.fromstring(parts[path])
    strings: list[str] = []
    for si in root.findall(f"{{{NS_MAIN}}}si"):
        strings.append("".join(t.text or "" for t in si.findall(f".//{{{NS_MAIN}}}t")))
    index = {text: i for i, text in enumerate(strings)}

    def add(text: Any) -> int:
        value = "" if text is None else str(text)
        if value in index:
            return index[value]
        si = ET.SubElement(root, f"{{{NS_MAIN}}}si")
        t = ET.SubElement(si, f"{{{NS_MAIN}}}t")
        if value[:1].isspace() or value[-1:].isspace():
            t.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
        t.text = value
        idx = len(strings)
        strings.append(value)
        index[value] = idx
        return idx

    def finalize() -> None:
        root.set("count", str(len(strings)))
        root.set("uniqueCount", str(len(strings)))
        parts[path] = (
            b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
            + ET.tostring(root, encoding="utf-8")
        )

    return add, finalize


def _clean_peer_view_xml(rows: list[dict[str, str]], ticker: str, add_shared_string) -> str:
    headers = [
        "Indication", "Drug", "Ticker/Owner", "Rating", "Status/Phase",
        "Mechanism", "Treatment Line", "N", "ORR", "CR", "PFS", "OS",
        "Safety", "Date/Source",
    ]
    max_row = max(8 + len(rows), 40)

    def cell(addr: str, value: Any, style: int = 0) -> str:
        sid = add_shared_string(value)
        style_attr = "" if style == 0 else f' s="{style}"'
        return f'<c r="{addr}"{style_attr} t="s"><v>{sid}</v></c>'

    def num_cell(addr: str, value: float, style: int = 0) -> str:
        style_attr = "" if style == 0 else f' s="{style}"'
        return f'<c r="{addr}"{style_attr}><v>{value}</v></c>'

    row_cells: dict[int, list[str]] = defaultdict(list)
    row_cells[5].append(cell("D5", f"{ticker.upper()} Peer View: Pipeline vs Indication Peers"))
    for idx, header in enumerate(headers, start=_col_num("D")):
        row_cells[7].append(cell(f"{_col_letter(idx)}7", header))

    for row_idx, row in enumerate(rows, start=8):
        values = [
            row["indication"], row["drug"], row["ticker"], row["rating"],
            row["status"], row["mechanism"], row["line"], row["n"],
            row["orr"], row["cr"], row["pfs"], row["os"], row["safety"],
            row["source"],
        ]
        for col_idx, value in enumerate(values, start=_col_num("D")):
            addr = f"{_col_letter(col_idx)}{row_idx}"
            if value in (None, ""):
                continue
            if isinstance(value, str) and re.fullmatch(r"-?\d+(?:\.\d+)?", value):
                row_cells[row_idx].append(num_cell(addr, float(value)))
            else:
                row_cells[row_idx].append(cell(addr, value))

    sheet_rows = []
    for row_idx in sorted(row_cells):
        sheet_rows.append(f'<row r="{row_idx}" spans="4:17" x14ac:dyDescent="0.25">{"".join(row_cells[row_idx])}</row>')

    cols = "".join([
        '<col min="1" max="3" width="5" customWidth="1"/>',
        '<col min="4" max="4" width="14" customWidth="1"/>',
        '<col min="5" max="5" width="22" customWidth="1"/>',
        '<col min="6" max="8" width="16" customWidth="1"/>',
        '<col min="9" max="9" width="42" customWidth="1"/>',
        '<col min="10" max="15" width="14" customWidth="1"/>',
        '<col min="16" max="17" width="24" customWidth="1"/>',
    ])
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
        'xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006" '
        'mc:Ignorable="x14ac xr xr2 xr3" '
        'xmlns:x14ac="http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac" '
        'xmlns:xr="http://schemas.microsoft.com/office/spreadsheetml/2014/revision" '
        'xmlns:xr2="http://schemas.microsoft.com/office/spreadsheetml/2015/revision2" '
        'xmlns:xr3="http://schemas.microsoft.com/office/spreadsheetml/2016/revision3" '
        'xr:uid="{00000000-0001-0000-0700-000000000000}">'
        '<sheetPr><tabColor rgb="FFFFC000"/></sheetPr>'
        f'<dimension ref="D5:Q{max(28, 7 + len(rows))}"/>'
        '<sheetViews><sheetView showGridLines="0" workbookViewId="0"/></sheetViews>'
        '<sheetFormatPr defaultColWidth="5.5703125" defaultRowHeight="15" x14ac:dyDescent="0.25"/>'
        f'<cols>{cols}</cols>'
        f'<sheetData>{"".join(sheet_rows)}</sheetData>'
        '<pageMargins left="0.7" right="0.7" top="0.75" bottom="0.75" header="0.511811023622047" footer="0.511811023622047"/>'
        '<pageSetup paperSize="9" orientation="portrait" horizontalDpi="300" verticalDpi="300"/>'
        '</worksheet>'
    )


def _clean_peer_views_datacenter_xml(
    drugs: dict[str, list[dict[str, str]]],
    metrics: dict[tuple[str, str], dict[str, str]],
    add_shared_string,
) -> str:
    """Render the raw Peer Views database as one row per peer drug/release."""
    headers = [
        "Section ID", "Section", "Source Col", "Drug", "Ticker/Owner",
        "Rating", "Status", "Mechanism", "Readout Phase", "Date",
        "Treatment Line", "N", "ORR", "CR", "DCR", "PFS", "OS",
        "Safety", "Latest Sale (MM USD)", "All Metrics JSON",
    ]
    row_cells: dict[int, list[str]] = defaultdict(list)

    def cell(addr: str, value: Any, style: int = 0) -> str:
        sid = add_shared_string(value)
        style_attr = "" if style == 0 else f' s="{style}"'
        return f'<c r="{addr}"{style_attr} t="s"><v>{sid}</v></c>'

    def num_cell(addr: str, value: float, style: int = 0) -> str:
        style_attr = "" if style == 0 else f' s="{style}"'
        return f'<c r="{addr}"{style_attr}><v>{value}</v></c>'

    def metric(row_metrics: dict[str, str], *names: str) -> str:
        for name in names:
            value = row_metrics.get(name)
            if value not in (None, "", "/"):
                return str(value)
        return ""

    row_cells[5].append(cell("D5", "Peer Views Data Center: one row per peer drug/release"))
    for idx, header in enumerate(headers, start=_col_num("D")):
        row_cells[7].append(cell(f"{_col_letter(idx)}7", header))

    all_rows: list[dict[str, Any]] = []
    for section in sorted(
        drugs,
        key=lambda s: int(drugs[s][0].get("section_id", "999999") or 999999),
    ):
        for drug in drugs[section]:
            row_metrics = metrics.get((section, drug["col"]), {})
            all_rows.append({
                "section_id": drug.get("section_id", ""),
                "section": section,
                "col": drug.get("col", ""),
                "drug": drug.get("drug", ""),
                "ticker": drug.get("ticker", ""),
                "rating": drug.get("rating", ""),
                "status": metric(row_metrics, "Result"),
                "mechanism": metric(row_metrics, "Innovation"),
                "phase": metric(row_metrics, "Readout Phase"),
                "date": _date_value(metric(row_metrics, "Date")),
                "line": metric(row_metrics, "Treatment Line"),
                "n": metric(row_metrics, "Evaluable Patients", "Patient Number"),
                "orr": metric(row_metrics, "ORR", "ORR (Confirmed)"),
                "cr": metric(row_metrics, "CR", "CR (Confirmed)"),
                "dcr": metric(row_metrics, "DCR"),
                "pfs": metric(row_metrics, "Median PFS", "Median rPFS"),
                "os": metric(row_metrics, "Median OS"),
                "safety": metric(row_metrics, "≥G3 SAE/Patients", "≥G3 clinical AE"),
                "sales": metric(row_metrics, "Latest Sale (MM USD)"),
                "json": json.dumps(row_metrics, sort_keys=True, ensure_ascii=False),
            })

    for row_idx, row in enumerate(all_rows, start=8):
        values = [
            row["section_id"], row["section"], row["col"], row["drug"],
            row["ticker"], row["rating"], row["status"], row["mechanism"],
            row["phase"], row["date"], row["line"], row["n"], row["orr"],
            row["cr"], row["dcr"], row["pfs"], row["os"], row["safety"],
            row["sales"], row["json"],
        ]
        for col_idx, value in enumerate(values, start=_col_num("D")):
            if value in (None, ""):
                continue
            addr = f"{_col_letter(col_idx)}{row_idx}"
            if isinstance(value, str) and re.fullmatch(r"-?\d+(?:\.\d+)?", value):
                row_cells[row_idx].append(num_cell(addr, float(value)))
            else:
                row_cells[row_idx].append(cell(addr, value))

    sheet_rows = []
    for row_idx in sorted(row_cells):
        sheet_rows.append(
            f'<row r="{row_idx}" spans="4:23" x14ac:dyDescent="0.25">'
            f'{"".join(row_cells[row_idx])}</row>'
        )

    cols = "".join([
        '<col min="1" max="3" width="5" customWidth="1"/>',
        '<col min="4" max="4" width="10" customWidth="1"/>',
        '<col min="5" max="5" width="26" customWidth="1"/>',
        '<col min="6" max="6" width="11" customWidth="1"/>',
        '<col min="7" max="9" width="18" customWidth="1"/>',
        '<col min="10" max="13" width="15" customWidth="1"/>',
        '<col min="14" max="20" width="12" customWidth="1"/>',
        '<col min="21" max="21" width="34" customWidth="1"/>',
        '<col min="22" max="22" width="18" customWidth="1"/>',
        '<col min="23" max="23" width="80" customWidth="1"/>',
    ])
    last_row = max(8, 7 + len(all_rows))
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
        'xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006" '
        'mc:Ignorable="x14ac xr xr2 xr3" '
        'xmlns:x14ac="http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac" '
        'xmlns:xr="http://schemas.microsoft.com/office/spreadsheetml/2014/revision" '
        'xmlns:xr2="http://schemas.microsoft.com/office/spreadsheetml/2015/revision2" '
        'xmlns:xr3="http://schemas.microsoft.com/office/spreadsheetml/2016/revision3" '
        'xr:uid="{00000000-0001-0000-0701-000000000000}">'
        '<sheetPr><tabColor rgb="FFFFC000"/></sheetPr>'
        f'<dimension ref="D5:W{last_row}"/>'
        '<sheetViews><sheetView showGridLines="0" workbookViewId="0"/></sheetViews>'
        '<sheetFormatPr defaultColWidth="5.5703125" defaultRowHeight="15" x14ac:dyDescent="0.25"/>'
        f'<cols>{cols}</cols>'
        f'<sheetData>{"".join(sheet_rows)}</sheetData>'
        '<autoFilter ref="D7:W7"/>'
        '<pageMargins left="0.7" right="0.7" top="0.75" bottom="0.75" header="0.511811023622047" footer="0.511811023622047"/>'
        '<pageSetup paperSize="9" orientation="landscape" horizontalDpi="300" verticalDpi="300"/>'
        '</worksheet>'
    )


def _load_datastore() -> tuple[dict[str, list[dict[str, str]]], dict[tuple[str, str], dict[str, str]]]:
    drugs: dict[str, list[dict[str, str]]] = defaultdict(list)
    metrics: dict[tuple[str, str], dict[str, str]] = defaultdict(dict)
    with (EXPORT / "peer_drug.csv").open(newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            drugs[row["section"]].append(row)
    with (EXPORT / "peer_metric.csv").open(newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            metrics[(row["section"], row["col"])][row["metric"]] = row.get("value", "")
    for section in drugs:
        drugs[section].sort(key=lambda r: _col_num(r["col"]))
    return drugs, metrics


def _pipeline_programs(path: Path, ticker: str) -> list[dict[str, str]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        ws = wb["Pipeline"]
        programs: list[dict[str, str]] = []
        current_drug = ""
        current_desc = ""
        for r in range(9, min(ws.max_row, 240) + 1):
            a = ws.cell(r, 1).value
            c = ws.cell(r, 3).value
            d = ws.cell(r, 4).value
            if a == "X" and (c is None or c == "") and isinstance(d, str) and not d.startswith("="):
                current_desc = d
                current_drug = d.split(" ", 1)[0].strip()
                continue
            # Indications are enumerated from the Market Share rows: the Pipeline
            # no longer carries visible TAM rows (TAM is inlined from the data
            # center). The indication label lives in the MS row's D formula suffix
            # (=D{n}&" <IND> Market Share") and the rating is the MS row's C label.
            if current_drug and isinstance(d, str) and "Market Share" in d:
                m = re.search(r'&\s*"\s*(.*?)\s*Market Share"', d)
                if not m:
                    continue
                indication = m.group(1).strip() or "All"
                rating = c.replace(" Growth", "").strip() if isinstance(c, str) else ""
                programs.append({
                    "indication": indication,
                    "drug": current_drug,
                    "ticker": f"{ticker.upper()} US Equity",
                    "rating": rating,
                    "status": "Clinical",
                    "mechanism": current_desc,
                    "line": "",
                    "n": "",
                    "orr": "",
                    "cr": "",
                    "pfs": "",
                    "os": "",
                    "safety": "",
                    "source": "Pipeline / approved assumptions",
                    "is_company": "1",
                })
        return programs
    finally:
        wb.close()


# ── Producer: company-row Peer View clinical data from research readouts ──────
# The target company's OWN drug rows would otherwise ship with blank clinical
# cells because _pipeline_programs hard-blanks line/n/orr/cr/pfs/os/safety.  The
# data already exists in the ticker's own PEER_VIEW readouts, so we materialize
# artifacts/{TICKER}/{TICKER}_peer_company_data.json here (BEFORE
# _company_peer_data reads it) and let _merge_company_data fill the real cells.
# A hand/LLM-authored JSON that is NEWER than the reports is never overwritten
# (mtime check), so the curated MOLN/CMPX data is preserved.

_BLANK_CELL_VALUES = {"", "/", "-", "—", "–"}
_NR_CELL_VALUES = {"N/A", "NA", "NR", "TBD", "PENDING", "UNKNOWN"}


def _cell_value(value: Any) -> str:
    """Raw readout value for a JSON cell; blanks/dashes collapse to ''."""
    v = str(value if value is not None else "").strip()
    return "" if v in _BLANK_CELL_VALUES else v


def _real_value(value: Any) -> str:
    """Value that counts as genuine clinical data (drops '/', 'NR', 'N/A', ...)."""
    v = _cell_value(value)
    return "" if v.upper() in _NR_CELL_VALUES else v


def _is_company_readout(readout: Any, ticker: str, asset_names: set[str]) -> bool:
    """True when a PEER_VIEW readout is the ticker's own asset (not a peer).

    Requires exact normalized-ticker equality (mirroring update_peer_database.
    _is_company_drug) so peers whose prose merely mentions the ticker are not
    captured; also accepts a readout whose drug is a known pipeline asset.
    """
    tk = _norm_section(getattr(readout, "ticker", ""))
    if tk and tk == _norm_section(ticker):
        return True
    dn = _norm_section(getattr(readout, "drug_name", ""))
    return bool(dn) and dn in asset_names


def _readout_to_company_row(readout: Any, indication: str) -> dict[str, str]:
    safety = (_cell_value(getattr(readout, "geq_g3_sae_pct", ""))
              or _cell_value(getattr(readout, "geq_g3_clinical_ae", "")))
    return {
        "drug": str(getattr(readout, "drug_name", "")).strip(),
        "indication": indication.strip(),
        "rating": _cell_value(getattr(readout, "rating", "")),
        "status": (_cell_value(getattr(readout, "result", ""))
                   or _cell_value(getattr(readout, "phase", ""))),
        "mechanism": _cell_value(getattr(readout, "innovation", "")),
        "line": _cell_value(getattr(readout, "treatment_line", "")),
        "n": _cell_value(getattr(readout, "n_patients", "")),
        "orr": _cell_value(getattr(readout, "orr", "")),
        "cr": _cell_value(getattr(readout, "cr", "")),
        "pfs": _cell_value(getattr(readout, "median_pfs", "")),
        "os": _cell_value(getattr(readout, "median_os", "")),
        "safety": safety,
        "source": _cell_value(getattr(readout, "source", "")),
    }


def _clinical_richness(readout: Any) -> int:
    """Count of genuinely-populated clinical fields (for best-readout ranking)."""
    fields = ("treatment_line", "n_patients", "orr", "cr", "median_pfs",
              "median_os", "geq_g3_sae_pct", "geq_g3_clinical_ae")
    return sum(1 for f in fields if _real_value(getattr(readout, f, "")))


def _company_rows_from_reports(reports: list[Path], ticker: str,
                               asset_names: set[str]) -> list[dict[str, str]]:
    """Pick the richest company readout per (drug, indication) -> override rows."""
    from fill.fill_peer_views import parse_peer_view_blocks

    # key -> (richness, readout, indication)
    best: dict[tuple[str, str], tuple[int, Any, str]] = {}
    for path in reports:
        try:
            text = path.read_text(encoding="utf-8", errors="ignore")
        except OSError:
            continue
        for indication, readouts in parse_peer_view_blocks(text).items():
            for readout in readouts:
                if not str(getattr(readout, "drug_name", "")).strip():
                    continue
                if not _is_company_readout(readout, ticker, asset_names):
                    continue
                key = (_norm_section(readout.drug_name), _norm_section(indication))
                score = _clinical_richness(readout)
                prev = best.get(key)
                if prev is None or score > prev[0]:
                    best[key] = (score, readout, indication)

    rows = [_readout_to_company_row(readout, indication)
            for _score, readout, indication in best.values()]
    rows.sort(key=lambda r: (r["drug"], r["indication"]))
    return rows


def ensure_company_peer_data_from_reports(
    report_dir: Path,
    ticker: str,
    asset_names: set[str] | None = None,
) -> Path | None:
    """Materialize company clinical rows before Pipeline generation.

    This report-only entry point lets the judgement pass consume the same
    ``peer_company_data`` later used by the delivered Peer View tab.  A newer
    curated JSON always wins and is never overwritten.
    """
    out_path = ARTIFACTS / ticker.upper() / f"{ticker.upper()}_peer_company_data.json"
    if not report_dir.exists():
        return out_path if out_path.exists() else None
    reports = sorted(set(
        report_dir.glob(f"{ticker}_*_research_*.md")
    ) | set(
        report_dir.glob(f"{ticker.upper()}_*_research_*.md")
    ))
    if not reports:
        return out_path if out_path.exists() else None
    newest_report = max(r.stat().st_mtime for r in reports)
    if out_path.exists() and out_path.stat().st_mtime >= newest_report:
        # A hand/LLM-authored (or previously generated) JSON at least as new as
        # the reports wins — do not clobber curated data.
        return out_path
    normalized_assets = {_norm_section(name) for name in (asset_names or set()) if name}
    rows = _company_rows_from_reports(reports, ticker, normalized_assets)
    if not rows:
        return out_path if out_path.exists() else None
    out_path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "source": (f"{ticker.upper()} company-row Peer View clinical data "
                   "(auto-generated from research PEER_VIEW readouts)"),
        "as_of": time.strftime("%Y-%m-%d"),
        "generated": True,
        "rows": rows,
    }
    out_path.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    print(f"Company peer data generated: {len(rows)} rows -> {out_path}")
    return out_path


def _ensure_company_peer_data(path: Path, ticker: str,
                              programs: list[dict[str, str]]) -> None:
    """Workbook-flow wrapper for the report-only company-data producer."""
    asset_names = {p.get("drug", "") for p in programs if p.get("drug")}
    ensure_company_peer_data_from_reports(
        path.parent / "pipeline_base4", ticker, asset_names
    )


def _company_peer_data(ticker: str) -> dict[tuple[str, str], dict[str, str]]:
    """Load researched company-row Peer View overrides for the current ticker."""
    path = ARTIFACTS / ticker.upper() / f"{ticker.upper()}_peer_company_data.json"
    if not path.exists():
        return {}
    data = json.loads(path.read_text(encoding="utf-8"))
    out: dict[tuple[str, str], dict[str, str]] = {}
    for row in data.get("rows", []):
        drug = str(row.get("drug", "")).strip()
        indication = str(row.get("indication", "")).strip()
        if not drug or not indication:
            continue
        clean = {k: "" if v is None else str(v) for k, v in row.items()}
        out[(drug.upper(), indication.upper())] = clean
    return out


def _merge_company_data(programs: list[dict[str, str]], ticker: str) -> list[dict[str, str]]:
    overrides = _company_peer_data(ticker)
    if not overrides:
        return programs
    merged: list[dict[str, str]] = []
    for program in programs:
        key = (program["drug"].upper(), program["indication"].upper())
        override = overrides.get(key)
        if override:
            updated = dict(program)
            for field in ("status", "mechanism", "line", "n", "orr", "cr", "pfs", "os", "safety", "source"):
                if override.get(field) not in (None, ""):
                    updated[field] = override[field]
            if override.get("rating") not in (None, ""):
                updated["rating"] = override["rating"]
            merged.append(updated)
        else:
            merged.append(program)
    return merged


def _metric(metrics: dict[str, str], *names: str) -> str:
    for name in names:
        value = metrics.get(name)
        if value not in (None, "", "/"):
            return str(value)
    return ""


def _date_value(value: str) -> str:
    if not value:
        return ""
    text = str(value)
    if re.fullmatch(r"\d+(?:\.0)?", text):
        try:
            # Excel serial date.
            from datetime import datetime, timedelta
            d = datetime(1899, 12, 30) + timedelta(days=float(text))
            return d.strftime("%Y-%m-%d")
        except Exception:
            return text
    return text[:10] if re.match(r"\d{4}-\d{2}-\d{2}", text) else text


def _norm_section(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(s or "").lower())


def _resolve_section(indication: str, drugs: dict[str, list[dict[str, str]]]) -> str | None:
    """Map a Pipeline indication to a datastore Peer View section generically.

    Works for ANY ticker (not just CMPX's BTC/SCLC): tries exact name, then the
    supplementary SECTION_ALIASES abbreviations, then normalized equality, then
    substring containment either direction (longest match wins).

    Substring containment is gated to codes of at least 3 normalized chars.  A
    bare 2-letter Pipeline code (e.g. BHVN's "AM"=Acute Migraine, "PD"=
    Parkinson's) spuriously matches unrelated stale sections ("Melanoma
    Post-PD1 (NCAM+)" normalizes to "...ncam", which contains "am"), leaking a
    prior ticker's peers.  Requiring len(ni) >= 3 blocks those false positives
    while preserving CMPX's real 3+ char containment codes (CRC/RCC/HCC/cHL/
    mUC/NSCLC/TNBC/mCRPC).  When nothing resolves the caller omits the row
    rather than emitting a blank placeholder.
    """
    if not indication:
        return None
    if indication in drugs:
        return indication
    alias = SECTION_ALIASES.get(indication)
    if alias and alias in drugs:
        return alias
    ni = _norm_section(indication)
    for sec in drugs:
        if _norm_section(sec) == ni:
            return sec
    if len(ni) >= 3:
        cands = [sec for sec in drugs
                 if ni in _norm_section(sec) or _norm_section(sec) in ni]
        if cands:
            return max(cands, key=lambda s: len(_norm_section(s)))
    return None


def _peer_rows_for_indication(indication: str, drugs: dict[str, list[dict[str, str]]],
                              metrics: dict[tuple[str, str], dict[str, str]],
                              ticker: str) -> list[dict[str, str]]:
    section = _resolve_section(indication, drugs)
    if not section or section not in drugs:
        # Unresolved: omit the row entirely instead of emitting an all-blank
        # "Database peer section pending" placeholder (req12: no blank key
        # cells).  The company's own program row is still written by the caller.
        return []
    out: list[dict[str, str]] = []
    for drug in drugs[section]:
        # The delivered table already starts each indication with the current
        # company's researched program row.  The shared datastore also carries
        # that program, so omit the same owner here to avoid a duplicate CMPX /
        # CMPX US Equity row while retaining every true comparator.
        owner = _norm_section(drug.get("ticker", ""))
        own = _norm_section(ticker)
        if owner in {own, own + "usequity"}:
            continue
        m = metrics.get((section, drug["col"]), {})
        out.append({
            "indication": indication,
            "drug": drug.get("drug", ""),
            "ticker": drug.get("ticker", ""),
            "rating": drug.get("rating", ""),
            "status": _metric(m, "Result", "Readout Phase"),
            "mechanism": _metric(m, "Innovation"),
            "line": _metric(m, "Treatment Line"),
            "n": _metric(m, "Evaluable Patients"),
            "orr": _metric(m, "ORR"),
            "cr": _metric(m, "CR"),
            "pfs": _metric(m, "Median PFS", "Median rPFS"),
            "os": _metric(m, "Median OS"),
            "safety": _metric(m, "≥G3 SAE/Patients", "≥G3 clinical AE"),
            "source": f"{section} / {_date_value(_metric(m, 'Date'))}",
            "is_company": "0",
        })
    return out


def build_peer_view(path: Path, ticker: str, make_backup: bool = True) -> int:
    if not (EXPORT / "peer_drug.csv").exists() or not (EXPORT / "peer_metric.csv").exists():
        raise FileNotFoundError(f"Missing datastore peer exports under {EXPORT}")
    raw_programs = _pipeline_programs(path, ticker)
    # Produce the company-row clinical JSON from the ticker's own PEER_VIEW
    # readouts before _merge_company_data consumes it (fills real cells; skips
    # when a newer curated JSON exists).
    _ensure_company_peer_data(path, ticker, raw_programs)
    programs = _merge_company_data(raw_programs, ticker)
    drugs, metrics = _load_datastore()

    rows: list[dict[str, str]] = []
    for program in programs:
        rows.append(program)
        rows.extend(_peer_rows_for_indication(
            program["indication"], drugs, metrics, ticker
        ))

    if make_backup:
        backup = path.with_name(f"{path.stem}_pre_peer_view_summary_{time.strftime('%Y%m%d_%H%M%S')}.xlsx")
        shutil.copy2(path, backup)
        print(f"Backup: {backup}")

    parts, order = _read_parts(path)
    peer_view_sheet_path = _sheet_zip_path(parts, "Peer View")
    peer_views_sheet_path = _optional_sheet_zip_path(parts, "Peer Views")
    add_shared_string, finalize_shared_strings = _shared_string_adder(parts)
    parts[peer_view_sheet_path] = _clean_peer_view_xml(rows, ticker, add_shared_string).encode("utf-8")
    if peer_views_sheet_path:
        parts[peer_views_sheet_path] = _clean_peer_views_datacenter_xml(
            drugs, metrics, add_shared_string
        ).encode("utf-8")
    finalize_shared_strings()
    _write_parts(path, parts, order)

    def to_windows_path(p: Path) -> str:
        text = str(p)
        m = re.match(r"^/mnt/([a-zA-Z])/(.*)$", text)
        if not m:
            return text
        rest = m.group(2).replace("/", "\\")
        return f"{m.group(1).upper()}:\\{rest}"

    try:
        subprocess.run(
            [
                "powershell.exe",
                "-NoProfile",
                "-ExecutionPolicy",
                "Bypass",
                "-File",
                str(Path(__file__).with_name("excel_repair_saveas.ps1")),
                "-Path",
                to_windows_path(path),
            ],
            check=True,
            timeout=600,
        )
    except Exception as exc:
        print(f"Excel COM repair skipped/failed ({exc}); continuing with OOXML repair later.")
    print(f"Peer View summary rebuilt: {len(programs)} company rows, {len(rows) - len(programs)} peer/gap rows")
    return len(rows)


def main() -> None:
    parser = argparse.ArgumentParser(description="Build ticker-specific Peer View summary")
    parser.add_argument("--ticker", required=True)
    parser.add_argument("--path")
    parser.add_argument("--no-backup", action="store_true")
    args = parser.parse_args()
    path = Path(args.path) if args.path else _default_path(args.ticker)
    if not path.exists():
        raise FileNotFoundError(path)
    build_peer_view(path, args.ticker, make_backup=not args.no_backup)


if __name__ == "__main__":
    main()
