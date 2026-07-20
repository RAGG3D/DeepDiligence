#!/usr/bin/env python3
"""One-time RJConv. terminology migration for current models and artifacts."""
from __future__ import annotations

import json
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import duckdb
from openpyxl import load_workbook


REPO = Path(__file__).resolve().parents[1]
if str(REPO) not in sys.path:
    sys.path.insert(0, str(REPO))

from datastore.research_fact_store import (  # noqa: E402
    DEFAULT_DB,
    DEFAULT_SEED,
    FACT_COLUMNS,
    _ensure_table,
    normalize_fact,
)
from tools.score_test_catalyst_event import frozen_sheet_digest  # noqa: E402


TICKERS = ("CMPX", "MOLN", "TARA")
OLD_METRIC = "highest_conviction_scenario_peak_test"
NEW_METRIC = "highest_rjconv_scenario_peak_test"
KEY_MAP = {
    "highest_conviction_scenario_test": "highest_rjconv_scenario_test",
    "highest_conviction_scenario_count": "highest_rjconv_scenario_count",
    "highest_total_conviction": "highest_rjconv",
    "total_conviction": "rjconv",
}


def read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, payload: Any) -> None:
    path.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )


def migrate_keys(value: Any) -> Any:
    if isinstance(value, list):
        return [migrate_keys(item) for item in value]
    if not isinstance(value, dict):
        if value == "max raw Table-3 conviction product; all exact ties retained":
            return "max RJConv. (raw Table-3 conviction product); all exact ties retained"
        return value
    migrated: dict[str, Any] = {}
    for key, item in value.items():
        migrated[KEY_MAP.get(key, key)] = migrate_keys(item)
    return migrated


def verify_workbook_labels(path: Path) -> None:
    wb = load_workbook(path, data_only=False, read_only=True, keep_links=True)
    try:
        for sheet_name in ("Catalyst", "Test-ASCO2026"):
            if sheet_name not in wb.sheetnames:
                raise RuntimeError(f"{path}: missing {sheet_name}")
            ws = wb[sheet_name]
            if ws["F7"].value != "RJConv.":
                raise RuntimeError(f"{path}: {sheet_name}!F7 is not RJConv.")
    finally:
        wb.close()


def migrate_artifacts(ticker: str, migrated_at: str) -> None:
    folder = REPO / "artifacts" / ticker
    catalyst_manifest_path = folder / f"{ticker}_catalyst_manifest.json"
    catalyst_manifest = read_json(catalyst_manifest_path)
    catalyst_manifest["framework_version"] = 7
    layout = catalyst_manifest["layout"]
    if "conviction_col" in layout:
        layout["rjconv_col"] = layout.pop("conviction_col")
    write_json(catalyst_manifest_path, catalyst_manifest)

    state_path = folder / f"{ticker}_catalyst_active_state.json"
    state = read_json(state_path)
    state["framework_version"] = 7
    write_json(state_path, state)

    workbook = Path(f"/mnt/c/Users/yzsun/Desktop/DD/{ticker}/DCF {ticker}.xlsx")
    verify_workbook_labels(workbook)

    test_manifest_path = folder / f"{ticker}_test_ASCO2026_manifest.json"
    test_manifest = migrate_keys(read_json(test_manifest_path))
    boundary = int(test_manifest["blind_snapshot"]["blind_sheet_original_max_row"])
    digest, digest_boundary = frozen_sheet_digest(workbook, "Test-ASCO2026", boundary)
    if digest_boundary != boundary:
        raise RuntimeError(f"{ticker}: frozen digest boundary changed")
    test_manifest["blind_snapshot"]["blind_sheet_digest"] = digest
    post_audit = test_manifest.get("post_release_scoring", {}).get("audit", {})
    if post_audit:
        post_audit["blind_sheet_digest"] = digest
    test_manifest["rjconv_naming_migration"] = {
        "migrated_at": migrated_at,
        "scope": "terminology_only",
        "old_main_label": "Conv.",
        "new_main_label": "RJConv.",
        "formula_or_value_change": False,
        "table3_target_conviction_label_retained": True,
    }
    write_json(test_manifest_path, test_manifest)

    score_path = folder / f"{ticker}_test_ASCO2026_post_release_score.json"
    write_json(score_path, migrate_keys(read_json(score_path)))


def migrate_database() -> tuple[int, int]:
    seed = read_json(DEFAULT_SEED)
    migrated: list[dict[str, Any]] = []
    changed: list[dict[str, Any]] = []
    for item in seed.get("facts", []):
        if item.get("metric") == OLD_METRIC:
            renamed = dict(item)
            renamed["metric"] = NEW_METRIC
            renamed.pop("fact_id", None)
            normalized = normalize_fact(renamed, str(renamed.get("context_ticker") or ""))
            migrated.append(normalized)
            changed.append(normalized)
        else:
            migrated.append(item)
    deduped = {item["fact_id"]: item for item in migrated}
    facts = sorted(deduped.values(), key=lambda item: (
        item["context_ticker"], item["subject"], item["metric"],
        item["as_of_date"], item["fact_id"],
    ))
    write_json(DEFAULT_SEED, {"facts": facts})

    connection = duckdb.connect(str(DEFAULT_DB))
    try:
        _ensure_table(connection)
        connection.execute("DELETE FROM research_fact WHERE metric=?", [OLD_METRIC])
        if changed:
            connection.executemany(
                "INSERT OR REPLACE INTO research_fact VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                [tuple(item[column] for column in FACT_COLUMNS) for item in changed],
            )
        old_count = connection.execute(
            "SELECT count(*) FROM research_fact WHERE metric=?", [OLD_METRIC]
        ).fetchone()[0]
        new_count = connection.execute(
            "SELECT count(*) FROM research_fact WHERE metric=?", [NEW_METRIC]
        ).fetchone()[0]
    finally:
        connection.close()
    return int(old_count), int(new_count)


def main() -> int:
    migrated_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    for ticker in TICKERS:
        migrate_artifacts(ticker, migrated_at)
    old_count, new_count = migrate_database()
    if old_count:
        raise RuntimeError(f"old database metric remains: {old_count}")
    print(json.dumps({
        "tickers": list(TICKERS),
        "old_database_metric_count": old_count,
        "new_database_metric_count": new_count,
    }, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
