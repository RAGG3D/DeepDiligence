#!/usr/bin/env python3
"""Finalize a DCF workbook by removing build-time data-center tabs.

The workbook can use data-center sheets while it is being assembled, but the
delivered model should not carry full TAM / Peer Views databases.  Before the
tabs are deleted this script:
  * replaces WELCOME! ticker links with a literal ticker label,
  * inlines legacy TAM maturity and COGS parameter references in Pipeline,
  * marks database-sourced Pipeline data in dark green, except Peer View.

It edits worksheet/styles XML directly, then uses Excel COM only for sheet
deletion so workbook relationships remain valid.
"""

from __future__ import annotations

import argparse
import html
import re
import shutil
import subprocess
import time
import zipfile
from copy import deepcopy
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl


NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
NS_CT = "http://schemas.openxmlformats.org/package/2006/content-types"
DELETE_SHEETS = ["WELCOME!", "Playgroud", "TAM Solid+MM", "TAM Solid", "TAM Blood", "Peer Views"]
GREEN = "FF006100"
COGS_RATE = 0.30
_PLATEAU = [1.056] * 21
MATURITY = {
    551: [0.193, 0.332, 0.423, 0.541, 0.669, 0.801, 0.913, 1.0] + _PLATEAU,
    552: [0.003, 0.078, 0.175, 0.342, 0.586, 0.795, 0.921, 1.0] + _PLATEAU,
    553: [0.046, 0.129, 0.233, 0.410, 0.554, 0.777, 0.930, 1.0] + _PLATEAU,
}


ET.register_namespace("", NS_MAIN)
ET.register_namespace("r", NS_R)


def _default_path(ticker: str) -> Path:
    return Path(f"/mnt/c/Users/yzsun/Desktop/DD/{ticker}/DCF {ticker}.xlsx")


def _to_windows_path(path: Path) -> str:
    text = str(path)
    m = re.match(r"^/mnt/([a-zA-Z])/(.*)$", text)
    if not m:
        return text
    return f"{m.group(1).upper()}:\\{m.group(2).replace('/', '\\')}"


def _col_letter(index: int) -> str:
    out = ""
    while index:
        index, rem = divmod(index - 1, 26)
        out = chr(65 + rem) + out
    return out


def _col_num(col: str) -> int:
    out = 0
    for ch in col:
        out = out * 26 + ord(ch) - 64
    return out


def _read_parts(path: Path) -> tuple[dict[str, bytes], list[str]]:
    with zipfile.ZipFile(path, "r") as zf:
        bad = zf.testzip()
        if bad:
            raise RuntimeError(f"zip CRC failure in {bad}")
        return {name: zf.read(name) for name in zf.namelist()}, zf.namelist()


def _write_parts(path: Path, parts: dict[str, bytes], order: list[str]) -> None:
    tmp = path.with_suffix(".~trimtabs.xlsx")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        written = set()
        for name in order:
            if name in parts:
                zout.writestr(name, parts[name])
                written.add(name)
        for name, data in parts.items():
            if name not in written:
                zout.writestr(name, data)
    tmp.replace(path)


def _sheet_paths(parts: dict[str, bytes]) -> dict[str, str]:
    wb = ET.fromstring(parts["xl/workbook.xml"])
    rels = ET.fromstring(parts["xl/_rels/workbook.xml.rels"])
    relmap = {rel.get("Id"): rel.get("Target") for rel in rels}
    out: dict[str, str] = {}
    for sheet in wb.findall(f".//{{{NS_MAIN}}}sheet"):
        name = html.unescape(sheet.get("name", ""))
        target = relmap.get(sheet.get(f"{{{NS_R}}}id"))
        if not target:
            continue
        target = target.lstrip("/")
        out[name] = target if target.startswith("xl/") else "xl/" + target
    return out


def _cell_style(xml: str, addr: str, default: str = "0") -> str:
    m = re.search(
        rf'<c\b(?=[^>]*\br="{addr}")[^>]*/>|'
        rf'<c\b(?=[^>]*\br="{addr}")[^>]*>.*?</c>',
        xml,
        re.S,
    )
    if not m:
        return default
    sm = re.search(r'\bs="(\d+)"', m.group(0))
    return sm.group(1) if sm else default


def _text_cell(addr: str, text: str, style: str) -> str:
    return f'<c r="{addr}" s="{style}" t="inlineStr"><is><t>{html.escape(text, quote=False)}</t></is></c>'


def _replace_welcome_refs(xml: str, ticker_label: str) -> tuple[str, int]:
    changed = 0

    def repl(match: re.Match[str]) -> str:
        nonlocal changed
        cell = match.group(0)
        if "WELCOME!" not in cell:
            return cell
        fm = re.search(r"<f[^>]*>(.*?)</f>", cell, re.S)
        if not fm:
            return cell
        formula = html.unescape(fm.group(1)).strip()
        if formula.startswith("="):
            formula = formula[1:]
        if formula != "'WELCOME!'!$B$18":
            return cell
        addr = re.search(r'\br="([^"]+)"', cell).group(1)
        changed += 1
        return _text_cell(addr, ticker_label, _cell_style(cell, addr, "0"))

    xml = re.sub(
        r'<c\b[^>]*?/>|<c\b[^>]*>.*?</c>', repl, xml, flags=re.S
    )
    return xml, changed


def _load_workbook_metadata(path: Path, ticker: str) -> tuple[str, dict[int, list[float]], float, list[str]]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ticker_label = f"{ticker.upper()} US Equity"
        if "WELCOME!" in wb.sheetnames and wb["WELCOME!"]["B18"].value:
            ticker_label = str(wb["WELCOME!"]["B18"].value)
        curves = {k: list(v) for k, v in MATURITY.items()}
        solid = "TAM Solid+MM" if "TAM Solid+MM" in wb.sheetnames else "TAM Solid" if "TAM Solid" in wb.sheetnames else None
        if solid:
            ws = wb[solid]
            for row in (551, 552, 553):
                vals = [ws.cell(row, c).value for c in range(6, 35)]
                if all(isinstance(v, (int, float)) for v in vals):
                    curves[row] = [float(v) for v in vals]
            cogs = ws["P562"].value
            cogs_rate = float(cogs) if isinstance(cogs, (int, float)) else COGS_RATE
        else:
            cogs_rate = COGS_RATE
        existing = list(wb.sheetnames)
        return ticker_label, curves, cogs_rate, existing
    finally:
        wb.close()


def _pipeline_database_coords(path: Path) -> list[str]:
    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    coords: list[str] = []
    try:
        if "Pipeline" not in wb.sheetnames:
            return coords
        ws = wb["Pipeline"]
        for row in range(1, ws.max_row + 1):
            d = ws.cell(row, 4).value
            if not isinstance(d, str):
                continue
            if " TAM" in d and "Market Share" not in d:
                coords.extend(f"{_col_letter(c)}{row}" for c in range(_col_num("F"), _col_num("AH") + 1))
            if "Market Share" in d:
                coords.append(f"C{row}")
        return coords
    finally:
        wb.close()


def _choose_formula(index_expr: str, values: list[float]) -> str:
    return "CHOOSE(" + index_expr + "," + ",".join(f"{v:.6g}" for v in values) + ")"


def _inline_tam_parameters(xml: str, curves: dict[int, list[float]], cogs_rate: float) -> tuple[str, int]:
    changed = 0
    idx_pattern = (
        r"INDEX\('TAM Solid(?:\+MM)?'!\$F\$(551|552|553):\$AH\$\1,"
        r"(MIN\(COLUMN\([A-Z]{1,3}1\)-MATCH\(5,\$F\$\d+:\$AH\$\d+,0\)-COLUMN\(\$F\$1\)\+2,29\))\)"
    )

    def formula_repl(match: re.Match[str]) -> str:
        nonlocal changed
        open_tag, raw, close_tag = match.group(1), match.group(2), match.group(3)
        formula = html.unescape(raw)

        def idx_repl(m: re.Match[str]) -> str:
            nonlocal changed
            row = int(m.group(1))
            changed += 1
            return _choose_formula(m.group(2), curves.get(row, MATURITY[row]))

        formula = re.sub(idx_pattern, idx_repl, formula)
        formula, n = re.subn(r"'TAM Solid(?:\+MM)?'!\$P\$562", f"{cogs_rate:.6g}", formula)
        changed += n
        return f"{open_tag}{html.escape(formula, quote=False)}{close_tag}"

    xml = re.sub(r"(<f\b(?![^>]*/>)[^>]*>)(.*?)(</f>)", formula_repl, xml, flags=re.S)
    return xml, changed


def _literal_pipeline_year_headers(xml: str) -> tuple[str, int]:
    """Detach Pipeline F7:AH7 from deleted TAM helper tabs."""
    changed = 0
    for offset, year in enumerate(range(2010, 2039)):
        addr = f"{_col_letter(6 + offset)}7"
        pattern = rf'<c\b(?=[^>]*\br="{addr}")[^>]*>.*?</c>'
        match = re.search(pattern, xml, re.S)
        if not match:
            continue
        cell = match.group(0)
        style = _cell_style(cell, addr, "0")
        replacement = f'<c r="{addr}" s="{style}"><v>{year}</v></c>'
        if cell != replacement:
            xml = xml[:match.start()] + replacement + xml[match.end():]
            changed += 1
    return xml, changed


def _green_style_map(styles_xml: str, style_ids: set[str]) -> tuple[str, dict[str, str]]:
    if not style_ids:
        return styles_xml, {}
    root = ET.fromstring(styles_xml)
    ns = f"{{{NS_MAIN}}}"
    fonts = root.find(f"{ns}fonts")
    cellxfs = root.find(f"{ns}cellXfs")
    if fonts is None or cellxfs is None:
        return styles_xml, {}

    out: dict[str, str] = {}
    for sid in sorted(style_ids, key=lambda x: int(x)):
        idx = int(sid)
        if idx >= len(cellxfs):
            continue
        xf = cellxfs[idx]
        font_id = int(xf.get("fontId", "0"))
        if font_id >= len(fonts):
            continue
        font = deepcopy(fonts[font_id])
        for child in list(font):
            if child.tag == f"{ns}color":
                font.remove(child)
        font.append(ET.Element(f"{ns}color", {"rgb": GREEN}))
        new_font_id = len(fonts)
        fonts.append(font)
        fonts.set("count", str(len(fonts)))

        new_xf = deepcopy(xf)
        new_xf.set("fontId", str(new_font_id))
        new_xf.set("applyFont", "1")
        new_style_id = len(cellxfs)
        cellxfs.append(new_xf)
        cellxfs.set("count", str(len(cellxfs)))
        out[sid] = str(new_style_id)

    xml = (
        b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        + ET.tostring(root, encoding="utf-8")
    ).decode("utf-8")
    return xml, out


def _set_cell_style(xml: str, addr: str, style_id: str) -> tuple[str, bool]:
    pattern = rf'(<c\b(?=[^>]*\br="{addr}")[^>]*)(/?>)'

    def repl(match: re.Match[str]) -> str:
        open_tag, end = match.group(1), match.group(2)
        if re.search(r'\bs="\d+"', open_tag):
            open_tag = re.sub(r'\bs="\d+"', f's="{style_id}"', open_tag, count=1)
        else:
            open_tag += f' s="{style_id}"'
        return open_tag + end

    new, n = re.subn(pattern, repl, xml, count=1)
    return new, bool(n)


def _mark_pipeline_database_cells(parts: dict[str, bytes], pipeline_path: str, coords: list[str]) -> int:
    if not coords:
        return 0
    xml = parts[pipeline_path].decode("utf-8", "ignore")
    old_styles = {_cell_style(xml, addr, "0") for addr in coords}
    styles_xml, style_map = _green_style_map(parts["xl/styles.xml"].decode("utf-8", "ignore"), old_styles)
    parts["xl/styles.xml"] = styles_xml.encode("utf-8")

    changed = 0
    for addr in coords:
        old = _cell_style(xml, addr, "0")
        new = style_map.get(old)
        if not new:
            continue
        xml, ok = _set_cell_style(xml, addr, new)
        changed += int(ok)
    parts[pipeline_path] = xml.encode("utf-8")
    return changed


def _welcome_ref_cells(parts: dict[str, bytes],
                       sheet_paths: dict[str, str]) -> list[str]:
    """Return ``Sheet!A1`` cells whose sole formula is WELCOME!$B$18."""
    refs: list[str] = []
    for sheet, sheet_path in sheet_paths.items():
        if sheet_path not in parts or not sheet_path.startswith("xl/worksheets/"):
            continue
        xml = parts[sheet_path].decode("utf-8", "ignore")
        for match in re.finditer(
            r'<c\b[^>]*?/>|<c\b[^>]*>.*?</c>', xml, flags=re.S
        ):
            cell = match.group(0)
            if "WELCOME!" not in cell:
                continue
            formula_match = re.search(r"<f[^>]*>(.*?)</f>", cell, re.S)
            addr_match = re.search(r'\br="([^"]+)"', cell)
            if not formula_match or not addr_match:
                continue
            formula = html.unescape(formula_match.group(1)).strip().lstrip("=")
            if formula == "'WELCOME!'!$B$18":
                refs.append(f"{sheet}!{addr_match.group(1)}")
    return refs


def _delete_sheets_with_excel(path: Path, sheets: list[str], ticker_label: str,
                               welcome_refs: list[str],
                               green_cells: list[str]) -> None:
    if not sheets:
        return
    script = Path(__file__).with_name("delete_model_tabs.ps1")
    subprocess.run(
        [
            "powershell.exe",
            "-NoProfile",
            "-ExecutionPolicy",
            "Bypass",
            "-File",
            str(script),
            "-Path",
            _to_windows_path(path),
            "-SheetsCsv",
            "|".join(sheets),
            "-TickerLabel",
            ticker_label,
            "-WelcomeCellsCsv",
            "|".join(welcome_refs),
            "-GreenCellsCsv",
            "|".join(green_cells),
        ],
        check=True,
        timeout=600,
    )


def _mark_full_calc(parts: dict[str, bytes]) -> None:
    wb_xml = parts["xl/workbook.xml"].decode("utf-8", "ignore")
    if "<calcPr" in wb_xml:
        wb_xml = re.sub(r"<calcPr\b[^>]*/>", '<calcPr calcId="191029" fullCalcOnLoad="1"/>', wb_xml, count=1)
    else:
        wb_xml = wb_xml.replace("</workbook>", '<calcPr calcId="191029" fullCalcOnLoad="1"/></workbook>')
    parts["xl/workbook.xml"] = wb_xml.encode("utf-8")

    parts.pop("xl/calcChain.xml", None)
    ct = parts.get("[Content_Types].xml", b"").decode("utf-8", "ignore")
    if ct:
        ct = re.sub(r'<Override[^>]*/xl/calcChain\.xml[^>]*/>', "", ct)
        parts["[Content_Types].xml"] = ct.encode("utf-8")
    rels = parts.get("xl/_rels/workbook.xml.rels", b"").decode("utf-8", "ignore")
    if rels:
        rels = re.sub(r'<Relationship[^>]*calcChain[^>]*/>', "", rels)
        parts["xl/_rels/workbook.xml.rels"] = rels.encode("utf-8")


def _relationship_target_to_part(target: str) -> str:
    target = target.lstrip("/")
    return target if target.startswith("xl/") else f"xl/{target}"


def _worksheet_rels_path(sheet_part: str) -> str:
    prefix, name = sheet_part.rsplit("/", 1)
    return f"{prefix}/_rels/{name}.rels"


def _delete_sheets_ooxml(parts: dict[str, bytes], order: list[str], sheets_to_delete: list[str]) -> int:
    if not sheets_to_delete:
        return 0
    delete_set = set(sheets_to_delete)
    ET.register_namespace("", NS_MAIN)
    ET.register_namespace("r", NS_R)

    wb_root = ET.fromstring(parts["xl/workbook.xml"])
    rels_root = ET.fromstring(parts["xl/_rels/workbook.xml.rels"])
    relmap = {rel.get("Id"): rel for rel in rels_root}
    sheets_el = wb_root.find(f"{{{NS_MAIN}}}sheets")
    if sheets_el is None:
        return 0

    before_names = [html.unescape(sheet.get("name", "")) for sheet in sheets_el.findall(f"{{{NS_MAIN}}}sheet")]
    delete_indices = {i for i, name in enumerate(before_names) if name in delete_set}
    removed_parts: set[str] = set()
    removed_rids: set[str] = set()
    deleted = 0

    for sheet in list(sheets_el.findall(f"{{{NS_MAIN}}}sheet")):
        name = html.unescape(sheet.get("name", ""))
        if name not in delete_set:
            continue
        rid = sheet.get(f"{{{NS_R}}}id")
        rel = relmap.get(rid)
        if rel is not None:
            target = rel.get("Target", "")
            part = _relationship_target_to_part(target)
            removed_parts.add(part)
            removed_parts.add(_worksheet_rels_path(part))
            rels_root.remove(rel)
            removed_rids.add(rid)
        sheets_el.remove(sheet)
        deleted += 1

    # Drop or shift local defined names so sheet-local metadata keeps pointing
    # to the same remaining sheet after deleted tabs are removed.
    defined_names = wb_root.find(f"{{{NS_MAIN}}}definedNames")
    if defined_names is not None:
        for dn in list(defined_names.findall(f"{{{NS_MAIN}}}definedName")):
            text = dn.text or ""
            local = dn.get("localSheetId")
            remove = any(f"'{name}'!" in text or f"{name}!" in text for name in delete_set)
            if local is not None:
                old_idx = int(local)
                if old_idx in delete_indices:
                    remove = True
                else:
                    dn.set("localSheetId", str(old_idx - sum(1 for i in delete_indices if i < old_idx)))
            if remove:
                defined_names.remove(dn)
        if len(list(defined_names)) == 0:
            wb_root.remove(defined_names)

    # Reset workbook view to a remaining first sheet.
    for view in wb_root.findall(f".//{{{NS_MAIN}}}workbookView"):
        view.set("activeTab", "0")
        if "firstSheet" in view.attrib:
            view.set("firstSheet", "0")

    ET.register_namespace("", NS_MAIN)
    ET.register_namespace("r", NS_R)
    parts["xl/workbook.xml"] = (
        b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        + ET.tostring(wb_root, encoding="utf-8")
    )
    ET.register_namespace("", NS_REL)
    parts["xl/_rels/workbook.xml.rels"] = (
        b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        + ET.tostring(rels_root, encoding="utf-8")
    )

    if "[Content_Types].xml" in parts:
        ct_root = ET.fromstring(parts["[Content_Types].xml"])
        removed_part_names = {"/" + part for part in removed_parts}
        for override in list(ct_root.findall(f"{{{NS_CT}}}Override")):
            if override.get("PartName") in removed_part_names:
                ct_root.remove(override)
        ET.register_namespace("", NS_CT)
        parts["[Content_Types].xml"] = (
            b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
            + ET.tostring(ct_root, encoding="utf-8")
        )

    for part in removed_parts:
        parts.pop(part, None)
    order[:] = [name for name in order if name in parts]
    _mark_full_calc(parts)
    return deleted


def _scan_for_deleted_refs(path: Path, deleted: list[str]) -> list[str]:
    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    hits: list[str] = []
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    value = cell.value
                    if not isinstance(value, str):
                        continue
                    for sheet in deleted:
                        if f"'{sheet}'!" in value or f"{sheet}!" in value:
                            hits.append(f"{ws.title}!{cell.coordinate}: {value[:160]}")
                            break
        return hits
    finally:
        wb.close()


def trim_model_tabs(path: Path, ticker: str, make_backup: bool = True) -> dict[str, int]:
    ticker_label, curves, cogs_rate, existing_sheets = _load_workbook_metadata(path, ticker)
    db_coords = _pipeline_database_coords(path)
    if make_backup:
        backup = path.with_name(f"{path.stem}_pre_trimtabs_{time.strftime('%Y%m%d_%H%M%S')}.xlsx")
        shutil.copy2(path, backup)
        print(f"Backup: {backup}")

    parts, order = _read_parts(path)
    sheet_paths = _sheet_paths(parts)
    report = {"welcome_refs": 0, "tam_param_refs": 0, "year_headers": 0,
              "green_cells": 0, "deleted_sheets": 0}

    welcome_refs = _welcome_ref_cells(parts, sheet_paths)
    report["welcome_refs"] = len(welcome_refs)

    if "Pipeline" in sheet_paths:
        pipeline_path = sheet_paths["Pipeline"]
        xml = parts[pipeline_path].decode("utf-8", "ignore")
        xml, n = _inline_tam_parameters(xml, curves, cogs_rate)
        xml, year_n = _literal_pipeline_year_headers(xml)
        parts[pipeline_path] = xml.encode("utf-8")
        report["tam_param_refs"] = n
        report["year_headers"] = year_n
    report["green_cells"] = len(db_coords)

    to_delete = [s for s in DELETE_SHEETS if s in existing_sheets and s != "Peer View"]
    # Keep all relationship/content-type graph surgery and the cell/style
    # mutations inside Excel.  Pure-OOXML sheet deletion, inlineStr conversion
    # of WELCOME links, and ElementTree style cloning each produced packages
    # that openpyxl/LibreOffice tolerated but Workbooks.Open rejected.
    # The current Pipeline already inlines TAM parameters (n==0); retain the
    # legacy formula-only fallback only when an old workbook still needs it.
    if report["tam_param_refs"] or report["year_headers"]:
        _mark_full_calc(parts)
        _write_parts(path, parts, order)
    _delete_sheets_with_excel(
        path, to_delete, ticker_label, welcome_refs, db_coords
    )
    report["deleted_sheets"] = len(to_delete)

    refs = _scan_for_deleted_refs(path, to_delete)
    if refs:
        print("References to deleted sheets remain:")
        for item in refs[:80]:
            print(f"  {item}")
        raise SystemExit(1)

    print(f"Trimmed model tabs: {report}")
    return report


def main() -> None:
    parser = argparse.ArgumentParser(description="Remove build-time data-center tabs from DCF workbook")
    parser.add_argument("--ticker", required=True)
    parser.add_argument("--path")
    parser.add_argument("--no-backup", action="store_true")
    args = parser.parse_args()
    path = Path(args.path) if args.path else _default_path(args.ticker)
    if not path.exists():
        raise FileNotFoundError(path)
    trim_model_tabs(path, args.ticker, make_backup=not args.no_backup)


if __name__ == "__main__":
    main()
