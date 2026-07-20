#!/usr/bin/env python3
"""
harden_formulas.py — wrap residual error-throwing formulas in IFERROR(…,0).

The DCF template's restated statements / valuation tabs carry many ratio and
lookup formulas that assume a revenue-generating company with complete data
(e.g. `=WorkingCapital/Revenue`, self-referential ramp helpers). For a
pre-commercial biotech these evaluate to 0/0 or a missed lookup and throw
#DIV/0! / #N/A, which cascade into VALUATION and Catalyst.

After the structural adaptation (adapt_ris + the Pipeline approval-range fix)
has made real revenue/expenses flow, this pass makes the *remaining* fragile
cells degrade gracefully to 0 so the whole model computes. It uses a LibreOffice
recalc to find the cells that actually error, then wraps only those — it never
touches cells that already compute, so real numbers are preserved.

    python tools/harden_formulas.py --ticker MOLN

Idempotent (skips cells already wrapped in IFERROR). Surgical XML patching.
"""
import argparse
import re
import shutil
import subprocess
import tempfile
import time
import warnings
import zipfile
from pathlib import Path

warnings.filterwarnings("ignore")
import openpyxl.drawing.text as _t  # noqa: E402
_t.Font.pitchFamily.max = 127
import openpyxl  # noqa: E402

TARGET_SHEETS = [
    "RIS", "RCFS", "RBS", "Schedules", "VALUATION", "Catalyst", "Pipeline",
    "FSA", "Historical Events", "FY DATA", "FY DATA K USD", "TAM Blood",
]
RECALC_XCU = """<?xml version="1.0" encoding="UTF-8"?>
<oor:items xmlns:oor="http://openoffice.org/2001/registry">
 <item oor:path="/org.openoffice.Office.Calc/Formula/Load"><prop oor:name="OOXMLRecalcMode" oor:op="fuse"><value>0</value></prop></item>
 <item oor:path="/org.openoffice.Office.Calc/Calculate/IterativeReference"><prop oor:name="Iteration" oor:op="fuse"><value>true</value></prop></item>
</oor:items>"""


def xesc(s):
    return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def recalc(path, workdir):
    prof = workdir / "prof"
    (prof / "user").mkdir(parents=True, exist_ok=True)
    (prof / "user" / "registrymodifications.xcu").write_text(RECALC_XCU)
    out = workdir / "out"
    subprocess.run(["soffice", "--headless", "--norestore", "--nolockcheck",
                    f"-env:UserInstallation=file://{prof}",
                    "--convert-to", "xlsx:Calc MS Excel 2007 XML",
                    "--outdir", str(out), str(path)],
                   check=True, capture_output=True, timeout=600)
    return out / path.name


def sheet_zip_path(zf, name):
    wbx = zf.read("xl/workbook.xml").decode("utf8", "ignore")
    rid = re.search(rf'<sheet name="{re.escape(name)}"[^>]*r:id="(rId\d+)"', wbx).group(1)
    rels = zf.read("xl/_rels/workbook.xml.rels").decode("utf8", "ignore")
    tgt = re.search(rf'Id="{rid}"[^>]*Target="([^"]+)"', rels).group(1).lstrip("/")
    return tgt if tgt.startswith("xl/") else "xl/" + tgt


def wrap_formula(xml, addr, inner):
    m = re.search(rf'<c r="{addr}"([^>]*?)>(.*?)</c>', xml, re.S)
    if not m:
        return xml, False
    sm = re.search(r'\ss="(\d+)"', m.group(1))
    style = f' s="{sm.group(1)}"' if sm else ""
    cell = f'<c r="{addr}"{style}><f>IFERROR({xesc(inner)},0)</f></c>'
    return re.sub(rf'<c r="{addr}"[^>]*?>.*?</c>', cell, xml, count=1, flags=re.S), True


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--ticker")
    ap.add_argument("--path")
    ap.add_argument("--max-passes", type=int, default=3)
    a = ap.parse_args()
    dcf = Path(a.path) if a.path else Path(
        f"/mnt/c/Users/yzsun/Desktop/DD/{a.ticker}/DCF {a.ticker}.xlsx")

    bak = dcf.with_name(f"{dcf.stem}_pre_harden_{time.strftime('%Y%m%d_%H%M%S')}.xlsx")
    shutil.copy2(dcf, bak)
    total = 0
    for p in range(1, a.max_passes + 1):
        with tempfile.TemporaryDirectory() as td:
            rc = recalc(dcf, Path(td))
            wv = openpyxl.load_workbook(rc, data_only=True)
            wf = openpyxl.load_workbook(dcf, data_only=False)
            # collect error formula cells per sheet
            todo = {}
            for sheet in TARGET_SHEETS:
                if sheet not in wv.sheetnames:
                    continue
                vs, fs = wv[sheet], wf[sheet]
                errs = []
                for row in vs.iter_rows():
                    for c in row:
                        if isinstance(c.value, str) and c.value.startswith("#"):
                            f = fs[c.coordinate].value
                            if isinstance(f, str) and f.startswith("=") \
                                    and not f[1:].lstrip().upper().startswith("IFERROR("):
                                errs.append((c.coordinate, f[1:]))
                if errs:
                    todo[sheet] = errs
            wv.close(); wf.close()
            if not todo:
                print(f"pass {p}: no more errors — clean.")
                break
            npass = sum(len(v) for v in todo.values())
            print(f"pass {p}: wrapping {npass} error cells across {list(todo)}")
            with zipfile.ZipFile(dcf) as zf:
                blobs = {n: zf.read(n) for n in zf.namelist()}
                paths = {s: sheet_zip_path(zf, s) for s in todo}
            for sheet, errs in todo.items():
                sp = paths[sheet]
                xml = blobs[sp].decode("utf8", "ignore")
                for addr, inner in errs:
                    xml, ok = wrap_formula(xml, addr, inner)
                    total += 1 if ok else 0
                blobs[sp] = xml.encode("utf8")
            with zipfile.ZipFile(dcf, "w", zipfile.ZIP_DEFLATED) as zo:
                for n, b in blobs.items():
                    zo.writestr(n, b)
    print(f"\nHardened {total} cells total. backup: {bak.name}")


if __name__ == "__main__":
    main()
