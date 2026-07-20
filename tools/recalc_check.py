#!/usr/bin/env python3
"""
recalc_check.py — headless LibreOffice recalc + check-cell / valuation readout.

`--convert-to` alone does NOT recompute formulas (LibreOffice keeps the cached
values). This forces recalc-on-load by writing a throwaway user profile whose
OOXML/ODF recalc mode = "Always (0)", then converting the workbook to a fresh
xlsx (which loads → recalculates → saves) and reading the computed values back.

    python tools/recalc_check.py --ticker MOLN
    python tools/recalc_check.py --path "/path/DCF X.xlsx"
"""
import argparse
import subprocess
import sys
import tempfile
import warnings
from pathlib import Path

warnings.filterwarnings("ignore")
import openpyxl.drawing.text as _t  # noqa: E402
_t.Font.pitchFamily.max = 127
import openpyxl  # noqa: E402
from openpyxl.utils import get_column_letter as cl  # noqa: E402
from openpyxl.utils.cell import range_boundaries  # noqa: E402
from openpyxl.worksheet.formula import DataTableFormula  # noqa: E402

RECALC_XCU = """<?xml version="1.0" encoding="UTF-8"?>
<oor:items xmlns:oor="http://openoffice.org/2001/registry" xmlns:xs="http://www.w3.org/2001/XMLSchema" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
 <item oor:path="/org.openoffice.Office.Calc/Formula/Load"><prop oor:name="OOXMLRecalcMode" oor:op="fuse"><value>0</value></prop></item>
 <item oor:path="/org.openoffice.Office.Calc/Formula/Load"><prop oor:name="ODFRecalcMode" oor:op="fuse"><value>0</value></prop></item>
 <item oor:path="/org.openoffice.Office.Calc/Calculate/IterativeReference"><prop oor:name="Iteration" oor:op="fuse"><value>true</value></prop></item>
</oor:items>
"""


def recalc(path: Path, outdir: Path, profile: Path) -> Path:
    (profile / "user").mkdir(parents=True, exist_ok=True)
    (profile / "user" / "registrymodifications.xcu").write_text(RECALC_XCU)
    cmd = [
        "soffice", "--headless", "--norestore", "--nolockcheck",
        f"-env:UserInstallation=file://{profile}",
        "--convert-to", "xlsx:Calc MS Excel 2007 XML",
        "--outdir", str(outdir), str(path),
    ]
    subprocess.run(cmd, check=True, capture_output=True, timeout=600)
    out = outdir / (path.stem + ".xlsx")
    if not out.exists():
        raise RuntimeError(f"recalc output not produced: {out}")
    return out


def val(ws, addr):
    v = ws[addr].value
    return v


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--ticker")
    ap.add_argument("--path")
    a = ap.parse_args()
    path = Path(a.path) if a.path else Path(
        f"/mnt/c/Users/yzsun/Desktop/DD/{a.ticker}/DCF {a.ticker}.xlsx")

    # LibreOffice does not implement Excel What-If Data Tables.  It can emit a
    # spurious #VALUE! in the table body or in the formula anchor immediately
    # above/left of it even when native Excel recalculates cleanly.  Exclude
    # only those cells from the LO error gate; all ordinary formulas remain
    # audited below.
    lo_unsupported: set[tuple[str, str]] = set()
    formula_wb = openpyxl.load_workbook(path, data_only=False, read_only=False)
    for formula_ws in formula_wb.worksheets:
        for row in formula_ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, DataTableFormula):
                    continue
                min_col, min_row, max_col, max_row = range_boundaries(cell.value.ref)
                for rr in range(min_row, max_row + 1):
                    for cc in range(min_col, max_col + 1):
                        lo_unsupported.add((formula_ws.title, f"{cl(cc)}{rr}"))
                if min_row > 1:
                    for cc in range(min_col, max_col + 1):
                        lo_unsupported.add((formula_ws.title, f"{cl(cc)}{min_row - 1}"))
                if min_col > 1:
                    for rr in range(min_row, max_row + 1):
                        lo_unsupported.add((formula_ws.title, f"{cl(min_col - 1)}{rr}"))
    formula_wb.close()

    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        print(f"Recalculating {path.name} via LibreOffice (headless)…")
        out = recalc(path, td / "out", td / "profile")
        wb = openpyxl.load_workbook(out, data_only=True)

        print("\n=== CHECK CELLS (must be 0) ===")
        ok = True
        checks = [("FY DATA K USD", [39, 49, 88, 100, 110],
                   ["R&D", "G&A", "BS", "PP&E", "Accrued"], range(6, 12)),
                  ("RBS", [47, 67], ["Acct-eq (NOA)", "Acct-eq (NFA)"], range(6, 12))]
        for sheet, rows, names, cols in checks:
            ws = wb[sheet]
            for r, nm in zip(rows, names):
                vals = [val(ws, f"{cl(c)}{r}") for c in cols]
                nums = [v for v in vals if isinstance(v, (int, float))]
                bad = [v for v in nums if abs(v) > 0.5]
                if bad:
                    ok = False
                flag = "OK" if not bad else "*** NONZERO ***"
                pretty = [f"{v:.1f}" if isinstance(v, (int, float)) else str(v) for v in vals]
                print(f"  {sheet}!R{r} {nm:14} {flag:16} {pretty}")
        print(f"\nALL CHECK CELLS ZERO: {ok}")

        print("\n=== VALUATION (DCF) ===")
        wv = wb["VALUATION"]
        for addr, lbl in [("C33", "WACC"), ("C36", "PV explicit"), ("C37", "PV terminal"),
                          ("C38", "Enterprise Value"), ("C45", "(+) Cash"),
                          ("C48", "Equity Value/■"), ("C52", "Last Price"), ("C53", "Consensus PT")]:
            print(f"  {lbl:20} {addr} = {val(wv, addr)}")

        print("\n=== CATALYST breakdown (row 7 drugs, row 9 value×LOA, W9 sum) ===")
        wc = wb["Catalyst"]
        for c in ("B", "F", "G", "I", "K", "M", "O", "Q", "S", "U", "W"):
            print(f"  {c}7={val(wc, c+'7')!r:22}  {c}9={val(wc, c+'9')!r}")

        # scan every sheet for error cells after recalc
        print("\n=== error cells after recalc (want none) ===")
        errs = 0
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cel in row:
                    if isinstance(cel.value, str) and cel.value.startswith("#") and cel.value.endswith(("!", "?", "A")):
                        if (ws.title, cel.coordinate) in lo_unsupported:
                            continue
                        if errs < 25:
                            print(f"  {ws.title}!{cel.coordinate} = {cel.value}")
                        errs += 1
        print(f"  total error cells: {errs}")
        wb.close()

    # Fail the build (nonzero exit) when any check cell is nonzero or any
    # #REF!/#VALUE! error cell survived the recalc, so a broken model can't
    # slip through as a success.
    code = 0 if ok and errs == 0 else 1
    sys.exit(code)


if __name__ == "__main__":
    main()
