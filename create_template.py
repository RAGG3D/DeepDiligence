#!/usr/bin/env python3
"""
create_template.py – Generate a blank DCF Excel template for a given ticker.

Creates two sheets:
  1. FY DATA K USD  – replicates the CMPX template structure & formatting
  2. Historical Events – daily price/catalyst framework with stock prices

Usage:
  python create_template.py --ticker BHVN
  python create_template.py --ticker BHVN --base-year 2020 --he-years 2022 2023 2024 2025
"""

import argparse
import logging
import re
import xml.etree.ElementTree as ET
import zipfile
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.colors import Color
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter


_EMPTY_CALC_CHAIN = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
    '<calcChain xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"></calcChain>'
)
logging.basicConfig(level=logging.INFO, format="%(asctime)s  %(levelname)-8s  %(message)s",
                    datefmt="%H:%M:%S")
logger = logging.getLogger(__name__)

# ══════════════════════════════════════════════════════════════════════
# FIXED FORMAT STYLES  (matching CMPX gold-standard template)
# ══════════════════════════════════════════════════════════════════════

# ── Theme Colors ─────────────────────────────────────────────────────
_CT0 = Color(theme=0)           # white
_CT1 = Color(theme=1)           # black/dark
_CB  = Color(rgb='FF0000FF')    # data-input blue
_CA  = Color(auto=True)         # border auto-color

# ── Fonts (explicit Calibri 11 everywhere) ───────────────────────────
_FD   = Font(name='Calibri', size=11, color=_CT1)           # default
_FB   = Font(name='Calibri', size=11, bold=True, color=_CT1) # bold
_FBL  = Font(name='Calibri', size=11, color=_CB)            # blue data
_FI   = Font(name='Calibri', size=11, italic=True, color=_CT1) # italic
_FIB  = Font(name='Calibri', size=11, italic=True, color=_CB) # italic blue
_FBI  = Font(name='Calibri', size=11, bold=True, italic=True, color=_CT1)
_FWB  = Font(name='Calibri', size=11, bold=True, color=_CT0) # white bold
_FWI  = Font(name='Calibri', size=11, italic=True, color=_CT0) # white italic
_FP   = Font(name='Calibri', size=11, color=Color(rgb='FF7030A0'))  # purple (SUMIFS)

# ── Fills ────────────────────────────────────────────────────────────
def _mf(theme=None, tint=0.0, rgb=None):
    f = PatternFill(patternType='solid')
    f.fgColor = Color(theme=theme, tint=tint) if theme is not None else Color(rgb=rgb)
    f.bgColor = Color(indexed=64)
    return f

_FIL_W = _mf(theme=0)                 # white  (data cells)
_FIL_K = _mf(theme=1)                 # black  (section headers)
_FIL_G = _mf(theme=0, tint=-0.05)     # light gray (check rows)
_FIL_Y = _mf(rgb='FFFFFFCC')          # yellow (HE price cells)
_FIL_N = _mf(rgb='FF002060')          # navy   (HE year headers)
_FIL_LG = _mf(theme=6, tint=0.8)      # light green (HE EVT/Category)

# ── Border components ────────────────────────────────────────────────
_ST = Side(style='thin', color=_CA)
_SM = Side(style='medium', color=_CA)

_BD  = Border(top=_SM, bottom=_SM, left=_SM)   # section header D
_BFK = Border(top=_SM, bottom=_SM)              # section header F-K
_BBT = Border(bottom=_ST)                       # sub-header bottom
_BLT = Border(left=_ST)                         # col-A left marker

# HE header borders
_BHB = Border(left=_SM, top=_SM, bottom=_SM)    # HE R5 col-B
_BHC = Border(top=_SM, bottom=_SM)              # HE R5 col C-X

# ── Alignments ───────────────────────────────────────────────────────
_AC  = Alignment(horizontal='center')
_AL  = Alignment(horizontal='left')
_AL1 = Alignment(horizontal='left', indent=1)
_ACC = Alignment(horizontal='centerContinuous')

# ── Number Formats ───────────────────────────────────────────────────
_HNF  = ';;;'                                    # hidden text (col A)
_DNF  = '#,##0_);\\(#,##0\\);'                   # col D  (hides zeros)
_NF   = '#,##0_);\\(#,##0\\);"—"'                # F-K data (dash zero)
_ENF  = '#,##0.00_);\\(#,##0.00\\);"—"'          # EPS
_BNF  = '0'                                      # col B counter
_DTNF = 'd/mm/yyyy;@'                            # date
_PNF  = '#,##0.00_);\\(#,##0.00\\);\\—'          # price
_DODNF = '#,##0.00%_);\\(#,##0.00%\\);\\—'       # DoD %

# ── Column widths ────────────────────────────────────────────────────
_KW = {'A':3.43,'B':1.86,'C':5.43,'D':67.57,'E':11.57,
       'F':11.14,'G':13.0,'H':13.0,'I':13.0,'J':13.0,'K':13.0,'L':6.86}

_HW = {'A':13.0,
       'B':21.29,'C':10.0,'D':8.43,'E':72.29,'F':11.0,'G':10.71,
       'H':10.43,'I':10.0,'J':13.57,'K':66.71,'L':11.0,'M':10.71,
       'N':10.43,'O':10.0,'P':13.29,'Q':66.71,'R':11.0,'S':10.71,
       'T':10.43,'U':10.0,'V':12.71,'W':66.71,'X':11.0}

_BFMT = ('=IF(AND($C{r}<>"",COUNTIFS($D:$D,$D{r},$C:$C,$C{r})>1),'
         'COUNTIFS($D$1:D{r},$D{r},$C$1:C{r},$C{r}),"")')
NY = 6  # year columns F-K

# ══════════════════════════════════════════════════════════════════════
# LOW-LEVEL CELL HELPERS
# ══════════════════════════════════════════════════════════════════════

def _sw(ws, w):
    for lt, v in w.items():
        ws.column_dimensions[lt].width = v

def _b(ws, r):
    c = ws.cell(row=r, column=2, value=_BFMT.format(r=r)); c.number_format = _BNF

def _c(ws, r, v):
    c = ws.cell(row=r, column=3, value=v); c.font = _FBL; c.alignment = _AC

def _d(ws, r, v, bold=False, indent=False):
    c = ws.cell(row=r, column=4, value=v)
    c.font = _FB if bold else _FD
    c.number_format = _DNF
    if indent:
        c.alignment = _AL1
    elif not bold:
        c.alignment = _AL
    # bold rows: leave alignment as None (matching CMPX)

_UNIT_LABEL = "[K USD]"  # overridden by create_template() for non-USD currencies

def _e(ws, r, v=None):
    """E column on data/formula/SUM rows — italic theme1, center, never bold."""
    if v is None:
        v = _UNIT_LABEL
    c = ws.cell(row=r, column=5, value=v)
    c.font = _FI; c.alignment = _AC

def _esb(ws, r):
    """E column on sub-header rows — italic blue, center."""
    c = ws.cell(row=r, column=5); c.font = _FIB; c.alignment = _AC

def _esp(ws, r):
    """E column on data-like spacer rows — italic blue."""
    c = ws.cell(row=r, column=5); c.font = _FIB

def _egap(ws, r):
    """E column on section-break gap rows — not italic."""
    c = ws.cell(row=r, column=5); c.font = _FD

def _fkd(ws, r, val=0, nf=None, brd=None):
    if nf is None: nf = _NF
    for i in range(NY):
        c = ws.cell(row=r, column=6+i, value=val)
        c.font = _FBL; c.fill = _FIL_W; c.number_format = nf
        if brd: c.border = brd

def _fkf(ws, r, tf, nf=None, brd=None):
    if nf is None: nf = _NF
    for i in range(NY):
        col = 6+i; cl = get_column_letter(col)
        f = tf.replace('F', cl) if i else tf
        c = ws.cell(row=r, column=col, value=f)
        c.font = _FD; c.number_format = nf
        if brd: c.border = brd

def _fks(ws, r, rs, re, nf=None):
    if nf is None: nf = _NF
    for i in range(NY):
        col = 6+i; cl = get_column_letter(col)
        c = ws.cell(row=r, column=col, value=f'=SUM({cl}{rs}:{cl}{re})')
        c.font = _FD; c.number_format = nf

def _fkw(ws, r):
    """Add white fill to F-K (for IS formula/SUM rows that have fill in CMPX)."""
    for i in range(NY):
        ws.cell(row=r, column=6+i).fill = _FIL_W

def _fkc(ws, r, r1, r2, s='-'):
    for i in range(NY):
        col = 6+i; cl = get_column_letter(col)
        c = ws.cell(row=r, column=col, value=f'={cl}{r1}{s}{cl}{r2}')
        c.font = _FD; c.fill = _FIL_G; c.number_format = _NF

def _fkh(ws, r, nf=None):
    if nf is None: nf = _NF
    for i in range(NY):
        c = ws.cell(row=r, column=6+i)
        c.font = _FB; c.fill = _FIL_K; c.border = _BFK; c.number_format = nf

def _fkn(ws, r, nf=None, brd=None):
    if nf is None: nf = _NF
    for i in range(NY):
        c = ws.cell(row=r, column=6+i); c.number_format = nf
        if brd: c.border = brd

def _am(ws, r, v=None):
    c = ws.cell(row=r, column=1, value=v); c.number_format = _HNF; c.border = _BLT

def _rh(ws, r, h=15.0):
    ws.row_dimensions[r].height = h

# ══════════════════════════════════════════════════════════════════════
# COMPOSITE ROW BUILDERS
# ══════════════════════════════════════════════════════════════════════

def _SH(ws, r, t, ca='X', fk_nf=None):
    """Section header row."""
    ws.cell(row=r, column=1, value=ca).number_format = _HNF
    d = ws.cell(row=r, column=4, value=t)
    d.font = _FWB; d.fill = _FIL_K; d.border = _BD; d.number_format = _DNF
    e = ws.cell(row=r, column=5)
    e.font = _FI; e.fill = _FIL_K; e.border = _BFK; e.alignment = _AC
    _b(ws, r); _fkh(ws, r, nf=fk_nf); _rh(ws, r, 15.75)

def _SB(ws, r, t, bdr=True):
    """Sub-header row. bdr=True adds bottom thin border (default for most)."""
    d = ws.cell(row=r, column=4, value=t)
    d.font = _FB; d.number_format = _DNF
    if bdr: d.border = _BBT
    _b(ws, r)
    e = ws.cell(row=r, column=5)
    e.font = _FIB; e.alignment = _AC  # italic blue, center
    if bdr: e.border = _BBT
    _fkn(ws, r, brd=_BBT if bdr else None)

def _DR(ws, r, cc, cd, bd=False, ind=False, nf=None, ce=None):
    """Data row."""
    _c(ws, r, cc); _d(ws, r, cd, bold=bd, indent=ind)
    _b(ws, r); _e(ws, r, v=ce); _fkd(ws, r, nf=nf)

def _FR(ws, r, cc, cd, tf, bd=True, nf=None, ce=None):
    """Formula row."""
    _c(ws, r, cc); _d(ws, r, cd, bold=bd)
    _b(ws, r); _e(ws, r, v=ce); _fkf(ws, r, tf, nf=nf)

def _SM(ws, r, cc, cd, rs, re, bd=True, ce=None):
    """SUM row."""
    _c(ws, r, cc); _d(ws, r, cd, bold=bd)
    _b(ws, r); _e(ws, r, v=ce); _fks(ws, r, rs, re)

def _CK(ws, r, cd, r1, r2, s='-', ce=None):
    """Check row — italic D/E, gray fill on D/E/F-K."""
    d = ws.cell(row=r, column=4, value=cd)
    d.font = _FI; d.fill = _FIL_G; d.number_format = _DNF
    _b(ws, r)
    e = ws.cell(row=r, column=5, value=ce)
    e.font = Font(name='Calibri', size=11, italic=True)  # color=none
    e.fill = _FIL_G; e.alignment = _AC
    _fkc(ws, r, r1, r2, s)

def _SP(ws, r, am=False, av=None):
    """Data-like spacer row (within a section). B='0', E=italic blue, F-K=DATA_NF."""
    ws.cell(row=r, column=2).number_format = _BNF
    _esp(ws, r)
    _fkn(ws, r)
    if am:
        ws.cell(row=r, column=4).number_format = _DNF
        ws.cell(row=r, column=4).alignment = _AL
        _am(ws, r, av)

def _SPG(ws, r):
    """Section-break gap spacer. B='0', E=not-italic, F-K=General."""
    ws.cell(row=r, column=2).number_format = _BNF
    _egap(ws, r)
    _fkn(ws, r, nf='General')

# ══════════════════════════════════════════════════════════════════════
# FY DATA K USD SHEET
# ══════════════════════════════════════════════════════════════════════

def _build_kusd_sheet(ws, ticker, base_year, fye_month_name):
    # Sheet-level properties
    ws.sheet_format.defaultRowHeight = 21.75
    ws.sheet_format.defaultColWidth = 5.5703125
    ws.sheet_properties.tabColor = Color(theme=5, tint=-0.25)
    ws.sheet_view.showGridLines = False
    _sw(ws, _KW)

    # ── R1-R4: Headers ────────────────────────────────────────────
    c2d = ws.cell(row=2, column=4, value=f'{ticker} US Equity')
    c2d.font = _FB; c2d.number_format = _DNF  # bold, no alignment
    c3d = ws.cell(row=3, column=4, value=f'Fiscal Year Ended {fye_month_name}.')
    c3d.font = Font(name='Calibri', size=11, italic=True)  # italic, color=none
    c3d.number_format = _DNF
    # R3 F-K: italic, center
    for i in range(NY):
        c = ws.cell(row=3, column=6+i)
        c.font = _FI; c.alignment = _AC
    _c(ws, 4, 'DATA')
    # R4 E: italic, center
    ws.cell(row=4, column=5).font = _FI
    ws.cell(row=4, column=5).alignment = _AC
    d4 = ws.cell(row=4, column=4, value='Fiscal Year')
    d4.font = _FB; d4.number_format = _DNF  # bold, no alignment
    f4 = ws.cell(row=4, column=6, value=base_year)
    f4.alignment = _AC; f4.number_format = '0'
    f4.font = Font(name='Calibri', size=11, bold=True)  # bold, color=none
    for i in range(1, NY):
        col = 6+i; prev = get_column_letter(col-1)
        c = ws.cell(row=4, column=col, value=f'={prev}4+1')
        c.alignment = _AC; c.number_format = '0'
        c.font = Font(name='Calibri', size=11, bold=True)  # bold, color=none

    # ── INCOME STATEMENT (R5-R24) ─────────────────────────────────
    _SH(ws, 5, 'Income (Loss) Statement', fk_nf='General')
    _SPG(ws, 6)
    # R7: separator row (bold D, bottom borders on D/E/F-K)
    ws.cell(row=7, column=2).number_format = _BNF
    d7 = ws.cell(row=7, column=4)
    d7.font = _FB; d7.number_format = _DNF; d7.border = _BBT
    e7 = ws.cell(row=7, column=5)
    e7.font = _FIB; e7.border = _BBT
    _fkn(ws, 7, nf='General', brd=_BBT)
    _DR(ws, 8, 'IS', 'Revenue')
    _DR(ws, 9, 'IS', 'Research And Development')
    _DR(ws, 10, 'IS', 'General And Administrative')
    _SM(ws, 11, 'IS', 'Total Operating Expenses', 9, 10); _fkw(ws, 11)
    _FR(ws, 12, 'IS', 'Loss From Operations', '=F8-F11'); _fkw(ws, 12)
    _SP(ws, 13, am=True)
    _DR(ws, 14, 'IS', 'Interest Income'); _am(ws, 14)
    _DR(ws, 15, 'IS', 'Interest Expense', bd=True); _am(ws, 15)
    _FR(ws, 16, 'IS', 'Total Other Income (Expense)', '=F14+F15'); _fkw(ws, 16)
    _am(ws, 16)
    _DR(ws, 17, 'IS', 'Income Tax Provision')
    _FR(ws, 18, 'IS', 'Net (Loss) Income', '=F12+F16-F17'); _fkw(ws, 18)
    _am(ws, 18)
    _FR(ws, 19, 'IS', 'Net (Loss) Income Per Common Share, Basic And Diluted',
         '=F18*1000/F20', nf=_ENF); _am(ws, 19, 'X'); _fkw(ws, 19)
    _DR(ws, 20, 'IS', 'Weighted-Average Number Of Common Shares, Basic And Diluted',
         bd=True); _am(ws, 20, 'X')
    _SP(ws, 21)
    _DR(ws, 22, 'IS', 'Foreign Currency Translation Adjustment', bd=True); _am(ws, 22)
    _SP(ws, 23, am=True)
    _FR(ws, 24, 'IS', 'Comprehensive (Loss) Income', '=F18+F22'); _am(ws, 24); _fkw(ws, 24)
    _SP(ws, 25)

    # ── IS NOTES (R26-R49) ────────────────────────────────────────
    _SH(ws, 26, 'Income (Loss) Statement Notes')
    _SB(ws, 27, 'Research And Development Expenses')
    for i, r in enumerate(range(28, 37)):
        _DR(ws, r, 'ISN', f'R&D Item {i+1}')
    _SM(ws, 37, 'ISN', 'Total Research And Development', 28, 36)
    _SP(ws, 38)
    _CK(ws, 39, 'Check - Research And Development', 37, 9)
    _SPG(ws, 40)
    _SB(ws, 41, 'General And Administrative Expenses')
    for i, r in enumerate(range(42, 47)):
        _DR(ws, r, 'ISN', f'G&A Item {i+1}')
    _SM(ws, 47, 'ISN', 'Total General And Administrative', 42, 46)
    _SP(ws, 48)
    _CK(ws, 49, 'Check - General And Administrative', 47, 10)
    _SPG(ws, 50)

    # ── BALANCE SHEET (R51-R88) ───────────────────────────────────
    _SH(ws, 51, 'Balance Sheet')
    _SB(ws, 52, 'Current Assets')
    _DR(ws, 53, 'BS', 'Cash And Cash Equivalents')
    _DR(ws, 54, 'BS', 'Marketable Securities')
    _DR(ws, 55, 'BS', 'Prepaid Expenses And Other Current Assets')
    _DR(ws, 56, 'BS', 'Research And Development Incentives Receivable')
    _SM(ws, 57, 'BS', 'Total Current Assets', 53, 56)
    _SP(ws, 58)
    _SB(ws, 59, 'Non-Current Assets')
    _DR(ws, 60, 'BS', 'Property And Equipment, Net')
    _DR(ws, 61, 'BS', 'Operating Lease Right-Of-Use Assets')
    _DR(ws, 62, 'BS', 'Other Assets')
    _FR(ws, 63, 'BS', 'Total Assets', '=F57+SUM(F60:F62)')
    _SP(ws, 64)
    _SB(ws, 65, "Liabilities And Stockholders' Equity", bdr=False)
    _SB(ws, 66, 'Current Liabilities')
    _DR(ws, 67, 'BS', 'Accounts Payable')
    _DR(ws, 68, 'BS', 'Accrued Expenses And Other Current Liabilities')
    _DR(ws, 69, 'BS', 'Operating Lease Liabilities, Current Portion')
    _DR(ws, 70, 'BS', 'Debt, Current Portion')
    _SM(ws, 71, 'BS', 'Total Current Liabilities', 67, 70)
    _SP(ws, 72)
    _SB(ws, 73, 'Long-Term Liabilities')
    _DR(ws, 74, 'BS', 'Long-Term Debt, Net Of Discount')
    _DR(ws, 75, 'BS', 'Operating Lease Liabilities, Net Of Current Portion')
    _DR(ws, 76, 'BS', 'Deferred Revenue, Net Of Current Portion')
    _DR(ws, 77, 'BS', 'Other Long-Term Liabilities')
    _FR(ws, 78, 'BS', 'Total Liabilities', '=F71+SUM(F74:F77)')
    _SP(ws, 79)
    _SB(ws, 80, "Stockholders' Equity (Deficit)")
    _DR(ws, 81, 'BS', "Ordinary Shares, \u00a30.01 Nominal Value")
    _DR(ws, 82, 'BS', 'Additional Paid-In Capital')
    _DR(ws, 83, 'BS', 'Accumulated Other Comprehensive (Loss) Income')
    _DR(ws, 84, 'BS', 'Accumulated Deficit')
    _SM(ws, 85, 'BS', "Total Stockholders' Equity (Deficit)", 81, 84)
    _FR(ws, 86, 'BS', "Total Liabilities And Stockholders' Equity", '=F78+F85')
    ws.cell(row=86, column=1, value='X').number_format = _HNF
    _SP(ws, 87)
    _CK(ws, 88, 'Check ', 63, 86)
    _SP(ws, 89)

    # ── BS NOTES (R90-R110) ───────────────────────────────────────
    _SH(ws, 90, 'Balance Sheet Notes')
    _SPG(ws, 91)
    _SB(ws, 92, 'Property And Equipment, Net')
    for i, r in enumerate(range(93, 98)):
        _DR(ws, r, 'BSN', f'PP&E Item {i+1}' if i < 4 else 'Accumulated Depreciation',
            ind=True)
    _SM(ws, 98, 'BSN', 'Property And Equipment, Net', 93, 97)
    _SP(ws, 99)
    _CK(ws, 100, 'Check - Property And Equipment', 98, 60)
    _SPG(ws, 101)
    _SB(ws, 102, 'Accrued Expenses And Other Current Liabilities')
    for i, r in enumerate(range(103, 108)):
        _DR(ws, r, 'BSN', f'Accrued Item {i+1}', ind=True)
    _SM(ws, 108, 'BSN', 'Accrued Expenses And Other Current Liabilities', 103, 107)
    _SP(ws, 109)
    _CK(ws, 110, 'Check - Accrued Expenses And Other Current Liabilities', 108, 68)
    _SPG(ws, 111)

    # ── CASH FLOW STATEMENT (R112-R121) ───────────────────────────
    _SH(ws, 112, 'Cash Flow Statement')
    _SP(ws, 113)
    _SB(ws, 114, 'Operating Activities')
    _FR(ws, 115, 'CFS', 'Net (Loss) Income', '=F18'); _fkw(ws, 115)
    _SP(ws, 116)
    _DR(ws, 117, 'CFS', 'Net Cash Used In Operating Activities', bd=True)
    _DR(ws, 118, 'CFS', 'Net Cash Provided by (Used In) Investing Activities', bd=True)
    _DR(ws, 119, 'CFS', 'Net Cash Provided By Financing Activities', bd=True)
    _DR(ws, 120, 'CFS', 'Effect Of Exchange Rate Changes On Cash And Cash Equivalents', bd=True)
    _SM(ws, 121, 'CFS', 'Net Change In Cash And Cash Equivalents', 117, 120)

    # ── Post: set ALL row heights ─────────────────────────────────
    tall = {4, 5, 26, 51, 90, 112}
    tall_sp = {25, 50, 89, 111}  # spacer rows before sections = 15.75
    for r in range(1, 122):
        if r in tall or r in tall_sp:
            _rh(ws, r, 15.75)
        elif r in (119, 120):
            _rh(ws, r, 15.95)
        else:
            _rh(ws, r, 15.0)

    # ── Post: col B numfmt on header rows ─────────────────────────
    for r in range(1, 5):
        ws.cell(row=r, column=2).number_format = _BNF

    # ── Conditional Formatting ────────────────────────────────────
    # Rule 1: Col B highlight duplicates
    rule_b = FormulaRule(formula=['B1<>""'])
    rule_b.dxf = DifferentialStyle(
        font=Font(bold=True, color=Color(theme=1, tint=0.5)),
        fill=PatternFill(bgColor=Color(theme=5, tint=0.8)))
    ws.conditional_formatting.add('B1:B1048576', rule_b)

    # Rule 2: Col C category highlight
    rule_c = FormulaRule(formula=['AND(AND(NOT(ISNUMBER(#REF!)),NOT(ISNUMBER($H1))),NOT(ISBLANK(C1)))'])
    rule_c.dxf = DifferentialStyle(
        fill=PatternFill(bgColor=Color(theme=5, tint=0.8)))
    ws.conditional_formatting.add('C1:C1048576', rule_c)

    # Rule 3+4: Check rows → red on ABS>0.1
    chk_dxf = DifferentialStyle(
        font=Font(bold=True, color=_CT0),
        fill=PatternFill(bgColor=Color(rgb='FFC00000')))
    rule_chk1 = FormulaRule(formula=['ABS(D39)>0.1'])
    rule_chk1.dxf = chk_dxf
    ws.conditional_formatting.add('D39:K39 D88:K88 D100:K100 D110:K110', rule_chk1)
    rule_chk2 = FormulaRule(formula=['ABS(D49)>0.1'])
    rule_chk2.dxf = chk_dxf
    ws.conditional_formatting.add('D49:K49', rule_chk2)

    logger.info("FY DATA K USD sheet created")


# ══════════════════════════════════════════════════════════════════════
# FY DATA (MM USD) SHEET — SUMIFS referencing FY DATA K USD
# ══════════════════════════════════════════════════════════════════════

# Rows that contain DATA cells (blue input values in K USD → SUMIFS in MM)
# All other rows have local formulas (SUM, subtraction, etc.) that work unchanged.
_SUMIFS_ROWS = {
    # IS data
    8, 9, 10, 14, 15, 17, 20, 22,
    # ISN R&D data (R28-R36)
    28, 29, 30, 31, 32, 33, 34, 35, 36,
    # ISN G&A data (R42-R46)
    42, 43, 44, 45, 46,
    # BS Current Assets
    53, 54, 55, 56,
    # BS Non-Current Assets
    60, 61, 62,
    # BS Current Liabilities
    67, 68, 69, 70,
    # BS Long-Term Liabilities
    74, 75, 76, 77,
    # BS Equity
    81, 82, 83, 84,
    # BSN PP&E (R93-R97)
    93, 94, 95, 96, 97,
    # BSN Accrued (R103-R107)
    103, 104, 105, 106, 107,
    # CFS data
    117, 118, 119, 120,
}

_KR = "'FY DATA K USD'"  # sheet reference constant for formulas


def _build_mm_sheet(ws, ticker, base_year, fye_month_name):
    """Build FY DATA sheet with SUMIFS/1000 from FY DATA K USD.

    Uses a 'build then overlay' approach:
      1) Build identical layout via _build_kusd_sheet (formatting, formulas)
      2) Overlay D column with formula references to K USD (auto-syncs renames)
      3) Change E column [K USD] → [MM USD]
      4) Replace data cells with SUMIFS/1000 formulas
      5) Year headers reference K USD
    """
    # Step 1: Build identical to K USD (all formatting, formulas, conditional fmt)
    _build_kusd_sheet(ws, ticker, base_year, fye_month_name)

    # Step 2: Tab color (distinct from K USD)
    ws.sheet_properties.tabColor = Color(theme=4, tint=-0.25)

    # Step 3: D column → formula references to K USD
    # This ensures D labels auto-sync when excel_writer renames rows in K USD
    for r in range(1, 122):
        cell = ws.cell(row=r, column=4)
        if cell.value is not None:
            cell.value = f"={_KR}!D{r}"

    # Step 4: E column [K ...] → [MM ...]
    for r in range(1, 122):
        cell = ws.cell(row=r, column=5)
        if cell.value and str(cell.value).startswith("[K "):
            cell.value = "[MM " + str(cell.value)[3:]

    # Step 5: Year headers (R4 F-K) → reference K USD years
    for i in range(NY):
        col = 6 + i
        cl = get_column_letter(col)
        ws.cell(row=4, column=col).value = f"={_KR}!{cl}4"

    # Step 6: Data cells → SUMIFS/1000 (default black font, not blue)
    for r in _SUMIFS_ROWS:
        for i in range(NY):
            col = 6 + i
            cl = get_column_letter(col)
            cell = ws.cell(row=r, column=col)
            cell.value = (
                f"=SUMIFS({_KR}!{cl}:{cl},"
                f"{_KR}!$D:$D,'FY DATA'!$D{r},"
                f"{_KR}!$C:$C,'FY DATA'!$C{r},"
                f"{_KR}!$B:$B,'FY DATA'!$B{r})/1000"
            )
            cell.font = _FP  # purple (SUMIFS formula)

    logger.info("FY DATA (MM USD SUMIFS) sheet created")


def _build_direct_mm_sheet(ws, ticker, base_year, fye_month_name):
    """Build FY DATA sheet for MM USD companies — blue data cells, no K USD intermediary.

    Identical to K USD layout but with [MM USD] labels and adjusted EPS formula.
    """
    _build_kusd_sheet(ws, ticker, base_year, fye_month_name)

    # Tab color
    ws.sheet_properties.tabColor = Color(theme=4, tint=-0.25)

    # E column: [K ...] → [MM ...]
    for r in range(1, 122):
        cell = ws.cell(row=r, column=5)
        if cell.value and str(cell.value).startswith("[K "):
            cell.value = "[MM " + str(cell.value)[3:]

    # EPS formula: *1000 → *1000000 (shares are raw count in MM mode)
    for i in range(NY):
        col = 6 + i
        cl = get_column_letter(col)
        ws.cell(row=19, column=col).value = f"={cl}18*1000000/{cl}20"

    logger.info("FY DATA (direct MM USD) sheet created")


# ══════════════════════════════════════════════════════════════════════
# HISTORICAL EVENTS SHEET
# ══════════════════════════════════════════════════════════════════════

def _build_he_sheet(ws, ticker, he_years, prices=None):
    if prices is None:
        prices = {}

    # Sheet properties
    ws.sheet_properties.tabColor = Color(theme=9, tint=-0.5)
    ws.sheet_view.showGridLines = False
    _sw(ws, _HW)

    block_starts = [2, 8, 14, 20]  # B, H, N, T

    # ── R1-R4: left-aligned ───────────────────────────────────────
    for r in range(1, 5):
        for col in range(1, 25):
            ws.cell(row=r, column=col).alignment = _AL

    # R3: ticker
    ws.cell(row=3, column=2, value=f'{ticker} US Equity').font = _FB

    # R4: subtitle (bold)
    ws.cell(row=4, column=2).font = _FB

    # ── R5: Section header bar (bold white B, italic white C-X) ──
    b5 = ws.cell(row=5, column=2, value='Historical Catalyst Study')
    b5.font = _FWB; b5.fill = _FIL_K; b5.border = _BHB
    for col in range(3, 25):  # C through X
        c = ws.cell(row=5, column=col)
        c.font = _FWI; c.fill = _FIL_K; c.border = _BHC

    # ── R7: Year headers (bold white on navy, centerContinuous) ──
    for idx, year in enumerate(he_years):
        if idx >= len(block_starts):
            break
        bc = block_starts[idx]
        for j in range(5):  # 5 columns per block
            c = ws.cell(row=7, column=bc+j)
            c.font = _FWB; c.fill = _FIL_N; c.alignment = _ACC
        ws.cell(row=7, column=bc).value = f'FA {year}'

    # ── R8: Column sub-headers (not bold, bottom thin border) ────
    headers = ['Date', 'Share Price', 'DoD Chg', 'EVT', 'Category']
    for idx, year in enumerate(he_years):
        if idx >= len(block_starts):
            break
        bc = block_starts[idx]
        for j, h in enumerate(headers):
            c = ws.cell(row=8, column=bc+j, value=h)
            c.font = _FD; c.border = _BBT
        # Date column header gets date numfmt
        ws.cell(row=8, column=bc).number_format = _DTNF

    # ── R9-R373: Data rows ────────────────────────────────────────
    for idx, year in enumerate(he_years):
        if idx >= len(block_starts):
            break
        bc = block_starts[idx]
        dc = bc          # date col
        pc = bc + 1      # price col
        dd = bc + 2      # dod col
        ec = bc + 3      # evt col
        cc = bc + 4      # category col
        dl = get_column_letter(dc)
        pl = get_column_letter(pc)

        # R9: start date
        ws.cell(row=9, column=dc, value=datetime(year,1,1)).number_format = _DTNF
        # R9 price
        p = prices.get(f'{year}-01-01')
        pc9 = ws.cell(row=9, column=pc, value=p)
        pc9.number_format = _PNF; pc9.font = _FBL; pc9.fill = _FIL_Y
        # R9 DoD (blank but with format)
        ws.cell(row=9, column=dd).number_format = _DODNF
        # R9 EVT (light green fill)
        ws.cell(row=9, column=ec).fill = _FIL_LG
        # R9 Category (bold, light green fill)
        c9cat = ws.cell(row=9, column=cc)
        c9cat.font = _FB; c9cat.fill = _FIL_LG

        # R10-R373
        for r in range(10, 374):
            # Date formula
            ws.cell(row=r, column=dc,
                    value=f'={dl}{r-1}+1').number_format = _DTNF
            # Price
            day_off = r - 9
            dt = datetime(year,1,1) + timedelta(days=day_off)
            p = prices.get(dt.strftime('%Y-%m-%d'))
            prc = ws.cell(row=r, column=pc, value=p)
            prc.number_format = _PNF; prc.font = _FBL; prc.fill = _FIL_Y
            # DoD formula
            ws.cell(row=r, column=dd,
                    value=f'=({pl}{r}-{pl}{r-1})/{pl}{r-1}').number_format = _DODNF
            # EVT (light green)
            ws.cell(row=r, column=ec).fill = _FIL_LG
            # Category (bold, light green)
            ccat = ws.cell(row=r, column=cc)
            ccat.font = _FB; ccat.fill = _FIL_LG

    # ── ColorScale conditional formatting on DoD columns ──────────
    # CMPX pattern: 5 rules total
    #   Rule 1: D9:D373 (first block alone)
    #   Rule 2: J9 P9 V9 (row-9 cells of blocks 2-4, grouped)
    #   Rule 3: J10:J373
    #   Rule 4: P10:P373
    #   Rule 5: V10:V373
    _cs_kw = dict(start_type='min', start_color='FF0033CC',
                  mid_type='num', mid_value=0, mid_color='FFFFFFFF',
                  end_type='max', end_color='FFC00000')
    dod_cols = [4, 10, 16, 22]  # D, J, P, V
    active = [dod_cols[i] for i in range(min(len(he_years), len(dod_cols)))]

    if active:
        # Rule 1: first DoD column full range
        c0 = get_column_letter(active[0])
        ws.conditional_formatting.add(f'{c0}9:{c0}373', ColorScaleRule(**_cs_kw))

    if len(active) > 1:
        # Rule 2: row-9 cells of remaining blocks grouped
        r9_parts = ' '.join(f'{get_column_letter(dc)}9' for dc in active[1:])
        ws.conditional_formatting.add(r9_parts, ColorScaleRule(**_cs_kw))

        # Rules 3-5: rows 10:373 for each remaining block
        for dc in active[1:]:
            cl = get_column_letter(dc)
            ws.conditional_formatting.add(f'{cl}10:{cl}373', ColorScaleRule(**_cs_kw))

    logger.info("Historical Events sheet created")


# ══════════════════════════════════════════════════════════════════════
# STOCK PRICE FETCHER
# ══════════════════════════════════════════════════════════════════════

def fetch_stock_prices(ticker, years):
    try:
        import yfinance as yf
        import pandas as pd
    except ImportError:
        logger.warning("yfinance/pandas not installed; skipping prices")
        return {}
    mn, mx = min(years), max(years)
    start, end = f'{mn}-01-01', f'{mx}-12-31'
    logger.info(f"Fetching {ticker} prices {start} to {end} ...")
    try:
        data = yf.download(ticker, start=start, end=end, progress=False)
        if data.empty:
            logger.warning(f"No price data for {ticker}"); return {}
        close = data['Close']
        if hasattr(close, 'columns'):
            close = close.iloc[:, 0]
        full_idx = pd.date_range(start=start, end=end, freq='D')
        prices = close.reindex(full_idx).ffill().bfill()
        result = {}
        for dt, price in prices.items():
            if pd.notna(price):
                result[dt.strftime('%Y-%m-%d')] = round(float(price), 2)
        logger.info(f"  {len(result)} daily prices")
        return result
    except Exception as e:
        logger.warning(f"Price fetch failed: {e}"); return {}


# ══════════════════════════════════════════════════════════════════════
# FYE DETECTION + MAIN
# ══════════════════════════════════════════════════════════════════════

MONTH_NAMES = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'May',6:'Jun',
               7:'Jul',8:'Aug',9:'Sep',10:'Oct',11:'Nov',12:'Dec'}

def detect_fye_month(ticker, cik=None):
    from core.sec_fetcher import SECFetcher
    f = SECFetcher()
    resolved = cik.zfill(10) if cik else f.get_cik(ticker)
    facts = f.get_company_facts(resolved)
    return f.detect_fye_month(facts)


# ══════════════════════════════════════════════════════════════════════
# TAM SHEET INJECTION (copy from base file with styles merge)
# ══════════════════════════════════════════════════════════════════════

_BASE_FILE = Path("/mnt/c/Users/yzsun/Desktop/DD/base/DCF Template 2020.xlsx")
_NS_SS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_NS_PKG = "http://schemas.openxmlformats.org/package/2006/relationships"
_NS_CT = "http://schemas.openxmlformats.org/package/2006/content-types"
_WS_REL_TYPE = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet"
_WS_CT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"
_COMMENT_CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.comments+xml"
_VML_CT = "application/vnd.openxmlformats-officedocument.vmlDrawing"


def _offset_xf(xf_elem, font_off, fill_off, border_off, numfmt_remap, xfid_off=0):
    """Adjust font/fill/border/numFmt/xfId references in an <xf> element."""
    for attr, off in [('fontId', font_off), ('fillId', fill_off),
                      ('borderId', border_off)]:
        old = int(xf_elem.get(attr, '0'))
        xf_elem.set(attr, str(old + off))
    old_nf = int(xf_elem.get('numFmtId', '0'))
    if old_nf in numfmt_remap:
        xf_elem.set('numFmtId', str(numfmt_remap[old_nf]))
    if xfid_off:
        old_xfid = int(xf_elem.get('xfId', '0'))
        xf_elem.set('xfId', str(old_xfid + xfid_off))


def _merge_styles(base_bytes: bytes, new_bytes: bytes) -> Tuple[bytes, int]:
    """Merge new (openpyxl) styles into base styles.
    Returns (merged_styles_bytes, cellxf_offset).
    TAM sheets use base indices 0..B-1; FY DATA uses B..B+N-1."""
    ET.register_namespace('', _NS_SS)
    # Preserve extra namespaces from base
    for ns_prefix in ('mc', 'x14ac', 'x16r2', 'xr'):
        ns_uri_m = re.search(rf'xmlns:{ns_prefix}="([^"]+)"', base_bytes.decode('utf-8'))
        if ns_uri_m:
            ET.register_namespace(ns_prefix, ns_uri_m.group(1))
    mc_uri = re.search(r'xmlns:mc="([^"]+)"', base_bytes.decode('utf-8'))
    if mc_uri:
        ET.register_namespace('mc', mc_uri.group(1))

    base = ET.fromstring(base_bytes)
    new = ET.fromstring(new_bytes)

    def _section(root, name):
        return root.find(f'{{{_NS_SS}}}{name}')

    def _count(elem):
        return int(elem.get('count', '0')) if elem is not None else 0

    # Base counts
    b_numfmts = _section(base, 'numFmts')
    b_fonts = _section(base, 'fonts')
    b_fills = _section(base, 'fills')
    b_borders = _section(base, 'borders')
    b_csxf = _section(base, 'cellStyleXfs')
    b_xf = _section(base, 'cellXfs')

    font_off = _count(b_fonts)
    fill_off = _count(b_fills)
    border_off = _count(b_borders)
    csxf_off = _count(b_csxf)
    xf_off = _count(b_xf)

    # Max numFmtId in base (custom formats are 164+)
    max_nfid = 163
    if b_numfmts is not None:
        for nf in b_numfmts:
            max_nfid = max(max_nfid, int(nf.get('numFmtId', '0')))

    # Build numFmt remap for new file's custom formats
    n_numfmts = _section(new, 'numFmts')
    nf_remap: Dict[int, int] = {}
    if n_numfmts is not None:
        next_id = max_nfid + 1
        for nf in list(n_numfmts):
            old_id = int(nf.get('numFmtId', '0'))
            if old_id >= 164:
                nf_remap[old_id] = next_id
                nf.set('numFmtId', str(next_id))
                next_id += 1

    # Append new entries to base sections
    def _append(base_sect, new_sect):
        if base_sect is None or new_sect is None:
            return
        for child in list(new_sect):
            base_sect.append(child)
        base_sect.set('count', str(len(list(base_sect))))

    _append(b_numfmts, n_numfmts)
    _append(b_fonts, _section(new, 'fonts'))
    _append(b_fills, _section(new, 'fills'))
    _append(b_borders, _section(new, 'borders'))

    # CellStyleXfs -- offset internal refs
    n_csxf = _section(new, 'cellStyleXfs')
    if n_csxf is not None and b_csxf is not None:
        for xf in list(n_csxf):
            _offset_xf(xf, font_off, fill_off, border_off, nf_remap)
            b_csxf.append(xf)
        b_csxf.set('count', str(len(list(b_csxf))))

    # CellXfs -- offset internal refs + xfId
    n_xf = _section(new, 'cellXfs')
    if n_xf is not None and b_xf is not None:
        for xf in list(n_xf):
            _offset_xf(xf, font_off, fill_off, border_off, nf_remap, csxf_off)
            b_xf.append(xf)
        b_xf.set('count', str(len(list(b_xf))))

    # DXFs -- append new conditional formatting styles
    b_dxfs = _section(base, 'dxfs')
    n_dxfs = _section(new, 'dxfs')
    if n_dxfs is not None:
        if b_dxfs is None:
            base.append(n_dxfs)
        else:
            _append(b_dxfs, n_dxfs)

    merged = ET.tostring(base, xml_declaration=True, encoding='UTF-8')
    return merged, xf_off


def _offset_sheet_styles(sheet_bytes: bytes, offset: int) -> bytes:
    """Add offset to all s="N" style attributes in a sheet XML."""
    if offset == 0:
        return sheet_bytes
    text = sheet_bytes.decode('utf-8')

    def _repl(m):
        return f's="{int(m.group(1)) + offset}"'

    text = re.sub(r's="(\d+)"', _repl, text)
    return text.encode('utf-8')


def _resolve_base_path(base_path: Optional[Path] = None) -> Optional[Path]:
    """Find the base file, with fallback to .~tam_patch.xlsx variant."""
    if base_path is None:
        base_path = _BASE_FILE
    if base_path.exists():
        return base_path
    alt = base_path.with_suffix('.~tam_patch.xlsx')
    if alt.exists():
        return alt
    return None


def _collect_tam_files(base_path: Path) -> Optional[Tuple[Dict[str, str], Dict[str, bytes]]]:
    """Read TAM sheet paths and all files to inject from base.
    Returns (tam_sheets {name: zip_path}, inject {zip_path: bytes}) or None."""
    with zipfile.ZipFile(base_path) as zf:
        wb_et = ET.fromstring(zf.read('xl/workbook.xml'))
        rels_et = ET.fromstring(zf.read('xl/_rels/workbook.xml.rels'))
        base_names = set(zf.namelist())

    rid_to_path: Dict[str, str] = {}
    for rel in rels_et:
        if 'worksheet' in rel.get('Type', ''):
            rid = rel.get('Id', '')
            tgt = rel.get('Target', '').lstrip('/')
            rid_to_path[rid] = f'xl/{tgt}' if not tgt.startswith('xl/') else tgt

    tam_sheets: Dict[str, str] = {}
    for s in wb_et.findall(f'.//{{{_NS_SS}}}sheet'):
        name = s.get('name', '')
        if name in ('TAM Solid', 'TAM Blood'):
            rid = s.get(f'{{{_NS_REL}}}id', '')
            if rid in rid_to_path:
                tam_sheets[name] = rid_to_path[rid]

    if len(tam_sheets) != 2:
        logger.warning(f"Found {len(tam_sheets)} TAM sheets in base; expected 2")
        return None

    inject: Dict[str, bytes] = {}
    with zipfile.ZipFile(base_path) as zf:
        for name, zip_path in tam_sheets.items():
            inject[zip_path] = zf.read(zip_path)
            sheet_fname = zip_path.split('/')[-1]
            rels_path = f'xl/worksheets/_rels/{sheet_fname}.rels'
            if rels_path in base_names:
                inject[rels_path] = zf.read(rels_path)
                rels_content = inject[rels_path].decode('utf-8')
                for ref_m in re.finditer(r'Target="([^"]+)"', rels_content):
                    ref_target = ref_m.group(1)
                    if ref_target.startswith('../'):
                        ref_fpath = f'xl/{ref_target[3:]}'
                    else:
                        ref_fpath = f'xl/worksheets/{ref_target}'
                    if ref_fpath in base_names:
                        inject[ref_fpath] = zf.read(ref_fpath)

    return tam_sheets, inject


def _register_and_write(target_path: Path, tam_sheets: Dict[str, str],
                        inject: Dict[str, bytes],
                        extra_modified: Optional[Dict[str, bytes]] = None):
    """Register TAM sheets in workbook metadata and write the patched zip."""
    with zipfile.ZipFile(target_path) as zf:
        tgt_wb_xml = zf.read('xl/workbook.xml').decode('utf-8')
        tgt_rels_xml = zf.read('xl/_rels/workbook.xml.rels').decode('utf-8')
        tgt_ct_xml = zf.read('[Content_Types].xml').decode('utf-8')

    # Check if TAM sheets already exist
    if 'TAM Solid' in tgt_wb_xml and 'TAM Blood' in tgt_wb_xml:
        logger.info("TAM sheets already registered in workbook; skipping")
        return

    max_sid = max((int(m.group(1)) for m in re.finditer(r'sheetId="(\d+)"', tgt_wb_xml)),
                  default=0)
    max_rid = max((int(m.group(1)) for m in re.finditer(r'Id="rId(\d+)"', tgt_rels_xml)),
                  default=0)

    # Detect r: namespace prefix (varies by xlsx generator)
    r_prefix = 'r'
    r_ns_m = re.search(r'xmlns:(\w+)="http://schemas\.openxmlformats\.org/officeDocument/2006/relationships"',
                        tgt_wb_xml)
    if r_ns_m:
        r_prefix = r_ns_m.group(1)
    needs_xmlns = r_ns_m is None  # openpyxl puts xmlns:r on each <sheet>

    for i, (name, zip_path) in enumerate(tam_sheets.items()):
        sid = max_sid + i + 1
        rid = f'rId{max_rid + i + 1}'
        rel_target = '/' + zip_path

        if needs_xmlns:
            sheet_tag = (f'<sheet xmlns:r="{_NS_REL}" '
                         f'name="{name}" sheetId="{sid}" state="visible" r:id="{rid}"/>')
        else:
            sheet_tag = (f'<sheet name="{name}" sheetId="{sid}" '
                         f'{r_prefix}:id="{rid}"/>')
        tgt_wb_xml = tgt_wb_xml.replace('</sheets>', f'{sheet_tag}</sheets>')

        rel_tag = f'<Relationship Type="{_WS_REL_TYPE}" Target="{rel_target}" Id="{rid}"/>'
        tgt_rels_xml = tgt_rels_xml.replace('</Relationships>',
                                            f'{rel_tag}</Relationships>')

        if zip_path not in tgt_ct_xml:
            ct_tag = f'<Override PartName="{rel_target}" ContentType="{_WS_CT_TYPE}"/>'
            tgt_ct_xml = tgt_ct_xml.replace('</Types>', f'{ct_tag}</Types>')

    # Register supporting files
    if 'Extension="vml"' not in tgt_ct_xml:
        vml_default = f'<Default Extension="vml" ContentType="{_VML_CT}"/>'
        tgt_ct_xml = tgt_ct_xml.replace('</Types>', f'{vml_default}</Types>')
    for fpath in inject:
        if 'comments' in fpath and fpath not in tgt_ct_xml:
            ct_tag = f'<Override PartName="/{fpath}" ContentType="{_COMMENT_CT}"/>'
            tgt_ct_xml = tgt_ct_xml.replace('</Types>', f'{ct_tag}</Types>')

    modified: Dict[str, bytes] = {
        'xl/workbook.xml': tgt_wb_xml.encode('utf-8'),
        'xl/_rels/workbook.xml.rels': tgt_rels_xml.encode('utf-8'),
        '[Content_Types].xml': tgt_ct_xml.encode('utf-8'),
    }
    if extra_modified:
        modified.update(extra_modified)
    modified.update(inject)

    tmp = target_path.with_suffix('.~tam.xlsx')
    with zipfile.ZipFile(target_path, 'r') as zin:
        with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
            existing = set()
            for item in zin.infolist():
                existing.add(item.filename)
                if item.filename == "xl/calcChain.xml":
                    continue  # removed — references stripped below
                if item.filename in modified:
                    zout.writestr(item, modified[item.filename])
                else:
                    zout.writestr(item, zin.read(item.filename))
            for fname, data in modified.items():
                if fname not in existing:
                    zout.writestr(fname, data)

    tmp.replace(target_path)


def _inject_tam_sheets(target_path: Path, base_path: Path = None,
                       merge_styles_flag: bool = True):
    """Copy TAM Solid and TAM Blood from base into target.

    merge_styles_flag=True:  For openpyxl-created files (new templates).
        Merges base styles + offsets FY DATA style indices.
    merge_styles_flag=False: For existing base-derived files (DCF TARA.xlsx etc.).
        Just injects TAM sheets; existing styles already match base.
    """
    resolved = _resolve_base_path(base_path)
    if resolved is None:
        logger.warning(f"Base file not found: {base_path or _BASE_FILE}; TAM sheets skipped")
        return

    result = _collect_tam_files(resolved)
    if result is None:
        return
    tam_sheets, inject = result

    if merge_styles_flag:
        # Full merge: offset openpyxl styles to avoid collision with base TAM styles
        with zipfile.ZipFile(resolved) as zf:
            base_styles = zf.read('xl/styles.xml')
        with zipfile.ZipFile(target_path) as zf:
            new_styles = zf.read('xl/styles.xml')
            tgt_sheets = {i.filename: zf.read(i.filename)
                          for i in zf.infolist()
                          if i.filename.startswith('xl/worksheets/sheet')
                          and not i.filename.endswith('.rels')}

        merged_styles, xf_offset = _merge_styles(base_styles, new_styles)
        logger.info(f"Styles merged: base={xf_offset} cellXfs, offset applied")

        offset_sheets = {f: _offset_sheet_styles(d, xf_offset)
                         for f, d in tgt_sheets.items()}
        offset_sheets['xl/styles.xml'] = merged_styles
        _register_and_write(target_path, tam_sheets, inject, offset_sheets)
    else:
        # Simple inject: file already has base-compatible styles
        _register_and_write(target_path, tam_sheets, inject)

    logger.info(f"TAM sheets injected from {resolved.name}")


def create_template(ticker, base_year=2020, he_years=None, fye_month=None,
                    output_path=None, cik=None, fetch_prices_flag=True,
                    reporting_unit="K", fydata_only=False, currency="USD"):
    global _UNIT_LABEL
    if he_years is None:
        he_years = [2022, 2023, 2024, 2025]
    if fye_month is None:
        logger.info(f"Detecting FYE month for {ticker}...")
        fye_month = detect_fye_month(ticker, cik=cik)
    fye_name = MONTH_NAMES[fye_month]
    logger.info(f"FYE month: {fye_name} ({fye_month})")
    logger.info(f"Reporting unit: {reporting_unit} {currency}")
    _UNIT_LABEL = f"[{reporting_unit} {currency}]"

    prices = {}
    if fetch_prices_flag and not fydata_only:
        prices = fetch_stock_prices(ticker, he_years)

    path = Path(output_path) if output_path else \
           Path(f'/mnt/c/Users/yzsun/Desktop/DD/{ticker}/DCF {ticker}.xlsx')
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    if reporting_unit == "K":
        # K USD mode: FY DATA (SUMIFS) + FY DATA K USD + Historical Events
        ws_mm = wb.active
        ws_mm.title = 'FY DATA'
        _build_mm_sheet(ws_mm, ticker, base_year, fye_name)

        ws_k = wb.create_sheet('FY DATA K USD')
        _build_kusd_sheet(ws_k, ticker, base_year, fye_name)

        if not fydata_only:
            ws_he = wb.create_sheet('Historical Events')
            _build_he_sheet(ws_he, ticker, he_years, prices)
    else:
        # MM USD mode: FY DATA (direct, blue data cells) + Historical Events
        ws_mm = wb.active
        ws_mm.title = 'FY DATA'
        _build_direct_mm_sheet(ws_mm, ticker, base_year, fye_name)

        if not fydata_only:
            ws_he = wb.create_sheet('Historical Events')
            _build_he_sheet(ws_he, ticker, he_years, prices)

    wb.save(str(path)); wb.close()
    logger.info(f"Template: {path}")

    # Inject TAM sheets from base file (with styles merge)
    if not fydata_only:
        _inject_tam_sheets(path)

    return path


def parse_args():
    p = argparse.ArgumentParser(description="Create blank DCF template.")
    p.add_argument('--ticker', default=None)
    p.add_argument('--base-year', type=int, default=2020)
    p.add_argument('--he-years', nargs='+', type=int, default=[2022,2023,2024,2025])
    p.add_argument('--fye-month', type=int, default=None)
    p.add_argument('--cik', default=None)
    p.add_argument('--path', default=None)
    p.add_argument('--no-prices', action='store_true')
    p.add_argument('--unit', choices=['K', 'MM'], default='K',
                   help='Reporting unit: K=thousands (default), MM=millions')
    p.add_argument('--currency', default='USD',
                   help='Currency code: USD (default), CHF, EUR, etc.')
    p.add_argument('--fydata-only', action='store_true',
                   help='Generate standalone file with FY DATA sheets only '
                        '(for inserting into existing workbook)')
    p.add_argument('--inject-tam', action='store_true',
                   help='Inject TAM sheets into an existing file (no FY DATA reset). '
                        'Requires --path (or --ticker to auto-find DCF file)')
    return p.parse_args()


if __name__ == '__main__':
    args = parse_args()

    if args.inject_tam:
        # ── Inject TAM sheets into an existing file ──────────────
        if args.path:
            target = Path(args.path)
        elif args.ticker:
            t = args.ticker.upper().strip()
            target = Path(f'/mnt/c/Users/yzsun/Desktop/DD/{t}/DCF {t}.xlsx')
        else:
            print("ERROR: --inject-tam requires --path or --ticker", flush=True)
            raise SystemExit(1)
        if not target.exists():
            print(f"ERROR: File not found: {target}", flush=True)
            raise SystemExit(1)

        # Auto-detect: if file has < base's cellXfs, it needs style merge
        with zipfile.ZipFile(target) as zf:
            sxml = zf.read('xl/styles.xml').decode('utf-8')
        xf_m = re.search(r'<cellXfs count="(\d+)"', sxml)
        existing_xf = int(xf_m.group(1)) if xf_m else 0
        need_merge = existing_xf < 700  # base has 1038; anything << that needs merge
        logger.info(f"Injecting TAM into: {target}  (cellXfs={existing_xf}, merge={'yes' if need_merge else 'no'})")
        _inject_tam_sheets(target, merge_styles_flag=need_merge)
        print(f"\n\u2714 TAM sheets injected into: {target}\n")
    else:
        # ── Normal template creation ─────────────────────────────
        if not args.ticker:
            print("ERROR: --ticker is required (unless using --inject-tam)", flush=True)
            raise SystemExit(1)
        t = args.ticker.upper().strip()
        path = create_template(t, base_year=args.base_year,
                               he_years=sorted(args.he_years),
                               fye_month=args.fye_month, output_path=args.path,
                               cik=args.cik, fetch_prices_flag=not args.no_prices,
                               reporting_unit=args.unit, fydata_only=args.fydata_only,
                               currency=args.currency)
        print(f"\n\u2714 Template created: {path}\n")
