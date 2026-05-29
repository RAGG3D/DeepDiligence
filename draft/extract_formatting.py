#!/usr/bin/env python3
"""Extract EVERY formatting detail from DCF CMPX.xlsx for cell-by-cell comparison."""

import openpyxl
from openpyxl.utils import get_column_letter
import json
import sys

FILEPATH = "/mnt/c/Users/yzsun/Desktop/DD/BHVN/DCF BHVN.xlsx"

def color_to_dict(color):
    """Convert an openpyxl Color object to a readable dict."""
    if color is None:
        return None
    result = {}
    if color.type == "theme":
        result["type"] = "theme"
        result["theme"] = color.theme
        result["tint"] = color.tint
    elif color.type == "rgb" or color.rgb:
        result["type"] = "rgb"
        result["rgb"] = str(color.rgb) if color.rgb else None
    elif color.type == "indexed":
        result["type"] = "indexed"
        result["indexed"] = color.indexed
    else:
        result["type"] = str(color.type)
        result["value"] = str(color.value) if hasattr(color, 'value') else None
    if hasattr(color, 'tint') and color.tint and color.tint != 0:
        result["tint"] = color.tint
    return result

def border_side_to_dict(side):
    """Convert a border Side to dict."""
    if side is None:
        return None
    return {
        "style": side.style,
        "color": color_to_dict(side.color) if side.color else None
    }

def extract_cell_format(cell):
    """Extract ALL formatting properties from a cell."""
    info = {}

    # Value
    val = cell.value
    if val is not None:
        info["value"] = str(val) if not isinstance(val, (int, float)) else val
        info["data_type"] = cell.data_type
    else:
        info["value"] = None
        info["data_type"] = cell.data_type

    # Font
    f = cell.font
    info["font"] = {
        "name": f.name,
        "size": f.size,
        "bold": f.bold,
        "italic": f.italic,
        "underline": f.underline,
        "strikethrough": f.strikethrough,
        "color": color_to_dict(f.color) if f.color else None,
    }

    # Fill
    fl = cell.fill
    info["fill"] = {
        "patternType": fl.patternType,
        "fgColor": color_to_dict(fl.fgColor) if fl.fgColor else None,
        "bgColor": color_to_dict(fl.bgColor) if fl.bgColor else None,
    }

    # Border
    b = cell.border
    info["border"] = {
        "left": border_side_to_dict(b.left),
        "right": border_side_to_dict(b.right),
        "top": border_side_to_dict(b.top),
        "bottom": border_side_to_dict(b.bottom),
        "diagonal": border_side_to_dict(b.diagonal),
        "diagonalUp": b.diagonalUp,
        "diagonalDown": b.diagonalDown,
    }

    # Alignment
    a = cell.alignment
    info["alignment"] = {
        "horizontal": a.horizontal,
        "vertical": a.vertical,
        "wrap_text": a.wrapText,
        "shrink_to_fit": a.shrinkToFit,
        "indent": a.indent,
        "text_rotation": a.textRotation,
    }

    # Number format
    info["number_format"] = cell.number_format

    # Protection
    p = cell.protection
    info["protection"] = {
        "locked": p.locked,
        "hidden": p.hidden,
    }

    return info


def extract_sheet_data(ws, row_range, col_range, sheet_name):
    """Extract all formatting from specified rows/cols of a sheet."""
    print(f"\n{'='*80}")
    print(f"SHEET: {sheet_name}")
    print(f"{'='*80}")

    # Sheet-level properties
    print(f"\n--- Sheet Properties ---")
    print(f"  Tab color: {ws.sheet_properties.tabColor}")
    print(f"  Default row height: {ws.sheet_format.defaultRowHeight}")
    print(f"  Default col width: {ws.sheet_format.defaultColWidth}")

    # Merged cells
    print(f"\n--- Merged Cells ---")
    for mc in sorted(ws.merged_cells.ranges, key=str):
        print(f"  {mc}")

    # Column widths
    print(f"\n--- Column Widths ---")
    for col_idx in col_range:
        col_letter = get_column_letter(col_idx)
        if col_letter in ws.column_dimensions:
            cd = ws.column_dimensions[col_letter]
            print(f"  Col {col_letter}: width={cd.width}, hidden={cd.hidden}, bestFit={cd.bestFit}, min={cd.min}, max={cd.max}, customWidth={cd.customWidth}")
        else:
            print(f"  Col {col_letter}: (default)")

    # Conditional formatting
    print(f"\n--- Conditional Formatting ---")
    for cf in ws.conditional_formatting:
        print(f"  Range: {cf}")
        for rule in cf.rules:
            print(f"    Rule: type={rule.type}, operator={rule.operator}, formula={rule.formula}")
            if rule.dxf:
                dxf = rule.dxf
                if dxf.font:
                    print(f"      DXF Font: bold={dxf.font.bold}, color={color_to_dict(dxf.font.color) if dxf.font.color else None}")
                if dxf.fill:
                    print(f"      DXF Fill: fgColor={color_to_dict(dxf.fill.fgColor) if dxf.fill.fgColor else None}")
                if dxf.border:
                    print(f"      DXF Border: present")

    # Row heights and hidden status
    print(f"\n--- Row Properties ---")
    for row_idx in row_range:
        rd = ws.row_dimensions.get(row_idx)
        if rd:
            print(f"  Row {row_idx}: height={rd.height}, hidden={rd.hidden}, customHeight={rd.customHeight}")
        else:
            print(f"  Row {row_idx}: (default)")

    # Cell-by-cell data
    print(f"\n--- Cell Data (row by row) ---")
    for row_idx in row_range:
        row_has_content = False
        row_data = {}
        for col_idx in col_range:
            cell = ws.cell(row=row_idx, column=col_idx)
            cell_info = extract_cell_format(cell)
            col_letter = get_column_letter(col_idx)
            row_data[col_letter] = cell_info
            if cell.value is not None:
                row_has_content = True

        # Print all rows (even empty ones for formatting)
        print(f"\n  ROW {row_idx}:")
        for col_idx in col_range:
            col_letter = get_column_letter(col_idx)
            ci = row_data[col_letter]
            # Always print if there's a value, non-default font, non-default fill, or non-default border
            has_value = ci["value"] is not None
            has_font = (ci["font"]["bold"] or ci["font"]["italic"] or
                       ci["font"]["name"] != "Calibri" or
                       (ci["font"]["size"] is not None and ci["font"]["size"] != 11))
            has_fill = ci["fill"]["patternType"] is not None and ci["fill"]["patternType"] != "none"
            has_border = any(
                ci["border"][side] and ci["border"][side]["style"] is not None
                for side in ["left", "right", "top", "bottom"]
            )
            has_alignment = (ci["alignment"]["horizontal"] is not None or
                           ci["alignment"]["vertical"] is not None or
                           ci["alignment"]["wrap_text"] or
                           ci["alignment"]["indent"] != 0)
            has_numfmt = ci["number_format"] != "General"

            if has_value or has_font or has_fill or has_border or has_alignment or has_numfmt:
                print(f"    {col_letter}{row_idx}:")
                print(f"      value={ci['value']}  dtype={ci['data_type']}")
                print(f"      font: name={ci['font']['name']}, sz={ci['font']['size']}, B={ci['font']['bold']}, I={ci['font']['italic']}, U={ci['font']['underline']}, color={ci['font']['color']}")
                print(f"      fill: pattern={ci['fill']['patternType']}, fg={ci['fill']['fgColor']}, bg={ci['fill']['bgColor']}")
                brd = ci["border"]
                print(f"      border: L={brd['left']}, R={brd['right']}, T={brd['top']}, B={brd['bottom']}")
                print(f"      align: h={ci['alignment']['horizontal']}, v={ci['alignment']['vertical']}, wrap={ci['alignment']['wrap_text']}, indent={ci['alignment']['indent']}, rot={ci['alignment']['text_rotation']}")
                print(f"      numfmt={ci['number_format']}")


def main():
    print("Loading workbook (data_only=False for formulas)...")
    wb = openpyxl.load_workbook(FILEPATH, data_only=False)

    print(f"Sheet names: {wb.sheetnames}")

    # ========== SHEET 1: FY DATA K USD ==========
    ws1_name = "FY DATA K USD"
    ws1 = wb[ws1_name]

    # Rows 1-130, Columns A-L (1-12)
    extract_sheet_data(
        ws1,
        range(1, 131),
        range(1, 13),  # A=1 through L=12
        ws1_name
    )

    # ========== SHEET 2: Historical Events ==========
    ws2_name = "Historical Events"
    ws2 = wb[ws2_name]

    # First find how many columns are used
    max_col = ws2.max_column
    max_row = ws2.max_row
    print(f"\n\nSheet 2 dimensions: max_row={max_row}, max_col={max_col}")

    # Rows 1-20 fully, plus sample rows
    sample_rows = list(range(1, 21))
    for r in [50, 100, 373]:
        if r <= max_row:
            sample_rows.append(r)

    col_end = min(max_col + 1, 25)  # up to X (col 24)

    extract_sheet_data(
        ws2,
        sample_rows,
        range(1, col_end),
        ws2_name
    )

    wb.close()
    print("\n\nDone.")


if __name__ == "__main__":
    main()
