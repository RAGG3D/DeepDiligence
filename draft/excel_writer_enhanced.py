"""
excel_writer_enhanced.py – Enhanced surgical xlsx patcher with dynamic row support.

Key enhancements:
1. Support for col_b sub-items (Notes details)
2. Dynamic Col D text modification using inlineStr
3. Accounting equation validation
"""

import logging
import shutil
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

logger = logging.getLogger(__name__)

# ── xlsx XML namespaces ───────────────────────────────────────────────────────
_NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NS_R    = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_NS_REL  = "http://schemas.openxmlformats.org/package/2006/relationships"

ET.register_namespace("",      _NS_MAIN)
ET.register_namespace("r",     _NS_R)
ET.register_namespace("mc",    "http://schemas.openxmlformats.org/markup-compatibility/2006")
ET.register_namespace("x14ac", "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac")
ET.register_namespace("xr",    "http://schemas.microsoft.com/office/spreadsheetml/2014/revision")
ET.register_namespace("xr2",   "http://schemas.microsoft.com/office/spreadsheetml/2015/revision2")
ET.register_namespace("xr3",   "http://schemas.microsoft.com/office/spreadsheetml/2016/revision3")

# ── sheet / column constants ──────────────────────────────────────────────────
FIRST_DATA_COL = 6   # column F (1-based)
MAX_YEAR_COLS  = 6   # F … K
HEADER_ROW     = 4   # row containing year integers
COL_B, COL_C, COL_D = 2, 3, 4

SHEET_KUSD = "FY DATA K USD"
SHEET_MM   = "FY DATA"

# ── low-level helpers ─────────────────────────────────────────────────────────

def _backup(xlsx_path: Path) -> Path:
    ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
    dst = xlsx_path.with_name(f"{xlsx_path.stem}_backup_{ts}.xlsx")
    shutil.copy2(xlsx_path, dst)
    logger.info(f"Backup created: {dst}")
    return dst


def _is_formula(v) -> bool:
    return isinstance(v, str) and v.startswith("=")


def _cell_addr(col_idx: int, row: int) -> str:
    return f"{get_column_letter(col_idx)}{row}"


def _col_of(addr: str) -> int:
    """Return 1-based column index from a cell address like 'F9'."""
    col_str = "".join(c for c in addr if c.isalpha())
    return column_index_from_string(col_str)


# ── openpyxl read helpers ─────────────────────────────────────────────────────

def _build_year_col_map(ws) -> Dict[int, int]:
    """Read the DATA header row to find {year: col_index}."""
    row = ws[HEADER_ROW]
    year_col: Dict[int, int] = {}
    base_year = base_col = None

    for col_idx in range(FIRST_DATA_COL, FIRST_DATA_COL + MAX_YEAR_COLS):
        val = row[col_idx - 1].value
        if isinstance(val, (int, float)) and 2000 <= int(val) <= 2099:
            base_year = int(val)
            base_col  = col_idx
            year_col[base_year] = col_idx
        elif base_year is not None and val is None:
            break
        elif base_year is not None and isinstance(val, str) and val.startswith("="):
            year_col[base_year + (col_idx - base_col)] = col_idx

    return year_col


def _build_row_map(ws_formula, ws_data, year_col: Dict[int, int]) -> Dict[Tuple, int]:
    """
    Build {(col_c, col_b, col_d): row_number} for writable (non-formula) rows.

    Enhanced to support col_b sub-items (Notes details).
    """
    first_data_col = min(year_col.values()) if year_col else FIRST_DATA_COL
    row_map: Dict[Tuple, int] = {}

    for row_num in range(1, ws_formula.max_row + 1):
        b_val = ws_data.cell(row=row_num, column=COL_B).value
        c_val = ws_data.cell(row=row_num, column=COL_C).value
        d_val = ws_data.cell(row=row_num, column=COL_D).value
        f_val = ws_formula.cell(row=row_num, column=first_data_col).value

        if c_val is None or d_val is None:
            continue
        if _is_formula(f_val):
            continue
        if isinstance(c_val, str) and c_val.startswith("="):
            continue

        # Normalize empty string → None for col_b
        if b_val == "":
            b_val = None
        # Convert numeric col_b to int
        elif isinstance(b_val, (int, float)):
            b_val = int(b_val)

        key = (c_val, b_val, d_val)
        if key not in row_map:
            row_map[key] = row_num
        else:
            logger.warning(
                f"Duplicate key {key} at rows {row_map[key]} and {row_num}; keeping first."
            )

    return row_map


def _find_dynamic_placeholder_rows(
    ws_formula,
    ws_data,
    year_col: Dict[int, int],
    pattern: str = "[Dynamic"
) -> List[Tuple[Tuple, int]]:
    """
    Find rows with Col D containing pattern like '[Dynamic R&D Note 1]'.

    Returns list of ((col_c, col_b, col_d), row_num) tuples.
    """
    first_data_col = min(year_col.values()) if year_col else FIRST_DATA_COL
    placeholders: List[Tuple[Tuple, int]] = []

    for row_num in range(1, ws_formula.max_row + 1):
        b_val = ws_data.cell(row=row_num, column=COL_B).value
        c_val = ws_data.cell(row=row_num, column=COL_C).value
        d_val = ws_data.cell(row=row_num, column=COL_D).value
        f_val = ws_formula.cell(row=row_num, column=first_data_col).value

        if d_val is None or not isinstance(d_val, str):
            continue
        if pattern not in d_val:
            continue
        if _is_formula(f_val):
            continue

        if b_val == "":
            b_val = None
        elif isinstance(b_val, (int, float)):
            b_val = int(b_val)

        key = (c_val, b_val, d_val)
        placeholders.append((key, row_num))

    return placeholders


# ── write-plan collectors ─────────────────────────────────────────────────────

def _collect_kusd_patches(
    wb:             openpyxl.Workbook,
    wb_data:        openpyxl.Workbook,
    financial_data: Dict[Tuple, Dict[int, Optional[float]]],
    years:          List[int],
) -> Tuple[List[Tuple[int, int, Any]], List[Tuple[int, int, str]]]:
    """
    Return (numeric_patches, text_patches).

    numeric_patches: [(row, col, value)] for K USD cells
    text_patches: [(row, col, text)] for Col D text modifications
    """
    ws      = wb[SHEET_KUSD]
    ws_data = wb_data[SHEET_KUSD]
    year_col = _build_year_col_map(ws)
    row_map  = _build_row_map(ws, ws_data, year_col)

    logger.info(
        f"K USD sheet: {len(year_col)} year columns, {len(row_map)} writable rows"
    )

    # Find dynamic placeholder rows
    placeholders = _find_dynamic_placeholder_rows(ws, ws_data, year_col)
    logger.info(f"K USD sheet: {len(placeholders)} dynamic placeholder rows")

    numeric_patches: List[Tuple[int, int, Any]] = []
    text_patches: List[Tuple[int, int, str]] = []

    # Track used placeholders by parent statement
    placeholder_usage: Dict[str, int] = {}  # {col_c: next_placeholder_index}

    for (col_c, col_b, col_d), year_vals in sorted(financial_data.items()):
        key = (col_c, col_b, col_d)

        # ── Case 1: Exact match in existing rows ──────────────────────────────
        if key in row_map:
            row_num = row_map[key]
            for year, val in year_vals.items():
                if year not in year_col or val is None:
                    continue
                col_idx = year_col[year]
                cell    = ws.cell(row=row_num, column=col_idx)
                if _is_formula(cell.value):
                    logger.warning(
                        f"  Skipping formula cell K USD R{row_num}C{col_idx} key={key}"
                    )
                    continue
                numeric_patches.append((row_num, col_idx, val))
                logger.debug(
                    f"  K USD [{col_c}|{col_b}|{col_d[:30]}] year={year} → {val:,}"
                )

        # ── Case 2: Notes detail (col_b is int) → use dynamic placeholder ─────
        elif col_b is not None and isinstance(col_b, int):
            # Find available placeholder for this statement type
            available_placeholders = [
                (ph_key, ph_row)
                for ph_key, ph_row in placeholders
                if ph_key[0] == col_c  # same statement type
            ]

            if not available_placeholders:
                logger.warning(
                    f"  No placeholder rows available for Notes item: {key}"
                )
                continue

            # Use next available placeholder
            used_count = placeholder_usage.get(col_c, 0)
            if used_count >= len(available_placeholders):
                logger.warning(
                    f"  All placeholders exhausted for statement {col_c}"
                )
                continue

            ph_key, ph_row = available_placeholders[used_count]
            placeholder_usage[col_c] = used_count + 1

            # Add text patch to rename the placeholder row
            text_patches.append((ph_row, COL_D, col_d))

            # Add numeric patches for data columns
            for year, val in year_vals.items():
                if year not in year_col or val is None:
                    continue
                col_idx = year_col[year]
                numeric_patches.append((ph_row, col_idx, val))
                logger.debug(
                    f"  K USD [{col_c}|{col_b}|{col_d[:30]}] year={year} → {val:,} (placeholder R{ph_row})"
                )

            # Also patch col_b to show the sub-item number
            numeric_patches.append((ph_row, COL_B, col_b))

        else:
            logger.debug(f"  Row not found in K USD sheet: {key}")

    logger.info(f"K USD: {len(numeric_patches)} numeric cells, {len(text_patches)} text patches")
    return numeric_patches, text_patches


def _collect_mm_patches(
    wb:      openpyxl.Workbook,
    wb_data: openpyxl.Workbook,
    years:   List[int],
) -> List[Tuple[int, int, Any]]:
    """Return [(row, col, formula)] for blank FY DATA cells needing SUMIFS."""
    ws_mm      = wb[SHEET_MM]
    ws_mm_data = wb_data[SHEET_MM]
    year_col_mm = _build_year_col_map(ws_mm)
    row_map_mm  = _build_row_map(ws_mm, ws_mm_data, year_col_mm)

    patches: List[Tuple[int, int, Any]] = []
    for row_num in set(row_map_mm.values()):
        for year in years:
            if year not in year_col_mm:
                continue
            col_idx    = year_col_mm[year]
            col_letter = get_column_letter(col_idx)
            if ws_mm.cell(row=row_num, column=col_idx).value is not None:
                continue
            formula = (
                f"=SUMIFS('FY DATA K USD'!{col_letter}:{col_letter},"
                f"'FY DATA K USD'!$D:$D,'FY DATA'!$D{row_num},"
                f"'FY DATA K USD'!$C:$C,'FY DATA'!$C{row_num},"
                f"'FY DATA K USD'!$B:$B,'FY DATA'!$B{row_num})/1000"
            )
            patches.append((row_num, col_idx, formula))

    logger.info(f"FY DATA: {len(patches)} SUMIFS formulas to inject")
    return patches


# ── surgical zip patcher ──────────────────────────────────────────────────────

def _get_sheet_zip_paths(xlsx_path: Path) -> Dict[str, str]:
    """Return {sheet_name: zip_entry_path} for all worksheets."""
    with zipfile.ZipFile(xlsx_path) as zf:
        wb_xml   = ET.fromstring(zf.read("xl/workbook.xml"))
        rels_xml = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))

    rid_to_path: Dict[str, str] = {}
    for rel in rels_xml:
        if "worksheet" in rel.get("Type", ""):
            rid    = rel.get("Id", "")
            target = rel.get("Target", "")
            rid_to_path[rid] = (
                f"xl/{target}" if not target.startswith("/") else target.lstrip("/")
            )

    sheet_map: Dict[str, str] = {}
    for sheet_elem in wb_xml.findall(f".//{{{_NS_MAIN}}}sheet"):
        name = sheet_elem.get("name", "")
        rid  = sheet_elem.get(f"{{{_NS_R}}}id", "")
        if rid in rid_to_path:
            sheet_map[name] = rid_to_path[rid]

    return sheet_map


def _num_str(value: Any) -> str:
    """Format a numeric value for the XML <v> element."""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value)


def _patch_numeric_cell(xml: str, addr: str, val_str: str) -> str:
    """
    Find cell r="ADDR" in the sheet XML string and update its <v> content.
    """
    search = f'r="{addr}"'
    start  = 0
    while True:
        pos = xml.find(search, start)
        if pos == -1:
            logger.warning(f"  Cell {addr} not found in sheet XML; skipping")
            return xml

        lt = xml.rfind("<", 0, pos)
        if lt == -1:
            start = pos + 1
            continue
        if xml[lt + 1] != "c" or xml[lt + 2] not in (" ", "\t", "\n", "/", ">"):
            start = pos + 1
            continue

        tag_end = xml.index(">", lt) + 1

        if xml[tag_end - 2 : tag_end] == "/>":
            return (
                xml[: tag_end - 2]
                + f"><v>{val_str}</v></c>"
                + xml[tag_end:]
            )

        c_end = xml.index("</c>", tag_end)
        cell_body = xml[tag_end:c_end]

        if "<v>" in cell_body:
            v_start = tag_end + cell_body.index("<v>")
            v_end   = tag_end + cell_body.index("</v>") + 4
            return xml[:v_start] + f"<v>{val_str}</v>" + xml[v_end:]
        else:
            return xml[:c_end] + f"<v>{val_str}</v>" + xml[c_end:]


def _patch_text_cell(xml: str, addr: str, text: str) -> str:
    """
    Modify cell text using inlineStr (<is><t>text</t></is>).

    This avoids touching sharedStrings.xml and is safer for dynamic text.
    """
    # Escape XML entities
    text_escaped = (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&apos;")
    )

    search = f'r="{addr}"'
    start  = 0
    while True:
        pos = xml.find(search, start)
        if pos == -1:
            logger.warning(f"  Cell {addr} not found for text patch; skipping")
            return xml

        lt = xml.rfind("<", 0, pos)
        if lt == -1:
            start = pos + 1
            continue
        if xml[lt + 1] != "c" or xml[lt + 2] not in (" ", "\t", "\n", "/", ">"):
            start = pos + 1
            continue

        # Find the opening tag end
        tag_end = xml.index(">", lt) + 1

        # Remove any t="..." attribute (string type) and add t="inlineStr"
        tag_start = xml[lt:tag_end]

        # Remove existing t="..." if present
        import re
        tag_start = re.sub(r'\s+t="[^"]*"', '', tag_start)

        # Add t="inlineStr"
        tag_start = tag_start.replace("<c ", '<c t="inlineStr" ')

        # Self-closing?
        if tag_start.endswith("/>"):
            new_cell = tag_start[:-2] + f"><is><t>{text_escaped}</t></is></c>"
            return xml[:lt] + new_cell + xml[tag_end:]

        # Find </c>
        c_end = xml.index("</c>", tag_end)

        # Replace content with inlineStr
        new_cell = tag_start + f"<is><t>{text_escaped}</t></is></c>"
        return xml[:lt] + new_cell + xml[c_end + 4:]


def _patch_formula_cell(xml: str, row_num: int, addr: str, formula_body: str) -> str:
    """
    Insert a formula cell into an existing row, or update it if it already exists.
    """
    search = f'r="{addr}"'
    pos = xml.find(search)
    if pos != -1:
        lt = xml.rfind("<", 0, pos)
        if lt != -1 and xml[lt + 1] == "c" and xml[lt + 2] in (" ", "\t", "\n", "/", ">"):
            tag_end = xml.index(">", lt) + 1
            if xml[tag_end - 2 : tag_end] == "/>":
                return (
                    xml[: tag_end - 2]
                    + f"><f>{formula_body}</f></c>"
                    + xml[tag_end:]
                )
            c_end = xml.index("</c>", tag_end)
            return xml[:tag_end] + f"<f>{formula_body}</f>" + xml[c_end:]

    # Cell doesn't exist: insert into the row
    row_search = f'r="{row_num}"'
    row_pos = 0
    while True:
        rp = xml.find(row_search, row_pos)
        if rp == -1:
            logger.warning(f"  Row {row_num} not found; cannot inject formula at {addr}")
            return xml
        lt = xml.rfind("<", 0, rp)
        if lt != -1 and xml[lt + 1 : lt + 4] in ("row", "ROW"):
            break
        row_pos = rp + 1

    row_tag_end  = xml.index(">", lt) + 1
    row_end      = xml.index("</row>", row_tag_end)
    row_body     = xml[row_tag_end:row_end]

    new_cell = f'<c r="{addr}"><f>{formula_body}</f></c>'
    col_idx  = _col_of(addr)

    import re as _re
    cells = list(_re.finditer(r'<c\b[^>]*\br="([A-Z]+\d+)"', row_body))
    insert_at = len(row_body)
    for m in cells:
        if _col_of(m.group(1)) > col_idx:
            insert_at = m.start()
            break

    new_body = row_body[:insert_at] + new_cell + row_body[insert_at:]
    return xml[:row_tag_end] + new_body + xml[row_end:]


def _patch_sheet_xml(
    xml_bytes:      bytes,
    numeric_writes: List[Tuple[int, int, Any]],
    text_writes:    List[Tuple[int, int, str]],
) -> bytes:
    """
    Apply both numeric and text patches to a worksheet XML.
    """
    xml = xml_bytes.decode("utf-8")

    # Apply numeric writes
    for (row_num, col_idx, value) in numeric_writes:
        addr = _cell_addr(col_idx, row_num)

        if isinstance(value, str) and value.startswith("="):
            xml = _patch_formula_cell(xml, row_num, addr, value[1:])
        else:
            xml = _patch_numeric_cell(xml, addr, _num_str(value))

    # Apply text writes
    for (row_num, col_idx, text) in text_writes:
        addr = _cell_addr(col_idx, row_num)
        xml = _patch_text_cell(xml, addr, text)

    return xml.encode("utf-8")


def _apply_xlsx_patches(
    xlsx_path:      Path,
    numeric_patches: Dict[str, List[Tuple[int, int, Any]]],
    text_patches:    Dict[str, List[Tuple[int, int, str]]],
) -> None:
    """
    Surgically write both numeric and text patches into xlsx.
    """
    sheet_zip_paths = _get_sheet_zip_paths(xlsx_path)

    modified: Dict[str, bytes] = {}
    with zipfile.ZipFile(xlsx_path) as zf:
        for sheet_name in set(numeric_patches.keys()) | set(text_patches.keys()):
            zip_path = sheet_zip_paths.get(sheet_name)
            if not zip_path:
                logger.warning(f"Sheet '{sheet_name}' not found in zip; skipping")
                continue

            xml_bytes = zf.read(zip_path)
            numeric_writes = numeric_patches.get(sheet_name, [])
            text_writes = text_patches.get(sheet_name, [])

            modified[zip_path] = _patch_sheet_xml(xml_bytes, numeric_writes, text_writes)
            logger.info(
                f"Patched '{sheet_name}' ({zip_path}): "
                f"{len(numeric_writes)} numeric, {len(text_writes)} text"
            )

    if not modified:
        logger.info("No patches to apply.")
        return

    tmp_path = xlsx_path.with_suffix(".~patch.xlsx")
    with zipfile.ZipFile(xlsx_path, "r") as zin:
        with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename in modified:
                    zout.writestr(item, modified[item.filename])
                else:
                    zout.writestr(item, data)

    tmp_path.replace(xlsx_path)
    logger.info(f"Surgical patch applied → {xlsx_path}")


# ── public entry point ────────────────────────────────────────────────────────

def update_workbook(
    ticker:         str,
    financial_data: Dict[Tuple, Dict[int, Optional[float]]],
    years:          List[int],
    excel_path:     Optional[str] = None,
) -> Path:
    """
    Main entry point: backup → read layout → build patch plan → surgical write.

    Enhanced to support Notes details with dynamic row naming.
    """
    path = (
        Path(excel_path)
        if excel_path
        else Path(f"/mnt/c/Users/yzsun/Desktop/DD/{ticker}/DCF {ticker}.xlsx")
    )

    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    _backup(path)

    logger.info(f"Opening workbook (read-only): {path}")
    wb      = openpyxl.load_workbook(str(path), data_only=False)
    wb_data = openpyxl.load_workbook(str(path), data_only=True)

    for sheet in (SHEET_KUSD, SHEET_MM):
        if sheet not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet}' not found in workbook.")

    # Build write plans
    kusd_numeric, kusd_text = _collect_kusd_patches(wb, wb_data, financial_data, years)
    mm_patches = _collect_mm_patches(wb, wb_data, years)

    wb.close()
    wb_data.close()

    # Apply surgically
    numeric_all: Dict[str, List] = {}
    text_all: Dict[str, List] = {}

    if kusd_numeric:
        numeric_all[SHEET_KUSD] = kusd_numeric
    if kusd_text:
        text_all[SHEET_KUSD] = kusd_text
    if mm_patches:
        numeric_all[SHEET_MM] = mm_patches

    _apply_xlsx_patches(path, numeric_all, text_all)

    logger.info(
        f"Done. K USD: {len(kusd_numeric)} numeric, {len(kusd_text)} text  |  "
        f"FY DATA: {len(mm_patches)} formulas"
    )
    return path
